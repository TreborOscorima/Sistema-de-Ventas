"""Mixin de exportación e importación masiva de inventario."""
import reflex as rx
import io
import logging
from typing import Any, List
from decimal import Decimal

from sqlmodel import select
from sqlalchemy.orm import selectinload
from app.utils.pricing import resolve_effective_price as _resolve_export_price

from app.models import (
    Product,
    ProductVariant,
    ProductBatch,
    Category,
    StockMovement,
)
from app.utils.exports import (
    create_excel_workbook,
    style_header_row,
    auto_adjust_column_widths,
    add_company_header,
    add_totals_row_with_formulas,
    add_notes_section,
    CURRENCY_FORMAT,
    PERCENT_FORMAT,
    THIN_BORDER,
    POSITIVE_FILL,
    NEGATIVE_FILL,
    WARNING_FILL,
)

logger = logging.getLogger(__name__)

# Plantilla de importación (encabezados + filas de ejemplo). Compartida por las
# descargas CSV y XLSX. Los encabezados coinciden con los alias de _parse_import_file.
_IMPORT_TEMPLATE_HEADERS = [
    "codigo", "descripcion", "categoria", "stock", "unidad", "precio compra", "precio venta",
]
_IMPORT_TEMPLATE_ROWS = [
    # 1ª fila: precio de venta explícito.
    ["7791234567890", "Coca Cola 500ml", "Bebidas", "50", "Unidad", "8.50", "12.00"],
    # 2ª fila: precio de venta VACÍO → se calcula con el margen global configurado.
    ["DUR-001", "Tornillo 3/8", "Ferreteria", "500", "Unidad", "0.10", ""],
]


class ExportMixin:
    """Exportación a Excel e importación masiva CSV/Excel de inventario."""

    @rx.event
    def export_inventory_to_excel(self):
        if not self.current_user["privileges"]["export_data"]:
            return rx.toast("No tiene permisos para exportar datos.", duration=3000)
        company_id = self._company_id()
        branch_id = self._branch_id()
        if not company_id or not branch_id:
            return rx.toast("Empresa no definida.", duration=3000)

        currency_label = self._currency_symbol_clean()
        currency_format = self._currency_excel_format()

        company_name = getattr(self, "company_name", "") or "EMPRESA"
        today = self._display_now().strftime("%d/%m/%Y")

        wb, ws = create_excel_workbook("Inventario Valorizado")

        # Encabezado profesional
        row = add_company_header(
            ws,
            company_name,
            "INVENTARIO VALORIZADO ACTUAL",
            f"Al {today}",
            columns=12,
            generated_at=self._display_now(),
        )

        headers = [
            "Código/SKU",
            "Descripción del Producto",
            "Categoría",
            "Stock Actual",
            "Unidad",
            f"Costo Unitario ({currency_label})",
            f"Precio Venta ({currency_label})",
            f"Margen Unitario ({currency_label})",
            "Margen (%)",
            f"Valor al Costo ({currency_label})",
            f"Valor a Venta ({currency_label})",
            "Estado Stock",
        ]

        with rx.session() as session:
            session.info["tenant_bypass"] = True
            products = session.exec(
                select(Product)
                .where(Product.company_id == company_id)
                .where(Product.branch_id == branch_id)
                .order_by(Product.description)
                .options(selectinload(Product.variants))
            ).all()

        def _variant_label(variant: ProductVariant) -> str:
            parts: list[str] = []
            if variant.size:
                parts.append(str(variant.size).strip())
            if variant.color:
                parts.append(str(variant.color).strip())
            return " ".join([p for p in parts if p]).strip()

        _exp_gm = float(getattr(self, "effective_profit_margin_decimal", 0.0) or 0.0)
        export_rows: list[dict[str, Any]] = []
        for product in products:
            variants = list(product.variants or [])
            if variants:
                for variant in variants:
                    label = _variant_label(variant)
                    description = product.description or "Sin descripción"
                    if label:
                        description = f"{description} ({label})"
                    export_rows.append(
                        {
                            "sku": variant.sku or product.barcode or "S/C",
                            "description": description,
                            "category": product.category or "Sin categoría",
                            "unit": product.unit or "Unid.",
                            "stock": variant.stock or 0,
                            "purchase_price": float(product.purchase_price or 0),
                            "sale_price": float(_resolve_export_price(product, variant, _exp_gm)),
                        }
                    )
            else:
                export_rows.append(
                    {
                        "sku": product.barcode or "S/C",
                        "description": product.description or "Sin descripción",
                        "category": product.category or "Sin categoría",
                        "unit": product.unit or "Unid.",
                        "stock": product.stock or 0,
                        "purchase_price": float(product.purchase_price or 0),
                        "sale_price": float(_resolve_export_price(product, global_margin=_exp_gm)),
                    }
                )

        total_items = len(export_rows)
        total_units = sum(float(item.get("stock", 0) or 0) for item in export_rows)
        total_cost_value = sum(
            float(item.get("stock", 0) or 0) * float(item.get("purchase_price", 0) or 0)
            for item in export_rows
        )
        total_sale_value = sum(
            float(item.get("stock", 0) or 0) * float(item.get("sale_price", 0) or 0)
            for item in export_rows
        )
        stock_zero = sum(1 for item in export_rows if float(item.get("stock", 0) or 0) == 0)
        stock_critical = sum(
            1 for item in export_rows if 0 < float(item.get("stock", 0) or 0) <= 5
        )
        stock_low = sum(
            1 for item in export_rows if 5 < float(item.get("stock", 0) or 0) <= 10
        )

        row += 1
        ws.cell(row=row, column=1, value="RESUMEN EJECUTIVO")
        row += 1
        ws.cell(row=row, column=1, value="Total SKUs:")
        ws.cell(row=row, column=2, value=total_items)
        row += 1
        ws.cell(row=row, column=1, value="Total unidades en stock:")
        ws.cell(row=row, column=2, value=total_units)
        row += 1
        ws.cell(row=row, column=1, value=f"Valor total al costo ({currency_label}):")
        ws.cell(row=row, column=2, value=total_cost_value).number_format = currency_format
        row += 1
        ws.cell(row=row, column=1, value=f"Valor total a venta ({currency_label}):")
        ws.cell(row=row, column=2, value=total_sale_value).number_format = currency_format
        row += 1
        ws.cell(row=row, column=1, value="Productos sin stock:")
        ws.cell(row=row, column=2, value=stock_zero)
        row += 1
        ws.cell(row=row, column=1, value="Productos críticos (1-5):")
        ws.cell(row=row, column=2, value=stock_critical)
        row += 1
        ws.cell(row=row, column=1, value="Productos bajos (6-10):")
        ws.cell(row=row, column=2, value=stock_low)
        row += 2

        style_header_row(ws, row, headers)
        data_start = row + 1
        row += 1

        for row_data in export_rows:
            barcode = row_data["sku"]
            description = row_data["description"]
            category = row_data["category"]
            unit = row_data["unit"]
            stock = row_data["stock"] or 0
            purchase_price = row_data["purchase_price"]
            sale_price = row_data["sale_price"]

            # Estado del stock
            if stock == 0:
                status = "SIN STOCK"
            elif stock <= 5:
                status = "CRÍTICO"
            elif stock <= 10:
                status = "BAJO"
            else:
                status = "NORMAL"

            ws.cell(row=row, column=1, value=barcode)
            ws.cell(row=row, column=2, value=description)
            ws.cell(row=row, column=3, value=category)
            ws.cell(row=row, column=4, value=stock)
            ws.cell(row=row, column=5, value=unit)
            ws.cell(row=row, column=6, value=purchase_price).number_format = currency_format
            ws.cell(row=row, column=7, value=sale_price).number_format = currency_format
            # Margen Unitario = Fórmula: Precio - Costo
            ws.cell(row=row, column=8, value=f"=G{row}-F{row}").number_format = currency_format
            # Margen % = (Margen / Costo) si Costo > 0; blanco si Costo no está definido
            ws.cell(row=row, column=9, value=f'=IF(F{row}>0,H{row}/F{row},"")').number_format = PERCENT_FORMAT
            # Valor al Costo = Fórmula: Stock × Costo
            ws.cell(row=row, column=10, value=f"=D{row}*F{row}").number_format = currency_format
            # Valor a Venta = Fórmula: Stock × Precio
            ws.cell(row=row, column=11, value=f"=D{row}*G{row}").number_format = currency_format
            ws.cell(row=row, column=12, value=status)

            # Color según estado
            status_cell = ws.cell(row=row, column=12)
            if "SIN STOCK" in status:
                status_cell.fill = NEGATIVE_FILL
            elif "CRÍTICO" in status:
                status_cell.fill = NEGATIVE_FILL
            elif "BAJO" in status:
                status_cell.fill = WARNING_FILL
            else:
                status_cell.fill = POSITIVE_FILL

            for col in range(1, 13):
                ws.cell(row=row, column=col).border = THIN_BORDER
            row += 1

        # Fila de totales
        totals_row = row
        add_totals_row_with_formulas(ws, totals_row, data_start, [
            {"type": "label", "value": "TOTALES"},
            {"type": "text", "value": ""},
            {"type": "text", "value": ""},
            {"type": "sum", "col_letter": "D"},
            {"type": "text", "value": ""},
            {"type": "text", "value": ""},
            {"type": "text", "value": ""},
            {"type": "text", "value": ""},
            {"type": "text", "value": ""},
            {"type": "sum", "col_letter": "J", "number_format": currency_format},
            {"type": "sum", "col_letter": "K", "number_format": currency_format},
            {"type": "text", "value": ""},
        ])

        # Notas explicativas
        add_notes_section(ws, totals_row, [
            "Costo Unitario: Precio al que se compró el producto al proveedor.",
            "Precio Venta: Precio de venta al público.",
            "Margen Unitario = Precio Venta - Costo Unitario (ganancia por unidad).",
            "Margen % = Margen Unitario ÷ Costo Unitario × 100. Productos sin costo definido muestran blanco.",
            "Valor al Costo: Inversión total = Stock × Costo Unitario.",
            "Valor a Venta: Potencial de ventas = Stock × Precio Venta.",
            "SIN STOCK: Producto agotado. CRÍTICO: ≤5 unidades. BAJO: ≤10 unidades.",
        ], columns=12)

        auto_adjust_column_widths(ws)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return rx.download(data=output.getvalue(), filename="inventario_valorizado.xlsx")

    # ══════════════════════════════════════════════════════════
    # IMPORTACIÓN MASIVA CSV / EXCEL
    # ══════════════════════════════════════════════════════════

    @rx.event
    def open_import_modal(self):
        if not self.current_user["privileges"].get("edit_inventario", False):
            return rx.toast("No tiene permisos para importar inventario.", duration=3000)
        self.import_modal_open = True
        self.import_preview_rows = []
        self.import_errors = []
        self.import_stats = {
            "new": 0, "updated": 0, "errors": 0, "total": 0,
            "stock_locked": 0, "margin_missing": 0,
        }
        self.import_processing = False
        self.import_file_name = ""

    @rx.event
    def close_import_modal(self):
        self.import_modal_open = False
        self.import_preview_rows = []
        self.import_errors = []
        self.import_file_name = ""

    @rx.event
    def download_import_template_csv(self):
        """Descarga la plantilla en CSV (encabezados esperados + filas de ejemplo).

        Los encabezados coinciden con los alias que acepta ``_parse_import_file``.
        Nota: Excel en locale español puede mostrar el CSV en una sola columna
        (usa ``;`` como separador de listas); el importador lo carga igual. Para
        editar cómodo en Excel, usar la plantilla Excel.
        """
        contenido = "\n".join(
            [",".join(_IMPORT_TEMPLATE_HEADERS)]
            + [",".join(fila) for fila in _IMPORT_TEMPLATE_ROWS]
        )
        # utf-8-sig (BOM) → Excel abre el CSV respetando los acentos.
        return rx.download(
            data=contenido.encode("utf-8-sig"),
            filename="plantilla_importar_productos.csv",
        )

    @rx.event
    def download_import_template_xlsx(self):
        """Descarga la plantilla en Excel (.xlsx) — se abre siempre en columnas."""
        import openpyxl
        from openpyxl.styles import Font

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Productos"
        ws.append(_IMPORT_TEMPLATE_HEADERS)
        for celda in ws[1]:
            celda.font = Font(bold=True)
        for fila in _IMPORT_TEMPLATE_ROWS:
            ws.append(fila)
        anchos = [16, 26, 16, 8, 10, 14, 12]
        for i, ancho in enumerate(anchos, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho
        buffer = io.BytesIO()
        wb.save(buffer)
        return rx.download(
            data=buffer.getvalue(),
            filename="plantilla_importar_productos.xlsx",
        )

    @rx.event
    async def handle_import_upload(self, files: list[rx.UploadFile]):
        """Procesa el archivo subido y genera preview."""
        if not files:
            return
        file = files[0]
        self.import_file_name = file.filename or "archivo"
        file_bytes = await file.read()

        try:
            rows = self._parse_import_file(file.filename or "", file_bytes)
        except Exception as e:
            self.import_errors = [f"Error al leer archivo: {str(e)}"]
            self.import_preview_rows = []
            return

        if not rows:
            self.import_errors = ["El archivo está vacío o no tiene filas de datos."]
            return

        company_id = self._company_id()
        branch_id = self._branch_id()
        if not company_id or not branch_id:
            self.import_errors = ["Empresa no configurada."]
            return

        # Cargar productos existentes para detectar nuevos vs actualizados y,
        # además, cuáles tienen su stock GESTIONADO por variantes o lotes
        # (product.stock es derivado; una importación no debe pisarlo).
        existing_barcodes: set[str] = set()
        managed_barcodes: set[str] = set()
        with rx.session() as session:
            session.info["tenant_bypass"] = True
            prod_rows = session.exec(
                select(Product.id, Product.barcode)
                .where(Product.company_id == company_id)
                .where(Product.branch_id == branch_id)
            ).all()
            existing_barcodes = {str(b) for _pid, b in prod_rows}
            id_by_barcode = {str(b): pid for pid, b in prod_rows}

            var_pids = set(session.exec(
                select(ProductVariant.product_id)
                .where(ProductVariant.company_id == company_id)
                .where(ProductVariant.branch_id == branch_id)
            ).all())
            bat_pids = {
                pid for pid in session.exec(
                    select(ProductBatch.product_id)
                    .where(ProductBatch.company_id == company_id)
                    .where(ProductBatch.branch_id == branch_id)
                ).all() if pid is not None
            }
            managed_pids = var_pids | bat_pids
            managed_barcodes = {
                bc for bc, pid in id_by_barcode.items() if pid in managed_pids
            }

        preview = []
        errors = []
        new_count = 0
        update_count = 0
        error_count = 0
        locked_count = 0
        margin_missing = 0  # P.Venta vacía pero sin margen global → quedaría en $0
        # Margen global vigente del tenant (sucursal → empresa). Si es 0, dejar la
        # columna P.Venta vacía daría precio $0 (se avisa en el preview).
        global_margin = float(getattr(self, "effective_profit_margin_decimal", 0.0) or 0.0)

        for idx, row in enumerate(rows, start=2):
            barcode = str(row.get("barcode", "") or "").strip()
            description = str(row.get("description", "") or "").strip()

            if not barcode:
                errors.append(f"Fila {idx}: código de barras vacío.")
                error_count += 1
                continue
            if not description:
                errors.append(f"Fila {idx} ({barcode}): descripción vacía.")
                error_count += 1
                continue

            # P.Venta VACÍA → precio dinámico por margen (sale_price = NULL al
            # importar). Con un número (incluido 0) → precio explícito.
            raw_sale = row.get("sale_price", None)
            price_from_margin = raw_sale is None or str(raw_sale).strip() == ""
            try:
                stock = float(row.get("stock", 0) or 0)
                purchase_price = float(row.get("purchase_price", 0) or 0)
                sale_price = 0.0 if price_from_margin else float(raw_sale)
            except (ValueError, TypeError):
                errors.append(f"Fila {idx} ({barcode}): valores numéricos inválidos.")
                error_count += 1
                continue

            is_new = barcode not in existing_barcodes
            if is_new:
                new_count += 1
            else:
                update_count += 1

            # Sólo un producto NUEVO sin precio y sin margen quedaría en $0. En un
            # producto EXISTENTE, precio vacío = preservar su precio actual (no $0).
            if is_new and price_from_margin and global_margin <= 0:
                margin_missing += 1

            # Producto existente cuyo stock lo gobiernan variantes/lotes:
            # se actualizan los demás campos pero NO el stock.
            stock_locked = (not is_new) and (barcode in managed_barcodes)
            if stock_locked:
                locked_count += 1

            preview.append({
                "row_num": idx,
                "barcode": barcode,
                "description": description,
                "category": str(row.get("category", "GENERAL") or "GENERAL").strip().upper(),
                "stock": stock,
                "unit": str(row.get("unit", "Unidad") or "Unidad").strip(),
                "purchase_price": purchase_price,
                "sale_price": sale_price,
                "price_from_margin": price_from_margin,
                "status": "Nuevo" if is_new else "Actualizar",
                "stock_locked": stock_locked,
            })

        self.import_preview_rows = preview[:200]  # Limitar preview a 200 filas
        self.import_errors = errors[:50]  # Limitar errores visibles
        self.import_stats = {
            "new": new_count,
            "updated": update_count,
            "errors": error_count,
            "total": len(rows),
            "stock_locked": locked_count,
            "margin_missing": margin_missing,
        }

    def _parse_import_file(self, filename: str, data: bytes) -> list[dict]:
        """Parsea CSV o Excel y retorna lista de dicts normalizados."""
        COLUMN_MAP = {
            "codigo": "barcode", "código": "barcode", "barcode": "barcode",
            "sku": "barcode", "cod": "barcode", "codigo/sku": "barcode",
            "código/sku": "barcode",
            "descripcion": "description", "descripción": "description",
            "description": "description", "producto": "description",
            "nombre": "description", "descripción del producto": "description",
            "categoria": "category", "categoría": "category",
            "category": "category",
            "stock": "stock", "stock actual": "stock", "cantidad": "stock",
            "unidad": "unit", "unit": "unit",
            "costo": "purchase_price", "costo unitario": "purchase_price",
            "purchase_price": "purchase_price", "precio compra": "purchase_price",
            "precio": "sale_price", "precio venta": "sale_price",
            "sale_price": "sale_price", "precio de venta": "sale_price",
            "pvp": "sale_price",
        }

        ext = (filename.rsplit(".", 1)[-1] if "." in filename else "").lower()

        if ext in ("xlsx", "xls"):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            header_raw = next(rows_iter, None)
            if not header_raw:
                return []
            headers = []
            for h in header_raw:
                h_clean = str(h or "").strip().lower()
                headers.append(COLUMN_MAP.get(h_clean, h_clean))
            result = []
            for row_values in rows_iter:
                if all(v is None for v in row_values):
                    continue
                row_dict = {}
                for i, val in enumerate(row_values):
                    if i < len(headers):
                        row_dict[headers[i]] = val
                result.append(row_dict)
            wb.close()
            return result
        else:
            # CSV
            import csv
            text = data.decode("utf-8-sig")
            # Detectar delimitador
            sniffer = csv.Sniffer()
            try:
                dialect = sniffer.sniff(text[:2048])
            except csv.Error:
                dialect = csv.excel
            reader = csv.DictReader(text.splitlines(), dialect=dialect)
            result = []
            for row in reader:
                normalized = {}
                for key, val in row.items():
                    k_clean = str(key or "").strip().lower()
                    mapped = COLUMN_MAP.get(k_clean, k_clean)
                    normalized[mapped] = val
                result.append(normalized)
            return result

    @rx.event
    def confirm_import(self):
        """Ejecuta la importación confirmada a la base de datos."""
        if not self.current_user["privileges"].get("edit_inventario", False):
            return rx.toast("No tiene permisos.", duration=3000)
        if not self.import_preview_rows:
            return rx.toast("No hay datos para importar.", duration=3000)

        company_id = self._company_id()
        branch_id = self._branch_id()
        user_id = self.current_user.get("id")
        if not company_id or not branch_id:
            return rx.toast("Empresa no configurada.", duration=3000)

        self.import_processing = True
        yield

        imported = 0
        updated = 0
        stock_skipped = 0
        errors = []

        with rx.session() as session:
            session.info["tenant_bypass"] = True
            # Pre-cargar productos existentes por barcode
            existing = session.exec(
                select(Product)
                .where(Product.company_id == company_id)
                .where(Product.branch_id == branch_id)
            ).all()
            products_by_barcode: dict[str, Product] = {
                p.barcode: p for p in existing
            }

            # Productos cuyo stock es DERIVADO (tienen variantes o lotes):
            # se actualizan sus datos pero su stock no se pisa desde el archivo.
            var_pids = set(session.exec(
                select(ProductVariant.product_id)
                .where(ProductVariant.company_id == company_id)
                .where(ProductVariant.branch_id == branch_id)
            ).all())
            bat_pids = {
                pid for pid in session.exec(
                    select(ProductBatch.product_id)
                    .where(ProductBatch.company_id == company_id)
                    .where(ProductBatch.branch_id == branch_id)
                ).all() if pid is not None
            }
            managed_pids = var_pids | bat_pids

            # Pre-cargar categorías
            cats = session.exec(
                select(Category)
                .where(Category.company_id == company_id)
                .where(Category.branch_id == branch_id)
            ).all()
            existing_categories: dict[str, Category] = {
                c.name: c for c in cats
            }

            try:
                for row in self.import_preview_rows:
                    barcode = row["barcode"]
                    description = row["description"]
                    category_name = row.get("category", "General") or "General"
                    stock = Decimal(str(row.get("stock", 0) or 0))
                    unit = row.get("unit", "Unidad") or "Unidad"
                    purchase_price = Decimal(str(row.get("purchase_price", 0) or 0))
                    # P.Venta vacía en el archivo → NULL → precio dinámico por margen
                    # (la cascada de precios usa precio_compra × (1 + margen global)).
                    if row.get("price_from_margin"):
                        sale_price = None
                    else:
                        sale_price = Decimal(str(row.get("sale_price", 0) or 0))

                    # Auto-crear categoría si no existe
                    if category_name not in existing_categories:
                        new_cat = Category(
                            name=category_name,
                            company_id=company_id,
                            branch_id=branch_id,
                        )
                        session.add(new_cat)
                        session.flush()
                        existing_categories[category_name] = new_cat

                    product = products_by_barcode.get(barcode)
                    if product:
                        # Actualizar datos comunes (sin tocar stock todavía).
                        product.description = description
                        product.category = category_name
                        product.unit = unit
                        product.purchase_price = purchase_price
                        # Precio: sólo se pisa si el archivo trae un precio EXPLÍCITO.
                        # Si P.Venta vino vacía en un producto EXISTENTE, NO se toca su
                        # precio actual → preserva el custom/margen que ya tenía.
                        if not row.get("price_from_margin"):
                            product.sale_price = sale_price

                        if product.id in managed_pids:
                            # Stock gobernado por variantes/lotes → NO se pisa,
                            # para no romper el invariante stock == SUM(lotes/variantes).
                            stock_skipped += 1
                        else:
                            stock_delta = stock - Decimal(str(product.stock or 0))
                            product.stock = stock
                            if stock_delta != 0:
                                session.add(StockMovement(
                                    product_id=product.id,
                                    user_id=user_id,
                                    type="Importacion",
                                    quantity=stock_delta,
                                    description=f"Importación masiva (ajuste): {description}",
                                    timestamp=self._event_timestamp(),
                                    company_id=company_id,
                                    branch_id=branch_id,
                                ))
                        session.add(product)
                        updated += 1
                    else:
                        # Crear
                        product = Product(
                            barcode=barcode,
                            description=description,
                            category=category_name,
                            stock=stock,
                            unit=unit,
                            purchase_price=purchase_price,
                            sale_price=sale_price,
                            company_id=company_id,
                            branch_id=branch_id,
                        )
                        session.add(product)
                        session.flush()
                        products_by_barcode[barcode] = product

                        # Registrar movimiento de stock
                        if stock > 0:
                            movement = StockMovement(
                                product_id=product.id,
                                user_id=user_id,
                                type="Importacion",
                                quantity=stock,
                                description=f"Importación masiva: {description}",
                                timestamp=self._event_timestamp(),
                                company_id=company_id,
                                branch_id=branch_id,
                            )
                            session.add(movement)
                        imported += 1

                session.commit()
            except Exception as e:
                session.rollback()
                logger.exception("Error en importación masiva")
                self.import_processing = False
                errors.append(f"Error de base de datos: {str(e)}")
                self.import_errors = errors
                return rx.toast(
                    "Error al importar. Verifique los datos e intente nuevamente.",
                    duration=5000,
                )

        self.import_processing = False
        self.close_import_modal()
        self._inventory_update_trigger += 1
        self.load_categories()
        msg = f"Importación exitosa: {imported} nuevos, {updated} actualizados."
        if stock_skipped:
            msg += (
                f" ({stock_skipped} con stock gestionado por lotes/variantes: "
                "no se modificó su stock)."
            )
        return rx.toast(msg, duration=5000)

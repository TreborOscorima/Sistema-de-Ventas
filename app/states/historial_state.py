import logging
import reflex as rx
from typing import Dict, Any
import datetime
import io
import unicodedata
from decimal import Decimal
from sqlmodel import select
import sqlalchemy as sa
from sqlalchemy.orm import selectinload
from app.enums import PaymentMethodType, ReceiptType, FiscalStatus, ReturnReason, SaleStatus
from app.i18n import MSG
from app.utils.payment import payment_method_label
from app.utils.sanitization import escape_like
from app.models import (
    Sale,
    SaleItem,
    SaleReturn,
    SaleReturnItem,
    Category,
    Product,
    CashboxLog,
    PaymentMethod,
    SalePayment,
    User,
    Company,
    CompanyBillingConfig,
    FiscalDocument,
)
from .mixin_state import MixinState
from app.utils.exports import (
    create_excel_workbook,
    style_header_row,
    add_data_rows,
    auto_adjust_column_widths,
    add_company_header,
    add_totals_row_with_formulas,
    add_notes_section,
    CURRENCY_FORMAT,
    NUMBER_FORMAT,
    THIN_BORDER,
    POSITIVE_FILL,
    NEGATIVE_FILL,
    WARNING_FILL,
)
from app.utils.tenant import tenant_bypass

REPORT_METHOD_KEYS = [
    "cash",
    "debit",
    "credit",
    "yape",
    "plin",
    "transfer",
    "mixed",
    "other",
]
REPORT_SOURCE_OPTIONS = [
    ["Todos", "Todos"],
    ["Ventas", "Ventas"],
    ["Cobranzas", "Cobranzas"],
]
REPORT_CASHBOX_ACTIONS = {
    "Cobranza",
    "Cobro de Cuota",
    "Pago Cuota",
    "Cobro Cuota",
    "Ingreso Cuota",
    "Amortizacion",
    "Pago Credito",
}

logger = logging.getLogger(__name__)


class HistorialState(MixinState):
    """Estado del historial de ventas y reportes financieros.

    Gestiona la vista de historial con filtros por fecha, método de pago,
    categoría y vendedor. Genera reportes exportables a Excel.
    """

    # history: List[Movement] = [] # Eliminado a favor de la BD
    history_filter_type: str = "Todos"
    history_filter_product: str = ""
    history_filter_category: str = "Todas"
    history_filter_start_date: str = ""
    history_filter_end_date: str = ""
    staged_history_filter_type: str = "Todos"
    staged_history_filter_product: str = ""
    staged_history_filter_category: str = "Todas"
    staged_history_filter_start_date: str = ""
    staged_history_filter_end_date: str = ""
    available_category_options: list[list[str]] = [["Todas", "Todas"]]
    report_active_tab: str = "metodos"
    report_filter_start_date: str = ""
    report_filter_end_date: str = ""
    report_filter_method: str = "Todos"
    report_filter_source: str = "Todos"
    report_filter_user: str = "Todos"
    staged_report_filter_start_date: str = ""
    staged_report_filter_end_date: str = ""
    staged_report_filter_method: str = "Todos"
    staged_report_filter_source: str = "Todos"
    staged_report_filter_user: str = "Todos"
    available_report_method_options: list[list[str]] = [["Todos", "Todos"]]
    available_report_source_options: list[list[str]] = REPORT_SOURCE_OPTIONS
    available_report_user_options: list[list[str]] = [["Todos", "Todos"]]
    _report_update_trigger: int = 0
    report_detail_current_page: int = 1
    report_detail_items_per_page: int = 10
    report_closing_current_page: int = 1
    report_closing_items_per_page: int = 10
    current_page_history: int = 1
    items_per_page: int = 10
    _history_update_trigger: int = 0
    sale_detail_modal_open: bool = False
    selected_sale_id: str = ""
    selected_sale_summary: dict = {}
    selected_sale_items: list[dict] = []
    filtered_history: list[dict] = []
    total_pages: int = 1
    report_method_summary: list[dict] = []
    report_detail_rows: list[dict] = []
    report_closing_rows: list[dict] = []
    payment_stats: Dict[str, float] = {
        "efectivo": 0.0,
        "debito": 0.0,
        "credito": 0.0,
        "yape": 0.0,
        "plin": 0.0,
        "transferencia": 0.0,
        "mixto": 0.0,
    }
    total_credit: float = 0.0
    credit_outstanding: float = 0.0
    dynamic_payment_cards: list[dict] = []
    productos_mas_vendidos: list[dict] = []
    productos_stock_bajo: list[Dict] = []
    sales_by_day: list[dict] = []

    # ── Devoluciones ──
    return_modal_open: bool = False
    return_sale_id: int = 0
    return_sale_summary: dict = {}
    return_items: list[dict] = []
    return_reason: str = "other"
    return_notes: str = ""

    # ── Reporte de devoluciones ──
    returns_report_rows: list[dict] = []
    returns_report_total: float = 0.0
    returns_report_count: int = 0
    returns_report_page: int = 1
    returns_report_per_page: int = 10
    returns_report_filter_start: str = ""
    returns_report_filter_end: str = ""

    def _history_date_range(self) -> tuple[datetime.datetime | None, datetime.datetime | None]:
        start_date = None
        end_date = None
        if self.history_filter_start_date:
            try:
                start_date, _ = self._company_day_bounds_utc_naive(
                    self.history_filter_start_date
                )
            except ValueError:
                start_date = None
        if self.history_filter_end_date:
            try:
                _, end_date = self._company_day_bounds_utc_naive(
                    self.history_filter_end_date
                )
            except ValueError:
                end_date = None
        return start_date, end_date

    def _load_category_options(self) -> None:
        categories: set[str] = set()
        company_id = self._company_id()
        branch_id = self._branch_id()
        if not company_id or not branch_id:
            self.available_category_options = [["Todas", "Todas"]]
            return
        with rx.session() as session:
            for name in session.exec(
                select(Category.name)
                .where(Category.company_id == company_id)
                .where(Category.branch_id == branch_id)
            ).all():
                if name:
                    categories.add(str(name).strip())
            for name in session.exec(
                select(Product.category)
                .where(Product.company_id == company_id)
                .where(Product.branch_id == branch_id)
            ).all():
                if name:
                    categories.add(str(name).strip())
            for name in session.exec(
                select(SaleItem.product_category_snapshot)
                .join(Sale, SaleItem.sale_id == Sale.id)
                .where(Sale.company_id == company_id)
                .where(Sale.branch_id == branch_id)
            ).all():
                if name:
                    categories.add(str(name).strip())

        categories = {name for name in categories if name}
        if not categories:
            categories = {"General"}

        options = [["Todas", "Todas"]]
        options.extend([[name, name] for name in sorted(categories)])
        self.available_category_options = options

        valid_values = {option[1] for option in options}
        if self.history_filter_category not in valid_values:
            self.history_filter_category = "Todas"
        if self.staged_history_filter_category not in valid_values:
            self.staged_history_filter_category = "Todas"

    def _report_date_range(
        self,
    ) -> tuple[datetime.datetime | None, datetime.datetime | None]:
        start_date = None
        end_date = None
        if self.report_filter_start_date:
            try:
                start_date, _ = self._company_day_bounds_utc_naive(
                    self.report_filter_start_date
                )
            except ValueError:
                start_date = None
        if self.report_filter_end_date:
            try:
                _, end_date = self._company_day_bounds_utc_naive(
                    self.report_filter_end_date
                )
            except ValueError:
                end_date = None
        return start_date, end_date

    def _method_key_from_label(self, label: str) -> str:
        raw = (label or "").strip().lower()
        if not raw:
            return "other"
        normalized = unicodedata.normalize("NFKD", raw)
        normalized = "".join(
            ch for ch in normalized if not unicodedata.combining(ch)
        )
        if "mixto" in normalized or "mixed" in normalized:
            return "mixed"
        if "yape" in normalized:
            return "yape"
        if "plin" in normalized:
            return "plin"
        if "transfer" in normalized or "banco" in normalized:
            return "transfer"
        if "debito" in normalized or "debit" in normalized:
            return "debit"
        if "credito" in normalized or "credit" in normalized or "tarjeta" in normalized:
            return "credit"
        if "efectivo" in normalized or normalized == "cash":
            return "cash"
        return "other"

    def _load_report_options(self) -> None:
        method_options = [["Todos", "Todos"]]
        for key in REPORT_METHOD_KEYS:
            method_options.append([self._payment_method_label(key), key])
        self.available_report_method_options = method_options
        self.available_report_source_options = [option[:] for option in REPORT_SOURCE_OPTIONS]

        company_id = self._company_id()
        if not company_id:
            self.available_report_user_options = [["Todos", "Todos"]]
            return
        with rx.session() as session:
            usernames = session.exec(
                select(User.username).where(User.company_id == company_id)
            ).all()

        user_values = sorted(
            {
                str(name).strip()
                for name in usernames
                if name and str(name).strip()
            }
        )
        user_options = [["Todos", "Todos"]]
        user_options.extend([[name, name] for name in user_values])
        self.available_report_user_options = user_options

        method_values = {option[1] for option in self.available_report_method_options}
        if self.report_filter_method not in method_values:
            self.report_filter_method = "Todos"
        if self.staged_report_filter_method not in method_values:
            self.staged_report_filter_method = "Todos"

        source_values = {option[1] for option in self.available_report_source_options}
        if self.report_filter_source not in source_values:
            self.report_filter_source = "Todos"
        if self.staged_report_filter_source not in source_values:
            self.staged_report_filter_source = "Todos"

        user_values_set = {option[1] for option in self.available_report_user_options}
        if self.report_filter_user not in user_values_set:
            self.report_filter_user = "Todos"
        if self.staged_report_filter_user not in user_values_set:
            self.staged_report_filter_user = "Todos"

    def _build_sale_user_lookup(
        self, session, sales: list[Sale]
    ) -> dict[int, str]:
        """Obtiene usernames por user_id sin restringir por branch_id."""
        company_id = self._company_id()
        if not company_id:
            return {}
        user_ids = {
            int(sale.user_id)
            for sale in sales
            if sale is not None and getattr(sale, "user_id", None) is not None
        }
        if not user_ids:
            return {}

        query = (
            select(User.id, User.username)
            .where(User.id.in_(user_ids))
            .where(User.company_id == company_id)
        )
        with tenant_bypass():
            rows = session.exec(query.execution_options(tenant_bypass=True)).all()

        lookup: dict[int, str] = {}
        for row in rows:
            try:
                user_id, username = row[0], row[1]
            except Exception:
                user_id = getattr(row, "id", None)
                username = getattr(row, "username", None)
            try:
                user_id_int = int(user_id)
            except (TypeError, ValueError):
                continue
            user_name = str(username or "").strip()
            if user_name:
                lookup[user_id_int] = user_name
        return lookup

    def _sale_username(
        self,
        sale: Sale | None,
        user_lookup: dict[int, str] | None = None,
        default: str = MSG.FALLBACK_UNKNOWN,
    ) -> str:
        if sale and sale.user and getattr(sale.user, "username", None):
            user_name = str(sale.user.username).strip()
            if user_name:
                return user_name
        if sale and user_lookup:
            try:
                user_id = int(getattr(sale, "user_id", None) or 0)
            except (TypeError, ValueError):
                user_id = 0
            if user_id and user_id in user_lookup:
                return user_lookup[user_id]
        return default

    def _build_report_entries(self) -> list[dict]:
        start_date, end_date = self._report_date_range()
        method_filter = self.report_filter_method or "Todos"
        source_filter = self.report_filter_source or "Todos"
        user_filter = self.report_filter_user or "Todos"

        rows: list[dict] = []
        company_id = self._company_id()
        branch_id = self._branch_id()
        if not company_id or not branch_id:
            return rows
        with rx.session() as session:
            if source_filter in {"Todos", "Ventas"}:
                payment_query = (
                    select(SalePayment)
                    .join(Sale, SalePayment.sale_id == Sale.id)
                    .where(Sale.company_id == company_id)
                    .where(Sale.branch_id == branch_id)
                    .options(selectinload(SalePayment.sale).selectinload(Sale.user))
                )
                if start_date:
                    payment_query = payment_query.where(
                        SalePayment.created_at >= start_date
                    )
                if end_date:
                    payment_query = payment_query.where(
                        SalePayment.created_at <= end_date
                    )
                payments = session.exec(payment_query).all()
                sales_for_lookup = [
                    payment.sale
                    for payment in payments
                    if payment is not None and getattr(payment, "sale", None) is not None
                ]
                sale_user_lookup = self._build_sale_user_lookup(
                    session, sales_for_lookup
                )

                for payment in payments:
                    sale = payment.sale
                    if sale and sale.status == SaleStatus.cancelled:
                        continue
                    user_name = self._sale_username(
                        sale, sale_user_lookup, MSG.FALLBACK_UNKNOWN
                    )
                    if user_filter != "Todos" and user_name != user_filter:
                        continue
                    method_key = self._payment_method_key(
                        getattr(payment, "method_type", None)
                    )
                    if not method_key:
                        method_key = "other"
                    if method_filter != "Todos" and method_key != method_filter:
                        continue
                    amount = Decimal(str(getattr(payment, "amount", 0) or 0))
                    timestamp = (
                        getattr(payment, "created_at", None)
                        or (sale.timestamp if sale else None)
                    )
                    reference_code = getattr(payment, "reference_code", "") or ""
                    sale_id = sale.id if sale else None
                    reference = (
                        reference_code
                        or (f"Venta #{sale_id}" if sale_id else "-")
                    )
                    rows.append(
                        {
                            "timestamp": timestamp,
                            "timestamp_display": self._format_company_datetime(
                                timestamp
                            )
                            if timestamp
                            else "",
                            "source": "Venta",
                            "method_key": method_key,
                            "method_label": self._normalize_wallet_label(method_key),
                            "amount": self._round_currency(amount),
                            "user": user_name,
                            "reference": reference,
                        }
                    )

            if source_filter in {"Todos", "Cobranzas"}:
                log_query = (
                    select(CashboxLog, User.username)
                    .join(User, User.id == CashboxLog.user_id, isouter=True)
                    .where(CashboxLog.company_id == company_id)
                    .where(CashboxLog.branch_id == branch_id)
                    .where(CashboxLog.action.in_(REPORT_CASHBOX_ACTIONS))
                    .where(CashboxLog.is_voided == False)
                )
                if start_date:
                    log_query = log_query.where(
                        CashboxLog.timestamp >= start_date
                    )
                if end_date:
                    log_query = log_query.where(CashboxLog.timestamp <= end_date)
                logs = session.exec(log_query).all()

                for log, username in logs:
                    user_name = username or MSG.FALLBACK_UNKNOWN
                    if user_filter != "Todos" and user_name != user_filter:
                        continue
                    method_key = self._method_key_from_label(
                        getattr(log, "payment_method", "")
                    )
                    if method_filter != "Todos" and method_key != method_filter:
                        continue
                    amount = Decimal(str(getattr(log, "amount", 0) or 0))
                    timestamp = getattr(log, "timestamp", None)
                    rows.append(
                        {
                            "timestamp": timestamp,
                            "timestamp_display": self._format_company_datetime(
                                timestamp
                            )
                            if timestamp
                            else "",
                            "source": "Cobranza",
                            "method_key": method_key,
                            "method_label": self._normalize_wallet_label(
                                getattr(log, "payment_method", "") or method_key
                            ),
                            "amount": self._round_currency(amount),
                            "user": user_name,
                            "reference": (log.notes or "").strip()
                            or "Cobranza registrada",
                        }
                    )

        rows.sort(
            key=lambda item: item.get("timestamp") or datetime.datetime.min,
            reverse=True,
        )
        return rows

    def _build_report_closings(self) -> list[dict]:
        start_date, end_date = self._report_date_range()
        user_filter = self.report_filter_user or "Todos"

        rows: list[dict] = []
        company_id = self._company_id()
        branch_id = self._branch_id()
        if not company_id or not branch_id:
            return rows
        with rx.session() as session:
            log_query = (
                select(CashboxLog, User.username)
                .join(User, User.id == CashboxLog.user_id, isouter=True)
                .where(CashboxLog.company_id == company_id)
                .where(CashboxLog.branch_id == branch_id)
                .where(CashboxLog.action.in_(["apertura", "cierre"]))
                .where(CashboxLog.is_voided == False)
            )
            if start_date:
                log_query = log_query.where(
                    CashboxLog.timestamp >= start_date
                )
            if end_date:
                log_query = log_query.where(CashboxLog.timestamp <= end_date)
            logs = session.exec(log_query).all()

            for log, username in logs:
                user_name = username or MSG.FALLBACK_UNKNOWN
                if user_filter != "Todos" and user_name != user_filter:
                    continue
                action_label = (
                    "Apertura" if log.action == "apertura" else "Cierre"
                )
                timestamp = getattr(log, "timestamp", None)
                amount = Decimal(str(getattr(log, "amount", 0) or 0))
                rows.append(
                    {
                        "timestamp": timestamp,
                        "timestamp_display": self._format_company_datetime(
                            timestamp
                        )
                        if timestamp
                        else "",
                        "action": action_label,
                        "amount": self._round_currency(amount),
                        "user": user_name,
                        "notes": (log.notes or "").strip(),
                    }
                )

        rows.sort(
            key=lambda item: item.get("timestamp") or datetime.datetime.min,
            reverse=True,
        )
        return rows

    def _apply_sales_filters(self, query):
        start_date, end_date = self._history_date_range()
        company_id = self._company_id()
        branch_id = self._branch_id()
        if not company_id or not branch_id:
            return query.where(sa.false())
        query = query.where(Sale.company_id == company_id)
        query = query.where(Sale.branch_id == branch_id)
        if start_date:
            query = query.where(Sale.timestamp >= start_date)
        if end_date:
            query = query.where(Sale.timestamp <= end_date)

        if self.history_filter_type not in {"Todos", "Venta"}:
            query = query.where(sa.false())

        search = (self.history_filter_product or "").strip()
        if search:
            like_search = f"%{escape_like(search)}%"
            sale_ids = (
                select(SaleItem.sale_id)
                .join(Sale, SaleItem.sale_id == Sale.id)
                .where(SaleItem.product_name_snapshot.ilike(like_search))
                .where(Sale.company_id == company_id)
                .where(Sale.branch_id == branch_id)
                .distinct()
            )
            query = query.where(Sale.id.in_(sale_ids))

        category = (self.history_filter_category or "").strip()
        if category and category != "Todas":
            sale_ids = (
                select(SaleItem.sale_id)
                .join(Sale, SaleItem.sale_id == Sale.id)
                .where(SaleItem.product_category_snapshot == category)
                .where(Sale.company_id == company_id)
                .where(Sale.branch_id == branch_id)
                .distinct()
            )
            query = query.where(Sale.id.in_(sale_ids))
        return query

    def _sales_query(self):
        company_id = self._company_id()
        branch_id = self._branch_id()
        if not company_id or not branch_id:
            return select(Sale).where(sa.false())
        query = (
            select(Sale)
            .where(Sale.status != SaleStatus.cancelled)
            .where(Sale.company_id == company_id)
            .where(Sale.branch_id == branch_id)
            .options(
                selectinload(Sale.payments),
                selectinload(Sale.items),
                selectinload(Sale.user),
                selectinload(Sale.client),
            )
        )
        query = self._apply_sales_filters(query)
        return query.order_by(Sale.timestamp.desc())

    def _payment_method_key(self, method_type: Any) -> str:
        if isinstance(method_type, PaymentMethodType):
            key = method_type.value
        elif hasattr(method_type, "value"):
            key = str(method_type.value).strip().lower()
        else:
            key = str(method_type or "").strip().lower()
        if key == "card":
            return "credit"
        if key == "wallet":
            return "yape"
        return key

    def _payment_method_label(self, method_key: str) -> str:
        return payment_method_label(method_key)

    def _payment_method_abbrev(self, method_key: str) -> str:
        _abbr = MSG.HIST_PAY_ABBR
        mapping = {
            "cash": _abbr["efectivo"],
            "debit": _abbr["tarjeta_debito"],
            "credit": _abbr["tarjeta_credito"],
            "yape": _abbr["yape"],
            "plin": _abbr["plin"],
            "transfer": _abbr["transferencia"],
            "mixed": _abbr["mixto"],
            "other": _abbr["otro"],
        }
        return mapping.get(method_key, _abbr["otro"])

    def _sorted_payment_keys(self, keys: list[str]) -> list[str]:
        order = [
            "cash",
            "debit",
            "credit",
            "yape",
            "plin",
            "transfer",
            "mixed",
            "other",
        ]
        ordered = [key for key in order if key in keys]
        for key in keys:
            if key not in ordered:
                ordered.append(key)
        return ordered

    def _payment_summary_from_payments(self, payments: list[Any]) -> str:
        if not payments:
            return "-"
        totals: dict[str, Decimal] = {}
        for payment in payments:
            key = self._payment_method_key(getattr(payment, "method_type", None))
            if not key:
                continue
            amount = Decimal(str(getattr(payment, "amount", 0) or 0))
            totals[key] = totals.get(key, Decimal("0.00")) + amount
        if not totals:
            return "-"
        parts = []
        for key in self._sorted_payment_keys(list(totals.keys())):
            label = self._payment_method_label(key)
            parts.append(f"{label}: {self._format_currency(totals[key])}")
        return ", ".join(parts)

    def _payment_method_display(self, payments: list[Any]) -> str:
        if not payments:
            return "-"
        keys: list[str] = []
        for payment in payments:
            key = self._payment_method_key(getattr(payment, "method_type", None))
            if key and key not in keys:
                keys.append(key)
        if not keys:
            return "-"
        if len(keys) == 1:
            return self._payment_method_label(keys[0])
        abbrevs = [
            self._payment_method_abbrev(key)
            for key in self._sorted_payment_keys(keys)
        ]
        return f"{self._payment_method_label('mixed')} ({'/'.join(abbrevs)})"

    def _credit_sale_payment_label(
        self, sale: Sale, payments: list[Any]
    ) -> str | None:
        condition = (sale.payment_condition or "").strip().lower()
        if condition != "credito":
            return None
        paid_total = sum(
            Decimal(str(getattr(payment, "amount", 0) or 0))
            for payment in payments
        )
        if paid_total > 0:
            return "Crédito c/ Inicial"
        return "Crédito"

    def _sale_payment_method_label(
        self, sale: Sale, payments: list[Any], fallback: str
    ) -> str:
        explicit = (getattr(sale, "payment_method", "") or "").strip()
        credit_label = self._credit_sale_payment_label(sale, payments)
        if credit_label:
            if explicit and explicit not in {"-", MSG.FALLBACK_NOT_SPECIFIED}:
                normalized = explicit.lower()
                if normalized.startswith("credito") or normalized.startswith("crédito"):
                    return explicit
            return credit_label
        if explicit and explicit not in {"-", MSG.FALLBACK_NOT_SPECIFIED}:
            return explicit
        return fallback

    def _normalize_credit_label(self, label: str) -> str:
        normalized = unicodedata.normalize("NFKD", label or "")
        normalized = "".join(
            ch for ch in normalized if not unicodedata.combining(ch)
        )
        normalized = "".join(
            ch for ch in normalized if ch.isalpha() or ch.isspace()
        )
        return normalized.lower().strip()

    def _is_credit_label(self, label: str) -> bool:
        normalized = self._normalize_credit_label(label)
        return (
            "credito" in normalized
            or "creito" in normalized
            or "crdito" in normalized
        )

    def _sale_log_payment_info(
        self, session, sale_ids: list[int]
    ) -> dict[int, dict[str, str]]:
        if not sale_ids:
            return {}
        company_id = self._company_id()
        branch_id = self._branch_id()
        if not company_id or not branch_id:
            return {}
        conditions = [CashboxLog.sale_id.in_(sale_ids)]
        conditions.extend(
            CashboxLog.notes.like(f"%Venta%{sale_id}%")
            for sale_id in sale_ids
        )
        if not conditions:
            return {}
        # FIX 38d: add branch_id filter for branch-level isolation
        logs = session.exec(
            select(CashboxLog)
            .where(CashboxLog.company_id == company_id)
            .where(CashboxLog.branch_id == branch_id)
            .where(CashboxLog.action.in_(["Venta", "Inicial Credito"]))
            .where(CashboxLog.is_voided == False)
            .where(sa.or_(*conditions))
        ).all()
        info: dict[int, dict[str, str]] = {}
        import re
        for log in logs:
            notes = log.notes or ""
            sale_id = getattr(log, "sale_id", None)
            if sale_id is None:
                match = re.search(r"Venta[^0-9]*(\d+)", notes, re.IGNORECASE)
                if not match:
                    continue
                try:
                    sale_id = int(match.group(1))
                except ValueError:
                    continue
            if sale_id not in sale_ids or sale_id in info:
                continue
            info[sale_id] = {
                "payment_method": (log.payment_method or MSG.FALLBACK_NOT_SPECIFIED).strip()
                or MSG.FALLBACK_NOT_SPECIFIED,
                "payment_details": notes,
            }
        return info

    def _fetch_sales_history(
        self, offset: int | None = None, limit: int | None = None
    ) -> list[dict]:
        with rx.session() as session:
            query = self._sales_query()
            if offset is not None:
                query = query.offset(offset)
            if limit is not None:
                query = query.limit(limit)
            sales = session.exec(query).all()
            sale_ids = [sale.id for sale in sales if sale and sale.id is not None]
            log_payment_info = self._sale_log_payment_info(session, sale_ids)
            sale_user_lookup = self._build_sale_user_lookup(session, sales)

            rows: list[dict] = []
            for sale in sales:
                payments = sale.payments or []
                payment_method = self._payment_method_display(payments)
                payment_details = self._payment_summary_from_payments(payments)
                if payment_method.strip() in {"", "-"}:
                    payment_method = MSG.FALLBACK_NOT_SPECIFIED
                payment_method = self._sale_payment_method_label(
                    sale, payments, payment_method
                )
                fallback = log_payment_info.get(sale.id)
                if fallback and payment_method == MSG.FALLBACK_NOT_SPECIFIED:
                    payment_method = fallback.get("payment_method", payment_method)
                    payment_details = fallback.get("payment_details", payment_details)
                client_name = (
                    sale.client.name if sale.client else MSG.FALLBACK_NO_CLIENT
                )
                user_name = self._sale_username(
                    sale, sale_user_lookup, MSG.FALLBACK_UNKNOWN
                )
                total_amount = self._round_currency(sale.total_amount or 0)
                rows.append(
                    {
                        "sale_id": str(sale.id),
                        "timestamp": self._format_company_datetime(sale.timestamp)
                        if sale.timestamp
                        else "",
                        "client_name": client_name,
                        "total": total_amount,
                        "payment_method": payment_method,
                        "payment_details": self._payment_details_text(payment_details),
                        "user": user_name,
                    }
                )
            return rows

    def _sales_total_count(self) -> int:
        company_id = self._company_id()
        branch_id = self._branch_id()
        if not company_id or not branch_id:
            return 0
        with rx.session() as session:
            count_query = (
                select(sa.func.count())
                .select_from(Sale)
                .where(Sale.status != SaleStatus.cancelled)
            )
            count_query = self._apply_sales_filters(count_query)
            return session.exec(count_query).one()

    def _refresh_history_cache(self):
        if not self.current_user["privileges"]["view_historial"]:
            self.filtered_history = []
            self.total_pages = 1
            return

        total_items = int(self._sales_total_count() or 0)
        total_pages = 1 if total_items == 0 else (total_items + self.items_per_page - 1) // self.items_per_page
        page = min(max(self.current_page_history, 1), total_pages)
        if page != self.current_page_history:
            self.current_page_history = page

        offset = (page - 1) * self.items_per_page
        self.filtered_history = self._fetch_sales_history(
            offset=offset,
            limit=self.items_per_page,
        )
        self.total_pages = total_pages

    def _refresh_report_cache(self):
        if not self.current_user["privileges"]["view_historial"]:
            self.report_method_summary = []
            self.report_detail_rows = []
            self.report_closing_rows = []
            return

        entries = self._build_report_entries()
        totals: dict[str, dict[str, Any]] = {}
        for entry in entries:
            key = entry.get("method_key") or "other"
            if key not in totals:
                totals[key] = {
                    "method_label": self._payment_method_label(key),
                    "count": 0,
                    "total": Decimal("0.00"),
                }
            totals[key]["count"] += 1
            totals[key]["total"] += Decimal(str(entry.get("amount", 0) or 0))

        summary: list[dict] = []
        for key in REPORT_METHOD_KEYS:
            if key in totals:
                item = totals[key]
                summary.append(
                    {
                        "method_label": item["method_label"],
                        "count": item["count"],
                        "total": self._round_currency(item["total"]),
                    }
                )
        for key, value in totals.items():
            if key not in REPORT_METHOD_KEYS:
                summary.append(
                    {
                        "method_label": value["method_label"],
                        "count": value["count"],
                        "total": self._round_currency(value["total"]),
                    }
                )

        closings = self._build_report_closings()
        self.report_method_summary = summary
        self.report_detail_rows = entries
        self.report_closing_rows = closings

        detail_total_pages = (
            1
            if len(entries) == 0
            else (len(entries) + self.report_detail_items_per_page - 1)
            // self.report_detail_items_per_page
        )
        if self.report_detail_current_page > detail_total_pages:
            self.report_detail_current_page = detail_total_pages

        closing_total_pages = (
            1
            if len(closings) == 0
            else (len(closings) + self.report_closing_items_per_page - 1)
            // self.report_closing_items_per_page
        )
        if self.report_closing_current_page > closing_total_pages:
            self.report_closing_current_page = closing_total_pages

    def _enabled_payment_kinds(self, session, company_id: int, branch_id: int) -> set[str]:
        enabled_kinds: set[str] = set()
        methods = session.exec(
            select(PaymentMethod)
            .where(PaymentMethod.enabled == True)
            .where(PaymentMethod.company_id == company_id)
            .where(PaymentMethod.branch_id == branch_id)
        ).all()
        for method in methods:
            kind = (method.kind or method.method_id or "other").strip().lower()
            if kind == "card":
                kind = "credit"
            elif kind == "wallet":
                kind = "yape"
            enabled_kinds.add(kind)
        return enabled_kinds

    def _build_dynamic_payment_cards_from_stats(
        self, stats: Dict[str, float], enabled_kinds: set[str]
    ) -> list[dict]:
        styles = {
            "cash": {"icon": "coins", "color": "blue"},
            "debit": {"icon": "credit-card", "color": "indigo"},
            "credit": {"icon": "credit-card", "color": "violet"},
            "yape": {"icon": "qr-code", "color": "pink"},
            "plin": {"icon": "qr-code", "color": "cyan"},
            "transfer": {"icon": "landmark", "color": "orange"},
            "mixed": {"icon": "layers", "color": "amber"},
            "other": {"icon": "circle-help", "color": "gray"},
        }
        order_index = {
            "cash": 0,
            "yape": 1,
            "plin": 2,
            "credit": 3,
            "debit": 4,
            "transfer": 5,
            "mixed": 6,
            "other": 7,
        }
        cards: list[dict] = []

        def _add_card(kind: str, label: str, stats_key: str) -> None:
            amount = stats.get(stats_key, 0.0)
            if amount == 0 and kind not in enabled_kinds:
                return
            style = styles.get(kind, styles["other"])
            cards.append(
                {
                    "name": label,
                    "amount": amount,
                    "icon": style["icon"],
                    "color": style["color"],
                    "_sort_key": kind,
                }
            )

        _add_card("cash", "Efectivo", "efectivo")
        _add_card("yape", "Yape", "yape")
        _add_card("plin", "Plin", "plin")
        _add_card("credit", "T. Credito", "credito")
        _add_card("debit", "T. Debito", "debito")
        _add_card("transfer", "Transferencia", "transferencia")
        mixed_label = "Pago Mixto" if "mixed" in enabled_kinds else "Otros"
        if stats.get("mixto", 0) > 0:
            _add_card("mixed", mixed_label, "mixto")

        cards.sort(
            key=lambda card: (
                order_index.get(card.get("_sort_key") or "", 99),
                (card.get("name") or "").strip().lower(),
            )
        )
        for card in cards:
            card.pop("_sort_key", None)
        return cards

    def _refresh_financial_cache(self):
        if not self.current_user["privileges"]["view_historial"]:
            self.payment_stats = {
                "efectivo": 0.0,
                "debito": 0.0,
                "credito": 0.0,
                "yape": 0.0,
                "plin": 0.0,
                "transferencia": 0.0,
                "mixto": 0.0,
            }
            self.total_credit = 0.0
            self.credit_outstanding = 0.0
            self.dynamic_payment_cards = []
            self.productos_mas_vendidos = []
            self.productos_stock_bajo = []
            self.sales_by_day = []
            return

        company_id = self._company_id()
        branch_id = self._branch_id()
        if not company_id or not branch_id:
            self.payment_stats = {
                "efectivo": 0.0,
                "debito": 0.0,
                "credito": 0.0,
                "yape": 0.0,
                "plin": 0.0,
                "transferencia": 0.0,
                "mixto": 0.0,
            }
            self.total_credit = 0.0
            self.credit_outstanding = 0.0
            self.dynamic_payment_cards = []
            self.productos_mas_vendidos = []
            self.productos_stock_bajo = []
            self.sales_by_day = []
            return

        start_date, end_date = self._history_date_range()
        stats = {
            "efectivo": Decimal("0.00"),
            "debito": Decimal("0.00"),
            "credito": Decimal("0.00"),
            "yape": Decimal("0.00"),
            "plin": Decimal("0.00"),
            "transferencia": Decimal("0.00"),
            "mixto": Decimal("0.00"),
        }
        total_credit = Decimal("0.00")
        pending_total = Decimal("0.00")

        with rx.session() as session:
            payment_query = (
                select(
                    SalePayment.method_type,
                    sa.func.sum(SalePayment.amount),
                )
                .join(Sale, SalePayment.sale_id == Sale.id)
                .where(Sale.status != SaleStatus.cancelled)
                .where(Sale.company_id == company_id)
                .where(Sale.branch_id == branch_id)
            )
            if start_date:
                payment_query = payment_query.where(
                    SalePayment.created_at >= start_date
                )
            if end_date:
                payment_query = payment_query.where(
                    SalePayment.created_at <= end_date
                )
            payment_query = payment_query.group_by(SalePayment.method_type)
            for method_type, amount in session.exec(payment_query).all():
                key = self._payment_method_key(method_type)
                if not key:
                    continue
                self._add_to_stats(stats, key, Decimal(str(amount or 0)))

            query_log = (
                select(
                    CashboxLog.payment_method,
                    sa.func.sum(CashboxLog.amount),
                )
                .where(CashboxLog.action.in_(REPORT_CASHBOX_ACTIONS))
                .where(CashboxLog.is_voided == False)
                .where(CashboxLog.company_id == company_id)
                .where(CashboxLog.branch_id == branch_id)
            )
            if start_date:
                query_log = query_log.where(CashboxLog.timestamp >= start_date)
            if end_date:
                query_log = query_log.where(CashboxLog.timestamp <= end_date)
            query_log = query_log.group_by(CashboxLog.payment_method)
            for payment_label, amount in session.exec(query_log).all():
                method_key = self._method_key_from_label(payment_label or "")
                self._add_to_stats(stats, method_key, Decimal(str(amount or 0)))

            credit_sales_query = (
                select(Sale)
                .where(Sale.status != SaleStatus.cancelled)
                .where(Sale.company_id == company_id)
                .where(Sale.branch_id == branch_id)
                .options(selectinload(Sale.payments), selectinload(Sale.installments))
            )
            if start_date:
                credit_sales_query = credit_sales_query.where(Sale.timestamp >= start_date)
            if end_date:
                credit_sales_query = credit_sales_query.where(Sale.timestamp <= end_date)

            sales = session.exec(credit_sales_query).all()
            for sale in sales:
                payments = sale.payments or []
                payment_method = self._payment_method_display(payments)
                if payment_method.strip() in {"", "-"}:
                    payment_method = MSG.FALLBACK_NOT_SPECIFIED
                payment_method = self._sale_payment_method_label(
                    sale, payments, payment_method
                )
                if self._is_credit_label(payment_method):
                    total_credit += Decimal(str(sale.total_amount or 0))

                payment_condition = (sale.payment_condition or "").strip().lower()
                is_credit = payment_condition == "credito" or self._is_credit_label(
                    payment_method
                )
                if not is_credit:
                    continue
                total_amount = Decimal(str(sale.total_amount or 0))
                paid_initial = sum(
                    Decimal(str(getattr(payment, "amount", 0) or 0))
                    for payment in payments
                )
                paid_installments = sum(
                    Decimal(str(getattr(installment, "paid_amount", 0) or 0))
                    for installment in (sale.installments or [])
                )
                pending = total_amount - paid_initial - paid_installments
                if pending > 0:
                    pending_total += pending

            enabled_kinds = self._enabled_payment_kinds(session, company_id, branch_id)

            top_products_stmt = (
                select(
                    SaleItem.product_name_snapshot,
                    sa.func.sum(SaleItem.quantity).label("total_qty"),
                )
                .join(Sale, SaleItem.sale_id == Sale.id)
                .where(Sale.status != SaleStatus.cancelled)
                .where(Sale.company_id == company_id)
                .where(Sale.branch_id == branch_id)
                .group_by(SaleItem.product_name_snapshot)
                .order_by(sa.desc("total_qty"))
                .limit(5)
            )
            top_products = session.exec(top_products_stmt).all()

            low_stock_products = session.exec(
                select(Product)
                .where(Product.company_id == company_id)
                .where(Product.branch_id == branch_id)
                .where(Product.stock <= 10)
                .order_by(Product.stock)
            ).all()

            date_col = sa.func.date(Sale.timestamp)
            sales_day_query = (
                select(date_col, sa.func.sum(Sale.total_amount))
                .where(Sale.status != SaleStatus.cancelled)
                .where(Sale.company_id == company_id)
                .where(Sale.branch_id == branch_id)
                .group_by(date_col)
                .order_by(date_col.desc())
                .limit(7)
            )
            sales_day_rows = list(reversed(session.exec(sales_day_query).all()))

        rounded_stats = {k: self._round_currency(v) for k, v in stats.items()}
        self.payment_stats = rounded_stats
        self.total_credit = self._round_currency(total_credit)
        self.credit_outstanding = self._round_currency(pending_total)
        self.dynamic_payment_cards = self._build_dynamic_payment_cards_from_stats(
            rounded_stats, enabled_kinds
        )
        self.productos_mas_vendidos = [
            {"description": name, "cantidad_vendida": qty}
            for name, qty in top_products
        ]
        self.productos_stock_bajo = [
            {
                "barcode": p.barcode,
                "description": p.description,
                "stock": p.stock,
                "unit": p.unit,
            }
            for p in low_stock_products
        ]
        self.sales_by_day = [
            {
                "date": day.strftime("%Y-%m-%d")
                if hasattr(day, "strftime")
                else str(day or ""),
                "total": self._round_currency(total or 0),
            }
            for day, total in sales_day_rows
        ]

    def _refresh_historial_cache(self):
        self._refresh_history_cache()
        self._refresh_report_cache()
        self._refresh_financial_cache()

    @rx.var(cache=True)
    def report_detail_total_pages(self) -> int:
        _ = self._report_update_trigger
        total_items = len(self.report_detail_rows)
        if total_items == 0:
            return 1
        return (
            total_items + self.report_detail_items_per_page - 1
        ) // self.report_detail_items_per_page

    @rx.var(cache=True)
    def paginated_report_detail_rows(self) -> list[dict]:
        _ = self._report_update_trigger
        if not self.current_user["privileges"]["view_historial"]:
            return []
        total_pages = self.report_detail_total_pages
        page = min(
            max(self.report_detail_current_page, 1), total_pages
        )
        per_page = max(self.report_detail_items_per_page, 1)
        offset = (page - 1) * per_page
        return self.report_detail_rows[offset : offset + per_page]

    @rx.var(cache=True)
    def report_closing_total_pages(self) -> int:
        _ = self._report_update_trigger
        total_items = len(self.report_closing_rows)
        if total_items == 0:
            return 1
        return (
            total_items + self.report_closing_items_per_page - 1
        ) // self.report_closing_items_per_page

    @rx.var(cache=True)
    def paginated_report_closing_rows(self) -> list[dict]:
        _ = self._report_update_trigger
        if not self.current_user["privileges"]["view_historial"]:
            return []
        total_pages = self.report_closing_total_pages
        page = min(
            max(self.report_closing_current_page, 1), total_pages
        )
        per_page = max(self.report_closing_items_per_page, 1)
        offset = (page - 1) * per_page
        return self.report_closing_rows[offset : offset + per_page]

    @rx.event
    def set_history_page(self, page_num: int):
        if 1 <= page_num <= self.total_pages:
            self.current_page_history = page_num
            self._refresh_history_cache()


    @rx.event
    def apply_history_filters(self):
        self.history_filter_type = self.staged_history_filter_type
        self.history_filter_product = self.staged_history_filter_product
        self.history_filter_category = self.staged_history_filter_category
        self.history_filter_start_date = self.staged_history_filter_start_date
        self.history_filter_end_date = self.staged_history_filter_end_date
        self.current_page_history = 1
        self._history_update_trigger += 1
        self._refresh_history_cache()
        self._refresh_financial_cache()

    @rx.event
    def reload_history(self):
        self._load_category_options()
        self._load_report_options()
        self._history_update_trigger += 1
        self._report_update_trigger += 1
        self._refresh_historial_cache()
        self._load_returns_report()

    @rx.event(background=True)
    async def reload_history_background(self):
        """Recarga el historial en segundo plano para evitar bloquear la navegación."""
        async with self:
            self.reload_history()

    @rx.event
    def reset_history_filters(self):
        self.staged_history_filter_type = "Todos"
        self.staged_history_filter_product = ""
        self.staged_history_filter_category = "Todas"
        self.staged_history_filter_start_date = ""
        self.staged_history_filter_end_date = ""
        self.apply_history_filters()

    @rx.event
    def set_staged_history_filter_type(self, value: str):
        self.staged_history_filter_type = value or "Todos"

    @rx.event
    def set_staged_history_filter_product(self, value: str):
        self.staged_history_filter_product = value or ""

    @rx.event
    def set_staged_history_filter_category(self, value: str):
        self.staged_history_filter_category = value or "Todas"

    @rx.event
    def set_staged_history_filter_start_date(self, value: str):
        self.staged_history_filter_start_date = value or ""

    @rx.event
    def set_staged_history_filter_end_date(self, value: str):
        self.staged_history_filter_end_date = value or ""

    @rx.event
    def set_report_tab(self, value: str):
        self.report_active_tab = value or "metodos"
        self.report_detail_current_page = 1
        self.report_closing_current_page = 1
        self._refresh_report_cache()

    @rx.event
    def apply_report_filters(self):
        self.report_filter_start_date = self.staged_report_filter_start_date
        self.report_filter_end_date = self.staged_report_filter_end_date
        self.report_filter_method = self.staged_report_filter_method
        self.report_filter_source = self.staged_report_filter_source
        self.report_filter_user = self.staged_report_filter_user
        self.report_detail_current_page = 1
        self.report_closing_current_page = 1
        self._report_update_trigger += 1
        self._refresh_report_cache()

    @rx.event
    def reset_report_filters(self):
        self.staged_report_filter_start_date = ""
        self.staged_report_filter_end_date = ""
        self.staged_report_filter_method = "Todos"
        self.staged_report_filter_source = "Todos"
        self.staged_report_filter_user = "Todos"
        self.report_detail_current_page = 1
        self.report_closing_current_page = 1
        self.apply_report_filters()

    @rx.event
    def set_staged_report_filter_start_date(self, value: str):
        self.staged_report_filter_start_date = value or ""

    @rx.event
    def set_staged_report_filter_end_date(self, value: str):
        self.staged_report_filter_end_date = value or ""

    @rx.event
    def set_staged_report_filter_method(self, value: str):
        self.staged_report_filter_method = value or "Todos"

    @rx.event
    def set_staged_report_filter_source(self, value: str):
        self.staged_report_filter_source = value or "Todos"

    @rx.event
    def set_staged_report_filter_user(self, value: str):
        self.staged_report_filter_user = value or "Todos"

    @rx.event
    def set_report_detail_page(self, page_num: int):
        if 1 <= page_num <= self.report_detail_total_pages:
            self.report_detail_current_page = page_num

    @rx.event
    def next_report_detail_page(self):
        if self.report_detail_current_page < self.report_detail_total_pages:
            self.report_detail_current_page += 1

    @rx.event
    def prev_report_detail_page(self):
        if self.report_detail_current_page > 1:
            self.report_detail_current_page -= 1

    @rx.event
    def set_report_closing_page(self, page_num: int):
        if 1 <= page_num <= self.report_closing_total_pages:
            self.report_closing_current_page = page_num

    @rx.event
    def next_report_closing_page(self):
        if self.report_closing_current_page < self.report_closing_total_pages:
            self.report_closing_current_page += 1

    @rx.event
    def prev_report_closing_page(self):
        if self.report_closing_current_page > 1:
            self.report_closing_current_page -= 1

    @rx.event
    def open_sale_detail(self, sale_id: str):
        if not sale_id:
            return rx.toast(MSG.SALE_NOT_FOUND, duration=3000)
        try:
            sale_db_id = int(sale_id)
        except ValueError:
            return rx.toast(MSG.SALE_NOT_FOUND, duration=3000)
        company_id = self._company_id()
        branch_id = self._branch_id()
        if not company_id or not branch_id:
            return rx.toast(MSG.VAL_COMPANY_UNDEFINED, duration=3000)
        with rx.session() as session:
            # FIX 38c: add branch_id filter for branch-level isolation
            sale = session.exec(
                select(Sale)
                .where(Sale.id == sale_db_id)
                .where(Sale.company_id == company_id)
                .where(Sale.branch_id == branch_id)
                .options(
                    selectinload(Sale.items),
                    selectinload(Sale.payments),
                    selectinload(Sale.user),
                    selectinload(Sale.client),
                )
            ).first()
            if not sale:
                return rx.toast(MSG.SALE_NOT_FOUND, duration=3000)
            payments = sale.payments or []
            payment_method = self._payment_method_display(payments)
            payment_details = self._payment_summary_from_payments(payments)
            if payment_method.strip() in {"", "-"}:
                payment_method = MSG.FALLBACK_NOT_SPECIFIED
            payment_method = self._sale_payment_method_label(
                sale, payments, payment_method
            )
            if payment_method == MSG.FALLBACK_NOT_SPECIFIED:
                log_info = self._sale_log_payment_info(session, [sale.id]).get(
                    sale.id, {}
                )
                if log_info:
                    payment_method = log_info.get("payment_method", payment_method)
                    payment_details = log_info.get(
                        "payment_details", payment_details
                    )
            sale_user_lookup = self._build_sale_user_lookup(session, [sale])
            self.selected_sale_items = [
                {
                    "description": item.product_name_snapshot,
                    "quantity": float(item.quantity or 0),
                    "unit_price": self._round_currency(item.unit_price or 0),
                    "subtotal": self._round_currency(item.subtotal or 0),
                }
                for item in (sale.items or [])
            ]
            self.selected_sale_id = str(sale.id)
            self.selected_sale_summary = {
                "sale_id": str(sale.id),
                "timestamp": self._format_company_datetime(sale.timestamp)
                if sale.timestamp
                else "",
                "client_name": sale.client.name if sale.client else MSG.FALLBACK_NO_CLIENT,
                "user": self._sale_username(
                    sale, sale_user_lookup, MSG.FALLBACK_UNKNOWN
                ),
                "payment_method": payment_method,
                "payment_details": self._payment_details_text(payment_details),
                "total": self._round_currency(sale.total_amount or 0),
            }
        self.sale_detail_modal_open = True

    @rx.event
    def close_sale_detail(self):
        self.sale_detail_modal_open = False
        self.selected_sale_id = ""
        self.selected_sale_summary = {}
        self.selected_sale_items = []

    @rx.event
    def set_sale_detail_modal_open(self, open_state: bool):
        if open_state:
            self.sale_detail_modal_open = True
        else:
            self.close_sale_detail()

    @rx.var(cache=True)
    def selected_sale_items_view(self) -> list[dict]:
        return list(self.selected_sale_items or [])

    @rx.event
    def export_to_excel(self):
        if not self.current_user["privileges"]["export_data"]:
            return rx.toast(MSG.PERM_EXPORT, duration=3000)

        company_id = self._company_id()
        if not company_id:
            return rx.toast(MSG.VAL_COMPANY_UNDEFINED, duration=3000)

        currency_label = self._currency_symbol_clean()
        currency_format = self._currency_excel_format()
        company_name = getattr(self, "company_name", "") or "EMPRESA"
        start_dt, end_dt = self._history_date_range()
        period_start = (
            self._format_company_datetime(start_dt, "%d/%m/%Y")
            if start_dt
            else "Inicio"
        )
        period_end = (
            self._format_company_datetime(end_dt, "%d/%m/%Y")
            if end_dt
            else "Actual"
        )
        period_label = f"Período: {period_start} a {period_end}"

        wb, ws = create_excel_workbook("Historial de Ventas")

        # Encabezado profesional
        row = add_company_header(
            ws,
            company_name,
            "HISTORIAL DE MOVIMIENTOS Y VENTAS",
            period_label,
            columns=11,
            generated_at=self._display_now(),
        )

        headers = [
            "Fecha y Hora",
            "Nº Venta",
            "Cliente",
            "Vendedor",
            "Método de Pago",
            MSG.FALLBACK_PRODUCT,
            "Variante",
            "Categoría",
            "Cantidad",
            f"Precio Unitario ({currency_label})",
            f"Subtotal ({currency_label})",
        ]
        style_header_row(ws, row, headers)
        data_start = row + 1
        row += 1

        with rx.session() as session:
            query = (
                select(Sale)
                .where(Sale.status != SaleStatus.cancelled)
                .options(
                    selectinload(Sale.items).selectinload(SaleItem.product_variant),
                    selectinload(Sale.payments),
                    selectinload(Sale.installments),
                    selectinload(Sale.user),
                    selectinload(Sale.client),
                )
            )
            query = self._apply_sales_filters(query).order_by(
                Sale.timestamp.desc()
            )
            sales = session.exec(query).all()
            sale_ids = [sale.id for sale in sales if sale and sale.id is not None]
            log_payment_info = self._sale_log_payment_info(session, sale_ids)
            sale_user_lookup = self._build_sale_user_lookup(session, sales)

            invalid_labels = {"", "-", "no especificado"}
            for sale in sales:
                payments = sale.payments or []
                payment_method = self._payment_method_display(payments)
                payment_details = self._payment_summary_from_payments(payments)
                if payment_method.strip() in {"", "-"}:
                    payment_method = MSG.FALLBACK_NOT_SPECIFIED
                payment_method = self._sale_payment_method_label(
                    sale, payments, payment_method
                )
                fallback = log_payment_info.get(sale.id)
                if fallback and payment_method == MSG.FALLBACK_NOT_SPECIFIED:
                    payment_method = fallback.get("payment_method", payment_method)
                    payment_details = fallback.get(
                        "payment_details", payment_details
                    )
                payment_type = (
                    getattr(sale, "payment_type", "") or ""
                ).strip().lower()
                payment_condition = (sale.payment_condition or "").strip().lower()
                is_credit = (
                    bool(getattr(sale, "is_credit", False))
                    or payment_type == "credit"
                    or payment_condition in {"credito", "credit"}
                )
                paid_total = sum(
                    Decimal(str(getattr(payment, "amount", 0) or 0))
                    for payment in payments
                )
                installments_paid = sum(
                    Decimal(str(getattr(installment, "paid_amount", 0) or 0))
                    for installment in sale.installments or []
                )
                amount_paid_total = paid_total + installments_paid
                total_amount_value = Decimal(str(sale.total_amount or 0))
                if is_credit:
                    payment_method = MSG.HIST_CREDIT_SALE
                    if total_amount_value > 0 and amount_paid_total >= total_amount_value:
                        payment_details = "Crédito (Completado)"
                    elif amount_paid_total > 0:
                        payment_details = (
                            f"Crédito (Adelanto: {self._format_currency(amount_paid_total)})"
                        )
                    else:
                        payment_details = "Crédito (Pendiente Total)"
                else:
                    if (payment_method or "").strip().lower() in invalid_labels:
                        payment_method = MSG.FALLBACK_NOT_SPECIFIED
                    if (payment_details or "").strip().lower() in invalid_labels:
                        if (payment_method or "").strip().lower() not in invalid_labels:
                            payment_details = f"Pago en {payment_method}"
                        else:
                            payment_details = MSG.HIST_PAYMENT_REGISTERED

                client_name = sale.client.name if sale.client else MSG.HIST_CASH_SALE
                user_name = self._sale_username(sale, sale_user_lookup, MSG.FALLBACK_SYSTEM)
                if is_credit:
                    method_display = payment_method
                else:
                    method_display = self._normalize_wallet_label(payment_method)

                sale_items = sale.items or []
                if not sale_items:
                    sale_items = [None]

                for item in sale_items:
                    if item is None:
                        product_name = MSG.FALLBACK_NO_PRODUCTS
                        variant_label = "-"
                        category = MSG.FALLBACK_NO_CATEGORY
                        quantity = Decimal("0")
                        unit_price = Decimal("0")
                        subtotal = Decimal("0")
                    else:
                        variant_label = "-"
                        if item.product_variant:
                            parts: list[str] = []
                            if item.product_variant.size:
                                parts.append(f"Talla {str(item.product_variant.size).strip()}")
                            if item.product_variant.color:
                                parts.append(str(item.product_variant.color).strip())
                            label = " / ".join([p for p in parts if p])
                            variant_label = label or "-"
                        else:
                            snapshot = (item.product_name_snapshot or "").strip()
                            if snapshot and snapshot.endswith(")") and "(" in snapshot:
                                variant_label = snapshot.rsplit("(", 1)[-1].rstrip(")") or "-"
                        name_snapshot = (item.product_name_snapshot or "").strip() or MSG.FALLBACK_PRODUCT
                        if variant_label != "-" and name_snapshot.endswith(")") and "(" in name_snapshot:
                            product_name = name_snapshot.rsplit("(", 1)[0].strip() or name_snapshot
                        else:
                            product_name = name_snapshot
                        category = (item.product_category_snapshot or "").strip()
                        if not category:
                            category = "Servicios" if item.product_id is None else "General"
                        quantity = Decimal(str(item.quantity or 0))
                        unit_price = Decimal(str(item.unit_price or 0))
                        subtotal = Decimal(str(item.subtotal or 0))

                    ws.cell(
                        row=row,
                        column=1,
                        value=self._format_company_datetime(
                            sale.timestamp,
                            "%d/%m/%Y %H:%M",
                        ) if sale.timestamp else "",
                    )
                    ws.cell(row=row, column=2, value=str(sale.id))
                    ws.cell(row=row, column=3, value=client_name)
                    ws.cell(row=row, column=4, value=user_name)
                    ws.cell(row=row, column=5, value=method_display)
                    ws.cell(row=row, column=6, value=product_name)
                    ws.cell(row=row, column=7, value=variant_label)
                    ws.cell(row=row, column=8, value=category)
                    ws.cell(row=row, column=9, value=quantity).number_format = NUMBER_FORMAT
                    ws.cell(row=row, column=10, value=unit_price).number_format = currency_format
                    ws.cell(row=row, column=11, value=subtotal).number_format = currency_format

                    for col in range(1, 12):
                        ws.cell(row=row, column=col).border = THIN_BORDER
                    row += 1

        # Fila de totales
        totals_row = row
        add_totals_row_with_formulas(ws, totals_row, data_start, [
            {"type": "label", "value": "TOTALES"},
            {"type": "text", "value": ""},
            {"type": "text", "value": ""},
            {"type": "text", "value": ""},
            {"type": "text", "value": ""},
            {"type": "text", "value": ""},
            {"type": "text", "value": ""},
            {"type": "text", "value": ""},
            {"type": "sum", "col_letter": "I", "number_format": NUMBER_FORMAT},
            {"type": "text", "value": ""},
            {"type": "sum", "col_letter": "K", "number_format": currency_format},
        ])

        # Notas explicativas
        add_notes_section(ws, totals_row, [
            "Nº Venta: Identificador único de la transacción en el sistema.",
            "Subtotal: Cantidad x Precio Unitario por cada ítem.",
            "Crédito (Completado): El cliente pagó la totalidad del crédito.",
            "Crédito (Adelanto): El cliente realizó un pago parcial.",
            "Crédito (Pendiente Total): No se ha recibido ningún pago aún.",
            "Venta al contado: Cliente no identificado, pago inmediato.",
        ], columns=11)

        auto_adjust_column_widths(ws)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return rx.download(data=output.getvalue(), filename="historial_ventas.xlsx")

    @rx.event
    def export_report_data(self):
        if not self.current_user["privileges"]["export_data"]:
            return rx.toast(MSG.PERM_EXPORT, duration=3000)

        currency_label = self._currency_symbol_clean()
        currency_format = self._currency_excel_format()
        company_name = getattr(self, "company_name", "") or "EMPRESA"
        start_dt, end_dt = self._report_date_range()
        period_start = (
            self._format_company_datetime(start_dt, "%d/%m/%Y")
            if start_dt
            else "Inicio"
        )
        period_end = (
            self._format_company_datetime(end_dt, "%d/%m/%Y")
            if end_dt
            else "Actual"
        )
        period_label = f"Período: {period_start} a {period_end}"

        active_tab = self.report_active_tab or "metodos"
        if active_tab == "cierres":
            rows = self._build_report_closings()
            if not rows:
                return rx.toast(MSG.HIST_NO_CLOSINGS_EXPORT, duration=3000)

            wb, ws = create_excel_workbook(MSG.REPORT_CLOSINGS_SHEET)

            # Encabezado profesional
            row = add_company_header(
                ws,
                company_name,
                "HISTORIAL DE CIERRES DE CAJA",
                period_label,
                columns=5,
                generated_at=self._display_now(),
            )

            headers = [
                "Fecha y Hora",
                "Tipo Operación",
                "Responsable",
                f"Monto ({currency_label})",
                "Observaciones",
            ]
            style_header_row(ws, row, headers)
            data_start = row + 1
            row += 1

            for item in rows:
                action = item.get("action", "")
                action_display = "Apertura de Caja" if action.lower() == "apertura" else "Cierre de Caja" if action.lower() == "cierre" else action.capitalize()

                ws.cell(row=row, column=1, value=item.get("timestamp_display", ""))
                ws.cell(row=row, column=2, value=action_display)
                ws.cell(row=row, column=3, value=item.get("user", MSG.FALLBACK_UNKNOWN))
                ws.cell(row=row, column=4, value=item.get("amount", 0) or 0).number_format = currency_format
                ws.cell(row=row, column=5, value=item.get("notes", "") or MSG.FALLBACK_NO_OBS)

                for col in range(1, 6):
                    ws.cell(row=row, column=col).border = THIN_BORDER
                row += 1

            # Totales
            totals_row = row
            add_totals_row_with_formulas(ws, totals_row, data_start, [
                {"type": "label", "value": "TOTAL"},
                {"type": "text", "value": ""},
                {"type": "text", "value": ""},
                {"type": "sum", "col_letter": "D", "number_format": currency_format},
                {"type": "text", "value": ""},
            ])

            add_notes_section(ws, totals_row, [
                "Apertura de Caja: Monto inicial del día.",
                "Cierre de Caja: Monto contado al finalizar.",
            ], columns=5)

            auto_adjust_column_widths(ws)

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return rx.download(data=output.getvalue(), filename="historial_cierres.xlsx")

        entries = self._build_report_entries()
        if not entries:
            return rx.toast(MSG.HIST_NO_INCOMES_EXPORT, duration=3000)

        if active_tab == "detalle":
            wb, ws = create_excel_workbook(MSG.REPORT_PAYMENTS_SHEET)

            # Encabezado profesional
            row = add_company_header(
                ws,
                company_name,
                MSG.REPORT_PAYMENTS_TITLE,
                period_label,
                columns=6,
                generated_at=self._display_now(),
            )

            detail_headers = [
                "Fecha y Hora",
                "Origen/Tipo",
                "Método de Pago",
                f"Monto ({currency_label})",
                "Responsable",
                "Referencia",
            ]
            style_header_row(ws, row, detail_headers)
            data_start = row + 1
            row += 1

            for entry in entries:
                ws.cell(row=row, column=1, value=entry.get("timestamp_display", ""))
                ws.cell(row=row, column=2, value=entry.get("source", "Venta"))
                ws.cell(row=row, column=3, value=entry.get("method_label", MSG.FALLBACK_NOT_SPECIFIED))
                ws.cell(row=row, column=4, value=entry.get("amount", 0) or 0).number_format = currency_format
                ws.cell(row=row, column=5, value=entry.get("user", MSG.FALLBACK_SYSTEM))
                ws.cell(row=row, column=6, value=entry.get("reference", "") or MSG.FALLBACK_NO_REFERENCE)

                for col in range(1, 7):
                    ws.cell(row=row, column=col).border = THIN_BORDER
                row += 1

            # Totales
            totals_row = row
            add_totals_row_with_formulas(ws, totals_row, data_start, [
                {"type": "label", "value": MSG.REPORT_TOTAL_INCOME},
                {"type": "text", "value": ""},
                {"type": "text", "value": ""},
                {"type": "sum", "col_letter": "D", "number_format": currency_format},
                {"type": "text", "value": ""},
                {"type": "text", "value": ""},
            ])

            add_notes_section(ws, totals_row, [
                "Origen: Tipo de transacción (Venta, Cobro de Cuota, Reserva, etc.).",
                "Referencia: Información adicional del pago.",
            ], columns=6)

            auto_adjust_column_widths(ws)

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return rx.download(data=output.getvalue(), filename="detalle_cobros.xlsx")

        # Tab por defecto: métodos de pago
        summary_totals: dict[str, dict[str, Any]] = {}
        for entry in entries:
            key = entry.get("method_key") or "other"
            if key not in summary_totals:
                summary_totals[key] = {
                    "method_label": self._payment_method_label(key),
                    "count": 0,
                    "total": Decimal("0.00"),
                }
            summary_totals[key]["count"] += 1
            summary_totals[key]["total"] += Decimal(
                str(entry.get("amount", 0) or 0)
            )

        wb, ws = create_excel_workbook("Resumen por Método")

        # Encabezado profesional
        row = add_company_header(
            ws,
            company_name,
            "INGRESOS POR MÉTODO DE PAGO",
            period_label,
            columns=4,
            generated_at=self._display_now(),
        )

        summary_headers = [
            "Método de Pago",
            "Nº Operaciones",
            f"Total Recaudado ({currency_label})",
            "Participación (%)",
        ]
        style_header_row(ws, row, summary_headers)
        data_start = row + 1
        row += 1

        for key in REPORT_METHOD_KEYS:
            if key in summary_totals:
                summary = summary_totals[key]
                ws.cell(row=row, column=1, value=summary["method_label"])
                ws.cell(row=row, column=2, value=summary["count"])
                ws.cell(row=row, column=3, value=summary["total"]).number_format = currency_format
                # Participación se calculará después
                ws.cell(row=row, column=4, value=summary["total"]).number_format = currency_format

                for col in range(1, 5):
                    ws.cell(row=row, column=col).border = THIN_BORDER
                row += 1

        for key, summary in summary_totals.items():
            if key not in REPORT_METHOD_KEYS:
                ws.cell(row=row, column=1, value=summary["method_label"])
                ws.cell(row=row, column=2, value=summary["count"])
                ws.cell(row=row, column=3, value=summary["total"]).number_format = currency_format
                ws.cell(row=row, column=4, value=summary["total"]).number_format = currency_format

                for col in range(1, 5):
                    ws.cell(row=row, column=col).border = THIN_BORDER
                row += 1

        # Totales
        totals_row = row
        add_totals_row_with_formulas(ws, totals_row, data_start, [
            {"type": "label", "value": "TOTAL RECAUDADO"},
            {"type": "sum", "col_letter": "B"},
            {"type": "sum", "col_letter": "C", "number_format": currency_format},
            {"type": "text", "value": "100.00%"},
        ])

        # Actualizar participación con fórmulas
        from openpyxl.styles import Font
        PERCENT_FORMAT_LOCAL = '0.00%'
        for r in range(data_start, totals_row):
            ws.cell(row=r, column=4, value=f"=IF($C${totals_row}>0,C{r}/$C${totals_row},0)").number_format = PERCENT_FORMAT_LOCAL

        add_notes_section(ws, totals_row, [
            "Total Recaudado: Suma de todos los pagos por método.",
            "Participación = Monto del Método ÷ Total General × 100.",
        ], columns=4)

        auto_adjust_column_widths(ws)

        # Segunda hoja: Detalle
        detail_ws = wb.create_sheet(MSG.REPORT_PAYMENTS_SHEET)
        row = add_company_header(
            detail_ws,
            company_name,
            MSG.REPORT_PAYMENTS_TITLE,
            period_label,
            columns=6,
            generated_at=self._display_now(),
        )

        detail_headers = [
            "Fecha y Hora",
            "Origen/Tipo",
            "Método de Pago",
            f"Monto ({currency_label})",
            "Responsable",
            "Referencia",
        ]
        style_header_row(detail_ws, row, detail_headers)
        detail_data_start = row + 1
        row += 1

        for entry in entries:
            detail_ws.cell(row=row, column=1, value=entry.get("timestamp_display", ""))
            detail_ws.cell(row=row, column=2, value=entry.get("source", "Venta"))
            detail_ws.cell(row=row, column=3, value=entry.get("method_label", MSG.FALLBACK_NOT_SPECIFIED))
            detail_ws.cell(row=row, column=4, value=entry.get("amount", 0) or 0).number_format = currency_format
            detail_ws.cell(row=row, column=5, value=entry.get("user", MSG.FALLBACK_SYSTEM))
            detail_ws.cell(row=row, column=6, value=entry.get("reference", "") or MSG.FALLBACK_NO_REFERENCE)

            for col in range(1, 7):
                detail_ws.cell(row=row, column=col).border = THIN_BORDER
            row += 1

        # Totales en detalle
        detail_totals_row = row
        add_totals_row_with_formulas(detail_ws, detail_totals_row, detail_data_start, [
            {"type": "label", "value": "TOTAL"},
            {"type": "text", "value": ""},
            {"type": "text", "value": ""},
            {"type": "sum", "col_letter": "D", "number_format": currency_format},
            {"type": "text", "value": ""},
            {"type": "text", "value": ""},
        ])

        auto_adjust_column_widths(detail_ws)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return rx.download(data=output.getvalue(), filename="ingresos_por_metodo.xlsx")

    def _parse_payment_amount(self, text: str, keyword: str) -> Decimal:
        """Extrae el monto de un keyword en un texto de pago mixto como 'Efectivo S/ 15.00'."""
        import re
        try:
            # Normalizar separadores.
            # Formato en BD: "Pagos Mixtos - Efectivo S/ 15.00 / Plin S/ 20.00 / Montos completos."

            # Reemplazar " / " (barra con espacios) por pipe
            text = text.replace(" / ", "|")
            # Reemplazar " - " (guion con espacios) por pipe
            text = text.replace(" - ", "|")
            # Reemplazar "/" (barra sin espacios) por si acaso
            text = text.replace("/", "|")

            parts = text.split("|")

            for part in parts:
                part = part.strip()
                if keyword.lower() in part.lower():
                    try:
                        # Extraer numero (maneja comas y decimales)
                        match = re.search(r"([0-9]+(?:[.,][0-9]{3})*(?:[.,][0-9]+)?)", part)
                        if match:
                            num_str = match.group(1)
                            if "," in num_str and "." in num_str:
                                num_str = num_str.replace(",", "")
                            elif "," in num_str and "." not in num_str:
                                num_str = num_str.replace(",", ".")
                            return Decimal(num_str)
                    except IndexError:
                        continue
        except Exception:
            logger.warning(
                "_parse_payment_amount: could not parse | text=%r keyword=%r",
                text[:80],
                keyword,
            )
        return Decimal("0.00")

    def _add_to_stats(self, stats: dict, key: str, amount: Decimal):
        """Helper para sumar montos usando las keys del Enum."""
        if key == PaymentMethodType.cash.value:
            stats["efectivo"] += amount
        elif key == PaymentMethodType.debit.value:
            stats["debito"] += amount
        elif key == PaymentMethodType.credit.value:
            stats["credito"] += amount
        elif key == PaymentMethodType.yape.value:
            stats["yape"] += amount
        elif key == PaymentMethodType.plin.value:
            stats["plin"] += amount
        elif key == PaymentMethodType.transfer.value:
            stats["transferencia"] += amount
        else:
            stats["mixto"] += amount

    @rx.var(cache=True)
    def total_ventas_efectivo(self) -> float:
        return self.payment_stats["efectivo"]

    @rx.var(cache=True)
    def total_ventas_debito(self) -> float:
        return self.payment_stats["debito"]

    @rx.var(cache=True)
    def total_ventas_credito(self) -> float:
        return self.payment_stats["credito"]

    @rx.var(cache=True)
    def total_ventas_yape(self) -> float:
        return self.payment_stats["yape"]

    @rx.var(cache=True)
    def total_ventas_plin(self) -> float:
        return self.payment_stats["plin"]

    @rx.var(cache=True)
    def total_ventas_tarjeta(self) -> float:
        return self._round_currency(
            self.payment_stats["debito"] + self.payment_stats["credito"]
        )

    @rx.var(cache=True)
    def total_ventas_transferencia(self) -> float:
        return self.payment_stats["transferencia"]

    @rx.var(cache=True)
    def total_ventas_mixtas(self) -> float:
        return self.payment_stats["mixto"]

    # ── Devoluciones: computed properties ──

    @rx.var(cache=True)
    def return_reason_options(self) -> list[list[str]]:
        return [
            [r.value, r.display_label]
            for r in ReturnReason
        ]

    @rx.var(cache=True)
    def return_refund_total(self) -> str:
        total = sum(
            (item.get("return_qty", 0) or 0) * (item.get("unit_price", 0) or 0)
            for item in self.return_items
        )
        return self._format_currency(total)

    @rx.var(cache=True)
    def return_has_selection(self) -> bool:
        return any(
            (item.get("return_qty", 0) or 0) > 0
            for item in self.return_items
        )

    # ── Devoluciones: event handlers ──

    @rx.event
    def open_return_modal(self, sale_id: str):
        """Abre el modal de devolución cargando los ítems de la venta."""
        if not self.current_user["privileges"].get("delete_sales", False):
            return rx.toast("No tiene permisos para procesar devoluciones.", duration=3000)
        company_id = self._company_id()
        branch_id = self._branch_id()
        if not company_id or not branch_id:
            return rx.toast("Empresa no configurada.", duration=3000)

        try:
            db_sale_id = int(sale_id)
        except (ValueError, TypeError):
            return rx.toast("ID de venta inválido.", duration=3000)

        with rx.session() as session:
            sale = session.exec(
                select(Sale)
                .options(selectinload(Sale.items))
                .where(Sale.id == db_sale_id)
                .where(Sale.company_id == company_id)
                .where(Sale.branch_id == branch_id)
            ).first()

            if not sale:
                return rx.toast("Venta no encontrada.", duration=3000)
            if sale.status == SaleStatus.cancelled:
                return rx.toast("La venta fue anulada, no se puede devolver.", duration=3000)
            if sale.status == SaleStatus.returned:
                return rx.toast("La venta ya fue devuelta completamente.", duration=3000)

            # Cargar devoluciones previas
            existing_returns = session.exec(
                select(SaleReturnItem)
                .join(SaleReturn)
                .where(SaleReturn.original_sale_id == db_sale_id)
                .where(SaleReturn.company_id == company_id)
            ).all()
            already_returned: dict[int, Decimal] = {}
            for er in existing_returns:
                already_returned[er.sale_item_id] = (
                    already_returned.get(er.sale_item_id, Decimal("0")) + er.quantity
                )

            self.return_sale_id = db_sale_id
            self.return_sale_summary = {
                "id": sale.id,
                "total": float(sale.total_amount or 0),
                "date": self._format_company_datetime(
                    sale.timestamp, "%d/%m/%Y %H:%M"
                ) if sale.timestamp else "",
            }
            items = []
            for si in sale.items:
                prev_returned = float(already_returned.get(si.id, Decimal("0")))
                available = float(si.quantity or 0) - prev_returned
                if available <= 0:
                    continue
                unit_price = float(si.unit_price or 0)
                items.append({
                    "sale_item_id": si.id,
                    "product_name": si.product_name_snapshot or "Producto",
                    "unit_price": unit_price,
                    "original_qty": float(si.quantity or 0),
                    "already_returned": prev_returned,
                    "available_qty": available,
                    "return_qty": available,
                    "refund_line": round(available * unit_price, 2),
                })
            if not items:
                return rx.toast("Todos los ítems ya fueron devueltos.", duration=3000)

            self.return_items = items
            self.return_reason = "other"
            self.return_notes = ""
            self.return_modal_open = True

    @rx.event
    def close_return_modal(self):
        self.return_modal_open = False
        self.return_sale_id = 0
        self.return_sale_summary = {}
        self.return_items = []
        self.return_reason = "other"
        self.return_notes = ""

    @rx.event
    def set_return_reason(self, value: str):
        self.return_reason = value

    @rx.event
    def set_return_notes(self, value: str):
        from app.utils.sanitization import sanitize_notes_preserve_spaces
        self.return_notes = sanitize_notes_preserve_spaces(value)

    @rx.event
    def set_return_item_qty(self, sale_item_id: str, value: str):
        """Actualiza la cantidad a devolver de un ítem."""
        try:
            qty = float(value) if value else 0
            if qty < 0:
                qty = 0
        except (ValueError, TypeError):
            qty = 0

        item_id = int(sale_item_id)
        updated = []
        for item in self.return_items:
            if item["sale_item_id"] == item_id:
                available = item["available_qty"]
                item = dict(item)
                item["return_qty"] = min(qty, available)
                item["refund_line"] = round(item["return_qty"] * item["unit_price"], 2)
            updated.append(item)
        self.return_items = updated

    @rx.event
    def select_all_return_items(self):
        """Selecciona la cantidad máxima para todos los ítems."""
        updated = []
        for item in self.return_items:
            item = dict(item)
            item["return_qty"] = item["available_qty"]
            item["refund_line"] = round(item["return_qty"] * item["unit_price"], 2)
            updated.append(item)
        self.return_items = updated

    @rx.event
    async def confirm_return(self):
        """Procesa la devolución y emite nota de crédito si aplica."""
        if not self.current_user["privileges"].get("delete_sales", False):
            yield rx.toast("No tiene permisos para procesar devoluciones.", duration=3000)
            return
        company_id = self._company_id()
        branch_id = self._branch_id()
        user_id = self.current_user.get("id")
        if not company_id or not branch_id or not user_id:
            yield rx.toast("Sesión inválida.", duration=3000)
            return

        if not self.return_sale_id:
            yield rx.toast("No hay venta seleccionada.", duration=3000)
            return

        from app.services.return_service import process_return, ReturnItemRequest

        items_to_return = []
        for item in self.return_items:
            qty = item.get("return_qty", 0) or 0
            if qty > 0:
                items_to_return.append(ReturnItemRequest(
                    sale_item_id=item["sale_item_id"],
                    quantity=Decimal(str(qty)),
                ))

        if not items_to_return:
            yield rx.toast("Seleccione al menos un ítem para devolver.", duration=3000)
            return

        # Capture values before close_return_modal clears state
        sale_id = self.return_sale_id
        reason = self.return_reason

        with rx.session() as session:
            try:
                result = process_return(
                    session,
                    sale_id=sale_id,
                    company_id=company_id,
                    branch_id=branch_id,
                    user_id=user_id,
                    reason=reason,
                    notes=self.return_notes,
                    items=items_to_return,
                    timestamp=self._event_timestamp(),
                )
                if not result.success:
                    yield rx.toast(result.error, duration=4000)
                    return
                session.commit()
            except Exception:
                session.rollback()
                logger.exception("Error al procesar devolución venta #%s", sale_id)
                yield rx.toast(
                    "Error al procesar la devolución. Intente nuevamente.",
                    duration=4000,
                )
                return

        self.close_return_modal()
        self._history_update_trigger += 1
        refund_display = self._format_currency(float(result.refund_amount))
        self._load_returns_report()
        yield rx.toast(
            f"Devolución procesada: {result.items_returned} ítem(s), "
            f"reembolso {refund_display}",
            duration=5000,
        )

        # ── Nota de crédito automática ──────────────────────────
        await self._try_emit_return_credit_note(
            sale_id=sale_id,
            company_id=company_id,
            branch_id=branch_id,
            reason=reason,
        )

    # ── Nota de crédito automática al devolver ──────────────

    async def _try_emit_return_credit_note(
        self,
        *,
        sale_id: int,
        company_id: int,
        branch_id: int,
        reason: str,
    ) -> None:
        """Emite nota de crédito automática si la venta tiene comprobante fiscal autorizado.

        Se ejecuta después de una devolución exitosa. No lanza excepción
        si falla — solo loggea y muestra toast informativo.
        """
        try:
            with rx.session() as session:
                # 1. Verificar que la empresa tiene facturación electrónica activa
                company = session.exec(
                    select(Company).where(Company.id == company_id)
                ).first()
                if not company or not company.has_electronic_billing:
                    return

                billing_config = session.exec(
                    select(CompanyBillingConfig).where(
                        CompanyBillingConfig.company_id == company_id,
                        CompanyBillingConfig.is_active == True,  # noqa: E712
                    )
                ).first()
                if not billing_config:
                    return

                # 2. Buscar comprobante fiscal autorizado de la venta original
                original_doc = session.exec(
                    select(FiscalDocument).where(
                        FiscalDocument.sale_id == sale_id,
                        FiscalDocument.company_id == company_id,
                        FiscalDocument.fiscal_status == FiscalStatus.authorized,
                        FiscalDocument.receipt_type.in_([
                            ReceiptType.boleta,
                            ReceiptType.factura,
                        ]),
                    )
                ).first()
                if not original_doc:
                    return

                # 3. Verificar que no exista ya una nota de crédito
                nc_exists = session.exec(
                    select(FiscalDocument).where(
                        FiscalDocument.sale_id == sale_id,
                        FiscalDocument.company_id == company_id,
                        FiscalDocument.receipt_type == ReceiptType.nota_credito,
                    )
                ).first()
                if nc_exists:
                    return

                # Capturar datos del comprobante original
                orig_doc_id = original_doc.id
                buyer_doc_type = original_doc.buyer_doc_type
                buyer_doc_number = original_doc.buyer_doc_number
                buyer_name = original_doc.buyer_name

            # 4. Emitir nota de crédito (fuera de la sesión de lectura)
            from app.services.billing_service import emit_fiscal_document

            try:
                reason_label = ReturnReason(reason).display_label
            except ValueError:
                reason_label = reason
            credit_note_reason = f"DEVOLUCIÓN: {reason_label}"

            result = await emit_fiscal_document(
                sale_id=sale_id,
                company_id=company_id,
                branch_id=branch_id,
                receipt_type_override=ReceiptType.nota_credito,
                buyer_doc_type=buyer_doc_type,
                buyer_doc_number=buyer_doc_number,
                buyer_name=buyer_name,
                credit_note_reason=credit_note_reason,
                original_fiscal_doc_id=orig_doc_id,
            )

            if result and result.fiscal_status == FiscalStatus.authorized:
                logger.info(
                    "Nota de crédito %s emitida automáticamente para venta #%s",
                    result.full_number, sale_id,
                )
            elif result:
                logger.warning(
                    "Nota de crédito para venta #%s quedó en estado %s",
                    sale_id, result.fiscal_status,
                )
            else:
                logger.warning(
                    "No se pudo emitir nota de crédito para venta #%s (billing inactivo o cuota agotada)",
                    sale_id,
                )

        except Exception:
            logger.exception(
                "Error al emitir nota de crédito automática para venta #%s", sale_id,
            )

    # ── Reporte de devoluciones: métodos ──

    def _returns_date_range(self) -> tuple[datetime.datetime | None, datetime.datetime | None]:
        start_date = None
        end_date = None
        if self.returns_report_filter_start:
            try:
                start_date, _ = self._company_day_bounds_utc_naive(
                    self.returns_report_filter_start
                )
            except ValueError:
                start_date = None
        if self.returns_report_filter_end:
            try:
                _, end_date = self._company_day_bounds_utc_naive(
                    self.returns_report_filter_end
                )
            except ValueError:
                end_date = None
        return start_date, end_date

    def _load_returns_report(self) -> None:
        """Carga el reporte de devoluciones desde la BD."""
        company_id = self._company_id()
        branch_id = self._branch_id()
        if not company_id or not branch_id:
            self.returns_report_rows = []
            self.returns_report_total = 0.0
            self.returns_report_count = 0
            return

        start_dt, end_dt = self._returns_date_range()

        with rx.session() as session:
            query = (
                select(SaleReturn)
                .options(
                    selectinload(SaleReturn.items),
                )
                .where(
                    SaleReturn.company_id == company_id,
                    SaleReturn.branch_id == branch_id,
                )
            )
            if start_dt:
                query = query.where(SaleReturn.timestamp >= start_dt)
            if end_dt:
                query = query.where(SaleReturn.timestamp <= end_dt)

            query = query.order_by(SaleReturn.timestamp.desc())
            returns = session.exec(query).all()

            # Build user lookup for return users
            user_ids = {
                int(r.user_id)
                for r in returns
                if r.user_id is not None
            }
            user_lookup: dict[int, str] = {}
            if user_ids:
                with tenant_bypass():
                    user_rows = session.exec(
                        select(User.id, User.username)
                        .where(User.id.in_(user_ids))
                        .where(User.company_id == company_id)
                        .execution_options(tenant_bypass=True)
                    ).all()
                for row in user_rows:
                    try:
                        uid, uname = row[0], row[1]
                        user_lookup[int(uid)] = str(uname or "").strip()
                    except (TypeError, ValueError):
                        continue

            # Build product lookup for item names
            product_ids = set()
            for ret in returns:
                for item in (ret.items or []):
                    if item.product_id:
                        product_ids.add(item.product_id)
            product_lookup: dict[int, str] = {}
            if product_ids:
                prod_rows = session.exec(
                    select(Product.id, Product.name)
                    .where(Product.id.in_(product_ids))
                ).all()
                for row in prod_rows:
                    try:
                        pid, pname = row[0], row[1]
                        product_lookup[int(pid)] = str(pname or "Producto")
                    except (TypeError, ValueError):
                        continue

            rows = []
            total_refund = 0.0
            for ret in returns:
                # Build items summary
                items_detail = []
                for item in (ret.items or []):
                    prod_name = product_lookup.get(item.product_id, "Producto") if item.product_id else "Producto"
                    items_detail.append(
                        f"{prod_name} x{float(item.quantity or 0):.0f}"
                    )

                reason_label = ret.reason or "other"
                try:
                    reason_label = ReturnReason(ret.reason).display_label
                except (ValueError, KeyError):
                    reason_label = ret.reason or "Otro"

                refund_amt = float(ret.refund_amount or 0)
                total_refund += refund_amt

                rows.append({
                    "id": ret.id,
                    "timestamp": self._format_company_datetime(
                        ret.timestamp, "%d/%m/%Y %H:%M"
                    ) if ret.timestamp else "",
                    "original_sale_id": ret.original_sale_id,
                    "reason": reason_label,
                    "notes": ret.notes or "",
                    "items_summary": ", ".join(items_detail) if items_detail else "Sin ítems",
                    "items_count": len(items_detail),
                    "refund_amount": refund_amt,
                    "user": user_lookup.get(int(ret.user_id), "Desconocido") if ret.user_id else "Desconocido",
                })

            self.returns_report_rows = rows
            self.returns_report_total = round(total_refund, 2)
            self.returns_report_count = len(rows)

    @rx.var(cache=True)
    def returns_report_total_pages(self) -> int:
        if not self.returns_report_rows:
            return 1
        return max(1, -(-len(self.returns_report_rows) // self.returns_report_per_page))

    @rx.var(cache=True)
    def returns_report_paginated(self) -> list[dict]:
        start = (self.returns_report_page - 1) * self.returns_report_per_page
        end = start + self.returns_report_per_page
        return self.returns_report_rows[start:end]

    @rx.var(cache=True)
    def formatted_returns_total(self) -> str:
        return self._format_currency(self.returns_report_total)

    @rx.event
    def load_returns_report(self):
        """Event handler público para cargar el reporte de devoluciones."""
        self.returns_report_page = 1
        self._load_returns_report()

    @rx.event
    def set_returns_filter_start(self, value: str):
        self.returns_report_filter_start = value or ""

    @rx.event
    def set_returns_filter_end(self, value: str):
        self.returns_report_filter_end = value or ""

    @rx.event
    def set_returns_report_page(self, page: int):
        if 1 <= page <= self.returns_report_total_pages:
            self.returns_report_page = page

    @rx.event
    def export_returns_excel(self):
        """Exporta el reporte de devoluciones a Excel."""
        if not self.current_user["privileges"].get("export_data", False):
            return rx.toast(MSG.PERM_EXPORT, duration=3000)

        company_id = self._company_id()
        if not company_id:
            return rx.toast(MSG.VAL_COMPANY_UNDEFINED, duration=3000)

        if not self.returns_report_rows:
            return rx.toast("No hay devoluciones para exportar.", duration=3000)

        currency_label = self._currency_symbol_clean()
        company_name = getattr(self, "company_name", "") or "EMPRESA"
        start_dt, end_dt = self._returns_date_range()
        period_start = (
            self._format_company_datetime(start_dt, "%d/%m/%Y")
            if start_dt
            else "Inicio"
        )
        period_end = (
            self._format_company_datetime(end_dt, "%d/%m/%Y")
            if end_dt
            else "Actual"
        )
        period_label = f"Período: {period_start} a {period_end}"

        wb, ws = create_excel_workbook("Devoluciones")

        row = add_company_header(
            ws,
            company_name,
            "REPORTE DE DEVOLUCIONES",
            period_label,
            columns=7,
            generated_at=self._display_now(),
        )

        headers = [
            "Fecha y Hora",
            "Venta Original #",
            "Motivo",
            "Productos Devueltos",
            "Cantidad Ítems",
            f"Reembolso ({currency_label})",
            "Usuario",
        ]
        style_header_row(ws, row, headers)
        row += 1

        data_rows = []
        for ret in self.returns_report_rows:
            data_rows.append([
                ret.get("timestamp", ""),
                ret.get("original_sale_id", ""),
                ret.get("reason", ""),
                ret.get("items_summary", ""),
                ret.get("items_count", 0),
                ret.get("refund_amount", 0),
                ret.get("user", ""),
            ])

        if data_rows:
            add_data_rows(ws, row, data_rows)
            data_end = row + len(data_rows) - 1

            # Format currency column (F)
            currency_format = self._currency_excel_format()
            for r in range(row, data_end + 1):
                cell = ws.cell(row=r, column=6)
                cell.number_format = currency_format

            # Totals row
            totals_row = data_end + 1
            ws.cell(row=totals_row, column=1, value="TOTAL")
            ws.cell(row=totals_row, column=1).font = ws.cell(row=totals_row, column=1).font.copy(bold=True)
            ws.cell(row=totals_row, column=5, value=self.returns_report_count)
            ws.cell(row=totals_row, column=5).font = ws.cell(row=totals_row, column=5).font.copy(bold=True)
            total_cell = ws.cell(row=totals_row, column=6, value=self.returns_report_total)
            total_cell.number_format = currency_format
            total_cell.font = total_cell.font.copy(bold=True)
            total_cell.fill = NEGATIVE_FILL

        auto_adjust_column_widths(ws)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return rx.download(
            data=buffer.getvalue(),
            filename=f"devoluciones_{company_name.replace(' ', '_')}.xlsx",
        )

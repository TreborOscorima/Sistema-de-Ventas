"""Estado de Inventario - Gestión de productos y stock.

Este módulo maneja toda la lógica relacionada con el inventario:

Funcionalidades principales:
- CRUD de productos (crear, leer, actualizar, eliminar)
- Gestión de categorías
- Ajustes de inventario con registro de movimientos
- Búsqueda y filtrado de productos
- Verificación de inventario físico
- Exportación de reportes

Permisos requeridos:
- view_inventario: Ver listado de productos
- edit_inventario: Crear, editar, eliminar productos y categorías

Clases:
    InventoryState: Estado principal del módulo de inventario
"""
import reflex as rx
from typing import List, Dict, Any, Optional
import datetime
import uuid
import logging
import io
from decimal import Decimal, InvalidOperation
from sqlmodel import select
from sqlalchemy import and_, or_, func, exists, literal, union_all, case
from sqlalchemy.orm import selectinload
from app.constants import DEFAULT_ITEMS_PER_PAGE, INVENTORY_RECENT_LIMIT

logger = logging.getLogger(__name__)
from app.models import (
    Product,
    ProductAttribute,
    ProductBatch,
    ProductKit,
    ProductVariant,
    PriceTier,
    StockMovement,
    User as UserModel,
    Category,
    SaleItem,
)
from .types import InventoryAdjustment
from .mixin_state import MixinState
from app.utils.tenant import set_tenant_context
from app.utils.stock import recalculate_stock_totals
from app.utils.barcode import clean_barcode, validate_barcode
from app.utils.sanitization import escape_like
from app.utils.exports import (
    create_excel_workbook,
    style_header_row,
    add_data_rows,
    auto_adjust_column_widths,
    add_company_header,
    add_totals_row_with_formulas,
    add_notes_section,
    CURRENCY_FORMAT,
    PERCENT_FORMAT,
    THIN_BORDER,
    POSITIVE_FILL,
    NEGATIVE_FILL,
    WARNING_FILL,
)


DEFAULT_LOW_STOCK_THRESHOLD = 5
# Alias para compatibilidad con dashboard_state y otros consumidores
LOW_STOCK_THRESHOLD = DEFAULT_LOW_STOCK_THRESHOLD


class InventoryState(MixinState):
    """Estado de gestión de inventario y productos.

    Maneja productos, categorías, ajustes de stock y reportes.
    Los productos se persisten en BD, no en memoria de estado.

    Attributes:
        new_category_name: Nombre para nueva categoría
        inventory_search_term: Término de búsqueda actual
        inventory_current_page: Página de paginación
        editing_product: Producto en edición (dict temporal)
        is_editing_product: True si hay modal de edición abierto
        inventory_check_modal_open: Modal de verificación de inventario
        inventory_adjustment_item: Item siendo ajustado
    """
    # inventory: Dict[str, Product] = {} # Eliminado a favor de la BD
    # categories: List[str] = ["General"] # Reemplazado por BD
    new_category_name: str = ""
    new_category_input_key: int = 0
    inventory_search_term: str = ""
    inventory_current_page: int = 1
    inventory_items_per_page: int = DEFAULT_ITEMS_PER_PAGE
    inventory_recent_limit: int = INVENTORY_RECENT_LIMIT

    editing_product: Dict[str, Any] = { # Tipo cambiado a Dict para manejo de formularios
        "id": None,
        "barcode": "",
        "description": "",
        "category": "",
        "stock": 0,
        "unit": "",
        "purchase_price": 0,
        "sale_price": 0,
    }
    is_editing_product: bool = False
    show_variants: bool = False
    show_wholesale: bool = False
    show_batches: bool = False
    show_attributes: bool = False
    show_kit_components: bool = False
    confirm_disable_wholesale: bool = False
    variants: List[Dict[str, Any]] = []
    price_tiers: List[Dict[str, Any]] = []
    batches: List[Dict[str, Any]] = []
    attributes: List[Dict[str, Any]] = []
    kit_components: List[Dict[str, Any]] = []
    stock_details_open: bool = False
    stock_details_title: str = ""
    stock_details_mode: str = "simple"
    selected_product_details: List[Dict[str, Any]] = []

    inventory_check_modal_open: bool = False
    inventory_check_status: str = "perfecto"
    inventory_adjustment_notes: str = ""
    inventory_adjustment_item: InventoryAdjustment = {
        "temp_id": "",
        "barcode": "",
        "description": "",
        "category": "",
        "unit": "",
        "current_stock": 0,
        "adjust_quantity": 0,
        "reason": "",
        "product_id": None,
        "variant_id": None,
    }
    inventory_adjustment_items: List[InventoryAdjustment] = []
    inventory_adjustment_suggestions: List[Dict[str, Any]] = []
    categories: List[str] = ["General"]
    _categories_loaded_once: bool = rx.field(default=False, is_var=False)
    categories_panel_expanded: bool = False
    show_inactive_products: bool = False
    _inventory_update_trigger: int = 0
    inventory_list: list[dict] = []
    inventory_total_pages: int = 1
    inventory_total_products: int = 0
    inventory_in_stock_count: int = 0
    inventory_low_stock_count: int = 0
    inventory_out_of_stock_count: int = 0

    # ── Importación masiva ──
    import_modal_open: bool = False
    import_preview_rows: list[dict] = []
    import_errors: list[str] = []
    import_stats: dict = {"new": 0, "updated": 0, "errors": 0, "total": 0}
    import_processing: bool = False
    import_file_name: str = ""

    def _company_id(self) -> int | None:
        company_id, branch_id = self._tenant_ids()
        set_tenant_context(company_id, branch_id)
        return company_id

    def load_categories(self):
        company_id = self._company_id()
        branch_id = self._branch_id()
        if not company_id or not branch_id:
            self.categories = ["General"]
            self._categories_loaded_once = True
            return
        with rx.session() as session:
            cats = session.exec(
                select(Category)
                .where(Category.company_id == company_id)
                .where(Category.branch_id == branch_id)
                .order_by(Category.name)
            ).all()
            names = [c.name for c in cats]
            if "General" not in names:
                names.insert(0, "General")
            self.categories = names
            self._categories_loaded_once = True

    def _inventory_search_clause(self, search: str, company_id: int, branch_id: int):
        term = f"%{escape_like(search)}%"
        variant_match = (
            exists()
            .where(ProductVariant.product_id == Product.id)
            .where(ProductVariant.company_id == company_id)
            .where(ProductVariant.branch_id == branch_id)
            .where(
                or_(
                    ProductVariant.sku.ilike(term),
                    ProductVariant.size.ilike(term),
                    ProductVariant.color.ilike(term),
                )
            )
        )
        return or_(
            Product.description.ilike(term),
            Product.barcode.ilike(term),
            Product.category.ilike(term),
            variant_match,
        )

    def _inventory_row_from_product(self, product: Product) -> Dict[str, Any]:
        stock_value = float(product.stock or 0)
        purchase_value = float(product.purchase_price or 0)
        stock_total = self._round_currency(stock_value * purchase_value)
        min_alert = float(getattr(product, 'min_stock_alert', None) or DEFAULT_LOW_STOCK_THRESHOLD)
        return {
            "id": product.id,
            "variant_id": None,
            "is_variant": False,
            "is_active": getattr(product, 'is_active', True),
            "barcode": product.barcode,
            "description": product.description,
            "category": product.category,
            "stock": product.stock,
            "stock_is_low": stock_value <= min_alert,
            "stock_is_medium": min_alert < stock_value <= min_alert * 2,
            "unit": product.unit,
            "purchase_price": product.purchase_price,
            "sale_price": product.sale_price,
            "stock_total_display": f"{stock_total:.2f}",
        }

    def _inventory_row_from_variant(
        self, product: Product, variant: ProductVariant
    ) -> Dict[str, Any]:
        label = self._variant_label(variant)
        description = product.description or ""
        if label:
            description = f"{description} ({label})"
        stock_value = float(variant.stock or 0)
        purchase_value = float(product.purchase_price or 0)
        stock_total = self._round_currency(stock_value * purchase_value)
        # Umbral por variante con fallback al Product padre.
        # variant.min_stock_alert == None significa "heredar del producto raíz".
        variant_alert = getattr(variant, 'min_stock_alert', None)
        if variant_alert is not None:
            min_alert = float(variant_alert)
        else:
            min_alert = float(
                getattr(product, 'min_stock_alert', None) or DEFAULT_LOW_STOCK_THRESHOLD
            )
        return {
            "id": product.id,
            "variant_id": variant.id,
            "is_variant": True,
            "is_active": getattr(product, 'is_active', True),
            "barcode": variant.sku or product.barcode,
            "description": description,
            "category": product.category,
            "stock": variant.stock,
            "stock_is_low": stock_value <= min_alert,
            "stock_is_medium": min_alert < stock_value <= min_alert * 2,
            "unit": product.unit,
            "purchase_price": product.purchase_price,
            "sale_price": product.sale_price,
            "stock_total_display": f"{stock_total:.2f}",
        }

    def _inventory_search_count(
        self,
        session,
        search: str,
        company_id: int,
        branch_id: int,
    ) -> int:
        """Count total matching rows for search using SQL COUNT (no row loading)."""
        term = f"%{escape_like(search)}%"
        active_filter = [] if self.show_inactive_products else [Product.is_active == True]
        variant_count = int(
            session.exec(
                select(func.count(ProductVariant.id))
                .join(Product, ProductVariant.product_id == Product.id)
                .where(ProductVariant.company_id == company_id)
                .where(ProductVariant.branch_id == branch_id)
                .where(Product.company_id == company_id)
                .where(Product.branch_id == branch_id)
                .where(*active_filter)
                .where(
                    or_(
                        ProductVariant.sku.ilike(term),
                        ProductVariant.size.ilike(term),
                        ProductVariant.color.ilike(term),
                        Product.description.ilike(term),
                        Product.barcode.ilike(term),
                        Product.category.ilike(term),
                    )
                )
            ).one()
            or 0
        )
        no_variant = ~exists().where(ProductVariant.product_id == Product.id).where(
            ProductVariant.company_id == company_id
        ).where(ProductVariant.branch_id == branch_id)
        product_count = int(
            session.exec(
                select(func.count(Product.id))
                .where(Product.company_id == company_id)
                .where(Product.branch_id == branch_id)
                .where(no_variant)
                .where(*active_filter)
                .where(
                    or_(
                        Product.description.ilike(term),
                        Product.barcode.ilike(term),
                        Product.category.ilike(term),
                    )
                )
            ).one()
            or 0
        )
        return variant_count + product_count

    def _inventory_search_rows(
        self,
        session,
        search: str,
        company_id: int,
        branch_id: int,
        offset: int = 0,
        per_page: int = 20,
    ) -> List[Dict[str, Any]]:
        """Fetch one page of search results via SQL UNION ALL + OFFSET/LIMIT."""
        term = f"%{escape_like(search)}%"
        active_filter = [] if self.show_inactive_products else [Product.is_active == True]

        # Variant IDs matching search
        variant_ids_q = (
            select(
                Product.id.label("pid"),
                ProductVariant.id.label("vid"),
                Product.description.label("sort_desc"),
                func.coalesce(ProductVariant.sku, Product.barcode).label("sort_code"),
            )
            .join(Product, ProductVariant.product_id == Product.id)
            .where(ProductVariant.company_id == company_id)
            .where(ProductVariant.branch_id == branch_id)
            .where(Product.company_id == company_id)
            .where(Product.branch_id == branch_id)
            .where(*active_filter)
            .where(
                or_(
                    ProductVariant.sku.ilike(term),
                    ProductVariant.size.ilike(term),
                    ProductVariant.color.ilike(term),
                    Product.description.ilike(term),
                    Product.barcode.ilike(term),
                    Product.category.ilike(term),
                )
            )
        )

        # Standalone product IDs matching search (no variants)
        no_variant = ~exists().where(ProductVariant.product_id == Product.id).where(
            ProductVariant.company_id == company_id
        ).where(ProductVariant.branch_id == branch_id)
        product_ids_q = (
            select(
                Product.id.label("pid"),
                literal(None).label("vid"),
                Product.description.label("sort_desc"),
                Product.barcode.label("sort_code"),
            )
            .where(Product.company_id == company_id)
            .where(Product.branch_id == branch_id)
            .where(no_variant)
            .where(*active_filter)
            .where(
                or_(
                    Product.description.ilike(term),
                    Product.barcode.ilike(term),
                    Product.category.ilike(term),
                )
            )
        )

        # SQL UNION ALL with ORDER BY + OFFSET + LIMIT
        unioned = union_all(variant_ids_q, product_ids_q).subquery()
        page_q = (
            select(unioned.c.pid, unioned.c.vid)
            .order_by(unioned.c.sort_desc, unioned.c.sort_code)
            .offset(offset)
            .limit(per_page)
        )
        page_id_rows = session.exec(page_q).all()

        if not page_id_rows:
            return []

        # Batch fetch ORM objects for this page only
        product_ids_needed = list({r[0] for r in page_id_rows})
        variant_ids_needed = [r[1] for r in page_id_rows if r[1] is not None]

        products_map = {
            p.id: p
            for p in session.exec(
                select(Product).where(Product.id.in_(product_ids_needed))
            ).all()
        }
        variants_map = {}
        if variant_ids_needed:
            variants_map = {
                v.id: v
                for v in session.exec(
                    select(ProductVariant).where(
                        ProductVariant.id.in_(variant_ids_needed)
                    )
                ).all()
            }

        # Build rows preserving SQL sort order
        rows: List[Dict[str, Any]] = []
        for pid, vid in page_id_rows:
            product = products_map.get(pid)
            if not product:
                continue
            if vid is not None:
                variant = variants_map.get(vid)
                if variant:
                    rows.append(self._inventory_row_from_variant(product, variant))
            else:
                rows.append(self._inventory_row_from_product(product))
        return rows

    def _refresh_inventory_cache(self):
        privileges = self.current_user["privileges"]
        if not privileges.get("view_inventario"):
            self.inventory_list = []
            self.inventory_total_pages = 1
            self.inventory_total_products = 0
            self.inventory_in_stock_count = 0
            self.inventory_low_stock_count = 0
            self.inventory_out_of_stock_count = 0
            return

        company_id = self._company_id()
        branch_id = self._branch_id()
        if not company_id or not branch_id:
            self.inventory_list = []
            self.inventory_total_pages = 1
            self.inventory_total_products = 0
            self.inventory_in_stock_count = 0
            self.inventory_low_stock_count = 0
            self.inventory_out_of_stock_count = 0
            return

        search = (self.inventory_search_term or "").strip().lower()
        per_page = max(self.inventory_items_per_page, 1)
        page = max(self.inventory_current_page, 1)

        with rx.session() as session:
            if search:
                total_items = self._inventory_search_count(
                    session, search, company_id, branch_id
                )
                total_pages = (
                    1 if total_items == 0 else (total_items + per_page - 1) // per_page
                )
                if page > total_pages:
                    page = total_pages
                    self.inventory_current_page = page
                offset = (page - 1) * per_page
                page_rows = self._inventory_search_rows(
                    session, search, company_id, branch_id,
                    offset=offset, per_page=per_page,
                )
            else:
                active_filter = [] if self.show_inactive_products else [Product.is_active == True]
                total_items = int(
                    session.exec(
                        select(func.count(Product.id))
                        .where(Product.company_id == company_id)
                        .where(Product.branch_id == branch_id)
                        .where(*active_filter)
                    ).one()
                    or 0
                )
                total_pages = (
                    1 if total_items == 0 else (total_items + per_page - 1) // per_page
                )
                if page > total_pages:
                    page = total_pages
                    self.inventory_current_page = page
                offset = (page - 1) * per_page
                query = (
                    select(Product)
                    .where(Product.company_id == company_id)
                    .where(Product.branch_id == branch_id)
                    .where(*active_filter)
                    .order_by(Product.description, Product.id)
                    .offset(offset)
                    .limit(per_page)
                )
                page_rows = [
                    self._inventory_row_from_product(product)
                    for product in session.exec(query).all()
                ]

            # Single query con CASE para obtener los 4 contadores en 1 round-trip
            # Solo cuenta productos activos para las estadísticas.
            stock_stats = session.exec(
                select(
                    func.count(Product.id),
                    func.sum(case(
                        (Product.stock > Product.min_stock_alert, 1),
                        else_=0,
                    )),
                    func.sum(case(
                        (and_(Product.stock > 0, Product.stock <= Product.min_stock_alert), 1),
                        else_=0,
                    )),
                    func.sum(case(
                        (Product.stock <= 0, 1),
                        else_=0,
                    )),
                )
                .where(Product.company_id == company_id)
                .where(Product.branch_id == branch_id)
                .where(Product.is_active == True)
            ).one()
            total_products = int(stock_stats[0] or 0)
            in_stock_count = int(stock_stats[1] or 0)
            low_stock_count = int(stock_stats[2] or 0)
            out_of_stock_count = int(stock_stats[3] or 0)

        self.inventory_list = page_rows
        self.inventory_total_pages = total_pages
        self.inventory_total_products = total_products
        self.inventory_in_stock_count = in_stock_count
        self.inventory_low_stock_count = low_stock_count
        self.inventory_out_of_stock_count = out_of_stock_count

    @rx.event
    def refresh_inventory_cache(self):
        self.load_categories()
        self._refresh_inventory_cache()

    @rx.event
    def toggle_show_inactive_products(self, value: bool):
        """Muestra/oculta productos inactivos en el listado."""
        self.show_inactive_products = value
        self.inventory_current_page = 1
        self._refresh_inventory_cache()

    @rx.event
    def toggle_product_active(self, product_id: int):
        """Activa/desactiva un producto."""
        if not self.current_user["privileges"]["edit_inventario"]:
            return self.add_notification(
                "No tiene permisos para editar productos.", "error"
            )
        block = self._require_active_subscription()
        if block:
            return block
        company_id = self._company_id()
        branch_id = self._branch_id()
        if not company_id or not branch_id:
            return self.add_notification("Empresa no definida.", "error")
        with rx.session() as session:
            product = session.exec(
                select(Product)
                .where(Product.id == product_id)
                .where(Product.company_id == company_id)
                .where(Product.branch_id == branch_id)
            ).first()
            if not product:
                return self.add_notification("Producto no encontrado.", "error")
            product.is_active = not product.is_active
            session.add(product)
            session.commit()
            status = "activado" if product.is_active else "desactivado"
            self._refresh_inventory_cache()
            return self.add_notification(
                f"Producto '{product.description}' {status}.", "success"
            )

    @rx.event
    def update_new_category_name(self, value: str):
        self.new_category_name = value

    @rx.event
    def toggle_categories_panel(self):
        self.categories_panel_expanded = not self.categories_panel_expanded

    @rx.event
    def add_category(self):
        if not self.current_user["privileges"]["edit_inventario"]:
            return self.add_notification(
                "No tiene permisos para editar categorias.", "error"
            )
        block = self._require_active_subscription()
        if block:
            return block
        name = (self.new_category_name or "").strip()
        if not name:
            return
        company_id = self._company_id()
        branch_id = self._branch_id()
        if not company_id or not branch_id:
            return self.add_notification("Empresa no definida.", "error")

        self.is_loading = True

        try:
            with rx.session() as session:
                existing = session.exec(
                    select(Category)
                    .where(Category.name == name)
                    .where(Category.company_id == company_id)
                    .where(Category.branch_id == branch_id)
                ).first()
                if not existing:
                    session.add(
                        Category(name=name, company_id=company_id, branch_id=branch_id)
                    )
                    session.commit()
                    self.new_category_name = ""
                    self.new_category_input_key += 1
                    self.load_categories()
                    events = [self._emit_runtime_sync_event()]
                    notification = self.add_notification(
                        f"Categoría '{name}' agregada.", "success"
                    )
                    if notification is not None:
                        events.append(notification)
                    return events
                return self.add_notification("La categoría ya existe.", "warning")
        finally:
            self.is_loading = False

    @rx.event
    def remove_category(self, category: str):
        if not self.current_user["privileges"]["edit_inventario"]:
            return self.add_notification(
                "No tiene permisos para editar categorias.", "error"
            )
        block = self._require_active_subscription()
        if block:
            return block
        if category == "General":
            return self.add_notification(
                "No se puede eliminar la categoría General.", "error"
            )
        company_id = self._company_id()
        branch_id = self._branch_id()
        if not company_id or not branch_id:
            return self.add_notification("Empresa no definida.", "error")

        self.is_loading = True

        try:
            with rx.session() as session:
                cat = session.exec(
                    select(Category)
                    .where(Category.name == category)
                    .where(Category.company_id == company_id)
                    .where(Category.branch_id == branch_id)
                ).first()
                if cat:
                    session.delete(cat)
                    session.commit()
                    self.load_categories()
                    events = [self._emit_runtime_sync_event()]
                    notification = self.add_notification(
                        f"Categoría '{category}' eliminada.", "success"
                    )
                    if notification is not None:
                        events.append(notification)
                    return events
                return self.add_notification("Categoría no encontrada.", "warning")
        finally:
            self.is_loading = False

    @rx.event
    def handle_inventory_adjustment_change(self, field: str, value: Any):
        self.inventory_adjustment_item[field] = value

        # Buscar productos cuando se escribe en el campo descripción
        if field == "description":
            self._process_inventory_adjustment_search(value)

    @rx.event
    def process_inventory_adjustment_search_blur(self, value: Any):
        """Procesa el buscador al perder foco (ideal para lector de código)."""
        return self._process_inventory_adjustment_search(value)

    @rx.event
    def handle_inventory_adjustment_search_enter(self, key: str, input_id: str):
        """Detecta Enter y fuerza blur para capturar el valor completo."""
        if key == "Enter":
            return rx.call_script(
                f"const el=document.getElementById('{input_id}'); if(el) el.blur();"
            )

    @rx.var(cache=True)
    def inventory_display_page(self) -> int:
        if self.inventory_current_page < 1:
            return 1
        if self.inventory_current_page > self.inventory_total_pages:
            return self.inventory_total_pages
        return self.inventory_current_page

    @rx.event
    def set_inventory_search_term(self, value: str):
        self.inventory_search_term = value or ""
        self.inventory_current_page = 1
        self._refresh_inventory_cache()

    @rx.event
    def set_inventory_page(self, page_num: int):
        if 1 <= page_num <= self.inventory_total_pages:
            self.inventory_current_page = page_num
            self._refresh_inventory_cache()

    @rx.event
    def open_edit_product(self, product: Product | Dict[str, Any] | None = None):
        if not self.current_user["privileges"]["edit_inventario"]:
            return rx.toast("No tiene permisos para editar el inventario.", duration=3000)

        if product is None:
            self.editing_product = {
                "id": None,
                "barcode": "",
                "description": "",
                "category": "",
                "stock": 0,
                "unit": "",
                "purchase_price": 0,
                "sale_price": 0,
            }
            self.show_variants = False
            self.show_wholesale = False
            self.show_batches = False
            self.show_attributes = False
            self.show_kit_components = False
            self.variants = []
            self.price_tiers = []
            self.batches = []
            self.attributes = []
            self.kit_components = []
            self.is_editing_product = True
            return

        def _read_value(key: str, default: Any = "") -> Any:
            if isinstance(product, dict):
                return product.get(key, default)
            return getattr(product, key, default)

        product_id = _read_value("id", None)

        # Convertir modelo a dict para edicion
        self.editing_product = {
            "id": product_id,
            "barcode": _read_value("barcode", ""),
            "description": _read_value("description", ""),
            "category": _read_value("category", ""),
            "stock": _read_value("stock", 0),
            "unit": _read_value("unit", ""),
            "purchase_price": _read_value("purchase_price", 0),
            "sale_price": _read_value("sale_price", 0),
        }

        self.show_variants = False
        self.show_wholesale = False
        self.show_batches = False
        self.show_attributes = False
        self.show_kit_components = False
        self.variants = []
        self.price_tiers = []
        self.batches = []
        self.attributes = []
        self.kit_components = []

        company_id = self._company_id()
        branch_id = self._branch_id()
        if company_id and branch_id and product_id:
            with rx.session() as session:
                variants = session.exec(
                    select(ProductVariant)
                    .where(ProductVariant.product_id == product_id)
                    .where(ProductVariant.company_id == company_id)
                    .where(ProductVariant.branch_id == branch_id)
                    .order_by(ProductVariant.id)
                ).all()
                if variants:
                    self.variants = [
                        {
                            "sku": variant.sku,
                            "size": variant.size or "",
                            "color": variant.color or "",
                            "stock": float(variant.stock or 0),
                        }
                        for variant in variants
                    ]
                    total_stock = sum(
                        float(variant.stock or 0) for variant in variants
                    )
                    self.editing_product["stock"] = total_stock
                    self.show_variants = True

                tiers = session.exec(
                    select(PriceTier)
                    .where(PriceTier.product_id == product_id)
                    .where(PriceTier.company_id == company_id)
                    .where(PriceTier.branch_id == branch_id)
                    .order_by(PriceTier.min_quantity)
                ).all()
                if tiers:
                    self.price_tiers = [
                        {
                            "min_qty": tier.min_quantity,
                            "price": float(tier.unit_price or 0),
                        }
                        for tier in tiers
                    ]
                    self.show_wholesale = True

                # Cargar lotes (batches) — sin variant_id, scoped al producto
                product_batches = session.exec(
                    select(ProductBatch)
                    .where(ProductBatch.product_id == product_id)
                    .where(ProductBatch.product_variant_id.is_(None))
                    .where(ProductBatch.company_id == company_id)
                    .where(ProductBatch.branch_id == branch_id)
                    .order_by(ProductBatch.expiration_date)
                ).all()
                if product_batches:
                    self.batches = [
                        {
                            "id": batch.id,
                            "batch_number": batch.batch_number,
                            "expiration_date": (
                                batch.expiration_date.strftime("%Y-%m-%d")
                                if batch.expiration_date
                                else ""
                            ),
                            "stock": float(batch.stock or 0),
                        }
                        for batch in product_batches
                    ]
                    self.show_batches = True

                # Cargar atributos dinámicos (EAV)
                product_attrs = session.exec(
                    select(ProductAttribute)
                    .where(ProductAttribute.product_id == product_id)
                    .where(ProductAttribute.company_id == company_id)
                    .where(ProductAttribute.branch_id == branch_id)
                    .order_by(ProductAttribute.attribute_name)
                ).all()
                if product_attrs:
                    self.attributes = [
                        {
                            "id": attr.id,
                            "name": attr.attribute_name,
                            "value": attr.attribute_value,
                        }
                        for attr in product_attrs
                    ]
                    self.show_attributes = True

                # Cargar componentes de kit
                kit_comps = session.exec(
                    select(ProductKit)
                    .where(ProductKit.kit_product_id == product_id)
                    .where(ProductKit.company_id == company_id)
                    .where(ProductKit.branch_id == branch_id)
                ).all()
                if kit_comps:
                    comp_ids = [c.component_product_id for c in kit_comps]
                    comp_products = session.exec(
                        select(Product).where(Product.id.in_(comp_ids))
                    ).all()
                    comp_map = {p.id: p for p in comp_products}
                    self.kit_components = [
                        {
                            "id": c.id,
                            "component_barcode": (comp_map[c.component_product_id].barcode or "") if c.component_product_id in comp_map else "",
                            "component_name": (comp_map[c.component_product_id].description or "") if c.component_product_id in comp_map else "",
                            "component_product_id": c.component_product_id,
                            "quantity": float(c.quantity or 1),
                        }
                        for c in kit_comps
                    ]
                    self.show_kit_components = True

        self.is_editing_product = True

    @rx.event
    def open_create_product_modal(self):
        if not self.current_user["privileges"]["edit_inventario"]:
            return rx.toast("No tiene permisos para editar el inventario.", duration=3000)
        self.editing_product = {
            "id": None,
            "barcode": "",
            "description": "",
            "category": "",
            "stock": 0,
            "unit": "",
            "purchase_price": 0,
            "sale_price": 0,
        }
        self.show_variants = False
        self.show_wholesale = False
        self.show_batches = False
        self.show_attributes = False
        self.show_kit_components = False
        self.confirm_disable_wholesale = False
        self.variants = []
        self.price_tiers = []
        self.batches = []
        self.attributes = []
        self.kit_components = []
        self.is_editing_product = True

    @rx.event
    def open_stock_details(self, product: Product | Dict[str, Any]):
        if not self.current_user["privileges"]["view_inventario"]:
            return rx.toast("No tiene permisos para ver el inventario.", duration=3000)
        company_id = self._company_id()
        branch_id = self._branch_id()
        if not company_id or not branch_id:
            return rx.toast("Empresa no definida.", duration=3000)

        def _read_value(key: str, default: Any = "") -> Any:
            if isinstance(product, dict):
                return product.get(key, default)
            return getattr(product, key, default)

        product_id = _read_value("id", None)
        description = _read_value("description", "")
        if not product_id:
            return rx.toast("Producto no encontrado.", duration=3000)

        self.stock_details_title = str(description or "Detalle de stock")
        self.selected_product_details = []
        self.stock_details_mode = "simple"

        with rx.session() as session:
            variants = session.exec(
                select(ProductVariant)
                .where(ProductVariant.product_id == product_id)
                .where(ProductVariant.company_id == company_id)
                .where(ProductVariant.branch_id == branch_id)
                .order_by(ProductVariant.sku)
            ).all()
            if variants:
                self.selected_product_details = [
                    {
                        "sku": variant.sku,
                        "size": variant.size or "",
                        "color": variant.color or "",
                        "stock": variant.stock,
                    }
                    for variant in variants
                ]
                self.stock_details_mode = "variant"
            else:
                batches = session.exec(
                    select(ProductBatch)
                    .where(ProductBatch.product_id == product_id)
                    .where(ProductBatch.company_id == company_id)
                    .where(ProductBatch.branch_id == branch_id)
                    .order_by(ProductBatch.expiration_date)
                ).all()
                if batches:
                    self.selected_product_details = [
                        {
                            "batch_number": batch.batch_number,
                            "expiration_date": (
                                batch.expiration_date.strftime("%Y-%m-%d")
                                if batch.expiration_date
                                else ""
                            ),
                            "stock": batch.stock,
                        }
                        for batch in batches
                    ]
                    self.stock_details_mode = "batch"

        self.stock_details_open = True

    @rx.event
    def close_stock_details(self):
        self.stock_details_open = False
        self.stock_details_title = ""
        self.stock_details_mode = "simple"
        self.selected_product_details = []

    @rx.event
    def cancel_edit_product(self):
        self.is_editing_product = False
        self.editing_product = {
            "id": None,
            "barcode": "",
            "description": "",
            "category": "",
            "stock": 0,
            "unit": "",
            "purchase_price": 0,
            "sale_price": 0,
        }
        self.show_variants = False
        self.show_wholesale = False
        self.show_batches = False
        self.show_attributes = False
        self.show_kit_components = False
        self.confirm_disable_wholesale = False
        self.variants = []
        self.price_tiers = []
        self.batches = []
        self.attributes = []
        self.kit_components = []

    @rx.event
    def handle_edit_product_change(self, field: str, value: str):
        if field in ["stock", "purchase_price", "sale_price"]:
            try:
                if value == "":
                    self.editing_product[field] = 0
                else:
                    self.editing_product[field] = float(value)
            except ValueError:
                pass
        else:
            self.editing_product[field] = value

    @rx.var(cache=True)
    def variant_rows(self) -> List[Dict[str, Any]]:
        rows: List[Dict[str, Any]] = []
        for index, variant in enumerate(self.variants):
            row = dict(variant)
            row["index"] = index
            rows.append(row)
        return rows

    @rx.var(cache=True)
    def price_tier_rows(self) -> List[Dict[str, Any]]:
        rows: List[Dict[str, Any]] = []
        for index, tier in enumerate(self.price_tiers):
            row = dict(tier)
            row["index"] = index
            rows.append(row)
        return rows

    @rx.var(cache=True)
    def variants_stock_total(self) -> float:
        total = 0.0
        for variant in self.variants:
            try:
                total += float(variant.get("stock", 0) or 0)
            except (TypeError, ValueError):
                continue
        return total

    @rx.event
    def set_show_variants(self, value: bool | str):
        if isinstance(value, str):
            value = value.lower() in ["true", "1", "on", "yes"]
        self.show_variants = bool(value)
        if self.show_variants and not self.variants:
            self.add_variant_row()

    @rx.event
    def set_show_wholesale(self, value: bool | str):
        if isinstance(value, str):
            value = value.lower() in ["true", "1", "on", "yes"]
        if not bool(value) and self.price_tiers:
            # Pedir confirmación antes de desactivar si hay tiers configurados
            self.confirm_disable_wholesale = True
            return
        self.show_wholesale = bool(value)
        if self.show_wholesale and not self.price_tiers:
            self.add_tier_row()

    @rx.event
    def confirm_disable_wholesale_yes(self):
        """El usuario confirmó que quiere desactivar mayoreo."""
        self.show_wholesale = False
        self.price_tiers = []
        self.confirm_disable_wholesale = False

    @rx.event
    def confirm_disable_wholesale_no(self):
        """El usuario canceló la desactivación de mayoreo."""
        self.confirm_disable_wholesale = False

    @rx.event
    def add_variant_row(self):
        self.variants = [
            *self.variants,
            {"sku": "", "size": "", "color": "", "stock": 0},
        ]

    @rx.event
    def remove_variant_row(self, index: int):
        if index < 0 or index >= len(self.variants):
            return
        self.variants = [row for idx, row in enumerate(self.variants) if idx != index]

    @rx.event
    def update_variant_field(self, index: int, field: str, value: Any):
        if index < 0 or index >= len(self.variants):
            return
        variants = list(self.variants)
        row = dict(variants[index])
        if field == "stock":
            try:
                row[field] = float(value) if value not in ("", None) else 0
            except (TypeError, ValueError):
                return
        else:
            row[field] = value
        variants[index] = row
        self.variants = variants

    @rx.event
    def handle_variant_sku_keydown(self, key: str, index: int):
        if key in ("Enter", "NumpadEnter", "Tab"):
            return rx.call_script(
                f"document.getElementById('variant_sku_{index}').blur()"
            )

    @rx.event
    def add_tier_row(self):
        self.price_tiers = [
            *self.price_tiers,
            {"min_qty": 0, "price": 0.0},
        ]

    @rx.event
    def remove_tier_row(self, index: int):
        if index < 0 or index >= len(self.price_tiers):
            return
        self.price_tiers = [
            row for idx, row in enumerate(self.price_tiers) if idx != index
        ]

    @rx.event
    def update_tier_field(self, index: int, field: str, value: Any):
        if index < 0 or index >= len(self.price_tiers):
            return
        tiers = list(self.price_tiers)
        row = dict(tiers[index])
        if field == "min_qty":
            try:
                row[field] = int(float(value)) if value not in ("", None) else 0
            except (TypeError, ValueError):
                return
        elif field == "price":
            try:
                row[field] = float(value) if value not in ("", None) else 0
            except (TypeError, ValueError):
                return
        else:
            row[field] = value
        tiers[index] = row
        self.price_tiers = tiers

    # ─── Batches (lotes con vencimiento) — Farmacia / Supermercado ───
    @rx.event
    def set_show_batches(self, value: bool | str):
        if isinstance(value, str):
            value = value.lower() in ["true", "1", "on", "yes"]
        self.show_batches = bool(value)
        if self.show_batches and not self.batches:
            self.add_batch_row()

    @rx.event
    def add_batch_row(self):
        self.batches = [
            *self.batches,
            {
                "id": None,
                "batch_number": "",
                "expiration_date": "",
                "stock": 0.0,
            },
        ]

    @rx.event
    def remove_batch_row(self, index: int):
        if index < 0 or index >= len(self.batches):
            return
        self.batches = [
            row for idx, row in enumerate(self.batches) if idx != index
        ]

    @rx.event
    def update_batch_field(self, index: int, field: str, value: Any):
        if index < 0 or index >= len(self.batches):
            return
        batches = list(self.batches)
        row = dict(batches[index])
        if field == "stock":
            try:
                row[field] = float(value) if value not in ("", None) else 0.0
            except (TypeError, ValueError):
                return
        else:
            row[field] = value
        batches[index] = row
        self.batches = batches

    # ─── Attributes (EAV dinámicos) — Ferretería / Farmacia ───
    @rx.event
    def set_show_attributes(self, value: bool | str):
        if isinstance(value, str):
            value = value.lower() in ["true", "1", "on", "yes"]
        self.show_attributes = bool(value)
        if self.show_attributes and not self.attributes:
            self.add_attribute_row()

    @rx.event
    def add_attribute_row(self):
        self.attributes = [
            *self.attributes,
            {"id": None, "name": "", "value": ""},
        ]

    @rx.event
    def remove_attribute_row(self, index: int):
        if index < 0 or index >= len(self.attributes):
            return
        self.attributes = [
            row for idx, row in enumerate(self.attributes) if idx != index
        ]

    @rx.event
    def update_attribute_field(self, index: int, field: str, value: Any):
        if index < 0 or index >= len(self.attributes):
            return
        attrs = list(self.attributes)
        row = dict(attrs[index])
        row[field] = value
        attrs[index] = row
        self.attributes = attrs

    # ─── Kit Components (composición de kits/combos) ───
    @rx.event
    def set_show_kit_components(self, value: bool | str):
        if isinstance(value, str):
            value = value.lower() in ["true", "1", "on", "yes"]
        self.show_kit_components = bool(value)
        if self.show_kit_components and not self.kit_components:
            self.add_kit_component_row()

    @rx.event
    def add_kit_component_row(self):
        self.kit_components = [
            *self.kit_components,
            {"id": None, "component_barcode": "", "component_name": "", "component_product_id": None, "quantity": 1.0},
        ]

    @rx.event
    def remove_kit_component_row(self, index: int):
        if index < 0 or index >= len(self.kit_components):
            return
        self.kit_components = [
            row for idx, row in enumerate(self.kit_components) if idx != index
        ]

    @rx.event
    def update_kit_component_field(self, index: int, field: str, value: Any):
        if index < 0 or index >= len(self.kit_components):
            return
        comps = list(self.kit_components)
        row = dict(comps[index])
        if field == "quantity":
            try:
                row[field] = float(value) if value not in ("", None) else 1.0
            except (TypeError, ValueError):
                return
        else:
            row[field] = value
        comps[index] = row
        self.kit_components = comps

    @rx.event
    def resolve_kit_component(self, index: int, barcode: str):
        """Busca un producto por código de barras y lo asigna como componente del kit."""
        if index < 0 or index >= len(self.kit_components):
            return
        code = (barcode or "").strip()
        if not code:
            return
        company_id = self._company_id()
        branch_id = self._branch_id()
        if not company_id or not branch_id:
            return
        with rx.session() as session:
            p = session.exec(
                select(Product)
                .where(Product.barcode == code)
                .where(Product.company_id == company_id)
                .where(Product.branch_id == branch_id)
            ).first()
            if not p:
                return rx.toast(f"Producto con código '{code}' no encontrado.", duration=3000)
            comps = list(self.kit_components)
            row = dict(comps[index])
            row["component_barcode"] = p.barcode or ""
            row["component_name"] = p.description or ""
            row["component_product_id"] = p.id
            comps[index] = row
            self.kit_components = comps

    @rx.var(cache=True)
    def kit_component_rows(self) -> List[Dict[str, Any]]:
        rows: List[Dict[str, Any]] = []
        for index, comp in enumerate(self.kit_components):
            row = dict(comp)
            row["index"] = index
            rows.append(row)
        return rows

    @rx.var(cache=True)
    def batch_rows(self) -> List[Dict[str, Any]]:
        rows: List[Dict[str, Any]] = []
        for index, batch in enumerate(self.batches):
            row = dict(batch)
            row["index"] = index
            rows.append(row)
        return rows

    @rx.var(cache=True)
    def attribute_rows(self) -> List[Dict[str, Any]]:
        rows: List[Dict[str, Any]] = []
        for index, attr in enumerate(self.attributes):
            row = dict(attr)
            row["index"] = index
            rows.append(row)
        return rows

    @rx.var(cache=True)
    def batches_stock_total(self) -> float:
        total = 0.0
        for batch in self.batches:
            try:
                total += float(batch.get("stock", 0) or 0)
            except (TypeError, ValueError):
                continue
        return total

    @rx.event
    def save_edited_product(self):
        if not self.current_user["privileges"]["edit_inventario"]:
            return self.add_notification(
                "No tiene permisos para editar el inventario.", "error"
            )
        block = self._require_active_subscription()
        if block:
            return block

        product_data = self.editing_product
        product_id = product_data.get("id")
        barcode = product_data.get("barcode", "").strip()
        description = product_data.get("description", "").strip()
        company_id = self._company_id()
        branch_id = self._branch_id()
        if not company_id or not branch_id:
            return self.add_notification("Empresa no definida.", "error")

        if not description:
            return self.add_notification(
                "La descripción no puede estar vacía.", "error"
            )
        if not barcode:
            return self.add_notification(
                "El código de barras no puede estar vacío.", "error"
            )

        if self.show_wholesale and self.price_tiers:
            min_qtys = []
            for tier in self.price_tiers:
                mq = tier.get("min_qty", 0) or 0
                try:
                    mq = int(mq)
                except (TypeError, ValueError):
                    mq = 0
                if mq > 0:
                    if mq in min_qtys:
                        return self.add_notification(
                            f"Hay cantidades mínimas duplicadas ({mq}) en las escalas de mayoreo.",
                            "error",
                        )
                    min_qtys.append(mq)

        if self.show_variants:
            skus = [
                (variant.get("sku") or "").strip()
                for variant in self.variants
                if (variant.get("sku") or "").strip()
            ]
            if not skus:
                return self.add_notification(
                    "Agregue al menos una variante con SKU válido.", "error"
                )
            if len(skus) != len(set(skus)):
                return self.add_notification(
                    "Hay SKUs de variantes duplicados.", "error"
                )

        if self.show_batches:
            batch_numbers = [
                (b.get("batch_number") or "").strip()
                for b in self.batches
                if (b.get("batch_number") or "").strip()
            ]
            if not batch_numbers:
                return self.add_notification(
                    "Agregue al menos un lote con número válido.", "error"
                )
            if len(batch_numbers) != len(set(batch_numbers)):
                return self.add_notification(
                    "Hay números de lote duplicados.", "error"
                )
            for b in self.batches:
                exp = (b.get("expiration_date") or "").strip()
                if exp:
                    try:
                        datetime.datetime.strptime(exp, "%Y-%m-%d")
                    except ValueError:
                        return self.add_notification(
                            f"Fecha de vencimiento inválida: '{exp}' (use YYYY-MM-DD).",
                            "error",
                        )

        if self.show_attributes:
            names = [
                (a.get("name") or "").strip().lower()
                for a in self.attributes
                if (a.get("name") or "").strip()
            ]
            if not names:
                return self.add_notification(
                    "Agregue al menos un atributo con nombre válido.", "error"
                )
            if len(names) != len(set(names)):
                return self.add_notification(
                    "Hay nombres de atributo duplicados.", "error"
                )

        if self.show_kit_components:
            comp_ids = [
                c.get("component_product_id")
                for c in self.kit_components
                if c.get("component_product_id")
            ]
            if not comp_ids:
                return self.add_notification(
                    "Agregue al menos un componente válido al kit.", "error"
                )
            if len(comp_ids) != len(set(comp_ids)):
                return self.add_notification(
                    "Hay componentes duplicados en el kit.", "error"
                )

        self.is_loading = True
        yield

        msg = ""
        try:
            with rx.session() as session:
                # Verificar codigo de barras duplicado
                existing = session.exec(
                    select(Product)
                    .where(Product.barcode == barcode)
                    .where(Product.company_id == company_id)
                    .where(Product.branch_id == branch_id)
                ).first()

                if existing and (product_id is None or existing.id != product_id):
                    return self.add_notification(
                        "Ya existe un producto con ese código de barras.",
                        "error",
                    )

                if self.show_variants:
                    variant_skus = [
                        (variant.get("sku") or "").strip()
                        for variant in self.variants
                        if (variant.get("sku") or "").strip()
                    ]
                    if variant_skus:
                        duplicate_query = (
                            select(ProductVariant)
                            .where(ProductVariant.company_id == company_id)
                            .where(ProductVariant.branch_id == branch_id)
                            .where(ProductVariant.sku.in_(variant_skus))
                        )
                        if product_id:
                            duplicate_query = duplicate_query.where(
                                ProductVariant.product_id != product_id
                            )
                        duplicate_variant = session.exec(duplicate_query).first()
                        if duplicate_variant:
                            return self.add_notification(
                                f"El SKU de variante '{duplicate_variant.sku}' ya existe en otro producto.",
                                "error",
                            )

                if product_id:
                    # Actualizar
                    # FIX 39b: with_for_update to prevent TOCTOU on stock/price
                    product = session.exec(
                        select(Product)
                        .where(Product.id == product_id)
                        .where(Product.company_id == company_id)
                        .where(Product.branch_id == branch_id)
                        .with_for_update()
                    ).first()
                    if not product:
                        return self.add_notification("Producto no encontrado.", "error")

                    product.barcode = barcode
                    product.description = description
                    product.category = product_data.get("category", "General")
                    product.stock = (
                        self.variants_stock_total
                        if self.show_variants
                        else product_data.get("stock", 0)
                    )
                    product.unit = product_data.get("unit", "Unidad")
                    product.purchase_price = product_data.get("purchase_price", 0)
                    product.sale_price = product_data.get("sale_price", 0)

                    session.add(product)
                    msg = "Producto actualizado correctamente."
                else:
                    # Crear
                    new_product = Product(
                        barcode=barcode,
                        description=description,
                        category=product_data.get("category", "General"),
                        stock=(
                            self.variants_stock_total
                            if self.show_variants
                            else product_data.get("stock", 0)
                        ),
                        unit=product_data.get("unit", "Unidad"),
                        purchase_price=product_data.get("purchase_price", 0),
                        sale_price=product_data.get("sale_price", 0),
                        company_id=company_id,
                        branch_id=branch_id,
                    )
                    session.add(new_product)
                    session.flush()
                    product = new_product
                    msg = "Producto creado correctamente."

                if product_id:
                    session.flush()

                product_id = product.id if product else None

                if product_id:
                    if self.show_variants:
                        existing_variants = session.exec(
                            select(ProductVariant)
                            .where(ProductVariant.product_id == product_id)
                            .where(ProductVariant.company_id == company_id)
                            .where(ProductVariant.branch_id == branch_id)
                        ).all()
                        for variant in existing_variants:
                            session.delete(variant)
                        session.flush()

                        for variant in self.variants:
                            sku = (variant.get("sku") or "").strip()
                            if not sku:
                                continue
                            stock_value = variant.get("stock", 0) or 0
                            try:
                                stock_value = Decimal(str(stock_value))
                            except (TypeError, InvalidOperation):
                                stock_value = Decimal("0")
                            session.add(
                                ProductVariant(
                                    product_id=product_id,
                                    sku=sku,
                                    size=(variant.get("size") or "").strip() or None,
                                    color=(variant.get("color") or "").strip() or None,
                                    stock=stock_value,
                                    company_id=company_id,
                                    branch_id=branch_id,
                                )
                            )
                    else:
                        existing_variants = session.exec(
                            select(ProductVariant)
                            .where(ProductVariant.product_id == product_id)
                            .where(ProductVariant.company_id == company_id)
                            .where(ProductVariant.branch_id == branch_id)
                        ).all()
                        for variant in existing_variants:
                            session.delete(variant)

                    if self.show_wholesale:
                        existing_tiers = session.exec(
                            select(PriceTier)
                            .where(PriceTier.product_id == product_id)
                            .where(PriceTier.company_id == company_id)
                            .where(PriceTier.branch_id == branch_id)
                        ).all()
                        for tier in existing_tiers:
                            session.delete(tier)

                        for tier in self.price_tiers:
                            min_qty = tier.get("min_qty", 0) or 0
                            price_value = tier.get("price", 0) or 0
                            try:
                                min_qty = int(min_qty)
                            except (TypeError, ValueError):
                                min_qty = 0
                            try:
                                price_value = Decimal(str(price_value))
                            except (TypeError, InvalidOperation):
                                price_value = Decimal("0")
                            if min_qty <= 0:
                                continue
                            session.add(
                                PriceTier(
                                    product_id=product_id,
                                    min_quantity=min_qty,
                                    unit_price=price_value,
                                    company_id=company_id,
                                    branch_id=branch_id,
                                )
                            )
                    else:
                        existing_tiers = session.exec(
                            select(PriceTier)
                            .where(PriceTier.product_id == product_id)
                            .where(PriceTier.company_id == company_id)
                            .where(PriceTier.branch_id == branch_id)
                        ).all()
                        for tier in existing_tiers:
                            session.delete(tier)

                    # ─── Persistir lotes (batches) ───
                    if self.show_batches:
                        existing_batches = session.exec(
                            select(ProductBatch)
                            .where(ProductBatch.product_id == product_id)
                            .where(ProductBatch.product_variant_id.is_(None))
                            .where(ProductBatch.company_id == company_id)
                            .where(ProductBatch.branch_id == branch_id)
                        ).all()
                        for batch in existing_batches:
                            session.delete(batch)
                        session.flush()

                        for batch in self.batches:
                            batch_number = (batch.get("batch_number") or "").strip()
                            if not batch_number:
                                continue
                            stock_value = batch.get("stock", 0) or 0
                            try:
                                stock_value = Decimal(str(stock_value))
                            except (TypeError, InvalidOperation):
                                stock_value = Decimal("0")
                            expiration_raw = (batch.get("expiration_date") or "").strip()
                            expiration_dt = None
                            if expiration_raw:
                                try:
                                    expiration_dt = datetime.datetime.strptime(
                                        expiration_raw, "%Y-%m-%d"
                                    )
                                except ValueError:
                                    expiration_dt = None
                            session.add(
                                ProductBatch(
                                    product_id=product_id,
                                    batch_number=batch_number,
                                    expiration_date=expiration_dt,
                                    stock=stock_value,
                                    company_id=company_id,
                                    branch_id=branch_id,
                                )
                            )
                    else:
                        existing_batches = session.exec(
                            select(ProductBatch)
                            .where(ProductBatch.product_id == product_id)
                            .where(ProductBatch.product_variant_id.is_(None))
                            .where(ProductBatch.company_id == company_id)
                            .where(ProductBatch.branch_id == branch_id)
                        ).all()
                        for batch in existing_batches:
                            session.delete(batch)

                    # ─── Persistir atributos dinámicos (EAV) ───
                    if self.show_attributes:
                        existing_attrs = session.exec(
                            select(ProductAttribute)
                            .where(ProductAttribute.product_id == product_id)
                            .where(ProductAttribute.company_id == company_id)
                            .where(ProductAttribute.branch_id == branch_id)
                        ).all()
                        for attr in existing_attrs:
                            session.delete(attr)
                        session.flush()

                        for attr in self.attributes:
                            name = (attr.get("name") or "").strip()
                            value = (attr.get("value") or "").strip()
                            if not name:
                                continue
                            session.add(
                                ProductAttribute(
                                    product_id=product_id,
                                    attribute_name=name,
                                    attribute_value=value,
                                    company_id=company_id,
                                    branch_id=branch_id,
                                )
                            )
                    else:
                        existing_attrs = session.exec(
                            select(ProductAttribute)
                            .where(ProductAttribute.product_id == product_id)
                            .where(ProductAttribute.company_id == company_id)
                            .where(ProductAttribute.branch_id == branch_id)
                        ).all()
                        for attr in existing_attrs:
                            session.delete(attr)

                    # ─── Persistir componentes de kit ───
                    if self.show_kit_components:
                        existing_kit = session.exec(
                            select(ProductKit)
                            .where(ProductKit.kit_product_id == product_id)
                            .where(ProductKit.company_id == company_id)
                            .where(ProductKit.branch_id == branch_id)
                        ).all()
                        for k in existing_kit:
                            session.delete(k)
                        session.flush()

                        for comp in self.kit_components:
                            comp_pid = comp.get("component_product_id")
                            if not comp_pid:
                                continue
                            qty_value = comp.get("quantity", 1) or 1
                            try:
                                qty_value = Decimal(str(qty_value))
                            except (TypeError, InvalidOperation):
                                qty_value = Decimal("1")
                            session.add(
                                ProductKit(
                                    kit_product_id=product_id,
                                    component_product_id=int(comp_pid),
                                    quantity=qty_value,
                                    company_id=company_id,
                                    branch_id=branch_id,
                                )
                            )
                    else:
                        existing_kit = session.exec(
                            select(ProductKit)
                            .where(ProductKit.kit_product_id == product_id)
                            .where(ProductKit.company_id == company_id)
                            .where(ProductKit.branch_id == branch_id)
                        ).all()
                        for k in existing_kit:
                            session.delete(k)

                session.commit()
        except Exception:
            logger.exception(
                "save_edited_product failed | company=%s branch=%s",
                self._company_id(),
                self._branch_id(),
            )
            return self.add_notification(
                "No se pudo guardar el producto.", "error"
            )
        finally:
            self.is_loading = False

        self._inventory_update_trigger += 1
        self._refresh_inventory_cache()
        self.is_editing_product = False
        self.show_variants = False
        self.show_wholesale = False
        self.show_batches = False
        self.show_attributes = False
        self.show_kit_components = False
        self.variants = []
        self.price_tiers = []
        self.batches = []
        self.attributes = []
        self.kit_components = []
        return self.add_notification(msg, "success")

    @rx.event
    def delete_product(self, product_id: int):
        if not self.current_user["privileges"]["edit_inventario"]:
            return self.add_notification(
                "No tiene permisos para eliminar productos.", "error"
            )
        block = self._require_active_subscription()
        if block:
            return block
        company_id = self._company_id()
        branch_id = self._branch_id()
        if not company_id or not branch_id:
            return self.add_notification("Empresa no definida.", "error")

        self.is_loading = True
        yield
        deleted = False

        try:
            with rx.session() as session:
                product = session.exec(
                    select(Product)
                    .where(Product.id == product_id)
                    .where(Product.company_id == company_id)
                    .where(Product.branch_id == branch_id)
                ).first()
                if product:
                    # Verificar si tiene historial de ventas
                    # FIX 38a: add company_id filter for tenant isolation
                    has_sales = session.exec(
                        select(SaleItem)
                        .where(SaleItem.product_id == product_id)
                        .where(SaleItem.company_id == company_id)
                        .where(SaleItem.branch_id == branch_id)
                    ).first()
                    if has_sales:
                        return self.add_notification(
                            "No se puede eliminar un producto con historial de ventas. Edítelo para desactivarlo.",
                            "warning",
                        )
                    session.delete(product)
                    session.commit()
                    self._inventory_update_trigger += 1
                    deleted = True
                else:
                    return self.add_notification("Producto no encontrado.", "error")
        finally:
            self.is_loading = False

        if deleted:
            self._refresh_inventory_cache()
            return self.add_notification("Producto eliminado.", "success")

    @rx.event
    def open_inventory_check_modal(self):
        if not self.current_user["privileges"]["edit_inventario"]:
            return rx.toast(
                "No tiene permisos para registrar inventario.", duration=3000
            )
        self.inventory_check_modal_open = True
        self.inventory_check_status = "perfecto"
        self.inventory_adjustment_notes = ""
        self.inventory_adjustment_items = []
        self._reset_inventory_adjustment_form()

    @rx.event
    def close_inventory_check_modal(self):
        self.inventory_check_modal_open = False
        self.inventory_check_status = "perfecto"
        self.inventory_adjustment_notes = ""
        self.inventory_adjustment_items = []
        self._reset_inventory_adjustment_form()

    @rx.event
    def set_inventory_check_status(self, status: str):
        if status not in ["perfecto", "ajuste"]:
            return
        self.inventory_check_status = status
        if status == "perfecto":
            self.inventory_adjustment_items = []
            self._reset_inventory_adjustment_form()

    @rx.event
    def set_inventory_adjustment_notes(self, notes: str):
        self.inventory_adjustment_notes = notes

    def _reset_inventory_adjustment_form(self):
        self.inventory_adjustment_item = {
            "temp_id": "",
            "barcode": "",
            "description": "",
            "category": "",
            "unit": "",
            "current_stock": 0,
            "adjust_quantity": 0,
            "reason": "",
            "product_id": None,
            "variant_id": None,
        }
        self.inventory_adjustment_suggestions = []

    def _variant_label(self, variant: ProductVariant) -> str:
        parts: List[str] = []
        if variant.size:
            parts.append(str(variant.size).strip())
        if variant.color:
            parts.append(str(variant.color).strip())
        label = " ".join([p for p in parts if p])
        sku = (variant.sku or "").strip()
        if label and sku:
            return f"{label} ({sku})"
        return label or sku or "Variante"

    def _fill_inventory_adjustment_from_product(
        self, product: Product, variant: ProductVariant | None = None
    ):
        def _field(obj: Any, name: str, default: Any = "") -> Any:
            if isinstance(obj, dict):
                return obj.get(name, default)
            return getattr(obj, name, default)

        barcode = variant.sku if variant else _field(product, "barcode", "")
        description = _field(product, "description", "")
        if variant:
            label = self._variant_label(variant)
            if label:
                description = f"{description} ({label})"
        self.inventory_adjustment_item["barcode"] = barcode or ""
        self.inventory_adjustment_item["description"] = description or ""
        self.inventory_adjustment_item["category"] = _field(product, "category", "")
        self.inventory_adjustment_item["unit"] = _field(product, "unit", "Unidad")
        self.inventory_adjustment_item["current_stock"] = (
            variant.stock if variant else _field(product, "stock", 0)
        )
        self.inventory_adjustment_item["product_id"] = _field(product, "id", None)
        self.inventory_adjustment_item["variant_id"] = variant.id if variant else None
        self.inventory_adjustment_item["adjust_quantity"] = 0
        self.inventory_adjustment_item["reason"] = ""

    def _product_has_variants(
        self,
        session,
        product_id: int,
        company_id: int,
        branch_id: int,
    ) -> bool:
        if not product_id:
            return False
        return (
            session.exec(
                select(ProductVariant.id)
                .where(ProductVariant.product_id == product_id)
                .where(ProductVariant.company_id == company_id)
                .where(ProductVariant.branch_id == branch_id)
                .limit(1)
            ).first()
            is not None
        )

    def _get_batches_for_adjustment(
        self,
        session,
        product: Product,
        variant: ProductVariant | None,
        company_id: int,
        branch_id: int,
        lock: bool = False,
    ) -> list[ProductBatch]:
        if variant:
            query = select(ProductBatch).where(
                ProductBatch.product_variant_id == variant.id
            )
        else:
            query = (
                select(ProductBatch)
                .where(ProductBatch.product_id == product.id)
                .where(ProductBatch.product_variant_id.is_(None))
            )
        query = (
            query.where(ProductBatch.company_id == company_id)
            .where(ProductBatch.branch_id == branch_id)
            .order_by(
                ProductBatch.expiration_date.is_(None),
                ProductBatch.expiration_date.asc(),
                ProductBatch.id.asc(),
            )
        )
        if lock:
            query = query.with_for_update()
        return session.exec(query).all()

    def _find_adjustment_product(
        self,
        session,
        barcode: str,
        description: str,
        variant_id: int | None = None,
        product_id: int | None = None,
        for_update: bool = False,
    ) -> tuple[Product | None, ProductVariant | None]:
        company_id = self._company_id()
        branch_id = self._branch_id()
        if not company_id or not branch_id:
            return None, None
        def _maybe_lock(statement):
            return statement.with_for_update() if for_update else statement
        if variant_id:
            variant = session.exec(
                _maybe_lock(
                    select(ProductVariant)
                    .where(ProductVariant.id == variant_id)
                    .where(ProductVariant.company_id == company_id)
                    .where(ProductVariant.branch_id == branch_id)
                )
            ).first()
            if variant:
                product = session.exec(
                    _maybe_lock(
                        select(Product)
                        .where(Product.id == variant.product_id)
                        .where(Product.company_id == company_id)
                        .where(Product.branch_id == branch_id)
                    )
                ).first()
                return product, variant
        if product_id:
            product = session.exec(
                _maybe_lock(
                    select(Product)
                    .where(Product.id == product_id)
                    .where(Product.company_id == company_id)
                    .where(Product.branch_id == branch_id)
                )
            ).first()
            if product:
                return product, None
        if barcode:
            variant = session.exec(
                _maybe_lock(
                    select(ProductVariant)
                    .where(ProductVariant.sku == barcode)
                    .where(ProductVariant.company_id == company_id)
                    .where(ProductVariant.branch_id == branch_id)
                )
            ).first()
            if variant:
                product = session.exec(
                    _maybe_lock(
                        select(Product)
                        .where(Product.id == variant.product_id)
                        .where(Product.company_id == company_id)
                        .where(Product.branch_id == branch_id)
                    )
                ).first()
                return product, variant
            product = session.exec(
                _maybe_lock(
                    select(Product)
                    .where(Product.barcode == barcode)
                    .where(Product.company_id == company_id)
                    .where(Product.branch_id == branch_id)
                )
            ).first()
            if product:
                return product, None
        if description:
            product = session.exec(
                _maybe_lock(
                    select(Product)
                    .where(Product.description == description)
                    .where(Product.company_id == company_id)
                    .where(Product.branch_id == branch_id)
                )
            ).first()
            if product:
                return product, None
        return None, None

    def _process_inventory_adjustment_search(self, value: Any):
        term = str(value or "").strip()
        self.inventory_adjustment_item["description"] = term
        if not term:
            self.inventory_adjustment_suggestions = []
            return

        code = clean_barcode(term)
        if validate_barcode(code):
            with rx.session() as session:
                product, variant = self._find_adjustment_product(
                    session, code, "", None, None
                )
                if product:
                    self._fill_inventory_adjustment_from_product(product, variant)
                    self.inventory_adjustment_suggestions = []
                    return

        search_term = term.lower()
        if len(search_term) < 2:
            self.inventory_adjustment_suggestions = []
            return
        company_id = self._company_id()
        branch_id = self._branch_id()
        if not company_id or not branch_id:
            self.inventory_adjustment_suggestions = []
            return
        with rx.session() as session:
            search = f"%{escape_like(search_term)}%"
            products = session.exec(
                select(Product)
                .where(
                    or_(
                        Product.description.ilike(search),
                        Product.barcode.ilike(search),
                    )
                )
                .where(Product.company_id == company_id)
                .where(Product.branch_id == branch_id)
                .limit(8)
            ).all()
            variant_rows = session.exec(
                select(ProductVariant, Product)
                .join(Product, ProductVariant.product_id == Product.id)
                .where(
                    or_(
                        ProductVariant.sku.ilike(search),
                        ProductVariant.size.ilike(search),
                        ProductVariant.color.ilike(search),
                        Product.description.ilike(search),
                    )
                )
                .where(ProductVariant.company_id == company_id)
                .where(ProductVariant.branch_id == branch_id)
                .limit(8)
            ).all()

        suggestions: list[dict] = []
        for product in products:
            suggestions.append(
                {
                    "label": product.description,
                    "kind": "product",
                    "product_id": product.id,
                    "variant_id": None,
                }
            )
        for variant, parent in variant_rows:
            label = self._variant_label(variant)
            full_label = parent.description
            if label:
                full_label = f"{parent.description} ({label})"
            suggestions.append(
                {
                    "label": full_label,
                    "kind": "variant",
                    "product_id": parent.id,
                    "variant_id": variant.id,
                }
            )
        self.inventory_adjustment_suggestions = suggestions

    @rx.event
    def select_inventory_adjustment_product(self, description: Any):
        variant_id = None
        product_id = None
        if isinstance(description, dict):
            variant_id = description.get("variant_id")
            product_id = description.get("product_id")
            description = (
                description.get("value")
                or description.get("description")
                or description.get("label")
                or ""
            )
        description = str(description or "").strip()
        company_id = self._company_id()
        branch_id = self._branch_id()
        if not company_id or not branch_id:
            return

        code = clean_barcode(description)
        barcode = code if validate_barcode(code) else ""
        with rx.session() as session:
            product, variant = self._find_adjustment_product(
                session,
                barcode,
                description,
                int(variant_id) if variant_id else None,
                int(product_id) if product_id else None,
            )
            if product:
                self._fill_inventory_adjustment_from_product(product, variant)
        self.inventory_adjustment_suggestions = []

    @rx.event
    def set_inventory_adjustment_item_barcode(self, value: str):
        self.inventory_adjustment_item["barcode"] = value

    @rx.event
    def set_inventory_adjustment_item_quantity(self, value: str):
        try:
            self.inventory_adjustment_item["adjust_quantity"] = float(value)
        except ValueError:
            self.inventory_adjustment_item["adjust_quantity"] = 0

    @rx.event
    def set_inventory_adjustment_item_reason(self, value: str):
        self.inventory_adjustment_item["reason"] = value

    @rx.event
    def add_inventory_adjustment_item(self):
        if not self.current_user["privileges"]["edit_inventario"]:
            return rx.toast("No tiene permisos para ajustar inventario.", duration=3000)
        description = self.inventory_adjustment_item["description"].strip()
        barcode = (self.inventory_adjustment_item.get("barcode") or "").strip()
        if not description and not barcode:
            return rx.toast("Seleccione un producto para ajustar.", duration=3000)
        company_id = self._company_id()
        branch_id = self._branch_id()
        if not company_id or not branch_id:
            return rx.toast("Empresa no definida.", duration=3000)

        with rx.session() as session:
            if description and not barcode:
                duplicate_count = session.exec(
                    select(func.count(Product.id))
                    .where(Product.description == description)
                    .where(Product.company_id == company_id)
                    .where(Product.branch_id == branch_id)
                ).one()
                if duplicate_count and duplicate_count > 1:
                    return rx.toast(
                        "Descripcion duplicada en inventario. Use codigo de barras.",
                        duration=3000,
                    )
            product, variant = self._find_adjustment_product(
                session,
                barcode,
                description,
                self.inventory_adjustment_item.get("variant_id"),
                self.inventory_adjustment_item.get("product_id"),
            )
            if not product:
                return rx.toast("Producto no encontrado en el inventario.", duration=3000)

            if not variant and self._product_has_variants(
                session, product.id, company_id, branch_id
            ):
                return rx.toast(
                    "Producto con variantes. Seleccione la variante a ajustar.",
                    duration=3500,
                )

            try:
                quantity = Decimal(
                    str(self.inventory_adjustment_item.get("adjust_quantity", 0) or 0)
                )
            except (InvalidOperation, TypeError, ValueError):
                quantity = Decimal("0")
            if quantity <= 0:
                return rx.toast("Ingrese la cantidad a ajustar.", duration=3000)

            batches = self._get_batches_for_adjustment(
                session, product, variant, company_id, branch_id
            )
            if batches:
                available = sum(
                    (Decimal(str(batch.stock or 0)) for batch in batches),
                    Decimal("0"),
                )
            else:
                available = (
                    Decimal(str(variant.stock or 0))
                    if variant
                    else Decimal(str(product.stock or 0))
                )
            if quantity > available:
                return rx.toast(
                    "La cantidad supera el stock disponible.", duration=3000
                )

            item_copy = self.inventory_adjustment_item.copy()
            item_copy["temp_id"] = str(uuid.uuid4())
            item_copy["product_id"] = product.id
            item_copy["variant_id"] = variant.id if variant else None
            item_copy["adjust_quantity"] = self._normalize_quantity_value(
                item_copy.get("adjust_quantity", 0), item_copy.get("unit", "")
            )
            # Asegurar que la unidad se tome del producto si falta
            if not item_copy.get("unit"):
                item_copy["unit"] = product.unit
            if not item_copy.get("barcode"):
                item_copy["barcode"] = variant.sku if variant else product.barcode
            if not item_copy.get("description"):
                label = self._variant_label(variant) if variant else ""
                item_copy["description"] = (
                    f"{product.description} ({label})" if label else product.description
                )

            self.inventory_adjustment_items.append(item_copy)
            self._reset_inventory_adjustment_form()

    @rx.event
    def remove_inventory_adjustment_item(self, temp_id: str):
        self.inventory_adjustment_items = [
            item for item in self.inventory_adjustment_items if item["temp_id"] != temp_id
        ]

    @rx.event
    def submit_inventory_check(self):
        if not self.current_user["privileges"]["edit_inventario"]:
            return self.add_notification(
                "No tiene permisos para registrar inventario.", "error"
            )
        block = self._require_active_subscription()
        if block:
            return block
        status = (
            self.inventory_check_status
            if self.inventory_check_status in ["perfecto", "ajuste"]
            else "perfecto"
        )
        notes = self.inventory_adjustment_notes.strip()
        company_id = self._company_id()
        branch_id = self._branch_id()
        if not company_id or not branch_id:
            return self.add_notification("Empresa no definida.", "error")

        self.is_loading = True
        yield
        refresh_needed = False

        success_message = "Registro de inventario guardado."
        try:
            if status == "perfecto":
                success_message = "Inventario verificado como perfecto."
            else:
                if not self.inventory_adjustment_items:
                    return self.add_notification(
                        "Agregue los productos que requieren re ajuste.", "error"
                    )

                recorded = False
                def _to_decimal(value: Any, default: Decimal = Decimal("0")) -> Decimal:
                    try:
                        return Decimal(str(value))
                    except (InvalidOperation, TypeError, ValueError):
                        return default

                with rx.session() as session:
                    products_to_recalculate: set[int] = set()
                    products_recalc_batches: set[int] = set()
                    variants_recalc_batches: set[int] = set()

                    # ── Batch pre-load: reduce N+1 a queries fijas ──
                    _items = self.inventory_adjustment_items
                    _product_ids: set[int] = set()
                    _variant_ids: set[int] = set()
                    for _it in _items:
                        _pid = _it.get("product_id")
                        _vid = _it.get("variant_id")
                        if _pid:
                            _product_ids.add(int(_pid))
                        if _vid:
                            _variant_ids.add(int(_vid))

                    # Query 1: Productos (con lock)
                    _products_map: dict[int, Product] = {}
                    if _product_ids:
                        _prods = session.exec(
                            select(Product)
                            .where(Product.id.in_(_product_ids))
                            .where(Product.company_id == company_id)
                            .where(Product.branch_id == branch_id)
                            .with_for_update()
                        ).all()
                        _products_map = {p.id: p for p in _prods}

                    # Query 2: Variantes (con lock)
                    _variants_map: dict[int, ProductVariant] = {}
                    if _variant_ids:
                        _vars = session.exec(
                            select(ProductVariant)
                            .where(ProductVariant.id.in_(_variant_ids))
                            .where(ProductVariant.company_id == company_id)
                            .where(ProductVariant.branch_id == branch_id)
                            .with_for_update()
                        ).all()
                        _variants_map = {v.id: v for v in _vars}

                    # Query 3: Batches de variantes + productos (con lock)
                    _variant_batches: dict[int, list[ProductBatch]] = {}
                    _product_batches: dict[int, list[ProductBatch]] = {}
                    _batch_conditions = []
                    if _variant_ids:
                        _batch_conditions.append(
                            ProductBatch.product_variant_id.in_(_variant_ids)
                        )
                    if _product_ids:
                        _batch_conditions.append(
                            and_(
                                ProductBatch.product_id.in_(_product_ids),
                                ProductBatch.product_variant_id.is_(None),
                            )
                        )
                    if _batch_conditions:
                        _all_batches = session.exec(
                            select(ProductBatch)
                            .where(or_(*_batch_conditions))
                            .where(ProductBatch.company_id == company_id)
                            .where(ProductBatch.branch_id == branch_id)
                            .order_by(
                                ProductBatch.expiration_date.is_(None),
                                ProductBatch.expiration_date.asc(),
                                ProductBatch.id.asc(),
                            )
                            .with_for_update()
                        ).all()
                        for _b in _all_batches:
                            if _b.product_variant_id:
                                _variant_batches.setdefault(
                                    _b.product_variant_id, []
                                ).append(_b)
                            else:
                                _product_batches.setdefault(
                                    _b.product_id, []
                                ).append(_b)

                    # Query 4: Productos que tienen variantes (para skip)
                    _products_with_variants: set[int] = set()
                    if _product_ids:
                        _has_vars = session.exec(
                            select(ProductVariant.product_id)
                            .where(ProductVariant.product_id.in_(_product_ids))
                            .where(ProductVariant.company_id == company_id)
                            .where(ProductVariant.branch_id == branch_id)
                            .distinct()
                        ).all()
                        _products_with_variants = set(_has_vars)

                    for item in _items:
                        description = (item.get("description") or "").strip()
                        barcode = (item.get("barcode") or "").strip()
                        if not description and not barcode:
                            continue

                        # Lookup desde maps pre-cargados
                        _pid = item.get("product_id")
                        _vid = item.get("variant_id")
                        product = _products_map.get(int(_pid)) if _pid else None
                        variant = _variants_map.get(int(_vid)) if _vid else None

                        # Fallback a búsqueda individual (raro: item sin IDs)
                        if not product:
                            product, variant = self._find_adjustment_product(
                                session, barcode, description,
                                _vid, _pid, for_update=True,
                            )
                        if not product:
                            continue
                        if not variant and product.id in _products_with_variants:
                            continue

                        quantity = _to_decimal(item.get("adjust_quantity", 0) or 0)
                        if quantity <= 0:
                            continue

                        unit = product.unit or item.get("unit") or ""
                        # Batches desde maps pre-cargados
                        if variant:
                            batches = _variant_batches.get(variant.id, [])
                        else:
                            batches = _product_batches.get(product.id, [])
                        if batches:
                            available = sum(
                                (_to_decimal(batch.stock, Decimal("0")) for batch in batches),
                                Decimal("0"),
                            )
                            qty = quantity if quantity <= available else available
                            if qty <= 0:
                                continue
                            remaining = qty
                            for batch in batches:
                                if remaining <= 0:
                                    break
                                current_stock = _to_decimal(batch.stock, Decimal("0"))
                                if current_stock <= 0:
                                    continue
                                deduct = remaining if remaining <= current_stock else current_stock
                                batch.stock = current_stock - deduct
                                session.add(batch)
                                remaining -= deduct
                            total_after = sum(
                                (_to_decimal(batch.stock, Decimal("0")) for batch in batches),
                                Decimal("0"),
                            )
                            if variant:
                                variant.stock = self._normalize_quantity_value(
                                    total_after, unit
                                )
                                session.add(variant)
                                variants_recalc_batches.add(variant.id)
                                products_to_recalculate.add(product.id)
                            else:
                                product.stock = self._normalize_quantity_value(
                                    total_after, unit
                                )
                                session.add(product)
                                products_recalc_batches.add(product.id)
                        else:
                            available = _to_decimal(
                                (variant.stock if variant else product.stock) or 0
                            )
                            qty = quantity if quantity <= available else available
                            if qty <= 0:
                                continue

                            # Actualizar stock
                            if variant:
                                variant.stock = max(available - qty, Decimal("0"))
                                session.add(variant)
                                products_to_recalculate.add(product.id)
                            else:
                                product.stock = max(available - qty, Decimal("0"))
                                session.add(product)

                        # Crear StockMovement
                        detail_parts = []
                        if item.get("reason"):
                            detail_parts.append(item["reason"])
                        if notes:
                            detail_parts.append(notes)
                        details = (
                            " | ".join(part for part in detail_parts if part)
                            if detail_parts
                            else "Ajuste inventario"
                        )

                        movement = StockMovement(
                            product_id=product.id,
                            user_id=self.current_user.get("id"),
                            type="Re Ajuste Inventario",
                            quantity=-qty,
                            description=details,
                            timestamp=self._utc_now(),
                            company_id=company_id,
                            branch_id=branch_id,
                        )
                        session.add(movement)
                        recorded = True

                    # Recalcular totales de stock (3 fases) usando helper compartido
                    recalculate_stock_totals(
                        session=session,
                        company_id=company_id,
                        branch_id=branch_id,
                        variants_from_batches=variants_recalc_batches,
                        products_from_variants=products_to_recalculate,
                        products_from_batches=products_recalc_batches,
                        normalize_fn=lambda total, prod: self._normalize_quantity_value(
                            total, prod.unit or ""
                        ),
                    )

                    if recorded:
                        session.commit()
                        self._inventory_update_trigger += 1
                        refresh_needed = True
                    else:
                        return self.add_notification(
                            "No se pudo registrar el re ajuste. Verifique los productos.",
                            "error",
                        )
        finally:
            self.is_loading = False

        self.inventory_check_modal_open = False
        self.inventory_check_status = "perfecto"
        self.inventory_adjustment_notes = ""
        self.inventory_adjustment_items = []
        self._reset_inventory_adjustment_form()
        if refresh_needed:
            self._refresh_inventory_cache()
        return self.add_notification(success_message, "success")

    @rx.event
    def export_inventory_to_excel(self):
        if not self.current_user["privileges"]["export_data"]:
            return rx.toast("No tiene permisos para exportar datos.", duration=3000)
        company_id = self._company_id()
        branch_id = self._branch_id()
        if not company_id or not branch_id:
            return rx.toast("Empresa no definida.", duration=3000)

        currency_label = self._currency_symbol_clean()
        currency_format = self._currency_excel_format()

        company_name = getattr(self, "company_name", "") or "EMPRESA"
        today = self._display_now().strftime("%d/%m/%Y")

        wb, ws = create_excel_workbook("Inventario Valorizado")

        # Encabezado profesional
        row = add_company_header(
            ws,
            company_name,
            "INVENTARIO VALORIZADO ACTUAL",
            f"Al {today}",
            columns=12,
            generated_at=self._display_now(),
        )

        headers = [
            "Código/SKU",
            "Descripción del Producto",
            "Categoría",
            "Stock Actual",
            "Unidad",
            f"Costo Unitario ({currency_label})",
            f"Precio Venta ({currency_label})",
            f"Margen Unitario ({currency_label})",
            "Margen (%)",
            f"Valor al Costo ({currency_label})",
            f"Valor a Venta ({currency_label})",
            "Estado Stock",
        ]

        with rx.session() as session:
            products = session.exec(
                select(Product)
                .where(Product.company_id == company_id)
                .where(Product.branch_id == branch_id)
                .order_by(Product.description)
                .options(selectinload(Product.variants))
            ).all()

        def _variant_label(variant: ProductVariant) -> str:
            parts: list[str] = []
            if variant.size:
                parts.append(str(variant.size).strip())
            if variant.color:
                parts.append(str(variant.color).strip())
            return " ".join([p for p in parts if p]).strip()

        export_rows: list[dict[str, Any]] = []
        for product in products:
            variants = list(product.variants or [])
            if variants:
                for variant in variants:
                    label = _variant_label(variant)
                    description = product.description or "Sin descripción"
                    if label:
                        description = f"{description} ({label})"
                    export_rows.append(
                        {
                            "sku": variant.sku or product.barcode or "S/C",
                            "description": description,
                            "category": product.category or "Sin categoría",
                            "unit": product.unit or "Unid.",
                            "stock": variant.stock or 0,
                            "purchase_price": float(product.purchase_price or 0),
                            "sale_price": float(product.sale_price or 0),
                        }
                    )
            else:
                export_rows.append(
                    {
                        "sku": product.barcode or "S/C",
                        "description": product.description or "Sin descripción",
                        "category": product.category or "Sin categoría",
                        "unit": product.unit or "Unid.",
                        "stock": product.stock or 0,
                        "purchase_price": float(product.purchase_price or 0),
                        "sale_price": float(product.sale_price or 0),
                    }
                )

        total_items = len(export_rows)
        total_units = sum(float(item.get("stock", 0) or 0) for item in export_rows)
        total_cost_value = sum(
            float(item.get("stock", 0) or 0) * float(item.get("purchase_price", 0) or 0)
            for item in export_rows
        )
        total_sale_value = sum(
            float(item.get("stock", 0) or 0) * float(item.get("sale_price", 0) or 0)
            for item in export_rows
        )
        stock_zero = sum(1 for item in export_rows if float(item.get("stock", 0) or 0) == 0)
        stock_critical = sum(
            1 for item in export_rows if 0 < float(item.get("stock", 0) or 0) <= 5
        )
        stock_low = sum(
            1 for item in export_rows if 5 < float(item.get("stock", 0) or 0) <= 10
        )

        row += 1
        ws.cell(row=row, column=1, value="RESUMEN EJECUTIVO")
        row += 1
        ws.cell(row=row, column=1, value="Total SKUs:")
        ws.cell(row=row, column=2, value=total_items)
        row += 1
        ws.cell(row=row, column=1, value="Total unidades en stock:")
        ws.cell(row=row, column=2, value=total_units)
        row += 1
        ws.cell(row=row, column=1, value=f"Valor total al costo ({currency_label}):")
        ws.cell(row=row, column=2, value=total_cost_value).number_format = currency_format
        row += 1
        ws.cell(row=row, column=1, value=f"Valor total a venta ({currency_label}):")
        ws.cell(row=row, column=2, value=total_sale_value).number_format = currency_format
        row += 1
        ws.cell(row=row, column=1, value="Productos sin stock:")
        ws.cell(row=row, column=2, value=stock_zero)
        row += 1
        ws.cell(row=row, column=1, value="Productos críticos (1-5):")
        ws.cell(row=row, column=2, value=stock_critical)
        row += 1
        ws.cell(row=row, column=1, value="Productos bajos (6-10):")
        ws.cell(row=row, column=2, value=stock_low)
        row += 2

        style_header_row(ws, row, headers)
        data_start = row + 1
        row += 1

        for row_data in export_rows:
            barcode = row_data["sku"]
            description = row_data["description"]
            category = row_data["category"]
            unit = row_data["unit"]
            stock = row_data["stock"] or 0
            purchase_price = row_data["purchase_price"]
            sale_price = row_data["sale_price"]

            # Estado del stock
            if stock == 0:
                status = "SIN STOCK"
            elif stock <= 5:
                status = "CRÍTICO"
            elif stock <= 10:
                status = "BAJO"
            else:
                status = "NORMAL"

            ws.cell(row=row, column=1, value=barcode)
            ws.cell(row=row, column=2, value=description)
            ws.cell(row=row, column=3, value=category)
            ws.cell(row=row, column=4, value=stock)
            ws.cell(row=row, column=5, value=unit)
            ws.cell(row=row, column=6, value=purchase_price).number_format = currency_format
            ws.cell(row=row, column=7, value=sale_price).number_format = currency_format
            # Margen Unitario = Fórmula: Precio - Costo
            ws.cell(row=row, column=8, value=f"=G{row}-F{row}").number_format = currency_format
            # Margen % = Fórmula: (Margen / Costo) si Costo > 0
            ws.cell(row=row, column=9, value=f"=IF(F{row}>0,H{row}/F{row},0)").number_format = PERCENT_FORMAT
            # Valor al Costo = Fórmula: Stock × Costo
            ws.cell(row=row, column=10, value=f"=D{row}*F{row}").number_format = currency_format
            # Valor a Venta = Fórmula: Stock × Precio
            ws.cell(row=row, column=11, value=f"=D{row}*G{row}").number_format = currency_format
            ws.cell(row=row, column=12, value=status)

            # Color según estado
            status_cell = ws.cell(row=row, column=12)
            if "SIN STOCK" in status:
                status_cell.fill = NEGATIVE_FILL
            elif "CRÍTICO" in status:
                status_cell.fill = NEGATIVE_FILL
            elif "BAJO" in status:
                status_cell.fill = WARNING_FILL
            else:
                status_cell.fill = POSITIVE_FILL

            for col in range(1, 13):
                ws.cell(row=row, column=col).border = THIN_BORDER
            row += 1

        # Fila de totales
        totals_row = row
        add_totals_row_with_formulas(ws, totals_row, data_start, [
            {"type": "label", "value": "TOTALES"},
            {"type": "text", "value": ""},
            {"type": "text", "value": ""},
            {"type": "sum", "col_letter": "D"},
            {"type": "text", "value": ""},
            {"type": "text", "value": ""},
            {"type": "text", "value": ""},
            {"type": "text", "value": ""},
            {"type": "text", "value": ""},
            {"type": "sum", "col_letter": "J", "number_format": currency_format},
            {"type": "sum", "col_letter": "K", "number_format": currency_format},
            {"type": "text", "value": ""},
        ])

        # Notas explicativas
        add_notes_section(ws, totals_row, [
            "Costo Unitario: Precio al que se compró el producto al proveedor.",
            "Precio Venta: Precio de venta al público.",
            "Margen Unitario = Precio Venta - Costo Unitario (ganancia por unidad).",
            "Margen % = Margen Unitario ÷ Costo Unitario × 100.",
            "Valor al Costo: Inversión total = Stock × Costo Unitario.",
            "Valor a Venta: Potencial de ventas = Stock × Precio Venta.",
            "SIN STOCK: Producto agotado. CRÍTICO: ≤5 unidades. BAJO: ≤10 unidades.",
        ], columns=12)

        auto_adjust_column_widths(ws)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return rx.download(data=output.getvalue(), filename="inventario_valorizado.xlsx")

    # ══════════════════════════════════════════════════════════
    # IMPORTACIÓN MASIVA CSV / EXCEL
    # ══════════════════════════════════════════════════════════

    @rx.event
    def open_import_modal(self):
        if not self.current_user["privileges"].get("edit_inventario", False):
            return rx.toast("No tiene permisos para importar inventario.", duration=3000)
        self.import_modal_open = True
        self.import_preview_rows = []
        self.import_errors = []
        self.import_stats = {"new": 0, "updated": 0, "errors": 0, "total": 0}
        self.import_processing = False
        self.import_file_name = ""

    @rx.event
    def close_import_modal(self):
        self.import_modal_open = False
        self.import_preview_rows = []
        self.import_errors = []
        self.import_file_name = ""

    @rx.event
    async def handle_import_upload(self, files: list[rx.UploadFile]):
        """Procesa el archivo subido y genera preview."""
        if not files:
            return
        file = files[0]
        self.import_file_name = file.filename or "archivo"
        file_bytes = await file.read()

        try:
            rows = self._parse_import_file(file.filename or "", file_bytes)
        except Exception as e:
            self.import_errors = [f"Error al leer archivo: {str(e)}"]
            self.import_preview_rows = []
            return

        if not rows:
            self.import_errors = ["El archivo está vacío o no tiene filas de datos."]
            return

        company_id = self._company_id()
        branch_id = self._branch_id()
        if not company_id or not branch_id:
            self.import_errors = ["Empresa no configurada."]
            return

        # Cargar barcodes existentes para detectar nuevos vs actualizados
        existing_barcodes: set[str] = set()
        with rx.session() as session:
            products = session.exec(
                select(Product.barcode)
                .where(Product.company_id == company_id)
                .where(Product.branch_id == branch_id)
            ).all()
            existing_barcodes = {str(b) for b in products}

        preview = []
        errors = []
        new_count = 0
        update_count = 0
        error_count = 0

        for idx, row in enumerate(rows, start=2):
            barcode = str(row.get("barcode", "") or "").strip()
            description = str(row.get("description", "") or "").strip()

            if not barcode:
                errors.append(f"Fila {idx}: código de barras vacío.")
                error_count += 1
                continue
            if not description:
                errors.append(f"Fila {idx} ({barcode}): descripción vacía.")
                error_count += 1
                continue

            try:
                stock = float(row.get("stock", 0) or 0)
                purchase_price = float(row.get("purchase_price", 0) or 0)
                sale_price = float(row.get("sale_price", 0) or 0)
            except (ValueError, TypeError):
                errors.append(f"Fila {idx} ({barcode}): valores numéricos inválidos.")
                error_count += 1
                continue

            is_new = barcode not in existing_barcodes
            if is_new:
                new_count += 1
            else:
                update_count += 1

            preview.append({
                "row_num": idx,
                "barcode": barcode,
                "description": description,
                "category": str(row.get("category", "General") or "General").strip(),
                "stock": stock,
                "unit": str(row.get("unit", "Unidad") or "Unidad").strip(),
                "purchase_price": purchase_price,
                "sale_price": sale_price,
                "status": "Nuevo" if is_new else "Actualizar",
            })

        self.import_preview_rows = preview[:200]  # Limitar preview a 200 filas
        self.import_errors = errors[:50]  # Limitar errores visibles
        self.import_stats = {
            "new": new_count,
            "updated": update_count,
            "errors": error_count,
            "total": len(rows),
        }

    def _parse_import_file(self, filename: str, data: bytes) -> list[dict]:
        """Parsea CSV o Excel y retorna lista de dicts normalizados."""
        COLUMN_MAP = {
            "codigo": "barcode", "código": "barcode", "barcode": "barcode",
            "sku": "barcode", "cod": "barcode", "codigo/sku": "barcode",
            "código/sku": "barcode",
            "descripcion": "description", "descripción": "description",
            "description": "description", "producto": "description",
            "nombre": "description", "descripción del producto": "description",
            "categoria": "category", "categoría": "category",
            "category": "category",
            "stock": "stock", "stock actual": "stock", "cantidad": "stock",
            "unidad": "unit", "unit": "unit",
            "costo": "purchase_price", "costo unitario": "purchase_price",
            "purchase_price": "purchase_price", "precio compra": "purchase_price",
            "precio": "sale_price", "precio venta": "sale_price",
            "sale_price": "sale_price", "precio de venta": "sale_price",
            "pvp": "sale_price",
        }

        ext = (filename.rsplit(".", 1)[-1] if "." in filename else "").lower()

        if ext in ("xlsx", "xls"):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            header_raw = next(rows_iter, None)
            if not header_raw:
                return []
            headers = []
            for h in header_raw:
                h_clean = str(h or "").strip().lower()
                headers.append(COLUMN_MAP.get(h_clean, h_clean))
            result = []
            for row_values in rows_iter:
                if all(v is None for v in row_values):
                    continue
                row_dict = {}
                for i, val in enumerate(row_values):
                    if i < len(headers):
                        row_dict[headers[i]] = val
                result.append(row_dict)
            wb.close()
            return result
        else:
            # CSV
            import csv
            text = data.decode("utf-8-sig")
            # Detectar delimitador
            sniffer = csv.Sniffer()
            try:
                dialect = sniffer.sniff(text[:2048])
            except csv.Error:
                dialect = csv.excel
            reader = csv.DictReader(text.splitlines(), dialect=dialect)
            result = []
            for row in reader:
                normalized = {}
                for key, val in row.items():
                    k_clean = str(key or "").strip().lower()
                    mapped = COLUMN_MAP.get(k_clean, k_clean)
                    normalized[mapped] = val
                result.append(normalized)
            return result

    @rx.event
    def confirm_import(self):
        """Ejecuta la importación confirmada a la base de datos."""
        if not self.current_user["privileges"].get("edit_inventario", False):
            return rx.toast("No tiene permisos.", duration=3000)
        if not self.import_preview_rows:
            return rx.toast("No hay datos para importar.", duration=3000)

        company_id = self._company_id()
        branch_id = self._branch_id()
        user_id = self.current_user.get("id")
        if not company_id or not branch_id:
            return rx.toast("Empresa no configurada.", duration=3000)

        self.import_processing = True
        yield

        imported = 0
        updated = 0
        errors = []

        with rx.session() as session:
            # Pre-cargar productos existentes por barcode
            existing = session.exec(
                select(Product)
                .where(Product.company_id == company_id)
                .where(Product.branch_id == branch_id)
            ).all()
            products_by_barcode: dict[str, Product] = {
                p.barcode: p for p in existing
            }

            # Pre-cargar categorías
            cats = session.exec(
                select(Category)
                .where(Category.company_id == company_id)
                .where(Category.branch_id == branch_id)
            ).all()
            existing_categories: dict[str, Category] = {
                c.name: c for c in cats
            }

            try:
                for row in self.import_preview_rows:
                    barcode = row["barcode"]
                    description = row["description"]
                    category_name = row.get("category", "General") or "General"
                    stock = Decimal(str(row.get("stock", 0) or 0))
                    unit = row.get("unit", "Unidad") or "Unidad"
                    purchase_price = Decimal(str(row.get("purchase_price", 0) or 0))
                    sale_price = Decimal(str(row.get("sale_price", 0) or 0))

                    # Auto-crear categoría si no existe
                    if category_name not in existing_categories:
                        new_cat = Category(
                            name=category_name,
                            company_id=company_id,
                            branch_id=branch_id,
                        )
                        session.add(new_cat)
                        session.flush()
                        existing_categories[category_name] = new_cat

                    product = products_by_barcode.get(barcode)
                    if product:
                        # Actualizar
                        product.description = description
                        product.category = category_name
                        product.stock = stock
                        product.unit = unit
                        product.purchase_price = purchase_price
                        product.sale_price = sale_price
                        session.add(product)
                        updated += 1
                    else:
                        # Crear
                        product = Product(
                            barcode=barcode,
                            description=description,
                            category=category_name,
                            stock=stock,
                            unit=unit,
                            purchase_price=purchase_price,
                            sale_price=sale_price,
                            company_id=company_id,
                            branch_id=branch_id,
                        )
                        session.add(product)
                        session.flush()
                        products_by_barcode[barcode] = product

                        # Registrar movimiento de stock
                        if stock > 0:
                            movement = StockMovement(
                                product_id=product.id,
                                user_id=user_id,
                                type="Importacion",
                                quantity=stock,
                                description=f"Importación masiva: {description}",
                                timestamp=self._event_timestamp(),
                                company_id=company_id,
                                branch_id=branch_id,
                            )
                            session.add(movement)
                        imported += 1

                session.commit()
            except Exception as e:
                session.rollback()
                logger.exception("Error en importación masiva")
                self.import_processing = False
                errors.append(f"Error de base de datos: {str(e)}")
                self.import_errors = errors
                return rx.toast(
                    "Error al importar. Verifique los datos e intente nuevamente.",
                    duration=5000,
                )

        self.import_processing = False
        self.close_import_modal()
        self._inventory_update_trigger += 1
        self.load_categories()
        return rx.toast(
            f"Importación exitosa: {imported} nuevos, {updated} actualizados.",
            duration=5000,
        )

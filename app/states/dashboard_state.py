"""
Estado del Dashboard con métricas y KPIs del sistema.

Proporciona:
- Resumen de ventas (diarias, semanales, mensuales)
- KPIs principales
- Datos para gráficos
- Alertas del sistema
"""
import logging
from datetime import datetime, timedelta
from decimal import Decimal

import reflex as rx
from sqlmodel import select, func
from sqlalchemy import and_

logger = logging.getLogger(__name__)

from app.models import (
    Sale,
    SaleItem,
    Product,
    ProductVariant,
    ProductBatch,
    Client,
    SaleInstallment,
    CashboxLog,
    FieldReservation,
)
from app.utils.timezone import utc_now_naive
from .inventory import LOW_STOCK_THRESHOLD
from app.enums import SaleStatus, ReservationStatus
from app.i18n import MSG
from app.services.alert_service import get_alert_summary, BATCH_EXPIRING_DAYS
from .mixin_state import MixinState
from app.utils.exports import (
    create_excel_workbook,
    style_header_row,
    auto_adjust_column_widths,
    add_company_header,
    add_totals_row_with_formulas,
    add_notes_section,
    THIN_BORDER,
)


# Cantidad máxima de lotes que muestra el panel del dashboard.
EXPIRING_BATCHES_DISPLAY_LIMIT = 10


class DashboardState(MixinState):
    """Estado para el dashboard de métricas."""

    # Filtro de período
    selected_period: str = "month"  # Valores: today, week, month, custom
    custom_start_date: str = ""
    custom_end_date: str = ""

    # Datos de resumen (período seleccionado)
    period_sales: float = 0.0
    period_sales_count: int = 0
    period_reservations_count: int = 0
    period_prev_sales: float = 0.0  # Período anterior para comparación

    # Datos de resumen
    today_sales: float = 0.0
    today_sales_count: int = 0
    week_sales: float = 0.0
    week_sales_count: int = 0
    month_sales: float = 0.0
    month_sales_count: int = 0

    # Indicadores clave de rendimiento (KPIs)
    avg_ticket: float = 0.0
    total_clients: int = 0
    active_credits: int = 0
    pending_debt: float = 0.0
    low_stock_count: int = 0

    # Margen bruto del período
    period_gross_margin: float = 0.0
    period_total_cost: float = 0.0
    period_margin_percent: float = 0.0

    # Alertas
    alerts: list[dict] = []
    alert_count: int = 0

    # Datos para gráficos
    dash_sales_by_day: list[dict] = []       # Últimos 7 días
    dash_sales_by_category: list[dict] = []  # Por categoría
    dash_top_products: list[dict] = []       # Top 5 productos
    dash_payment_breakdown: list[dict] = []

    # Lotes por vencer (Farmacia / Supermercado)
    dash_expiring_batches: list[dict] = []
    expiring_batches_count: int = 0
    expired_batches_count: int = 0

    # Estado de carga
    dashboard_loading: bool = False
    last_refresh: str = ""
    _last_dashboard_load_ts: float = rx.field(default=0.0, is_var=False)
    _DASHBOARD_TTL: float = rx.field(default=30.0, is_var=False)

    def set_loading(self, loading: bool):
        """Establece el estado de carga."""
        self.dashboard_loading = loading

    def _load_dashboard_data(self) -> None:
        """Carga todos los datos del dashboard sin manejar UI state."""
        self._load_sales_summary()
        self._load_kpis()
        self._load_alerts()
        self._load_sales_by_day()
        self._load_top_products()
        self._load_sales_by_category()
        self._load_payment_breakdown()
        self._load_expiring_batches()
        self.last_refresh = self._display_now().strftime("%H:%M:%S")

    def _local_period_dates(self) -> tuple[datetime, datetime, datetime, datetime]:
        """Devuelve rangos locales según la zona horaria configurada."""
        now = self._display_now()
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)

        if self.selected_period == "today":
            start = today_start
            end = now
            prev_start = today_start - timedelta(days=1)
            prev_end = today_start
        elif self.selected_period == "week":
            start = today_start - timedelta(days=today_start.weekday())
            end = now
            prev_start = start - timedelta(days=7)
            prev_end = start
        elif (
            self.selected_period == "custom"
            and self.custom_start_date
            and self.custom_end_date
        ):
            start = datetime.strptime(self.custom_start_date, "%Y-%m-%d")
            end = datetime.strptime(self.custom_end_date, "%Y-%m-%d").replace(
                hour=23,
                minute=59,
                second=59,
            )
            delta = end - start
            prev_start = start - delta - timedelta(days=1)
            prev_end = start - timedelta(days=1)
        else:
            start = today_start.replace(day=1)
            end = now
            prev_month = start - timedelta(days=1)
            prev_start = prev_month.replace(day=1)
            prev_end = start
        return start, end, prev_start, prev_end

    def _get_period_dates(self) -> tuple[datetime, datetime, datetime, datetime]:
        """Convierte rangos locales a UTC-naive para consultar eventos persistidos."""
        start, end, prev_start, prev_end = self._local_period_dates()
        return (
            self._company_local_datetime_to_utc_naive(start),
            self._company_local_datetime_to_utc_naive(end),
            self._company_local_datetime_to_utc_naive(prev_start),
            self._company_local_datetime_to_utc_naive(prev_end),
        )

    def _get_reservation_period_end(self, today_start: datetime) -> datetime:
        """Devuelve fin de día actual (23:59:59) para contar reservas.

        Las reservas se agendan a futuro, así que una reserva hoy a las 20:00
        debe contarse incluso si ahora son las 15:00.  Para el período custom
        se respeta el end seleccionado por el usuario.
        """
        if self.selected_period == "custom" and self.custom_end_date:
            return datetime.strptime(self.custom_end_date, "%Y-%m-%d").replace(
                hour=23, minute=59, second=59
            )
        return today_start.replace(hour=23, minute=59, second=59)

    @rx.event
    def set_period(self, period: str):
        """Cambia el período seleccionado y recarga datos."""
        self.selected_period = period
        self.load_dashboard()

    @rx.event
    def set_custom_dates(self, start: str, end: str):
        """Establece fechas personalizadas."""
        self.custom_start_date = start
        self.custom_end_date = end
        self.selected_period = "custom"
        self.load_dashboard()

    @rx.event
    def load_dashboard(self):
        """Carga todos los datos del dashboard."""
        self.dashboard_loading = True

        try:
            self._load_dashboard_data()
        except Exception:
            logger.exception("load_dashboard failed")
        finally:
            self.dashboard_loading = False

    @rx.event(background=True)
    async def load_dashboard_background(self):
        """Carga el dashboard en segundo plano con TTL para evitar recargas innecesarias."""
        import time as _time
        async with self:
            now_ts = _time.time()
            if (now_ts - self._last_dashboard_load_ts) < self._DASHBOARD_TTL:
                return  # TTL vigente, no recargar
            self._last_dashboard_load_ts = now_ts
            self.dashboard_loading = True
            try:
                self._load_dashboard_data()
            except Exception:
                logger.exception("load_dashboard_background failed")
            finally:
                self.dashboard_loading = False

    def _load_sales_summary(self):
        """Carga resumen de ventas por período."""
        local_now = self._display_now()
        today_start_local = local_now.replace(hour=0, minute=0, second=0, microsecond=0)
        week_start_local = today_start_local - timedelta(days=today_start_local.weekday())
        month_start_local = today_start_local.replace(day=1)
        today_start = self._company_local_datetime_to_utc_naive(today_start_local)
        week_start = self._company_local_datetime_to_utc_naive(week_start_local)
        month_start = self._company_local_datetime_to_utc_naive(month_start_local)
        company_id = self._company_id()
        branch_id = self._branch_id()
        if not company_id or not branch_id:
            self.today_sales = 0.0
            self.today_sales_count = 0
            self.week_sales = 0.0
            self.week_sales_count = 0
            self.month_sales = 0.0
            self.month_sales_count = 0
            self.period_sales = 0.0
            self.period_sales_count = 0
            self.period_reservations_count = 0
            self.period_prev_sales = 0.0
            self.avg_ticket = 0.0
            return

        with rx.session() as session:
            # Ventas de hoy
            today_result = session.exec(
                select(
                    func.count(Sale.id),
                    func.coalesce(func.sum(Sale.total_amount), 0)
                )
                .where(
                    and_(
                        Sale.timestamp >= today_start,
                        Sale.status != SaleStatus.cancelled,
                        Sale.company_id == company_id,
                        Sale.branch_id == branch_id,
                    )
                )
            ).one()
            self.today_sales_count = today_result[0] or 0
            self.today_sales = float(today_result[1] or 0)

            # Ventas de la semana
            week_result = session.exec(
                select(
                    func.count(Sale.id),
                    func.coalesce(func.sum(Sale.total_amount), 0)
                )
                .where(
                    and_(
                        Sale.timestamp >= week_start,
                        Sale.status != SaleStatus.cancelled,
                        Sale.company_id == company_id,
                        Sale.branch_id == branch_id,
                    )
                )
            ).one()
            self.week_sales_count = week_result[0] or 0
            self.week_sales = float(week_result[1] or 0)

            # Ventas del mes
            month_result = session.exec(
                select(
                    func.count(Sale.id),
                    func.coalesce(func.sum(Sale.total_amount), 0)
                )
                .where(
                    and_(
                        Sale.timestamp >= month_start,
                        Sale.status != SaleStatus.cancelled,
                        Sale.company_id == company_id,
                        Sale.branch_id == branch_id,
                    )
                )
            ).one()
            self.month_sales_count = month_result[0] or 0
            self.month_sales = float(month_result[1] or 0)

            # Ventas del período seleccionado y período anterior
            period_start, period_end, prev_start, prev_end = self._get_period_dates()
            reservation_start, _, _, _ = self._local_period_dates()

            period_result = session.exec(
                select(
                    func.count(Sale.id),
                    func.coalesce(func.sum(Sale.total_amount), 0)
                )
                .where(
                    and_(
                        Sale.timestamp >= period_start,
                        Sale.timestamp <= period_end,
                        Sale.status != SaleStatus.cancelled,
                        Sale.company_id == company_id,
                        Sale.branch_id == branch_id,
                    )
                )
            ).one()
            self.period_sales_count = period_result[0] or 0
            self.period_sales = float(period_result[1] or 0)

            # Reservas del período seleccionado (agenda operativa).
            # Se usa end-of-day para incluir reservas futuras del día actual,
            # ya que start_datetime es la hora agendada (no la hora de creación).
            # Excluye estados cancelled y refunded.
            reservation_end = self._get_reservation_period_end(
                today_start_local
            )
            reservation_result = session.exec(
                select(func.count(FieldReservation.id))
                .where(
                    and_(
                        FieldReservation.start_datetime >= reservation_start,
                        FieldReservation.start_datetime <= reservation_end,
                        FieldReservation.status.notin_(
                            [ReservationStatus.CANCELLED, ReservationStatus.REFUNDED]
                        ),
                        FieldReservation.company_id == company_id,
                        FieldReservation.branch_id == branch_id,
                    )
                )
            ).one()
            self.period_reservations_count = int(reservation_result or 0)

            # Ticket promedio del período seleccionado
            if self.period_sales_count > 0:
                self.avg_ticket = self.period_sales / self.period_sales_count
            else:
                self.avg_ticket = 0.0

            # Margen bruto del período: ingresos - costo
            margin_result = session.exec(
                select(
                    func.coalesce(
                        func.sum(SaleItem.quantity * SaleItem.unit_price), 0
                    ),
                    func.coalesce(
                        func.sum(SaleItem.quantity * Product.purchase_price), 0
                    ),
                )
                .select_from(SaleItem)
                .join(Sale, SaleItem.sale_id == Sale.id)
                .join(Product, SaleItem.product_id == Product.id, isouter=True)
                .where(
                    and_(
                        Sale.timestamp >= period_start,
                        Sale.timestamp <= period_end,
                        Sale.status != SaleStatus.cancelled,
                        Sale.company_id == company_id,
                        Sale.branch_id == branch_id,
                    )
                )
            ).one()
            revenue = float(margin_result[0] or 0)
            cost = float(margin_result[1] or 0)
            self.period_total_cost = cost
            self.period_gross_margin = revenue - cost
            self.period_margin_percent = (
                ((revenue - cost) / revenue * 100) if revenue > 0 else 0.0
            )

            # Período anterior para comparación
            prev_result = session.exec(
                select(func.coalesce(func.sum(Sale.total_amount), 0))
                .where(
                    and_(
                        Sale.timestamp >= prev_start,
                        Sale.timestamp < prev_end,
                        Sale.status != SaleStatus.cancelled,
                        Sale.company_id == company_id,
                        Sale.branch_id == branch_id,
                    )
                )
            ).one()
            self.period_prev_sales = float(prev_result or 0)

    def _load_kpis(self):
        """Carga KPIs principales."""
        company_id = self._company_id()
        branch_id = self._branch_id()
        if not company_id or not branch_id:
            self.total_clients = 0
            self.active_credits = 0
            self.pending_debt = 0.0
            self.low_stock_count = 0
            return
        with rx.session() as session:
            # Total de clientes
            self.total_clients = session.exec(
                select(func.count())
                .select_from(Client)
                .where(Client.company_id == company_id)
                .where(Client.branch_id == branch_id)
            ).one() or 0

            # Créditos activos (ventas a crédito con cuotas pendientes)
            self.active_credits = session.exec(
                select(func.count(func.distinct(Sale.id)))
                .select_from(Sale)
                .join(SaleInstallment)
                .where(
                    and_(
                        Sale.status != SaleStatus.cancelled,
                        SaleInstallment.status == "pending",
                        Sale.company_id == company_id,
                        Sale.branch_id == branch_id,
                    )
                )
            ).one() or 0

            # Deuda pendiente total
            pending = session.exec(
                select(func.sum(SaleInstallment.amount - SaleInstallment.paid_amount))
                .select_from(SaleInstallment)
                .join(Sale)
                .where(
                    and_(
                        SaleInstallment.status == "pending",
                        Sale.company_id == company_id,
                        Sale.branch_id == branch_id,
                    )
                )
            ).one()
            self.pending_debt = float(pending or 0)

            # Productos con stock bajo (alineado con Inventario).
            # Cuenta productos raíz Y variantes con stock bajo:
            #   - Producto: stock > 0 AND stock <= min_stock_alert
            #   - Variante: stock > 0 AND stock <= COALESCE(variant.min_stock_alert, product.min_stock_alert)
            product_low = session.exec(
                select(func.count())
                .select_from(Product)
                .where(
                    and_(
                        Product.company_id == company_id,
                        Product.branch_id == branch_id,
                        Product.stock > 0,
                        Product.stock <= Product.min_stock_alert,
                    )
                )
            ).one() or 0
            variant_low = session.exec(
                select(func.count())
                .select_from(ProductVariant)
                .join(Product, ProductVariant.product_id == Product.id)
                .where(
                    and_(
                        ProductVariant.company_id == company_id,
                        ProductVariant.branch_id == branch_id,
                        ProductVariant.stock > 0,
                        ProductVariant.stock <= func.coalesce(
                            ProductVariant.min_stock_alert,
                            Product.min_stock_alert,
                        ),
                    )
                )
            ).one() or 0
            self.low_stock_count = int(product_low) + int(variant_low)

    def _load_alerts(self):
        """Carga alertas del sistema."""
        company_id = self._company_id()
        branch_id = self._branch_id()
        if not company_id or not branch_id:
            self.alerts = []
            self.alert_count = 0
            return
        settings = {}
        if hasattr(self, "_company_settings_snapshot"):
            settings = self._company_settings_snapshot()
        country_code = settings.get("country_code") or getattr(
            self, "selected_country_code", None
        )
        timezone = settings.get("timezone")
        summary = get_alert_summary(
            self.currency_symbol,
            company_id=company_id,
            branch_id=branch_id,
            country_code=country_code,
            timezone=timezone,
        )
        self.alerts = summary.get("alerts", [])
        self.alert_count = summary.get("total", 0)

    def _load_sales_by_day(self):
        """Carga ventas de los últimos 7 días para gráfico."""
        today_local = self._display_now().replace(hour=0, minute=0, second=0, microsecond=0)
        company_id = self._company_id()
        branch_id = self._branch_id()
        if not company_id or not branch_id:
            self.dash_sales_by_day = []
            return

        days_data = []
        day_names = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
        start_local = today_local - timedelta(days=6)
        end_local = today_local + timedelta(days=1)
        start_date = self._company_local_datetime_to_utc_naive(start_local)
        end_date = self._company_local_datetime_to_utc_naive(end_local)

        with rx.session() as session:
            results = session.exec(
                select(
                    Sale.timestamp,
                    Sale.total_amount,
                )
                .where(
                    and_(
                        Sale.timestamp >= start_date,
                        Sale.timestamp < end_date,
                        Sale.status != SaleStatus.cancelled,
                        Sale.company_id == company_id,
                        Sale.branch_id == branch_id,
                    )
                )
            ).all()

            totals_by_day: dict[datetime.date, float] = {}
            for row in results:
                timestamp = row[0]
                if not timestamp:
                    continue
                local_value = self._to_company_datetime(timestamp)
                if not local_value:
                    continue
                day_value = local_value.date()
                totals_by_day[day_value] = totals_by_day.get(day_value, 0.0) + float(
                    row[1] or 0
                )

            for i in range(6, -1, -1):
                day_start = today_local - timedelta(days=i)
                day_key = day_start.date()
                days_data.append({
                    "day": day_names[day_start.weekday()],
                    "date": day_start.strftime("%d/%m"),
                    "total": float(totals_by_day.get(day_key, 0)),
                })

        self.dash_sales_by_day = days_data

    def _load_top_products(self):
        """Carga los 5 productos más vendidos del período seleccionado."""
        period_start, period_end, _, _ = self._get_period_dates()
        company_id = self._company_id()
        branch_id = self._branch_id()
        if not company_id or not branch_id:
            self.dash_top_products = []
            return

        with rx.session() as session:
            results = session.exec(
                select(
                    Product.description,
                    func.sum(SaleItem.quantity).label("qty"),
                    func.sum(SaleItem.subtotal).label("revenue")
                )
                .select_from(SaleItem)
                .join(Product)
                .join(Sale)
                .where(
                    and_(
                        Sale.timestamp >= period_start,
                        Sale.timestamp <= period_end,
                        Sale.status != SaleStatus.cancelled,
                        Sale.company_id == company_id,
                        Sale.branch_id == branch_id,
                        Product.company_id == company_id,
                        Product.branch_id == branch_id,
                    )
                )
                .group_by(Product.id, Product.description)
                .order_by(func.sum(SaleItem.quantity).desc())
                .limit(10)
            ).all()

            self.dash_top_products = [
                {
                    "name": r[0][:25] + "..." if len(r[0]) > 25 else r[0],
                    "quantity": int(r[1] or 0),
                    "revenue": float(r[2] or 0),
                }
                for r in results
            ]

    def _load_expiring_batches(self):
        """Carga lotes vencidos y próximos a vencer (con stock > 0).

        Devuelve hasta EXPIRING_BATCHES_DISPLAY_LIMIT entradas combinando
        lotes ya vencidos primero y luego los que vencen en los próximos
        BATCH_EXPIRING_DAYS días, ordenados por fecha de vencimiento.
        Solo cuenta lotes con stock > 0 (los demás son irrelevantes).

        Aplica a verticales con productos perecederos (Farmacia, Supermercado).
        """
        company_id = self._company_id()
        branch_id = self._branch_id()
        if not company_id or not branch_id:
            self.dash_expiring_batches = []
            self.expiring_batches_count = 0
            self.expired_batches_count = 0
            return

        now = utc_now_naive()
        threshold = now + timedelta(days=BATCH_EXPIRING_DAYS)

        with rx.session() as session:
            # Lote + producto + variante (LEFT JOIN sobre variante porque
            # un lote puede pertenecer al producto raíz o a una variante).
            base_filters = and_(
                ProductBatch.company_id == company_id,
                ProductBatch.branch_id == branch_id,
                ProductBatch.expiration_date != None,
                ProductBatch.expiration_date <= threshold,
                ProductBatch.stock > 0,
            )

            rows = session.exec(
                select(
                    ProductBatch.id,
                    ProductBatch.batch_number,
                    ProductBatch.expiration_date,
                    ProductBatch.stock,
                    ProductBatch.product_id,
                    ProductBatch.product_variant_id,
                    Product.description,
                    ProductVariant.size,
                    ProductVariant.color,
                    ProductVariant.sku,
                )
                .select_from(ProductBatch)
                .outerjoin(Product, Product.id == ProductBatch.product_id)
                .outerjoin(
                    ProductVariant,
                    ProductVariant.id == ProductBatch.product_variant_id,
                )
                .where(base_filters)
                .order_by(ProductBatch.expiration_date.asc())
                .limit(EXPIRING_BATCHES_DISPLAY_LIMIT)
            ).all()

            expired = 0
            soon = 0
            display: list[dict] = []
            for r in rows:
                exp_dt = r[2]
                if exp_dt is None:
                    continue
                is_expired = exp_dt < now
                if is_expired:
                    expired += 1
                else:
                    soon += 1
                days_left = (exp_dt.date() - now.date()).days
                description = r[6] or "(Sin descripción)"
                variant_size = r[7]
                variant_color = r[8]
                if variant_size or variant_color:
                    label_parts = [p for p in (variant_size, variant_color) if p]
                    description = f"{description} ({' / '.join(label_parts)})"
                display.append(
                    {
                        "id": int(r[0]),
                        "batch_number": r[1] or "",
                        "expiration_date": exp_dt.strftime("%Y-%m-%d"),
                        "stock": float(r[3] or 0),
                        "product_id": int(r[4]) if r[4] is not None else None,
                        "product_variant_id": (
                            int(r[5]) if r[5] is not None else None
                        ),
                        "description": description,
                        "is_expired": is_expired,
                        "days_left": int(days_left),
                    }
                )

            # Conteos exactos (no truncados por limit) — para badges
            expired_total_query = (
                select(func.count())
                .select_from(ProductBatch)
                .where(
                    and_(
                        ProductBatch.company_id == company_id,
                        ProductBatch.branch_id == branch_id,
                        ProductBatch.expiration_date != None,
                        ProductBatch.expiration_date < now,
                        ProductBatch.stock > 0,
                    )
                )
            )
            soon_total_query = (
                select(func.count())
                .select_from(ProductBatch)
                .where(
                    and_(
                        ProductBatch.company_id == company_id,
                        ProductBatch.branch_id == branch_id,
                        ProductBatch.expiration_date != None,
                        ProductBatch.expiration_date >= now,
                        ProductBatch.expiration_date <= threshold,
                        ProductBatch.stock > 0,
                    )
                )
            )
            self.expired_batches_count = int(session.exec(expired_total_query).one() or 0)
            self.expiring_batches_count = int(session.exec(soon_total_query).one() or 0)
            self.dash_expiring_batches = display

    def _load_sales_by_category(self):
        """Carga ventas por categoría del período seleccionado."""
        self.dash_sales_by_category = self._query_sales_by_category(limit=10)

    def _query_sales_by_category(self, limit: int | None = None) -> list[dict]:
        """Obtiene ventas agrupadas por categoría para el período seleccionado."""
        period_start, period_end, _, _ = self._get_period_dates()
        company_id = self._company_id()
        branch_id = self._branch_id()
        if not company_id or not branch_id:
            return []

        with rx.session() as session:
            category_expr = func.coalesce(
                func.nullif(func.trim(SaleItem.product_category_snapshot), ""),
                Product.category,
                MSG.FALLBACK_NO_CATEGORY,
            )
            query = (
                select(
                    category_expr,
                    func.sum(SaleItem.subtotal).label("total"),
                )
                .select_from(SaleItem)
                .outerjoin(Product, Product.id == SaleItem.product_id)
                .join(Sale, Sale.id == SaleItem.sale_id)
                .where(
                    and_(
                        Sale.timestamp >= period_start,
                        Sale.timestamp <= period_end,
                        Sale.status != SaleStatus.cancelled,
                        Sale.company_id == company_id,
                        Sale.branch_id == branch_id,
                        SaleItem.company_id == company_id,
                        SaleItem.branch_id == branch_id,
                    )
                )
                .group_by(category_expr)
                .order_by(func.sum(SaleItem.subtotal).desc())
            )
            if limit is not None and int(limit) > 0:
                query = query.limit(int(limit))
            rows = session.exec(query).all()

        total_sales = sum(float(row[1] or 0) for row in rows)
        return [
            {
                "category": row[0] or MSG.FALLBACK_NO_CATEGORY,
                "total": float(row[1] or 0),
                "percentage": (
                    round((float(row[1] or 0) / total_sales * 100), 1)
                    if total_sales > 0
                    else 0
                ),
            }
            for row in rows
        ]

    def _load_payment_breakdown(self):
        """Carga desglose de métodos de pago del período seleccionado."""
        period_start, period_end, _, _ = self._get_period_dates()
        company_id = self._company_id()
        branch_id = self._branch_id()
        if not company_id or not branch_id:
            self.dash_payment_breakdown = []
            return

        with rx.session() as session:
            results = session.exec(
                select(
                    CashboxLog.action,
                    func.sum(CashboxLog.amount).label("total")
                )
                .where(
                    and_(
                        CashboxLog.timestamp >= period_start,
                        CashboxLog.timestamp <= period_end,
                        CashboxLog.is_voided == False,
                        CashboxLog.action.in_(["Venta", "Cobranza", "Cobro Cuota", "Reserva", "Adelanto"]),
                        CashboxLog.company_id == company_id,
                        CashboxLog.branch_id == branch_id,
                    )
                )
                .group_by(CashboxLog.action)
            ).all()

            self.dash_payment_breakdown = [
                {
                    "method": r[0],
                    "total": float(r[1] or 0),
                }
                for r in results
            ]

    @rx.var(cache=True)
    def has_critical_alerts(self) -> bool:
        """Indica si hay alertas críticas."""
        return any(a.get("severity") in ("critical", "error") for a in self.alerts)

    @rx.var(cache=True)
    def period_label(self) -> str:
        """Etiqueta del período seleccionado."""
        labels = {
            "today": "Hoy",
            "week": "Esta Semana",
            "month": "Este Mes",
            "custom": "Personalizado",
        }
        return labels.get(self.selected_period, "Este Mes")

    @rx.var(cache=True)
    def sales_change_percent(self) -> float:
        """Porcentaje de cambio vs período anterior."""
        if self.period_prev_sales > 0:
            return ((self.period_sales - self.period_prev_sales) / self.period_prev_sales) * 100
        return 0.0

    @rx.var(cache=True)
    def sales_trend_up(self) -> bool:
        """Indica si las ventas van en aumento."""
        return self.period_sales >= self.period_prev_sales

    @rx.var(cache=True)
    def formatted_sales_change(self) -> str:
        """Cambio formateado con signo."""
        change = self.sales_change_percent
        if change >= 0:
            return f"+{change:.1f}%"
        return f"{change:.1f}%"

    @rx.var(cache=True)
    def formatted_today_sales(self) -> str:
        return f"{self.currency_symbol}{self.today_sales:,.2f}"

    @rx.var(cache=True)
    def formatted_week_sales(self) -> str:
        return f"{self.currency_symbol}{self.week_sales:,.2f}"

    @rx.var(cache=True)
    def formatted_month_sales(self) -> str:
        return f"{self.currency_symbol}{self.month_sales:,.2f}"

    @rx.var(cache=True)
    def formatted_avg_ticket(self) -> str:
        return f"{self.currency_symbol}{self.avg_ticket:,.2f}"

    @rx.var(cache=True)
    def formatted_pending_debt(self) -> str:
        return f"{self.currency_symbol}{self.pending_debt:,.2f}"

    @rx.var(cache=True)
    def formatted_period_sales(self) -> str:
        return f"{self.currency_symbol}{self.period_sales:,.2f}"

    @rx.var(cache=True)
    def formatted_gross_margin(self) -> str:
        return f"{self.currency_symbol}{self.period_gross_margin:,.2f}"

    @rx.var(cache=True)
    def formatted_margin_percent(self) -> str:
        return f"{self.period_margin_percent:.1f}%"

    @rx.var(cache=True)
    def category_total_sales(self) -> float:
        """Total de ventas de todas las categorías."""
        return sum(c.get("total", 0) for c in self.dash_sales_by_category)

    @rx.var(cache=True)
    def formatted_category_total(self) -> str:
        return f"{self.currency_symbol}{self.category_total_sales:,.2f}"

    @rx.event
    def export_categories_excel(self):
        """Exporta ventas por categoría a Excel con formato profesional."""
        import io
        from openpyxl.styles import Alignment
        from openpyxl.chart import PieChart, Reference

        export_categories = self._query_sales_by_category(limit=None)
        currency_format = self._currency_excel_format()
        currency_label = self._currency_symbol_clean()
        company_name = getattr(self, "company_name", "") or "EMPRESA"
        percent_format = '0.0%'

        wb, ws = create_excel_workbook(MSG.REPORT_CAT_SALES_SHEET)

        row = add_company_header(
            ws,
            company_name,
            "REPORTE DE VENTAS POR CATEGORÍA",
            self.period_label,
            columns=4,
            generated_at=self._display_now(),
        )

        headers = ["#", "Categoría", f"Total Ventas ({currency_label})", "Participación"]
        style_header_row(ws, row, headers)
        data_start = row + 1
        row += 1

        total = sum(cat.get("total", 0) for cat in export_categories)
        for idx, cat in enumerate(export_categories, 1):
            pct = cat["total"] / total if total > 0 else 0

            ws.cell(row=row, column=1, value=idx).border = THIN_BORDER
            ws.cell(row=row, column=2, value=cat["category"]).border = THIN_BORDER

            cell_total = ws.cell(row=row, column=3, value=cat["total"])
            cell_total.number_format = currency_format
            cell_total.border = THIN_BORDER
            cell_total.alignment = Alignment(horizontal='right')

            cell_pct = ws.cell(row=row, column=4, value=pct)
            cell_pct.number_format = percent_format
            cell_pct.border = THIN_BORDER
            cell_pct.alignment = Alignment(horizontal='right')

            row += 1

        add_totals_row_with_formulas(ws, row, data_start, [
            {"type": "text", "value": ""},
            {"type": "label", "value": "TOTAL"},
            {"type": "sum", "col_letter": "C", "number_format": currency_format},
            {"type": "text", "value": 1, "number_format": percent_format},
        ])
        totals_row = row

        # Gráfico de torta
        if len(export_categories) > 0:
            chart = PieChart()
            chart.title = "Distribución de Ventas"
            data_ref = Reference(ws, min_col=3, min_row=data_start - 1, max_row=totals_row - 1)
            labels = Reference(ws, min_col=2, min_row=data_start, max_row=totals_row - 1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(labels)
            chart.width = 12
            chart.height = 8
            ws.add_chart(chart, "F4")

        add_notes_section(ws, totals_row, [
            "Total Ventas: Suma de ventas completadas en el período seleccionado.",
            "Participación: Porcentaje respecto al total general de ventas.",
        ], columns=4)

        auto_adjust_column_widths(ws)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"ventas_categoria_{self._display_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return rx.download(data=output.getvalue(), filename=filename)

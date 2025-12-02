import reflex as rx
from typing import TypedDict, Union
import datetime
import math
import calendar
import uuid
import logging
import bcrypt
import json
import time
from decimal import Decimal, ROUND_HALF_UP
from app.states.auth_state import (
    AuthState,
    User,
    Privileges,
    EMPTY_PRIVILEGES,
    SUPERADMIN_PRIVILEGES,
    )

# Import utility functions for reuse
from app.utils.formatting import round_currency as _round_currency_util
from app.utils.dates import get_today_str, get_current_month_str, get_current_week_str
from app.utils.barcode import clean_barcode, validate_barcode

TODAY_STR = get_today_str()
CURRENT_MONTH_STR = get_current_month_str()
CURRENT_WEEK_STR = get_current_week_str()


# ========================================
# region: TypedDict Definitions
# ========================================

class Product(TypedDict):
    id: str
    barcode: str
    description: str
    category: str
    stock: float
    unit: str
    purchase_price: float
    sale_price: float


class TransactionItem(TypedDict):
    temp_id: str
    barcode: str
    description: str
    category: str
    quantity: float
    unit: str
    price: float
    sale_price: float
    subtotal: float


class Movement(TypedDict):
    id: str
    timestamp: str
    type: str
    product_description: str
    quantity: float
    unit: str
    total: float
    payment_method: str
    payment_details: str
    user: str
    sale_id: str


class CurrencyOption(TypedDict):
    code: str
    name: str
    symbol: str


class PaymentMethodConfig(TypedDict):
    id: str
    name: str
    description: str
    kind: str
    enabled: bool


class PaymentBreakdownItem(TypedDict):
    label: str
    amount: float


class FieldPrice(TypedDict):
    id: str
    sport: str
    name: str
    price: float


class FieldReservation(TypedDict):
    id: str
    client_name: str
    dni: str
    phone: str
    sport: str
    sport_label: str
    field_name: str
    start_datetime: str
    end_datetime: str
    advance_amount: float
    total_amount: float
    paid_amount: float
    status: str
    created_at: str
    cancellation_reason: str
    delete_reason: str


class ServiceLogEntry(TypedDict):
    id: str
    timestamp: str
    type: str
    sport: str
    client_name: str
    field_name: str
    amount: float
    status: str
    notes: str
    reservation_id: str


class ReservationReceipt(TypedDict):
    cliente: str
    deporte: str
    campo: str
    horario: str
    monto_adelanto: str
    monto_total: str
    saldo: str
    estado: str


class CashboxSale(TypedDict):
    sale_id: str
    timestamp: str
    user: str
    payment_method: str
    payment_label: str
    payment_breakdown: list[PaymentBreakdownItem]
    payment_details: str
    total: float
    service_total: float
    items: list[TransactionItem]
    is_deleted: bool
    delete_reason: str


class CashboxSession(TypedDict):
    opening_amount: float
    opening_time: str
    closing_time: str
    is_open: bool
    opened_by: str


class CashboxLogEntry(TypedDict):
    id: str
    action: str
    timestamp: str
    user: str
    opening_amount: float
    closing_total: float
    totals_by_method: list[dict[str, float]]
    notes: str


class InventoryAdjustment(TypedDict):
    temp_id: str
    barcode: str
    description: str
    category: str
    unit: str
    current_stock: float
    adjust_quantity: float
    reason: str


class NewUser(TypedDict):
    username: str
    password: str
    confirm_password: str
    role: str
    privileges: Privileges

# endregion

# ========================================
# region: Constants & Default Values
# ========================================

DEFAULT_USER_PRIVILEGES: Privileges = {
    "view_ingresos": True,
    "create_ingresos": True,
    "view_ventas": True,
    "create_ventas": True,
    "view_inventario": True,
    "edit_inventario": True,
    "view_historial": True,
    "export_data": False,
    "view_cashbox": True,
    "manage_cashbox": True,
    "delete_sales": False,
    "manage_users": False,
    "view_servicios": True,
    "manage_reservations": True,
    "manage_config": False,
}

ADMIN_PRIVILEGES: Privileges = {
    "view_ingresos": True,
    "create_ingresos": True,
    "view_ventas": True,
    "create_ventas": True,
    "view_inventario": True,
    "edit_inventario": True,
    "view_historial": True,
    "export_data": True,
    "view_cashbox": True,
    "manage_cashbox": True,
    "delete_sales": True,
    "manage_users": True,
    "view_servicios": True,
    "manage_reservations": True,
    "manage_config": True,
}

CASHIER_PRIVILEGES: Privileges = {
    "view_ingresos": False,
    "create_ingresos": False,
    "view_ventas": True,
    "create_ventas": True,
    "view_inventario": True,
    "edit_inventario": False,
    "view_historial": False,
    "export_data": False,
    "view_cashbox": True,
    "manage_cashbox": True,
    "delete_sales": False,
    "manage_users": False,
    "view_servicios": False,
    "manage_reservations": False,
    "manage_config": False,
}

DEFAULT_ROLE_TEMPLATES: dict[str, Privileges] = {
    "Superadmin": SUPERADMIN_PRIVILEGES,
    "Administrador": ADMIN_PRIVILEGES,
    "Usuario": DEFAULT_USER_PRIVILEGES,
    "Cajero": CASHIER_PRIVILEGES,
}

# endregion

# ========================================
# region: State Class Definition
# ========================================


class State(AuthState):
    """
    Main application state class managing all business logic and data.
    
    Organized into logical sections:
    - Core & Navigation
    - Currency Management
    - Services & Reservations
    - Inventory Management
    - Entry (Ingreso)
    - Sales (Venta)
    - Payment Methods
    - Cashbox (Caja)
    - History & Reports
    - User Management
    """
    
    # ----------------------------------------
    # Core & Navigation Attributes
    # ----------------------------------------
    sidebar_open: bool = True
    current_page: str = "Ingreso"
    config_active_tab: str = "usuarios"
    
    # ----------------------------------------
    # Services & Reservations Attributes
    # ----------------------------------------
    service_active_tab: str = "campo"
    field_rental_sport: str = "futbol"
    schedule_view_mode: str = "dia"
    schedule_selected_date: str = TODAY_STR
    schedule_selected_week: str = CURRENT_WEEK_STR
    schedule_selected_month: str = CURRENT_MONTH_STR
    schedule_selected_slots: list[dict[str, str]] = []
    reservation_form: dict[str, str] = {
        "client_name": "",
        "dni": "",
        "phone": "",
        "field_name": "",
        "sport_label": "",
        "selected_price_id": "",
        "date": TODAY_STR,
        "start_time": "00:00",
        "end_time": "01:00",
        "advance_amount": "0",
        "total_amount": "0",
        "status": "pendiente",
    }
    service_reservations: list[FieldReservation] = []
    service_admin_log: list[ServiceLogEntry] = []
    reservation_payment_id: str = ""
    reservation_payment_amount: str = ""
    reservation_cancel_selection: str = ""
    reservation_cancel_reason: str = ""
    reservation_modal_open: bool = False
    reservation_modal_mode: str = "new"
    reservation_modal_reservation_id: str = ""
    reservation_search: str = ""
    reservation_filter_status: str = "todos"
    reservation_filter_start_date: str = ""
    reservation_filter_end_date: str = ""
    reservation_staged_search: str = ""
    reservation_staged_status: str = "todos"
    reservation_staged_start_date: str = ""
    reservation_staged_end_date: str = ""
    reservation_payment_routed: bool = False
    last_reservation_receipt: ReservationReceipt | None = None
    reservation_delete_selection: str = ""
    reservation_delete_reason: str = ""
    reservation_delete_modal_open: bool = False
    field_prices: list[FieldPrice] = []
    new_field_price_sport: str = ""
    new_field_price_name: str = ""
    new_field_price_amount: str = ""
    editing_field_price_id: str = ""
    service_log_filter_start_date: str = ""
    service_log_filter_end_date: str = ""
    service_log_filter_sport: str = "todos"
    service_log_filter_status: str = "todos"
    
    # ----------------------------------------
    # Configuration & Units Attributes
    # ----------------------------------------
    units: list[str] = ["Unidad", "Kg", "Litro", "Metro", "Caja"]
    new_unit_name: str = ""
    new_unit_allows_decimal: bool = False
    roles: list[str] = list(DEFAULT_ROLE_TEMPLATES.keys())
    role_privileges: dict[str, Privileges] = {
        name: template.copy() for name, template in DEFAULT_ROLE_TEMPLATES.items()
    }
    
    # ----------------------------------------
    # Currency Attributes
    # ----------------------------------------
    available_currencies: list[CurrencyOption] = [
        {"code": "PEN", "name": "Sol peruano (PEN)", "symbol": "S/"},
        {"code": "ARS", "name": "Peso argentino (ARS)", "symbol": "$"},
        {"code": "USD", "name": "Dolar estadounidense (USD)", "symbol": "US$"},
    ]
    selected_currency_code: str = "PEN"
    new_currency_name: str = ""
    new_currency_code: str = ""
    new_currency_symbol: str = ""
    decimal_units: set[str] = {
        "kg",
        "kilogramo",
        "kilogramos",
        "g",
        "gramo",
        "gramos",
        "l",
        "litro",
        "litros",
        "ml",
        "metro",
        "metros",
    }
    
    # ----------------------------------------
    # Inventory Management Attributes
    # ----------------------------------------
    inventory: dict[str, Product] = {}
    history: list[Movement] = []
    inventory_search_term: str = ""
    categories: list[str] = ["General"]
    new_category_name: str = ""
    
    # ----------------------------------------
    # History & Filters Attributes
    # ----------------------------------------
    history_filter_type: str = "Todos"
    history_filter_product: str = ""
    history_filter_start_date: str = ""
    history_filter_end_date: str = ""
    staged_history_filter_type: str = "Todos"
    staged_history_filter_product: str = ""
    staged_history_filter_start_date: str = ""
    staged_history_filter_end_date: str = ""
    current_page_history: int = 1
    items_per_page: int = 10
    
    # ----------------------------------------
    # Entry (Ingreso) Attributes
    # ----------------------------------------
    entry_form_key: int = 0
    new_entry_item: TransactionItem = {
        "temp_id": "",
        "barcode": "",
        "description": "",
        "category": "General",
        "quantity": 0,
        "unit": "Unidad",
        "price": 0,
        "sale_price": 0,
        "subtotal": 0,
    }
    new_entry_items: list[TransactionItem] = []
    entry_autocomplete_suggestions: list[str] = []
    
    # ----------------------------------------
    # Sales (Venta) Attributes
    # ----------------------------------------
    sale_form_key: int = 0
    new_sale_item: TransactionItem = {
        "temp_id": "",
        "barcode": "",
        "description": "",
        "category": "General",
        "quantity": 0,
        "unit": "Unidad",
        "price": 0,
        "sale_price": 0,
        "subtotal": 0,
    }
    new_sale_items: list[TransactionItem] = []
    autocomplete_suggestions: list[str] = []
    last_sale_receipt: list[TransactionItem] = []
    last_sale_total: float = 0
    last_sale_timestamp: str = ""
    sale_receipt_ready: bool = False
    last_sale_reservation_context: dict | None = None
    
    # ----------------------------------------
    # Payment Methods Attributes
    # ----------------------------------------
    payment_methods: list[PaymentMethodConfig] = [
        {
            "id": "cash",
            "name": "Efectivo",
            "description": "Billetes, Monedas",
            "kind": "cash",
            "enabled": True,
        },
        {
            "id": "card",
            "name": "Tarjeta",
            "description": "Credito, Debito",
            "kind": "card",
            "enabled": True,
        },
        {
            "id": "wallet",
            "name": "Billetera Digital / QR",
            "description": "Yape, Plin, Billeteras Bancarias",
            "kind": "wallet",
            "enabled": True,
        },
        {
            "id": "mixed",
            "name": "Pagos Mixtos",
            "description": "Combinacion de metodos",
            "kind": "mixed",
            "enabled": True,
        },
    ]
    payment_method: str = "Efectivo"
    payment_method_description: str = "Billetes, Monedas"
    payment_method_kind: str = "cash"
    payment_cash_amount: float = 0
    payment_cash_message: str = ""
    payment_cash_status: str = "neutral"
    payment_card_type: str = "Credito"
    payment_wallet_choice: str = "Yape"
    payment_wallet_provider: str = "Yape"
    payment_mixed_cash: float = 0
    payment_mixed_card: float = 0
    payment_mixed_wallet: float = 0
    payment_mixed_message: str = ""
    payment_mixed_status: str = "neutral"
    payment_mixed_notes: str = ""
    last_payment_summary: str = ""
    new_payment_method_name: str = ""
    new_payment_method_description: str = ""
    new_payment_method_kind: str = "other"
    
    # ----------------------------------------
    # Inventory Adjustment Attributes
    # ----------------------------------------
    inventory_check_modal_open: bool = False
    inventory_check_status: str = "perfecto"
    inventory_adjustment_notes: str = ""
    inventory_adjustment_item: InventoryAdjustment = {
        "temp_id": "",
        "barcode": "",
        "description": "",
        "category": "",
        "unit": "",
        "current_stock": 0,
        "adjust_quantity": 0,
        "reason": "",
    }
    
    editing_product: Product = {
        "id": "",
        "barcode": "",
        "description": "",
        "category": "",
        "stock": 0,
        "unit": "",
        "purchase_price": 0,
        "sale_price": 0,
    }
    is_editing_product: bool = False
    
    inventory_adjustment_items: list[InventoryAdjustment] = []
    inventory_adjustment_suggestions: list[str] = []
    
    # ----------------------------------------
    # Cashbox (Caja) Attributes
    # ----------------------------------------
    cashbox_sales: list[CashboxSale] = []
    cashbox_filter_start_date: str = ""
    cashbox_filter_end_date: str = ""
    cashbox_staged_start_date: str = ""
    cashbox_staged_end_date: str = ""
    cashbox_current_page: int = 1
    cashbox_items_per_page: int = 10
    show_cashbox_advances: bool = True
    sale_delete_modal_open: bool = False
    sale_to_delete: str = ""
    sale_delete_reason: str = ""
    cashbox_close_modal_open: bool = False
    cashbox_close_summary_totals: dict[str, float] = {}
    cashbox_close_summary_sales: list[CashboxSale] = []
    cashbox_close_summary_date: str = ""
    cashbox_sessions: dict[str, CashboxSession] = {}
    cashbox_open_amount_input: str = "0"
    cashbox_logs: list[CashboxLogEntry] = []
    cashbox_log_filter_start_date: str = ""
    cashbox_log_filter_end_date: str = ""
    cashbox_log_staged_start_date: str = ""
    cashbox_log_staged_end_date: str = ""
    cashbox_log_modal_open: bool = False
    cashbox_log_selected: CashboxLogEntry | None = None
    
    # ----------------------------------------
    # User Management Attributes
    # ----------------------------------------
    show_user_form: bool = False
    editing_user: User | None = None
    new_role_name: str = ""
    new_user_data: NewUser = {
        "username": "",
        "password": "",
        "confirm_password": "",
        "role": "Usuario",
        "privileges": DEFAULT_USER_PRIVILEGES.copy(),
    }

    # ========================================
    # region: Private Helper Methods
    # ========================================

    def _normalize_privileges(self, privileges: dict | None) -> Privileges:
        merged = EMPTY_PRIVILEGES.copy()
        if privileges:
            merged.update(privileges)
        return merged

    def _get_or_create_cashbox_session(self, username: str) -> CashboxSession:
        key = (username or "").strip().lower()
        if key not in self.cashbox_sessions:
            self.cashbox_sessions[key] = {
                "opening_amount": 0.0,
                "opening_time": "",
                "closing_time": "",
                "is_open": False,
                "opened_by": key,
            }
        return self.cashbox_sessions[key]

    def _require_cashbox_open(self):
        if not self.cashbox_is_open:
            return rx.toast("Debe aperturar la caja para operar.", duration=3000)
        return None
    
    def _find_role_key(self, role: str) -> str | None:
        """Return the stored role name matching a role string (case-insensitive)."""
        target = (role or "").strip().lower()
        for name in self.role_privileges.keys():
            if name.lower() == target:
                return name
        return None

    def _role_privileges(self, role: str) -> Privileges:
        key = self._find_role_key(role)
        template = self.role_privileges.get(key or "")
        if template:
            return self._normalize_privileges(template)
        return EMPTY_PRIVILEGES.copy()

    def _round_currency(self, value: float) -> float:
        return float(
            Decimal(str(value or 0)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        )

    @rx.var
    def currency_symbol(self) -> str:
        match = next(
            (c for c in self.available_currencies if c["code"] == self.selected_currency_code),
            None,
        )
        return f"{match['symbol']} " if match else "S/ "

    @rx.var
    def currency_name(self) -> str:
        match = next(
            (c for c in self.available_currencies if c["code"] == self.selected_currency_code),
            None,
        )
        return match["name"] if match else "Sol peruano (PEN)"

    def _format_currency(self, value: float) -> str:
        return f"{self.currency_symbol}{self._round_currency(value):.2f}"

    @rx.event
    def set_currency(self, code: str):
        code = (code or "").upper()
        match = next((c for c in self.available_currencies if c["code"] == code), None)
        if not match:
            return rx.toast("Moneda no soportada.", duration=3000)
        self.selected_currency_code = code
        self._refresh_payment_feedback()
        return rx.toast(f"Moneda cambiada a {match['name']}.", duration=2500)

    @rx.event
    def set_new_currency_code(self, value: str):
        self.new_currency_code = (value or "").upper()

    @rx.event
    def set_new_currency_name(self, value: str):
        self.new_currency_name = value

    @rx.event
    def set_new_currency_symbol(self, value: str):
        self.new_currency_symbol = value

    @rx.event
    def add_currency(self):
        code = (self.new_currency_code or "").strip().upper()
        name = (self.new_currency_name or "").strip()
        symbol = (self.new_currency_symbol or "").strip()
        if not code or not name or not symbol:
            return rx.toast("Complete codigo, nombre y simbolo.", duration=3000)
        if any(c["code"] == code for c in self.available_currencies):
            return rx.toast("La moneda ya existe.", duration=3000)
        self.available_currencies.append({"code": code, "name": name, "symbol": symbol})
        self.selected_currency_code = code
        self.new_currency_code = ""
        self.new_currency_name = ""
        self.new_currency_symbol = ""
        self._refresh_payment_feedback()
        return rx.toast(f"Moneda {name} agregada.", duration=2500)

    @rx.event
    def remove_currency(self, code: str):
        code = (code or "").upper()
        if len(self.available_currencies) <= 1:
            return rx.toast("Debe quedar al menos una moneda.", duration=3000)
        if not any(c["code"] == code for c in self.available_currencies):
            return
        self.available_currencies = [
            currency for currency in self.available_currencies if currency["code"] != code
        ]
        if self.selected_currency_code == code and self.available_currencies:
            self.selected_currency_code = self.available_currencies[0]["code"]
        self._refresh_payment_feedback()
        return rx.toast("Moneda eliminada.", duration=2500)

    def _unit_allows_decimal(self, unit: str) -> bool:
        return unit and unit.lower() in self.decimal_units

    def _normalize_quantity_value(self, value: float, unit: str) -> float:
        if self._unit_allows_decimal(unit):
            return float(
                Decimal(str(value or 0)).quantize(
                    Decimal("0.01"), rounding=ROUND_HALF_UP
                )
            )
        return int(
            Decimal(str(value or 0)).quantize(
                Decimal("1"), rounding=ROUND_HALF_UP
            )
        )

    @rx.event
    def set_new_unit_name(self, value: str):
        self.new_unit_name = value

    @rx.event
    def set_new_unit_allows_decimal(self, value: bool | str):
        if isinstance(value, str):
            value = value.lower() in ["true", "1", "on", "yes"]
        self.new_unit_allows_decimal = bool(value)

    @rx.event
    def add_unit(self):
        if not self.current_user["privileges"]["manage_config"]:
            return rx.toast("No tiene permisos para configurar el sistema.", duration=3000)
        name = (self.new_unit_name or "").strip()
        if not name:
            return rx.toast("Ingrese un nombre de unidad.", duration=3000)
        unit_lower = name.lower()
        if any(u.lower() == unit_lower for u in self.units):
            return rx.toast("La unidad ya existe.", duration=3000)
        self.units.append(name)
        if self.new_unit_allows_decimal:
            self.decimal_units.add(unit_lower)
        self.new_unit_name = ""
        self.new_unit_allows_decimal = False
        return rx.toast(f"Unidad {name} agregada.", duration=2500)

    @rx.event
    def remove_unit(self, unit: str):
        value = (unit or "").strip()
        unit_lower = value.lower()
        if not value:
            return
        if unit_lower == "unidad":
            return rx.toast("No puedes eliminar la unidad base.", duration=3000)
        if unit_lower not in [u.lower() for u in self.units]:
            return
        if any(p.get("unit", "").lower() == unit_lower for p in self.inventory.values()):
            return rx.toast("No puedes eliminar unidades usadas en inventario.", duration=3000)
        self.units = [u for u in self.units if u.lower() != unit_lower]
        self.decimal_units.discard(unit_lower)
        if self.new_entry_item["unit"].lower() == unit_lower:
            self.new_entry_item["unit"] = "Unidad"
        if self.new_sale_item["unit"].lower() == unit_lower:
            self.new_sale_item["unit"] = "Unidad"
        return rx.toast(f"Unidad {value} eliminada.", duration=2500)

    @rx.event
    def set_unit_decimal(self, unit: str, allows_decimal: bool | str):
        unit_lower = (unit or "").lower()
        if isinstance(allows_decimal, str):
            allows_decimal = allows_decimal.lower() in ["true", "1", "on", "yes"]
        if allows_decimal:
            self.decimal_units.add(unit_lower)
        else:
            self.decimal_units.discard(unit_lower)

    def _apply_item_rounding(self, item: TransactionItem):
        unit = item.get("unit", "")
        item["quantity"] = self._normalize_quantity_value(
            item.get("quantity", 0), unit
        )
        item["price"] = self._round_currency(item.get("price", 0))
        if "sale_price" in item:
            item["sale_price"] = self._round_currency(item.get("sale_price", 0))
        item["subtotal"] = self._round_currency(item["quantity"] * item["price"])

    @rx.var
    def current_user(self) -> User:
        if not self.token:
            return self._guest_user()
        user = self.users.get(self.token)
        if user:
            merged_privileges = self._normalize_privileges(user.get("privileges", {}))
            user["privileges"] = merged_privileges
            return user
        return self._guest_user()

    @rx.var
    def current_cashbox_session(self) -> CashboxSession:
        return self._get_or_create_cashbox_session(self.current_user["username"])

    @rx.var
    def cashbox_is_open(self) -> bool:
        return bool(self.current_cashbox_session.get("is_open"))

    @rx.var
    def cashbox_opening_amount(self) -> float:
        return float(self.current_cashbox_session.get("opening_amount", 0))

    @rx.var
    def cashbox_opening_time(self) -> str:
        return self.current_cashbox_session.get("opening_time", "")

    @rx.var
    def filtered_cashbox_logs(self) -> list[CashboxLogEntry]:
        if not self.current_user["privileges"]["view_cashbox"]:
            return []
        logs = sorted(
            self.cashbox_logs, key=lambda entry: entry.get("timestamp", ""), reverse=True
        )
        start_date = None
        end_date = None
        if self.cashbox_log_filter_start_date:
            try:
                start_date = datetime.datetime.strptime(
                    self.cashbox_log_filter_start_date, "%Y-%m-%d"
                ).date()
            except Exception as e:
                logging.exception(f"Error parsing cashbox log start date: {e}")
        if self.cashbox_log_filter_end_date:
            try:
                end_date = datetime.datetime.strptime(
                    self.cashbox_log_filter_end_date, "%Y-%m-%d"
                ).date()
            except Exception as e:
                logging.exception(f"Error parsing cashbox log end date: {e}")
        filtered: list[CashboxLogEntry] = []
        for log in logs:
            timestamp = log.get("timestamp", "")
            try:
                log_date = datetime.datetime.strptime(
                    timestamp.split(" ")[0], "%Y-%m-%d"
                ).date()
            except Exception:
                continue
            if start_date and log_date < start_date:
                continue
            if end_date and log_date > end_date:
                continue
            filtered.append(log)
        return filtered

    @rx.var
    def entry_subtotal(self) -> float:
        return self.new_entry_item["subtotal"]

    @rx.var
    def entry_total(self) -> float:
        return self._round_currency(
            sum((item["subtotal"] for item in self.new_entry_items))
        )

    @rx.var
    def sale_subtotal(self) -> float:
        return self.new_sale_item["subtotal"]

    @rx.var
    def sale_total(self) -> float:
        reservation = (
            self._find_reservation_by_id(self.reservation_payment_id)
            if self.reservation_payment_id
            else None
        )
        balance = 0.0
        if reservation:
            balance = max(reservation["total_amount"] - reservation["paid_amount"], 0)
        products_total = sum((item["subtotal"] for item in self.new_sale_items))
        return self._round_currency(products_total + balance)

    @rx.var
    def enabled_payment_methods(self) -> list[PaymentMethodConfig]:
        return [m for m in self.payment_methods if m.get("enabled", True)]

    @rx.var
    def unit_rows(self) -> list[dict[str, Union[str, bool]]]:
        return [
            {"name": unit, "allows_decimal": unit.lower() in self.decimal_units}
            for unit in self.units
        ]

    @rx.var
    def payment_summary(self) -> str:
        return self._generate_payment_summary()

    @rx.var
    def inventory_list(self) -> list[Product]:
        if not self.current_user["privileges"]["view_inventario"]:
            return []
        for product in self.inventory.values():
            if "barcode" not in product:
                product["barcode"] = ""
            if "category" not in product:
                product["category"] = self.categories[0] if self.categories else ""
        if self.inventory_search_term:
            search = self.inventory_search_term.lower()
            return sorted(
                [
                    p
                    for p in self.inventory.values()
                    if search in p["description"].lower()
                    or search in p.get("barcode", "").lower()
                    or search in p.get("category", "").lower()
                ],
                key=lambda p: p["description"],
            )
        return sorted(list(self.inventory.values()), key=lambda p: p["description"])

    @rx.event
    def set_inventory_search_term(self, value: str):
        self.inventory_search_term = value or ""

    @rx.var
    def filtered_history(self) -> list[Movement]:
        if not self.current_user["privileges"]["view_historial"]:
            return []
        movements = self.history
        if self.history_filter_type != "Todos":
            movements = [m for m in movements if m["type"] == self.history_filter_type]
        if self.history_filter_product:
            movements = [
                m
                for m in movements
                if self.history_filter_product.lower()
                in m["product_description"].lower()
            ]
        if self.history_filter_start_date:
            try:
                start_date = datetime.datetime.fromisoformat(
                    self.history_filter_start_date
                )
                movements = [
                    m
                    for m in movements
                    if datetime.datetime.fromisoformat(m["timestamp"].split()[0])
                    >= start_date
                ]
            except ValueError as e:
                logging.exception(f"Error parsing start date: {e}")
        if self.history_filter_end_date:
            try:
                end_date = datetime.datetime.fromisoformat(self.history_filter_end_date)
                movements = [
                    m
                    for m in movements
                    if datetime.datetime.fromisoformat(m["timestamp"].split()[0])
                    <= end_date
                ]
            except ValueError as e:
                logging.exception(f"Error parsing end date: {e}")
        return sorted(movements, key=lambda m: m["timestamp"], reverse=True)

    @rx.var
    def paginated_history(self) -> list[Movement]:
        start_index = (self.current_page_history - 1) * self.items_per_page
        end_index = start_index + self.items_per_page
        return self.filtered_history[start_index:end_index]

    @rx.var
    def total_pages(self) -> int:
        total_items = len(self.filtered_history)
        if total_items == 0:
            return 1
        return (total_items + self.items_per_page - 1) // self.items_per_page

    @rx.var
    def filtered_cashbox_sales(self) -> list[CashboxSale]:
        if not self.current_user["privileges"]["view_cashbox"]:
            return []
        sales = sorted(
            self.cashbox_sales, key=lambda s: s["timestamp"], reverse=True
        )
        if not self.show_cashbox_advances:
            sales = [sale for sale in sales if not self._is_advance_sale(sale)]
        if self.cashbox_filter_start_date:
            try:
                start_date = datetime.datetime.strptime(
                    self.cashbox_filter_start_date, "%Y-%m-%d"
                ).date()
                sales = [
                    sale
                    for sale in sales
                    if self._sale_date(sale) and self._sale_date(sale) >= start_date
                ]
            except ValueError as e:
                logging.exception(f"Error parsing cashbox start date: {e}")
        if self.cashbox_filter_end_date:
            try:
                end_date = datetime.datetime.strptime(
                    self.cashbox_filter_end_date, "%Y-%m-%d"
                ).date()
                sales = [
                    sale
                    for sale in sales
                    if self._sale_date(sale) and self._sale_date(sale) <= end_date
                ]
            except ValueError as e:
                logging.exception(f"Error parsing cashbox end date: {e}")
        
        for sale in sales:
            self._ensure_sale_payment_fields(sale)

        # Split mixed sales (Service + Products) for display
        final_sales = []
        for sale in sales:
            service_items = [i for i in sale["items"] if i.get("category") == "Servicios"]
            product_items = [i for i in sale["items"] if i.get("category") != "Servicios"]
            
            if service_items and product_items:
                # Service Part
                service_sale = sale.copy()
                service_sale["items"] = service_items
                service_total = self._round_currency(sum(i["subtotal"] for i in service_items))
                service_sale["service_total"] = service_total
                service_sale["total"] = service_total
                final_sales.append(service_sale)
                
                # Product Part - Split individually
                for item in product_items:
                    product_sale = sale.copy()
                    product_sale["items"] = [item]
                    product_total = self._round_currency(item["subtotal"])
                    product_sale["service_total"] = product_total
                    product_sale["total"] = product_total
                    
                    # Remove reservation specific details from product row
                    if " | Total:" in product_sale["payment_details"]:
                        product_sale["payment_details"] = product_sale["payment_details"].split(" | Total:")[0]
                    
                    final_sales.append(product_sale)
            else:
                final_sales.append(sale)
                
        return final_sales

    @rx.var
    def paginated_cashbox_sales(self) -> list[CashboxSale]:
        start_index = (self.cashbox_current_page - 1) * self.cashbox_items_per_page
        end_index = start_index + self.cashbox_items_per_page
        return self.filtered_cashbox_sales[start_index:end_index]

    @rx.var
    def cashbox_total_pages(self) -> int:
        total = len(self.filtered_cashbox_sales)
        if total == 0:
            return 1
        return (total + self.cashbox_items_per_page - 1) // self.cashbox_items_per_page

    @rx.var
    def cashbox_close_totals(self) -> list[dict[str, str]]:
        if not self.current_user["privileges"]["view_cashbox"]:
            return []
        return [
            {"method": method, "amount": self._format_currency(amount)}
            for method, amount in self.cashbox_close_summary_totals.items()
            if amount > 0
        ]

    @rx.var
    def cashbox_close_total_amount(self) -> str:
        total_value = sum(self.cashbox_close_summary_totals.values())
        return self._format_currency(total_value)

    @rx.var
    def cashbox_close_sales(self) -> list[CashboxSale]:
        if not self.current_user["privileges"]["view_cashbox"]:
            return []
        return self.cashbox_close_summary_sales

    @rx.var
    def total_ingresos(self) -> float:
        return self._round_currency(
            sum((m["total"] for m in self.history if m["type"] == "Ingreso"))
        )

    @rx.var
    def total_ventas(self) -> float:
        return self._round_currency(
            sum((m["total"] for m in self.history if m["type"] == "Venta"))
        )

    @rx.var
    def ganancia_bruta(self) -> float:
        return self._round_currency(self.total_ventas - self.total_ingresos)

    @rx.var
    def total_movimientos(self) -> int:
        return len(self.history)

    def _ventas_by_payment(self, match_fn) -> float:
        return self._round_currency(
            sum(
                (
                    m["total"]
                    for m in self.history
                    if m["type"] == "Venta" and match_fn(m)
                )
            )
        )

    @rx.var
    def total_ventas_efectivo(self) -> float:
        return self._ventas_by_payment(
            lambda m: m.get("payment_kind") == "cash"
            or m.get("payment_method", "").lower() == "efectivo"
        )

    @rx.var
    def total_ventas_yape(self) -> float:
        return self._ventas_by_payment(
            lambda m: m.get("payment_method", "").lower()
            == "pago qr / billetera digital"
            and "yape" in m.get("payment_details", "").lower()
            or (
                m.get("payment_kind") == "wallet"
                and "yape" in m.get("payment_details", "").lower()
            )
        )

    @rx.var
    def total_ventas_plin(self) -> float:
        return self._ventas_by_payment(
            lambda m: m.get("payment_method", "").lower()
            == "pago qr / billetera digital"
            and "plin" in m.get("payment_details", "").lower()
            or (
                m.get("payment_kind") == "wallet"
                and "plin" in m.get("payment_details", "").lower()
            )
        )

    @rx.var
    def total_ventas_mixtas(self) -> float:
        return self._ventas_by_payment(
            lambda m: m.get("payment_kind") == "mixed"
            or m.get("payment_method", "").lower() == "pagos mixtos"
        )

    @rx.var
    def productos_mas_vendidos(self) -> list[dict]:
        from collections import Counter

        sales = [m for m in self.history if m["type"] == "Venta"]
        product_counts = Counter((m["product_description"] for m in sales))
        top_products = product_counts.most_common(5)
        return [
            {"description": desc, "cantidad_vendida": qty} for desc, qty in top_products
        ]

    @rx.var
    def productos_stock_bajo(self) -> list[Product]:
        return sorted(
            [p for p in self.inventory.values() if p["stock"] <= 10],
            key=lambda p: p["stock"],
        )

    @rx.var
    def sales_by_day(self) -> list[dict]:
        from collections import defaultdict

        sales_data = defaultdict(lambda: {"ingresos": 0, "ventas": 0})
        for m in self.history:
            day = m["timestamp"].split(" ")[0]
            if m["type"] == "Ingreso":
                sales_data[day]["ingresos"] += m["total"]
            elif m["type"] == "Venta":
                sales_data[day]["ventas"] += m["total"]
        sorted_days = sorted(sales_data.keys())
        return [
            {
                "date": day,
                "ingresos": sales_data[day]["ingresos"],
                "ventas": sales_data[day]["ventas"],
            }
            for day in sorted_days
        ]

    @rx.var
    def user_list(self) -> list[User]:
        if not self.current_user["privileges"]["manage_users"]:
            return []
        normalized_users = []
        for user in self.users.values():
            merged_privileges = self._normalize_privileges(user.get("privileges", {}))
            if merged_privileges != user.get("privileges"):
                user["privileges"] = merged_privileges
            normalized_users.append(user)
        return sorted(normalized_users, key=lambda u: u["username"])

    def _navigation_items_config(self) -> list[dict[str, str]]:
        return [
            {"label": "Ingreso", "icon": "arrow-down-to-line", "page": "Ingreso"},
            {"label": "Venta", "icon": "arrow-up-from-line", "page": "Venta"},
            {
                "label": "Gestion de Caja",
                "icon": "wallet",
                "page": "Gestion de Caja",
            },
            {"label": "Inventario", "icon": "boxes", "page": "Inventario"},
            {"label": "Historial", "icon": "history", "page": "Historial"},
            {"label": "Servicios", "icon": "briefcase", "page": "Servicios"},
            {"label": "Configuracion", "icon": "settings", "page": "Configuracion"},
        ]

    def _page_permission_map(self) -> dict[str, str]:
        return {
            "Ingreso": "view_ingresos",
            "Venta": "view_ventas",
            "Gestion de Caja": "view_cashbox",
            "Inventario": "view_inventario",
            "Historial": "view_historial",
            "Servicios": "view_servicios",
            "Configuracion": "manage_users",
        }

    def _can_access_page(self, page: str) -> bool:
        required = self._page_permission_map().get(page)
        if not required:
            return True
        return bool(self.current_user["privileges"].get(required))

    @rx.var
    def navigation_items(self) -> list[dict[str, str]]:
        return [
            item
            for item in self._navigation_items_config()
            if self._can_access_page(item["page"])
        ]

    @rx.var
    def allowed_pages(self) -> list[str]:
        return [item["page"] for item in self.navigation_items]

    @rx.var
    def active_page(self) -> str:
        if self._can_access_page(self.current_page):
            return self.current_page
        if self.allowed_pages:
            return self.allowed_pages[0]
        return self.current_page

    @rx.event
    def set_page(self, page: str):
        if not self._can_access_page(page):
            return rx.toast("No tiene permisos para acceder a este modulo.", duration=3000)
        previous_page = self.current_page
        self.current_page = page
        if page == "Venta" and previous_page != "Venta":
            self._reset_sale_form()
            if not self.reservation_payment_routed:
                self.reservation_payment_id = ""
                self.reservation_payment_amount = ""
        if page != "Servicios":
            self.service_active_tab = "campo"
        if self.sidebar_open:
            pass
        # Siempre limpiar el flag de ruta hacia Venta
        self.reservation_payment_routed = False

    @rx.event
    def set_service_tab(self, tab: str):
        self.service_active_tab = tab

    def _sport_label(self, sport: str) -> str:
        mapping = {"futbol": "Futbol", "voley": "Voley"}
        normalized = (sport or "").lower()
        return mapping.get(normalized, sport or "Campo")

    def _reservation_default_form(self) -> dict[str, str]:
        base_date = self.schedule_selected_date or TODAY_STR
        return {
            "client_name": "",
            "dni": "",
            "phone": "",
            "field_name": "",
            "sport_label": self._sport_label(self.field_rental_sport),
            "selected_price_id": "",
            "date": base_date,
            "start_time": "00:00",
            "end_time": "01:00",
            "advance_amount": "0",
            "total_amount": "0",
            "status": "pendiente",
        }

    def _find_reservation_by_id(self, reservation_id: str) -> FieldReservation | None:
        for reservation in self.service_reservations:
            if reservation["id"] == reservation_id:
                return reservation
        return None

    def _update_reservation_status(self, reservation: FieldReservation):
        if reservation["status"] in ["cancelado", "eliminado"]:
            return
        if reservation["paid_amount"] >= reservation["total_amount"]:
            reservation["status"] = "pagado"
        else:
            reservation["status"] = "pendiente"

    def _log_service_action(
        self,
        reservation: FieldReservation,
        entry_type: str,
        amount: float = 0.0,
        notes: str = "",
        status: str | None = None,
    ):
        self.service_admin_log.insert(
            0,
            {
                "id": str(uuid.uuid4()),
                "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
                "type": entry_type,
                "sport": reservation["sport"],
                "client_name": reservation["client_name"],
                "field_name": reservation["field_name"],
                "amount": self._round_currency(amount),
                "status": status or reservation["status"],
                "notes": notes,
                "reservation_id": reservation["id"],
            },
        )

    def _set_last_reservation_receipt(self, reservation: FieldReservation | None):
        if not reservation:
            self.last_reservation_receipt = None
            return
        balance = max(reservation["total_amount"] - reservation["paid_amount"], 0)
        self.last_reservation_receipt = {
            "cliente": reservation["client_name"],
            "deporte": reservation.get("sport_label", self._sport_label(reservation["sport"])),
            "campo": reservation["field_name"],
            "horario": f"{reservation['start_datetime']} - {reservation['end_datetime']}",
            "monto_adelanto": self._format_currency(reservation["advance_amount"]),
            "monto_total": self._format_currency(reservation["total_amount"]),
            "saldo": self._format_currency(balance),
            "estado": reservation["status"],
        }

    @rx.var
    def service_reservations_for_sport(self) -> list[dict]:
        reservations = []
        search = (self.reservation_search or "").lower()
        filter_status = (self.reservation_filter_status or "todos").lower()
        start_date = (self.reservation_filter_start_date or "").strip()
        end_date = (self.reservation_filter_end_date or "").strip()
        # Compat: if old single-date filter exists, treat as start=end.
        legacy_date = getattr(self, "reservation_filter_date", "")
        if legacy_date and not start_date and not end_date:
            start_date = legacy_date
            end_date = legacy_date
        start_dt = None
        end_dt = None
        try:
            if start_date:
                start_dt = datetime.datetime.strptime(start_date, "%Y-%m-%d").date()
        except Exception:
            start_dt = None
        try:
            if end_date:
                end_dt = datetime.datetime.strptime(end_date, "%Y-%m-%d").date()
        except Exception:
            end_dt = None
        for reservation in self.service_reservations:
            if reservation["sport"] != self.field_rental_sport:
                continue
            if filter_status != "todos" and reservation["status"] != filter_status:
                continue
            if start_dt or end_dt:
                try:
                    res_date = datetime.datetime.strptime(
                        reservation.get("start_datetime", "").split(" ")[0], "%Y-%m-%d"
                    ).date()
                except Exception:
                    continue
                if start_dt and res_date < start_dt:
                    continue
                if end_dt and res_date > end_dt:
                    continue
            if search:
                text = " ".join(
                    [
                        reservation.get("client_name", ""),
                        reservation.get("field_name", ""),
                        reservation.get("start_datetime", ""),
                        reservation.get("end_datetime", ""),
                        reservation.get("status", ""),
                    ]
                ).lower()
                if search not in text:
                    continue
            record = reservation.copy()
            record["balance"] = self._round_currency(
                max(reservation["total_amount"] - reservation["paid_amount"], 0)
            )
            reservations.append(record)
        return reservations

    @rx.var
    def field_prices_for_current_sport(self) -> list[FieldPrice]:
        sport_cmp = (self.field_rental_sport or "").lower()
        return [
            price
            for price in self.field_prices
            if (price.get("sport", "") or "").lower() == sport_cmp
        ]
        # Si no hay precios definidos para el deporte actual, retorna lista vacía.

    def _fill_form_from_price(self, price: FieldPrice):
        self.reservation_form["field_name"] = price.get("name", "")
        self.reservation_form["sport_label"] = price.get("sport", "")
        self._apply_price_total(price)

    @rx.var
    def active_reservation_options(self) -> list[dict[str, str]]:
        options = []
        for reservation in self.service_reservations_for_sport:
            if reservation["status"] not in ["cancelado", "eliminado"]:
                label = f"{reservation['client_name']} - {reservation['start_datetime']}"
                options.append({"id": reservation["id"], "label": label})
        return options

    @rx.var
    def selected_reservation_balance(self) -> float:
        reservation = self._find_reservation_by_id(self.reservation_payment_id)
        if not reservation:
            return 0
        return self._round_currency(
            max(reservation["total_amount"] - reservation["paid_amount"], 0)
        )

    @rx.var
    def reservation_selected_for_payment(self) -> FieldReservation | None:
        return self._find_reservation_by_id(self.reservation_payment_id)

    @rx.var
    def reservation_selected_for_cancel(self) -> FieldReservation | None:
        return self._find_reservation_by_id(self.reservation_cancel_selection)

    @rx.var
    def reservation_selected_for_delete(self) -> FieldReservation | None:
        return self._find_reservation_by_id(self.reservation_delete_selection)

    @rx.var
    def reservation_delete_reason_filled(self) -> bool:
        return bool((self.reservation_delete_reason or "").strip())

    @rx.var
    def reservation_delete_button_disabled(self) -> bool:
        return not self.reservation_delete_reason_filled

    @rx.var
    def modal_reservation(self) -> FieldReservation | None:
        if not self.reservation_modal_reservation_id:
            return None
        return self._find_reservation_by_id(self.reservation_modal_reservation_id)

    @rx.var
    def filtered_service_admin_log(self) -> list[ServiceLogEntry]:
        entries = list(self.service_admin_log)
        if self.service_log_filter_sport != "todos":
            entries = [
                entry
                for entry in entries
                if entry["sport"] == self.service_log_filter_sport
            ]
        if self.service_log_filter_status != "todos":
            entries = [
                entry
                for entry in entries
                if entry["status"] == self.service_log_filter_status
            ]
        try:
            if self.service_log_filter_start_date:
                start = datetime.datetime.strptime(
                    self.service_log_filter_start_date, "%Y-%m-%d"
                )
                entries = [
                    entry
                    for entry in entries
                    if datetime.datetime.strptime(
                        entry["timestamp"].split(" ")[0], "%Y-%m-%d"
                    )
                    >= start
                ]
            if self.service_log_filter_end_date:
                end = datetime.datetime.strptime(
                    self.service_log_filter_end_date, "%Y-%m-%d"
                )
                entries = [
                    entry
                    for entry in entries
                    if datetime.datetime.strptime(
                        entry["timestamp"].split(" ")[0], "%Y-%m-%d"
                    )
                    <= end
                ]
        except ValueError:
            pass
        return entries

    def _slot_has_conflict(
        self, date_str: str, start_time: str, end_time: str, sport: str
    ) -> bool:
        try:
            slot_start = datetime.datetime.strptime(
                f"{date_str} {start_time}", "%Y-%m-%d %H:%M"
            )
            slot_end = datetime.datetime.strptime(
                f"{date_str} {end_time}", "%Y-%m-%d %H:%M"
            )
        except ValueError:
            return False
        for reservation in self.service_reservations:
            if reservation.get("status") in ["cancelado", "eliminado"]:
                continue
            if reservation.get("sport") != sport:
                continue
            if not reservation.get("start_datetime") or not reservation.get("end_datetime"):
                continue
            try:
                res_start = datetime.datetime.strptime(
                    reservation["start_datetime"], "%Y-%m-%d %H:%M"
                )
                res_end = datetime.datetime.strptime(
                    reservation["end_datetime"], "%Y-%m-%d %H:%M"
                )
            except ValueError:
                continue
            if res_start.date().strftime("%Y-%m-%d") != date_str:
                continue
            if slot_start < res_end and slot_end > res_start:
                return True
        return False

    @rx.var
    def schedule_week_days(self) -> list[dict[str, str]]:
        if not self.schedule_selected_week:
            return []
        try:
            year_str, week_str = self.schedule_selected_week.split("-W")
            base_date = datetime.datetime.strptime(
                f"{year_str}-W{week_str}-1", "%G-W%V-%u"
            )
            day_names = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
            days = []
            for offset in range(7):
                day = base_date + datetime.timedelta(days=offset)
                days.append(
                    {
                        "label": f"{day_names[offset]} {day.strftime('%d/%m')}",
                        "date": day.strftime("%Y-%m-%d"),
                    }
                )
            return days
        except ValueError:
            return []

    @rx.var
    def schedule_month_days(self) -> list[dict[str, str]]:
        if not self.schedule_selected_month:
            return []
        try:
            year, month = self.schedule_selected_month.split("-")
            year_int = int(year)
            month_int = int(month)
            _, days_in_month = calendar.monthrange(year_int, month_int)
            return [
                {
                    "label": f"{day:02d}",
                    "date": f"{year_int:04d}-{month_int:02d}-{day:02d}",
                }
                for day in range(1, days_in_month + 1)
            ]
        except (ValueError, IndexError):
            return []

    def _sorted_selected_slots(self) -> list[dict[str, str]]:
        return sorted(
            self.schedule_selected_slots, key=lambda slot: slot.get("start", "")
        )

    def _hours_for_current_selection(self) -> int:
        selection = self._selection_range()
        if self.schedule_selected_slots and selection:
            return max(len(self.schedule_selected_slots), 1)
        start = self.reservation_form.get("start_time", "00:00")
        end = self.reservation_form.get("end_time", "00:00")
        try:
            start_dt = datetime.datetime.strptime(start, "%H:%M")
            end_dt = datetime.datetime.strptime(end, "%H:%M")
        except ValueError:
            return 1
        minutes = int((end_dt - start_dt).total_seconds() / 60)
        if minutes <= 0:
            return 1
        return max(1, math.ceil(minutes / 60))

    def _selection_range(self) -> tuple[str, str] | None:
        slots = self._sorted_selected_slots()
        if not slots:
            return None
        start = slots[0]["start"]
        end = slots[0]["end"]
        for slot in slots[1:]:
            if slot.get("start") != end:
                return None
            end = slot.get("end", end)
        return start, end

    def _clear_schedule_selection(self):
        self.schedule_selected_slots = []

    def _apply_price_total(self, price: FieldPrice):
        hours = self._hours_for_current_selection()
        total = self._round_currency((price.get("price") or 0) * hours)
        self.reservation_form["total_amount"] = f"{total:.2f}"

    def _apply_selected_price_total(self):
        price_id = self.reservation_form.get("selected_price_id", "")
        target = next((p for p in self.field_prices if p["id"] == price_id), None)
        if not target:
            return
        self._apply_price_total(target)

    @rx.var
    def schedule_selected_slots_count(self) -> int:
        return len(self.schedule_selected_slots)

    @rx.var
    def schedule_selection_valid(self) -> bool:
        return self._selection_range() is not None

    @rx.var
    def schedule_selection_label(self) -> str:
        if not self.schedule_selected_slots:
            return "Sin horarios seleccionados"
        slots = self._sorted_selected_slots()
        start = slots[0]["start"]
        end = slots[-1]["end"]
        hours = len(slots)
        if not self._selection_range():
            return f"{start} - {end} (seleccion no consecutiva)"
        suffix = "hora" if hours == 1 else "horas"
        return f"{start} - {end} ({hours} {suffix})"

    @rx.var
    def schedule_slots(self) -> list[dict]:
        date_str = (
            self.schedule_selected_date
            or self.reservation_form.get("date", "")
            or TODAY_STR
        )
        slots: list[dict] = []
        for hour in range(24):
            start = f"{hour:02d}:00"
            end = "23:59" if hour == 23 else f"{hour + 1:02d}:00"
            reserved = self._slot_has_conflict(
                date_str, start, end, self.field_rental_sport
            )
            is_selected = any(
                selected.get("start") == start for selected in self.schedule_selected_slots
            )
            slots.append(
                {
                    "start": start,
                    "end": end,
                    "reserved": reserved,
                    "selected": is_selected,
                }
            )
        return slots

    @rx.event
    def set_field_rental_sport(self, sport: str):
        normalized = (sport or "").lower()
        if normalized not in ["futbol", "voley"]:
            return
        self.field_rental_sport = normalized
        self.reservation_payment_id = ""
        self.reservation_payment_amount = ""
        self.reservation_cancel_selection = ""
        self.reservation_modal_open = False
        self.reservation_modal_mode = "new"
        self.reservation_modal_reservation_id = ""
        self.schedule_selected_date = TODAY_STR
        self.schedule_selected_week = CURRENT_WEEK_STR
        self.schedule_selected_month = CURRENT_MONTH_STR
        self._clear_schedule_selection()
        self.reservation_form = self._reservation_default_form()

    @rx.event
    def set_schedule_view(self, view: str):
        normalized = (view or "").lower()
        if normalized in ["dia", "semana", "mes"]:
            self.schedule_view_mode = normalized

    @rx.event
    def set_schedule_date(self, date: str):
        self.schedule_selected_date = date or ""
        self.update_reservation_form("date", date)
        self._clear_schedule_selection()

    @rx.event
    def set_schedule_week(self, week: str):
        self.schedule_selected_week = week or ""
        self._clear_schedule_selection()

    @rx.event
    def set_schedule_month(self, month: str):
        self.schedule_selected_month = month or ""
        self._clear_schedule_selection()

    @rx.event
    def select_week_day(self, offset: int):
        if not self.schedule_selected_week:
            return rx.toast("Seleccione una semana primero.", duration=2500)
        try:
            year_str, week_str = self.schedule_selected_week.split("-W")
            base_date = datetime.datetime.strptime(
                f"{year_str}-W{week_str}-1", "%G-W%V-%u"
            )
            target = base_date + datetime.timedelta(days=int(offset))
            date_str = target.strftime("%Y-%m-%d")
            self.schedule_selected_date = date_str
            self.update_reservation_form("date", date_str)
            self._clear_schedule_selection()
        except ValueError:
            return rx.toast("Semana invalida.", duration=2500)

    @rx.event
    def select_month_day(self, date: str):
        if not date:
            return
        try:
            parsed = datetime.datetime.strptime(date, "%Y-%m-%d").date()
            self.schedule_selected_month = parsed.strftime("%Y-%m")
            self.schedule_selected_date = parsed.strftime("%Y-%m-%d")
            self.update_reservation_form("date", self.schedule_selected_date)
            self._clear_schedule_selection()
        except ValueError:
            return rx.toast("Dia invalido para el mes seleccionado.", duration=2500)

    @rx.event
    def select_time_slot(self, start_time: str):
        if not start_time:
            return
        try:
            hour_int = int(str(start_time).split(":")[0])
        except ValueError:
            return
        if hour_int < 0 or hour_int > 23:
            return
        start = f"{hour_int:02d}:00"
        end = "23:59" if hour_int == 23 else f"{hour_int + 1:02d}:00"
        date_str = self.schedule_selected_date or self.reservation_form.get("date", "")
        if not date_str:
            date_str = TODAY_STR
        if self._slot_has_conflict(date_str, start, end, self.field_rental_sport):
            return rx.toast("Este horario ya esta reservado. Elige otro.", duration=3000)
        self.reservation_form["start_time"] = start
        self.reservation_form["end_time"] = end
        self.reservation_form["date"] = date_str
        self.schedule_selected_date = date_str
        self.reservation_modal_open = True
        self.reservation_modal_mode = "new"
        self.reservation_modal_reservation_id = ""

    @rx.event
    def toggle_schedule_slot(self, start_time: str, end_time: str):
        if not start_time or not end_time:
            return
        date_str = (
            self.schedule_selected_date or self.reservation_form.get("date", "") or TODAY_STR
        )
        if self._slot_has_conflict(date_str, start_time, end_time, self.field_rental_sport):
            return rx.toast("Este horario ya esta reservado. Elige otro.", duration=3000)
        exists = any(slot.get("start") == start_time for slot in self.schedule_selected_slots)
        if exists:
            self.schedule_selected_slots = [
                slot for slot in self.schedule_selected_slots if slot.get("start") != start_time
            ]
        else:
            self.schedule_selected_slots.append({"start": start_time, "end": end_time})
            self.schedule_selected_slots = self._sorted_selected_slots()
        self.schedule_selected_date = date_str
        self.reservation_form["date"] = date_str
        if self.schedule_selected_slots:
            sorted_slots = self._sorted_selected_slots()
            self.reservation_form["start_time"] = sorted_slots[0]["start"]
            self.reservation_form["end_time"] = sorted_slots[-1]["end"]
            contiguous = self._selection_range()
            if contiguous:
                self.reservation_form["start_time"], self.reservation_form["end_time"] = contiguous
        self._apply_selected_price_total()

    @rx.event
    def clear_schedule_selection(self):
        self._clear_schedule_selection()

    @rx.event
    def open_selected_slots_modal(self):
        date_str = self.schedule_selected_date or TODAY_STR
        selection = self._selection_range()
        if not self.schedule_selected_slots:
            return rx.toast("Selecciona al menos un horario.", duration=2500)
        if not selection:
            return rx.toast("Selecciona horarios consecutivos para la misma reserva.", duration=3000)
        start_time, end_time = selection
        if self._slot_has_conflict(date_str, start_time, end_time, self.field_rental_sport):
            return rx.toast("El rango seleccionado tiene un cruce con otra reserva.", duration=3000)
        # Limpia el formulario antes de preparar una nueva reserva
        self.reservation_form = self._reservation_default_form()
        self.schedule_selected_date = date_str
        self.reservation_form["date"] = date_str
        self.reservation_form["start_time"] = start_time
        self.reservation_form["end_time"] = end_time
        self.reservation_form["sport_label"] = self._sport_label(self.field_rental_sport)
        self.reservation_form["selected_price_id"] = ""
        self.reservation_modal_mode = "new"
        self.reservation_modal_reservation_id = ""
        self.reservation_modal_open = True

    @rx.event
    def open_reservation_modal(self, start_time: str, end_time: str):
        date_str = self.schedule_selected_date or self.reservation_form.get("date", "") or TODAY_STR
        # Prepara un formulario limpio antes de decidir modo
        self.reservation_form = self._reservation_default_form()
        self.schedule_selected_date = date_str
        self.reservation_form["date"] = date_str
        self.reservation_form["start_time"] = start_time
        self.reservation_form["end_time"] = end_time
        self.reservation_form["sport_label"] = self._sport_label(self.field_rental_sport)
        self.reservation_form["selected_price_id"] = ""
        existing = None
        for reservation in self.service_reservations:
            if reservation.get("status") in ["cancelado", "eliminado"]:
                continue
            if reservation.get("sport") != self.field_rental_sport:
                continue
            if reservation.get("start_datetime", "").endswith(start_time) and reservation.get("start_datetime", "").startswith(date_str):
                existing = reservation
                break
        if existing:
            self.reservation_modal_mode = "view"
            self.reservation_modal_reservation_id = existing["id"]
            self.reservation_cancel_selection = existing["id"]
            self.reservation_cancel_reason = ""
        else:
            self.reservation_modal_mode = "new"
            self.reservation_modal_reservation_id = ""
        # Preselecciona el deporte actual en el selector si existe precio
        current_prices = self.field_prices_for_current_sport
        if current_prices:
            self.reservation_form["selected_price_id"] = current_prices[0]["id"]
            self.reservation_form["sport_label"] = current_prices[0].get("sport", self._sport_label(self.field_rental_sport))
        self.reservation_modal_open = True

    @rx.event
    def close_reservation_modal(self):
        self.reservation_modal_open = False
        # Limpia modo y formulario para evitar que datos de vista anterior se mantengan
        self.reservation_modal_mode = "new"
        self.reservation_modal_reservation_id = ""
        self.reservation_form = self._reservation_default_form()

    @rx.event
    def cancel_reservation_from_modal(self):
        if not self.reservation_modal_reservation_id:
            return rx.toast("No hay reserva seleccionada.", duration=2500)
        self.reservation_cancel_selection = self.reservation_modal_reservation_id
        if not self.reservation_cancel_reason:
            self.reservation_cancel_reason = "Cancelado desde planificador."
        return self.cancel_reservation()

    @rx.event
    def pay_reservation_from_modal(self):
        if not self.reservation_modal_reservation_id:
            return rx.toast("Selecciona una reserva primero.", duration=2500)
        self.select_reservation_for_payment(self.reservation_modal_reservation_id)
        return self.pay_reservation_balance()

    @rx.event
    def print_reservation_receipt(self):
        reservation = self.modal_reservation
        if not reservation:
            return rx.toast("No hay reserva seleccionada.", duration=2500)
        if reservation["status"] != "pagado":
            return rx.toast("Solo puedes imprimir cuando la reserva esta pagada.", duration=3000)
        self._set_last_reservation_receipt(reservation)
        return rx.toast("Comprobante generado para impresion.", duration=2500)

    @rx.event
    def update_reservation_form(self, field: str, value: str):
        if field not in self.reservation_form:
            return
        self.reservation_form[field] = value or ""
        if field in ["start_time", "end_time"]:
            self._apply_selected_price_total()

    @rx.event
    def create_field_reservation(self):
        if not self.current_user["privileges"]["manage_reservations"]:
            return rx.toast("No tiene permisos para gestionar reservas.", duration=3000)
        form = self.reservation_form
        name = form.get("client_name", "").strip()
        dni = form.get("dni", "").strip()
        phone = form.get("phone", "").strip()
        field_name = form.get("field_name", "").strip() or f"Campo {self._sport_label(self.field_rental_sport)}"
        date = form.get("date", "").strip()
        start_time = form.get("start_time", "").strip()
        end_time = form.get("end_time", "").strip()
        total_amount = self._safe_amount(form.get("total_amount", "0"))
        advance_amount = self._safe_amount(form.get("advance_amount", "0"))
        status = (form.get("status", "pendiente") or "pendiente").lower()
        if not name or not date or not start_time or not end_time:
            return rx.toast("Complete los datos obligatorios de la reserva.", duration=3000)
        if total_amount <= 0:
            return rx.toast("Ingrese el monto total de la reserva.", duration=3000)
        try:
            start_dt = datetime.datetime.strptime(f"{date} {start_time}", "%Y-%m-%d %H:%M")
            end_dt = datetime.datetime.strptime(f"{date} {end_time}", "%Y-%m-%d %H:%M")
            if end_dt <= start_dt:
                return rx.toast("La hora fin debe ser mayor a la hora inicio.", duration=3000)
        except ValueError:
            return rx.toast("Formato de fecha u hora invalido.", duration=3000)
        if self._slot_has_conflict(date, start_time, end_time, self.field_rental_sport):
            return rx.toast("El horario seleccionado ya esta reservado.", duration=3000)
        paid_amount = min(advance_amount, total_amount)
        if status not in ["pendiente", "pagado"]:
            status = "pendiente"
        if paid_amount >= total_amount:
            status = "pagado"
        elif status == "pagado":
            paid_amount = total_amount
        reservation: FieldReservation = {
            "id": str(uuid.uuid4()),
            "client_name": name,
            "dni": dni,
            "phone": phone,
            "sport": self.field_rental_sport,
            "sport_label": form.get("sport_label", self._sport_label(self.field_rental_sport)),
            "field_name": field_name,
            "start_datetime": start_dt.strftime("%Y-%m-%d %H:%M"),
            "end_datetime": end_dt.strftime("%Y-%m-%d %H:%M"),
            "advance_amount": self._round_currency(advance_amount),
            "total_amount": self._round_currency(total_amount),
            "paid_amount": self._round_currency(paid_amount),
            "status": status,
            "created_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
            "cancellation_reason": "",
            "delete_reason": "",
        }
        self.service_reservations.insert(0, reservation)
        self._log_service_action(reservation, "reserva", 0, notes="Reserva creada", status=reservation["status"])
        if advance_amount > 0:
            self._log_service_action(
                reservation,
                "adelanto",
                advance_amount,
                notes="Adelanto registrado al crear la reserva",
                status=reservation["status"],
            )
            self._register_reservation_advance_in_cashbox(reservation, advance_amount)
        # No preseleccionar pago en Venta hasta que el usuario lo solicite
        self.reservation_payment_id = ""
        self.reservation_payment_amount = ""
        self.reservation_cancel_selection = ""
        self._set_last_reservation_receipt(reservation)
        self._clear_schedule_selection()
        self.reservation_form = self._reservation_default_form()
        self.reservation_modal_open = False
        return rx.toast("Reserva registrada.", duration=3000)

    @rx.event
    def select_reservation_for_payment(self, reservation_id: str):
        self.reservation_payment_id = reservation_id or ""
        reservation = self._find_reservation_by_id(self.reservation_payment_id)
        if reservation:
            balance = max(reservation["total_amount"] - reservation["paid_amount"], 0)
            self.reservation_payment_amount = f"{balance:.2f}" if balance > 0 else ""

    @rx.event
    def go_to_sale_for_reservation(self, reservation_id: str):
        reservation = self._find_reservation_by_id(reservation_id)
        if not reservation:
            return rx.toast("Reserva no encontrada.", duration=3000)
        if reservation["status"] in ["cancelado", "eliminado"]:
            return rx.toast("No se puede cobrar una reserva cancelada o eliminada.", duration=3000)
        self.select_reservation_for_payment(reservation_id)
        self.reservation_payment_routed = True
        return self.set_page("Venta")

    @rx.event
    def set_reservation_payment_amount(self, value: str):
        self.reservation_payment_amount = value or ""

    @rx.event
    def view_reservation_details(self, reservation_id: str):
        reservation = self._find_reservation_by_id(reservation_id)
        if not reservation:
            return rx.toast("Reserva no encontrada.", duration=3000)
        # Llena el formulario con los datos de la reserva para mostrar en el modal.
        try:
            date_part, start_time = reservation.get("start_datetime", "").split(" ")
            _, end_time = reservation.get("end_datetime", "").split(" ")
        except ValueError:
            date_part = reservation.get("start_datetime", "").split(" ")[0] if reservation.get("start_datetime") else TODAY_STR
            start_time = reservation.get("start_datetime", "").split(" ")[1] if " " in reservation.get("start_datetime", "") else ""
            end_time = reservation.get("end_datetime", "").split(" ")[1] if " " in reservation.get("end_datetime", "") else ""
        self.reservation_form = {
            "client_name": reservation.get("client_name", ""),
            "dni": reservation.get("dni", ""),
            "phone": reservation.get("phone", ""),
            "field_name": reservation.get("field_name", ""),
            "sport_label": reservation.get("sport_label", self._sport_label(reservation.get("sport", ""))),
            "selected_price_id": reservation.get("selected_price_id", ""),
            "date": date_part or self.schedule_selected_date or TODAY_STR,
            "start_time": start_time,
            "end_time": end_time,
            "advance_amount": str(reservation.get("advance_amount", 0)),
            "total_amount": str(reservation.get("total_amount", 0)),
            "status": reservation.get("status", "pendiente"),
        }
        self.reservation_modal_reservation_id = reservation_id
        self.reservation_modal_mode = "view"
        self.reservation_cancel_selection = reservation_id
        self.reservation_cancel_reason = ""
        self.reservation_modal_open = True

    @rx.event
    def apply_reservation_payment(self):
        reservation = self._find_reservation_by_id(self.reservation_payment_id)
        if not reservation:
            return rx.toast("Seleccione una reserva para registrar el pago.", duration=3000)
        if reservation["status"] in ["cancelado", "eliminado"]:
            return rx.toast("No se pueden registrar pagos en una reserva cancelada o eliminada.", duration=3000)
        amount = self._safe_amount(self.reservation_payment_amount)
        if amount <= 0:
            return rx.toast("Ingrese un monto valido.", duration=3000)
        balance = max(reservation["total_amount"] - reservation["paid_amount"], 0)
        if balance <= 0:
            self.reservation_payment_amount = ""
            return rx.toast("La reserva ya esta pagada.", duration=3000)
        applied_amount = min(amount, balance)
        reservation["paid_amount"] = self._round_currency(
            reservation["paid_amount"] + applied_amount
        )
        self._update_reservation_status(reservation)
        entry_type = "pago" if reservation["paid_amount"] >= reservation["total_amount"] else "adelanto"
        notes = "Pago completado" if entry_type == "pago" else "Pago parcial registrado"
        self._log_service_action(
            reservation,
            entry_type,
            applied_amount,
            notes=notes,
            status=reservation["status"],
        )
        self.reservation_payment_amount = ""
        self._set_last_reservation_receipt(reservation)
        return rx.toast("Pago registrado correctamente.", duration=3000)

    @rx.event
    def pay_reservation_with_payment_method(self):
        reservation = self._find_reservation_by_id(self.reservation_payment_id)
        if not reservation:
            return rx.toast("Seleccione una reserva desde Servicios -> Pagar.", duration=3000)
        if reservation["status"] in ["cancelado", "eliminado"]:
            return rx.toast("No se puede cobrar una reserva cancelada o eliminada.", duration=3000)
        balance = max(reservation["total_amount"] - reservation["paid_amount"], 0)
        if balance <= 0:
            return rx.toast("La reserva ya esta pagada.", duration=3000)
        if not self.payment_method:
            return rx.toast("Seleccione un metodo de pago.", duration=3000)

        # Validaciones segun metodo elegido, usando el saldo como total objetivo.
        if self.payment_method_kind == "cash":
            self._update_cash_feedback(total_override=balance)
            if self.payment_cash_status not in ["exact", "change"]:
                message = self.payment_cash_message or "Ingrese un monto valido en efectivo."
                return rx.toast(message, duration=3000)
        if self.payment_method_kind == "mixed":
            self._update_mixed_message(total_override=balance)
            if self.payment_mixed_status not in ["exact", "change"]:
                message = self.payment_mixed_message or "Complete los montos del pago mixto."
                return rx.toast(message, duration=3000)

        applied_amount = balance
        reservation["paid_amount"] = self._round_currency(reservation["paid_amount"] + applied_amount)
        self._update_reservation_status(reservation)
        entry_type = "pago" if reservation["paid_amount"] >= reservation["total_amount"] else "adelanto"
        payment_summary = self._generate_payment_summary()
        self._log_service_action(
            reservation,
            entry_type,
            applied_amount,
            notes=payment_summary,
            status=reservation["status"],
        )
        self.reservation_payment_amount = ""
        self._set_last_reservation_receipt(reservation)
        # Limpia montos de la UI de metodo de pago.
        self.payment_cash_amount = 0
        self.payment_cash_message = ""
        self.payment_cash_status = "neutral"
        self.payment_mixed_cash = 0
        self.payment_mixed_card = 0
        self.payment_mixed_wallet = 0
        self.payment_mixed_message = ""
        self.payment_mixed_status = "neutral"
        return rx.toast("Pago registrado con metodo de pago.", duration=3000)

    @rx.event
    def pay_reservation_balance(self):
        reservation = self._find_reservation_by_id(self.reservation_payment_id)
        if not reservation:
            return rx.toast("Seleccione una reserva para pagar el saldo.", duration=3000)
        if reservation["status"] in ["cancelado", "eliminado"]:
            return rx.toast("No se pueden registrar pagos en una reserva cancelada o eliminada.", duration=3000)
        balance = max(reservation["total_amount"] - reservation["paid_amount"], 0)
        if balance <= 0:
            return rx.toast("La reserva ya esta pagada.", duration=3000)
        self.reservation_payment_amount = f"{balance:.2f}"
        return self.apply_reservation_payment()

    @rx.event
    def select_reservation_to_cancel(self, reservation_id: str):
        self.reservation_cancel_selection = reservation_id or ""

    @rx.event
    def set_reservation_cancel_reason(self, reason: str):
        self.reservation_cancel_reason = reason or ""

    @rx.event
    def start_reservation_delete(self, reservation_id: str):
        if not self.current_user["privileges"]["manage_reservations"]:
            return rx.toast("No tiene permisos para eliminar reservas.", duration=3000)
        reservation = self._find_reservation_by_id(reservation_id)
        if not reservation:
            return rx.toast("Reserva no encontrada.", duration=3000)
        if reservation["status"] == "eliminado":
            return rx.toast("La reserva ya esta eliminada.", duration=3000)
        self.reservation_delete_selection = reservation_id
        self.reservation_delete_reason = ""
        self.reservation_delete_modal_open = True

    @rx.event
    def set_reservation_delete_reason(self, reason: str):
        self.reservation_delete_reason = reason or ""

    @rx.event
    def close_reservation_delete_modal(self):
        self.reservation_delete_modal_open = False
        self.reservation_delete_selection = ""
        self.reservation_delete_reason = ""

    @rx.event
    def set_reservation_delete_modal_open(self, open_state: bool):
        """Control the delete modal open state (Radix on_open_change compatibility)."""
        if open_state:
            self.reservation_delete_modal_open = True
        else:
            self.close_reservation_delete_modal()

    @rx.event
    def confirm_reservation_delete(self):
        reservation = self._find_reservation_by_id(self.reservation_delete_selection)
        if not reservation:
            self.close_reservation_delete_modal()
            return rx.toast("Reserva no encontrada.", duration=3000)
        reason = (self.reservation_delete_reason or "").strip()
        if not reason:
            return rx.toast("Ingresa un sustento para eliminar la reserva.", duration=3000)
        if reservation["status"] == "eliminado":
            self.close_reservation_delete_modal()
            return rx.toast("La reserva ya esta eliminada.", duration=3000)
        reservation["status"] = "eliminado"
        reservation["delete_reason"] = reason
        reservation["cancellation_reason"] = reservation.get("cancellation_reason", "") or reason
        # Reassign list to trigger UI refresh and slot availability
        self.service_reservations = [res for res in self.service_reservations]
        self._log_service_action(
            reservation,
            "eliminacion",
            0,
            notes=reason,
            status="eliminado",
        )
        self.reservation_payment_id = ""
        self.reservation_payment_amount = ""
        self.reservation_cancel_selection = ""
        self.reservation_cancel_reason = ""
        self._set_last_reservation_receipt(reservation)
        self.close_reservation_delete_modal()
        return rx.toast("Reserva eliminada y marcada en el historial.", duration=3000)

    @rx.event
    def cancel_reservation(self):
        reservation = self._find_reservation_by_id(self.reservation_cancel_selection)
        if not reservation:
            return rx.toast("Seleccione una reserva a cancelar.", duration=3000)
        if reservation["status"] == "cancelado":
            return rx.toast("La reserva ya esta cancelada.", duration=3000)
        if reservation["status"] == "eliminado":
            return rx.toast("La reserva ya fue eliminada.", duration=3000)
        reservation["status"] = "cancelado"
        reservation["cancellation_reason"] = (
            self.reservation_cancel_reason.strip() or "Sin motivo especificado"
        )
        self._log_service_action(
            reservation,
            "cancelacion",
            0,
            notes=reservation["cancellation_reason"],
            status="cancelado",
        )
        self.reservation_cancel_selection = ""
        self.reservation_cancel_reason = ""
        self.reservation_payment_id = ""
        self.reservation_payment_amount = ""
        self._set_last_reservation_receipt(reservation)
        return rx.toast("Reserva cancelada y registrada en el historial.", duration=3000)

    @rx.event
    def set_service_log_filter_start_date(self, value: str):
        self.service_log_filter_start_date = value or ""

    @rx.event
    def set_service_log_filter_end_date(self, value: str):
        self.service_log_filter_end_date = value or ""

    @rx.event
    def set_service_log_filter_sport(self, value: str):
        self.service_log_filter_sport = (value or "todos").lower()

    @rx.event
    def set_service_log_filter_status(self, value: str):
        self.service_log_filter_status = (value or "todos").lower()

    @rx.event
    def reset_service_log_filters(self):
        self.service_log_filter_start_date = ""
        self.service_log_filter_end_date = ""
        self.service_log_filter_sport = "todos"
        self.service_log_filter_status = "todos"

    @rx.event
    def set_reservation_search(self, value: str):
        self.reservation_staged_search = value or ""

    @rx.event
    def set_reservation_filter_status(self, value: str):
        self.reservation_staged_status = (value or "todos").lower()

    @rx.event
    def set_reservation_filter_date(self, value: str):
        self.reservation_filter_start_date = value or ""
        self.reservation_filter_end_date = value or ""

    @rx.event
    def set_reservation_filter_start_date(self, value: str):
        self.reservation_staged_start_date = value or ""

    @rx.event
    def set_reservation_filter_end_date(self, value: str):
        self.reservation_staged_end_date = value or ""

    @rx.event
    def apply_reservation_filters(self):
        self.reservation_search = self.reservation_staged_search
        self.reservation_filter_status = self.reservation_staged_status
        self.reservation_filter_start_date = self.reservation_staged_start_date
        self.reservation_filter_end_date = self.reservation_staged_end_date

    @rx.event
    def reset_reservation_filters(self):
        self.reservation_search = ""
        self.reservation_filter_status = "todos"
        self.reservation_filter_start_date = ""
        self.reservation_filter_end_date = ""
        self.reservation_staged_search = ""
        self.reservation_staged_status = "todos"
        self.reservation_staged_start_date = ""
        self.reservation_staged_end_date = ""

    @rx.event
    def export_reservations_excel(self):
        try:
            reservations = self.service_reservations_for_sport
            if not reservations:
                return rx.toast("No hay reservas para exportar con los filtros actuales.", duration=3000)
            import openpyxl
            from io import BytesIO

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Reservas"
            sheet.append(
                [
                    "Cliente",
                    "DNI",
                    "Telefono",
                    "Campo",
                    "Deporte",
                    "Inicio",
                    "Fin",
                    "Estado",
                    "Adelanto",
                    "Pagado",
                    "Total",
                    "Saldo",
                    "Creado",
                    "Motivo cancelacion/eliminacion",
                ]
            )
            for r in reservations:
                balance = self._round_currency(r["total_amount"] - r["paid_amount"])
                sheet.append(
                    [
                        r.get("client_name", ""),
                        r.get("dni", ""),
                        r.get("phone", ""),
                        r.get("field_name", ""),
                        r.get("sport_label", r.get("sport", "")),
                        r.get("start_datetime", ""),
                        r.get("end_datetime", ""),
                        r.get("status", ""),
                        self._round_currency(r.get("advance_amount", 0)),
                        self._round_currency(r.get("paid_amount", 0)),
                        self._round_currency(r.get("total_amount", 0)),
                        balance,
                        r.get("created_at", ""),
                        r.get("cancellation_reason", "") or r.get("delete_reason", ""),
                    ]
                )
            file_stream = BytesIO()
            workbook.save(file_stream)
            file_stream.seek(0)
            return rx.download(data=file_stream.read(), filename="reservas.xlsx")
        except Exception as e:
            logging.exception(f"Error exportando reservas: {e}")
            return rx.toast("No se pudo exportar las reservas.", duration=3000)

    @rx.event
    def set_new_field_price_sport(self, value: str):
        self.new_field_price_sport = value or ""

    @rx.event
    def set_new_field_price_name(self, value: str):
        self.new_field_price_name = value or ""

    @rx.event
    def set_new_field_price_amount(self, value: Union[str, float, int]):
        try:
            self.new_field_price_amount = str(value) if value is not None else "0"
        except Exception:
            self.new_field_price_amount = "0"

    @rx.event
    def add_field_price(self):
        if not self.current_user["privileges"]["manage_config"]:
            return rx.toast("No tiene permisos para configurar el sistema.", duration=3000)
        name = self.new_field_price_name.strip()
        sport_raw = (self.new_field_price_sport or "").strip()
        sport_cmp = sport_raw.lower()
        amount = int(self._safe_amount(self.new_field_price_amount))
        if not name:
            return rx.toast("Ingrese un nombre para la modalidad.", duration=2500)
        for price in self.field_prices:
            if price["name"].lower() == name.lower() and price["sport"].lower() == sport_cmp:
                return rx.toast("Ya existe un precio para esta modalidad.", duration=2500)
        self.field_prices.append(
            {
                "id": str(uuid.uuid4()),
                "sport": sport_raw,
                "name": name,
                "price": amount,
            }
        )
        self.new_field_price_name = ""
        self.new_field_price_amount = ""
        self.new_field_price_sport = ""
        self.editing_field_price_id = ""
        return rx.toast("Precio de campo agregado.", duration=2500)

    @rx.event
    def remove_field_price(self, price_id: str):
        self.field_prices = [p for p in self.field_prices if p["id"] != price_id]

    @rx.event
    def update_field_price_amount(self, price_id: str, value: str):
        amount = int(self._safe_amount(value))
        for price in self.field_prices:
            if price["id"] == price_id:
                price["price"] = amount
                break

    @rx.event
    def update_field_price(self):
        if not self.current_user["privileges"]["manage_config"]:
            return rx.toast("No tiene permisos para configurar el sistema.", duration=3000)
        if not self.editing_field_price_id:
            return rx.toast("Seleccione un precio para editar primero.", duration=2500)

        name = self.new_field_price_name.strip()
        sport_raw = (self.new_field_price_sport or "").strip()
        sport_cmp = sport_raw.lower()
        amount = int(self._safe_amount(self.new_field_price_amount))

        if not name:
            return rx.toast("Ingrese un nombre para la modalidad.", duration=2500)

        target = next(
            (p for p in self.field_prices if p["id"] == self.editing_field_price_id),
            None,
        )
        if not target:
            self.editing_field_price_id = ""
            return rx.toast(
                "No se encontro el precio seleccionado.", duration=2500
            )

        for price in self.field_prices:
            if (
                price["id"] != self.editing_field_price_id
                and price["name"].lower() == name.lower()
                and price["sport"].lower() == sport_cmp
            ):
                return rx.toast(
                    "Ya existe un precio para esta modalidad.", duration=2500
                )

        target["name"] = name
        target["sport"] = sport_raw
        target["price"] = amount

        self.editing_field_price_id = ""
        self.new_field_price_name = ""
        self.new_field_price_amount = ""
        self.new_field_price_sport = ""

        return rx.toast("Precio de campo actualizado.", duration=2500)

    @rx.event
    def edit_field_price(self, price_id: str):
        target = next((p for p in self.field_prices if p["id"] == price_id), None)
        if not target:
            return
        self.editing_field_price_id = price_id
        self.new_field_price_sport = target.get("sport", "")
        self.new_field_price_name = target.get("name", "")
        self.new_field_price_amount = str(target.get("price", "0"))

    @rx.event
    def select_reservation_field_price(self, price_id: str):
        target = next((p for p in self.field_prices if p["id"] == price_id), None)
        if not target:
            return
        self.reservation_form["selected_price_id"] = price_id
        self._fill_form_from_price(target)

    @rx.event
    def set_reservation_sport_from_price(self, price_id: str):
        target = next((p for p in self.field_prices if p["id"] == price_id), None)
        if not target:
            self.reservation_form["sport_label"] = ""
            self.reservation_form["selected_price_id"] = ""
            return
        self.reservation_form["sport_label"] = target.get("sport", "")
        self.reservation_form["selected_price_id"] = price_id

    @rx.event
    def toggle_sidebar(self):
        self.sidebar_open = not self.sidebar_open

    @rx.event
    def set_config_tab(self, tab: str | dict):
        if isinstance(tab, dict):
            return
        self.config_active_tab = tab

    @rx.event
    def go_to_config_tab(self, tab: str):
        if not self._can_access_page("Configuracion"):
            return rx.toast("No tiene permisos para acceder a configuracion.", duration=3000)
        self.set_page("Configuracion")
        self.set_config_tab(tab)

    @rx.event
    def update_new_category_name(self, value: str):
        self.new_category_name = value

    @rx.event
    def add_category(self):
        name = self.new_category_name.strip()
        if not name:
            return rx.toast("El nombre de la categoría no puede estar vacío.", duration=3000)
        if name in self.categories:
            return rx.toast("La categoría ya existe.", duration=3000)
        self.categories.append(name)
        if not self.new_entry_item["category"]:
            self.new_entry_item["category"] = name
        self.new_category_name = ""
        return rx.toast(f"Categoría {name} creada.", duration=3000)

    @rx.event
    def remove_category(self, name: str):
        if name == "General":
            return rx.toast("No puedes eliminar la categoría predeterminada.", duration=3000)
        if name not in self.categories:
            return
        self.categories = [cat for cat in self.categories if cat != name]
        fallback = self.categories[0] if self.categories else ""
        if self.new_entry_item["category"] == name:
            self.new_entry_item["category"] = fallback
        if self.new_sale_item["category"] == name:
            self.new_sale_item["category"] = fallback
        for product in self.inventory.values():
            if product.get("category") == name:
                product["category"] = fallback
        return rx.toast(f"Categoría {name} eliminada.", duration=3000)

    @rx.event
    def handle_entry_change(self, field: str, value: str):
        try:
            if field in ["quantity", "price", "sale_price"]:
                numeric = float(value) if value else 0
                if field == "quantity":
                    self.new_entry_item[field] = self._normalize_quantity_value(
                        numeric, self.new_entry_item.get("unit", "")
                    )
                else:
                    self.new_entry_item[field] = self._round_currency(numeric)
            else:
                self.new_entry_item[field] = value
                if field == "unit":
                    self.new_entry_item["quantity"] = self._normalize_quantity_value(
                        self.new_entry_item["quantity"], value
                    )
            self.new_entry_item["subtotal"] = self._round_currency(
                self.new_entry_item["quantity"] * self.new_entry_item["price"]
            )
            if field == "description":
                if value:
                    search = str(value).lower()
                    self.entry_autocomplete_suggestions = [
                        p["description"]
                        for p in self.inventory.values()
                        if search in p["description"].lower()
                    ][:5]
                else:
                    self.entry_autocomplete_suggestions = []
            elif field == "barcode":
                # Si se borra el código de barras, limpiar todos los campos
                if not value or not str(value).strip():
                    self.new_entry_item["barcode"] = ""
                    self.new_entry_item["description"] = ""
                    self.new_entry_item["quantity"] = 0
                    self.new_entry_item["price"] = 0
                    self.new_entry_item["sale_price"] = 0
                    self.new_entry_item["subtotal"] = 0
                    self.entry_autocomplete_suggestions = []
                else:
                    # Limpiar el código de barras usando la utilidad
                    code = clean_barcode(str(value))
                    # Solo buscar si el código es válido
                    if validate_barcode(code):
                        product = self._find_product_by_barcode(code)
                        if product:
                            self._fill_entry_item_from_product(product)
                            self.entry_autocomplete_suggestions = []
                            return rx.toast(f"Producto '{product['description']}' cargado", duration=2000)
        except ValueError as e:
            logging.exception(f"Error parsing entry value: {e}")

    @rx.event
    def process_entry_barcode_from_input(self, barcode_value):
        """Procesa el barcode del input cuando pierde el foco"""
        # Actualizar el estado con el valor del input
        self.new_entry_item["barcode"] = str(barcode_value) if barcode_value else ""
        
        if not barcode_value or not str(barcode_value).strip():
            self.new_entry_item["barcode"] = ""
            self.new_entry_item["description"] = ""
            self.new_entry_item["quantity"] = 0
            self.new_entry_item["price"] = 0
            self.new_entry_item["sale_price"] = 0
            self.new_entry_item["subtotal"] = 0
            self.entry_autocomplete_suggestions = []
            return
        
        code = clean_barcode(str(barcode_value))
        if validate_barcode(code):
            product = self._find_product_by_barcode(code)
            if product:
                self._fill_entry_item_from_product(product)
                self.entry_autocomplete_suggestions = []
                return rx.toast(f"Producto '{product['description']}' cargado", duration=2000)

    @rx.event
    def handle_barcode_enter(self, key: str, input_id: str):
        """Detecta la tecla Enter y fuerza el blur del input para procesar"""
        if key == "Enter":
            return rx.call_script(f"document.getElementById('{input_id}').blur()")

    @rx.event
    def add_item_to_entry(self):
        if not self.current_user["privileges"]["create_ingresos"]:
            return rx.toast("No tiene permisos para crear ingresos.", duration=3000)
        if (
            not self.new_entry_item["description"]
            or not self.new_entry_item["category"]
            or self.new_entry_item["quantity"] <= 0
            or self.new_entry_item["price"] <= 0
            or self.new_entry_item["sale_price"] <= 0
        ):
            return rx.toast(
                "Por favor, complete todos los campos correctamente.", duration=3000
            )
        item_copy = self.new_entry_item.copy()
        item_copy["temp_id"] = str(uuid.uuid4())
        self._apply_item_rounding(item_copy)
        self.new_entry_items.append(item_copy)
        self._reset_entry_form()

    @rx.event
    def remove_item_from_entry(self, temp_id: str):
        self.new_entry_items = [
            item for item in self.new_entry_items if item["temp_id"] != temp_id
        ]

    @rx.event
    def update_entry_item_category(self, temp_id: str, category: str):
        for item in self.new_entry_items:
            if item["temp_id"] == temp_id:
                item["category"] = category
                break

    @rx.event
    def edit_item_from_entry(self, temp_id: str):
        for item in self.new_entry_items:
            if item["temp_id"] == temp_id:
                self.new_entry_item = item.copy()
                self.new_entry_items = [
                    entry for entry in self.new_entry_items if entry["temp_id"] != temp_id
                ]
                self.entry_autocomplete_suggestions = []
                return

    @rx.event
    def confirm_entry(self):
        if not self.current_user["privileges"]["create_ingresos"]:
            return rx.toast("No tiene permisos para crear ingresos.", duration=3000)
        if not self.new_entry_items:
            return rx.toast("No hay productos para ingresar.", duration=3000)
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for item in self.new_entry_items:
            product_id = item["description"].lower().strip()
            barcode = item.get("barcode", "").strip()
            sale_price = item.get("sale_price", 0)
            purchase_price = item["price"]
            if product_id in self.inventory:
                if "barcode" not in self.inventory[product_id]:
                    self.inventory[product_id]["barcode"] = ""
                self.inventory[product_id]["stock"] += item["quantity"]
                self.inventory[product_id]["stock"] = self._normalize_quantity_value(
                    self.inventory[product_id]["stock"],
                    self.inventory[product_id]["unit"],
                )
                self.inventory[product_id]["purchase_price"] = purchase_price
                if barcode:
                    self.inventory[product_id]["barcode"] = barcode
                if sale_price > 0:
                    self.inventory[product_id]["sale_price"] = sale_price
                self.inventory[product_id]["category"] = item["category"]
            else:
                default_sale_price = (
                    sale_price if sale_price > 0 else purchase_price * 1.25
                )
                self.inventory[product_id] = {
                    "id": product_id,
                    "barcode": barcode,
                    "description": item["description"],
                    "category": item["category"],
                    "stock": item["quantity"],
                    "unit": item["unit"],
                    "purchase_price": purchase_price,
                    "sale_price": self._round_currency(default_sale_price),
                }
            self.history.append(
                {
                    "id": str(uuid.uuid4()),
                    "timestamp": timestamp,
                    "type": "Ingreso",
                    "product_description": item["description"],
                    "quantity": item["quantity"],
                    "unit": item["unit"],
                    "total": item["subtotal"],
                    "payment_method": "",
                    "payment_details": "",
                    "user": self.current_user["username"],
                    "sale_id": "",
                }
            )
        self.new_entry_items = []
        self._reset_entry_form()
        return rx.toast("Ingreso de productos confirmado.", duration=3000)

    def _reset_entry_form(self):
        self.entry_form_key += 1
        self.new_entry_item = {
            "temp_id": "",
            "barcode": "",
            "description": "",
            "category": self.categories[0] if self.categories else "",
            "quantity": 0,
            "unit": "Unidad",
            "price": 0,
            "sale_price": 0,
            "subtotal": 0,
        }
        self.entry_autocomplete_suggestions = []

    def _find_product_by_barcode(self, barcode: str) -> Product | None:
        """Busca un producto por código de barras usando limpieza y validación"""
        code = clean_barcode(barcode)
        if not code or len(code) == 0:
            return None
        
        logging.info(f"[BUSCAR] Código limpio: '{code}' (long: {len(code)})")
        
        # Buscar coincidencia exacta
        for product in self.inventory.values():
            product_code = product.get("barcode", "")
            if product_code:
                clean_product_code = clean_barcode(product_code)
                if clean_product_code == code:
                    logging.info(f"[BUSCAR] ✓ Exacta: '{clean_product_code}' = '{code}'")
                    return product
        
        logging.warning(f"[BUSCAR] ✗ Sin coincidencia exacta para: '{code}'")
        return None

    def _sale_date(self, sale: CashboxSale):
        try:
            return datetime.datetime.strptime(
                sale["timestamp"], "%Y-%m-%d %H:%M:%S"
            ).date()
        except ValueError:
            return None

    def _is_advance_sale(self, sale: CashboxSale) -> bool:
        if sale.get("is_deleted"):
            return False
        if sale.get("is_advance"):
            return True
        label = (sale.get("payment_label") or "").lower()
        details = (sale.get("payment_details") or "").lower()
        description = " ".join(item.get("description", "") for item in sale.get("items", []))
        return (
            "adelanto" in label
            or "adelanto" in details
            or "adelanto" in description.lower()
        )

    def _payment_label_and_breakdown(self, sale_total: float) -> tuple[str, list[PaymentBreakdownItem]]:
        kind = (self.payment_method_kind or "other").lower()
        method_name = self._normalize_wallet_label(self.payment_method or "Metodo")
        breakdown: list[PaymentBreakdownItem] = []
        label = method_name
        if kind == "cash":
            label = "Efectivo"
            breakdown = [{"label": label, "amount": self._round_currency(sale_total)}]
        elif kind == "card":
            card_type = self.payment_card_type or "Tarjeta"
            label = f"{method_name} ({card_type})"
            breakdown = [{"label": label, "amount": self._round_currency(sale_total)}]
        elif kind == "wallet":
            provider = self.payment_wallet_provider or self.payment_wallet_choice or "Billetera"
            label = f"{method_name} ({provider})"
            breakdown = [{"label": label, "amount": self._round_currency(sale_total)}]
        elif kind == "mixed":
            paid_cash = self._round_currency(self.payment_mixed_cash)
            paid_card = self._round_currency(self.payment_mixed_card)
            paid_wallet = self._round_currency(self.payment_mixed_wallet)
            provider = self.payment_wallet_provider or self.payment_wallet_choice or "Billetera"
            remaining = self._round_currency(sale_total)
            parts: list[PaymentBreakdownItem] = []
            if paid_card > 0:
                applied_card = min(paid_card, remaining)
                if applied_card > 0:
                    parts.append(
                        {
                            "label": f"Tarjeta ({self.payment_card_type})",
                            "amount": self._round_currency(applied_card),
                        }
                    )
                    remaining = self._round_currency(remaining - applied_card)
            if paid_wallet > 0 and remaining > 0:
                applied_wallet = min(paid_wallet, remaining)
                if applied_wallet > 0:
                    parts.append({"label": provider, "amount": self._round_currency(applied_wallet)})
                    remaining = self._round_currency(remaining - applied_wallet)
            if paid_cash > 0 and remaining > 0:
                applied_cash = min(paid_cash, remaining)
                if applied_cash > 0:
                    parts.append({"label": "Efectivo", "amount": self._round_currency(applied_cash)})
                    remaining = self._round_currency(remaining - applied_cash)
            if not parts:
                breakdown = [{"label": method_name, "amount": self._round_currency(sale_total)}]
            else:
                if remaining > 0:
                    parts[0]["amount"] = self._round_currency(parts[0]["amount"] + remaining)
                breakdown = parts
            labels = [p["label"] for p in breakdown]
            detail = ", ".join(labels) if labels else method_name
            label = f"{method_name} ({detail})"
        else:
            label = method_name or "Otros"
            breakdown = [{"label": label, "amount": self._round_currency(sale_total)}]
        return label, breakdown

    def _normalize_wallet_label(self, label: str) -> str:
        value = (label or "").strip()
        if not value:
            return value
        lower = value.lower()
        if "pago qr" in lower or "billetera" in lower:
            patterns = [
                "Pago QR / Billetera Digital",
                "Pago QR / Billetera",
                "Pago QR/Billetera Digital",
                "Pago QR/Billetera",
            ]
            for old in patterns:
                value = value.replace(old, "Billetera Digital / QR")
            return value
        return value

    def _payment_category(self, method: str, kind: str = "") -> str:
        normalized_kind = (kind or "").lower()
        label = method.lower() if method else ""
        if normalized_kind == "mixed" or "mixto" in label:
            return "Pago Mixto"
        if normalized_kind == "card" or "tarjeta" in label:
            return "Tarjeta"
        if normalized_kind == "wallet" or "qr" in label or "billetera" in label or "yape" in label or "plin" in label:
            return "Billetera Digital / QR"
        if normalized_kind == "cash" or "efectivo" in label:
            return "Efectivo"
        return "Otros"

    def _get_day_sales(self, date: str) -> list[CashboxSale]:
        day_sales = [
            sale
            for sale in self.cashbox_sales
            if sale["timestamp"].startswith(date) and not sale.get("is_deleted")
        ]
        for sale in day_sales:
            self._ensure_sale_payment_fields(sale)
        return day_sales

    def _register_reservation_advance_in_cashbox(
        self, reservation: FieldReservation, advance_amount: float
    ):
        """Registra en caja un adelanto si la caja esta abierta."""
        amount = self._round_currency(advance_amount)
        if amount <= 0:
            return
        if not self.cashbox_is_open:
            return
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sale_id = str(uuid.uuid4())
        description = (
            f"Adelanto {reservation['field_name']} "
            f"({reservation['start_datetime']} - {reservation['end_datetime']})"
        )
        self.cashbox_sales.append(
            {
                "sale_id": sale_id,
                "timestamp": timestamp,
                "user": self.current_user["username"],
                "payment_method": "Efectivo",
                "payment_kind": "cash",
                "payment_label": "Efectivo (Adelanto)",
                "payment_breakdown": [{"label": "Efectivo", "amount": amount}],
                "payment_details": f"Adelanto registrado al crear la reserva. Monto {self._format_currency(amount)}",
                "total": amount,
                "service_total": amount,
                "items": [
                    {
                        "temp_id": reservation["id"],
                        "barcode": reservation["id"],
                        "description": description,
                        "category": "Servicios",
                        "quantity": 1,
                        "unit": "Servicio",
                        "price": amount,
                        "sale_price": amount,
                        "subtotal": amount,
                    }
                ],
                "is_deleted": False,
                "delete_reason": "",
                "is_advance": True,
            }
        )

    def _ensure_sale_payment_fields(self, sale: CashboxSale):
        if "payment_label" not in sale or not sale.get("payment_label"):
            sale["payment_label"] = sale.get("payment_method", "Metodo")
        sale["payment_label"] = self._normalize_wallet_label(sale.get("payment_label", ""))
        if (
            "payment_breakdown" not in sale
            or not isinstance(sale.get("payment_breakdown"), list)
            or len(sale.get("payment_breakdown") or []) == 0
        ):
            fallback_label = sale.get("payment_label", sale.get("payment_method", "Metodo"))
            sale["payment_breakdown"] = [
                {
                    "label": self._normalize_wallet_label(fallback_label),
                    "amount": self._round_currency(sale.get("total", 0)),
                }
            ]
        else:
            normalized_items: list[PaymentBreakdownItem] = []
            for item in sale.get("payment_breakdown", []):
                normalized_items.append(
                    {
                        "label": self._normalize_wallet_label(item.get("label", "")),
                        "amount": self._round_currency(item.get("amount", 0)),
                    }
                )
            target_total = self._round_currency(sale.get("total", 0))
            total_applied = sum(item["amount"] for item in normalized_items)
            if target_total > 0 and total_applied > target_total:
                factor = target_total / total_applied if total_applied else 0
                normalized_items = [
                    {
                        "label": item["label"],
                        "amount": self._round_currency(item["amount"] * factor),
                    }
                    for item in normalized_items
                ]
                total_applied = sum(item["amount"] for item in normalized_items)
            if target_total > 0 and normalized_items:
                diff = self._round_currency(target_total - total_applied)
                if diff != 0:
                    normalized_items[0]["amount"] = self._round_currency(
                        normalized_items[0]["amount"] + diff
                    )
            sale["payment_breakdown"] = normalized_items
        # Asegura campo de total de servicio para mostrar en caja
        service_total = sale.get("service_total")
        if service_total is None:
            sale["service_total"] = self._round_currency(sale.get("total", 0))

    def _build_cashbox_summary(self, sales: list[CashboxSale]) -> dict[str, float]:
        summary: dict[str, float] = {}
        for sale in sales:
            breakdown = sale.get("payment_breakdown") if isinstance(sale, dict) else []
            if breakdown:
                for item in breakdown:
                    method_label = self._normalize_wallet_label(
                        item.get("label") or sale.get("payment_label") or sale.get("payment_method", "Otros")
                    )
                    amount = self._round_currency(item.get("amount", 0))
                    summary[method_label] = self._round_currency(
                        summary.get(method_label, 0) + amount
                    )
            else:
                category = self._payment_category(
                    self._normalize_wallet_label(sale.get("payment_method", "")),
                    sale.get("payment_kind", ""),
                )
                if category not in summary:
                    summary[category] = 0.0
                summary[category] = self._round_currency(summary[category] + sale["total"])
        return summary

    def _reset_cashbox_close_summary(self):
        self.cashbox_close_modal_open = False
        self.cashbox_close_summary_totals = {}
        self.cashbox_close_summary_sales = []
        self.cashbox_close_summary_date = ""

    def _fill_entry_item_from_product(self, product: Product):
        # IMPORTANTE: Usar el código de barras completo del producto en el inventario
        product_barcode = product.get("barcode", "")
        
        self.new_entry_item["description"] = product["description"]
        self.new_entry_item["unit"] = product["unit"]
        self.new_entry_item["price"] = self._round_currency(product["purchase_price"])
        self.new_entry_item["sale_price"] = self._round_currency(product["sale_price"])
        self.new_entry_item["category"] = product.get("category", self.categories[0] if self.categories else "")
        self.new_entry_item["barcode"] = product_barcode  # Código completo del inventario
        self.new_entry_item["quantity"] = self._normalize_quantity_value(
            self.new_entry_item["quantity"], product["unit"]
        )
        self.new_entry_item["subtotal"] = self._round_currency(
            self.new_entry_item["quantity"] * self.new_entry_item["price"]
        )
        
        logging.info(f"[FILL-ENTRY] Código restaurado a: '{product_barcode}' (del inventario)")

    @rx.event
    def select_product_for_entry(self, description: str):
        if isinstance(description, dict):
            description = (
                description.get("value")
                or description.get("description")
                or description.get("label")
                or ""
            )
        product_id = description.lower().strip()
        if product_id in self.inventory:
            self._fill_entry_item_from_product(self.inventory[product_id])
        self.entry_autocomplete_suggestions = []

    def _fill_sale_item_from_product(
        self, product: Product, keep_quantity: bool = False
    ):
        # Usar el código de barras COMPLETO del inventario (no el parcial escaneado)
        product_barcode = product.get("barcode", "")
        
        quantity = (
            self.new_sale_item["quantity"]
            if keep_quantity and self.new_sale_item["quantity"] > 0
            else 1
        )
        
        # Reemplazar con el código completo del inventario
        self.new_sale_item["barcode"] = product_barcode
        self.new_sale_item["description"] = product["description"]
        self.new_sale_item["category"] = product.get("category", self.categories[0] if self.categories else "")
        self.new_sale_item["unit"] = product["unit"]
        self.new_sale_item["quantity"] = self._normalize_quantity_value(quantity, product["unit"])
        self.new_sale_item["price"] = self._round_currency(product["sale_price"])
        self.new_sale_item["sale_price"] = self._round_currency(product["sale_price"])
        self.new_sale_item["subtotal"] = self._round_currency(
            self.new_sale_item["quantity"] * self.new_sale_item["price"]
        )
        
        if not keep_quantity:
            self.autocomplete_suggestions = []
        
        logging.info(f"[FILL-SALE] Código corregido: escaneado incompleto → '{product_barcode}' completo (producto: {product['description']})")

    def _reset_sale_form(self):
        self.sale_form_key += 1
        self.new_sale_item = {
            "temp_id": "",
            "barcode": "",
            "description": "",
            "category": self.categories[0] if self.categories else "",
            "quantity": 0,
            "unit": "Unidad",
            "price": 0,
            "sale_price": 0,
            "subtotal": 0,
        }
        self.autocomplete_suggestions = []

    def _fill_inventory_adjustment_from_product(self, product: Product):
        stock = self._normalize_quantity_value(
            product.get("stock", 0), product.get("unit", "")
        )
        self.inventory_adjustment_item = {
            "temp_id": "",
            "barcode": product.get("barcode", ""),
            "description": product["description"],
            "category": product.get(
                "category", self.categories[0] if self.categories else ""
            ),
            "unit": product.get("unit", "Unidad"),
            "current_stock": stock,
            "adjust_quantity": 0,
            "reason": "",
        }

    def _reset_inventory_adjustment_form(self):
        self.inventory_adjustment_item = {
            "temp_id": "",
            "barcode": "",
            "description": "",
            "category": "",
            "unit": "",
            "current_stock": 0,
            "adjust_quantity": 0,
            "reason": "",
        }
        self.inventory_adjustment_suggestions = []

    @rx.event
    def handle_sale_change(self, field: str, value: Union[str, float]):
        try:
            if field in ["quantity", "price"]:
                numeric = float(value) if value else 0
                if field == "quantity":
                    self.new_sale_item[field] = self._normalize_quantity_value(
                        numeric, self.new_sale_item.get("unit", "")
                    )
                else:
                    self.new_sale_item[field] = self._round_currency(numeric)
            else:
                self.new_sale_item[field] = value
                if field == "unit":
                    self.new_sale_item["quantity"] = self._normalize_quantity_value(
                        self.new_sale_item["quantity"], value
                    )
            self.new_sale_item["subtotal"] = self._round_currency(
                self.new_sale_item["quantity"] * self.new_sale_item["price"]
            )
            if field == "description":
                if value:
                    self.autocomplete_suggestions = [
                        p["description"]
                        for p in self.inventory.values()
                        if str(value).lower() in p["description"].lower()
                    ][:5]
                else:
                    self.autocomplete_suggestions = []
            elif field == "barcode":
                if not value or not str(value).strip():
                    self.new_sale_item["barcode"] = ""
                    self.new_sale_item["description"] = ""
                    self.new_sale_item["quantity"] = 0
                    self.new_sale_item["price"] = 0
                    self.new_sale_item["subtotal"] = 0
                    self.autocomplete_suggestions = []
                else:
                    code = clean_barcode(str(value))
                    if validate_barcode(code):
                        product = self._find_product_by_barcode(code)
                        if product:
                            self._fill_sale_item_from_product(product, keep_quantity=False)
                            self.autocomplete_suggestions = []
                            return rx.toast(f"Producto '{product['description']}' cargado", duration=2000)
        except ValueError as e:
            logging.exception(f"Error parsing sale value: {e}")

    @rx.event
    def process_sale_barcode_from_input(self, barcode_value):
        """Procesa el barcode del input cuando pierde el foco"""
        # Actualizar el estado con el valor del input
        self.new_sale_item["barcode"] = str(barcode_value) if barcode_value else ""
        
        if not barcode_value or not str(barcode_value).strip():
            self.new_sale_item["description"] = ""
            self.new_sale_item["quantity"] = 0
            self.new_sale_item["price"] = 0
            self.new_sale_item["subtotal"] = 0
            self.autocomplete_suggestions = []
            return
        
        code = clean_barcode(str(barcode_value))
        if validate_barcode(code):
            product = self._find_product_by_barcode(code)
            if product:
                self._fill_sale_item_from_product(product, keep_quantity=False)
                self.autocomplete_suggestions = []
                return rx.toast(f"Producto '{product['description']}' cargado", duration=2000)

    @rx.event
    def handle_inventory_adjustment_change(self, field: str, value: Union[str, float]):
        try:
            if field == "adjust_quantity":
                amount = float(value) if value else 0
                if amount < 0:
                    amount = 0
                self.inventory_adjustment_item["adjust_quantity"] = (
                    self._normalize_quantity_value(
                        amount, self.inventory_adjustment_item.get("unit", "")
                    )
                )
            elif field == "reason":
                self.inventory_adjustment_item["reason"] = value
            elif field == "description":
                self.inventory_adjustment_item["description"] = value
                if value:
                    search = str(value).lower()
                    self.inventory_adjustment_suggestions = [
                        p["description"]
                        for p in self.inventory.values()
                        if search in p["description"].lower()
                    ][:5]
                else:
                    self.inventory_adjustment_suggestions = []
        except ValueError as e:
            logging.exception(f"Error parsing inventory adjustment value: {e}")

    @rx.event
    def select_product_for_sale(self, description: str):
        if isinstance(description, dict):
            description = (
                description.get("value")
                or description.get("description")
                or description.get("label")
                or ""
            )
        product_id = description.lower().strip()
        if product_id in self.inventory:
            product = self.inventory[product_id]
            self._fill_sale_item_from_product(product)
        self.autocomplete_suggestions = []

    @rx.event
    def select_inventory_adjustment_product(self, description: str):
        if isinstance(description, dict):
            description = (
                description.get("value")
                or description.get("description")
                or description.get("label")
                or ""
            )
        product_id = description.lower().strip()
        if product_id in self.inventory:
            self._fill_inventory_adjustment_from_product(self.inventory[product_id])
        self.inventory_adjustment_suggestions = []

    @rx.event
    def add_inventory_adjustment_item(self):
        if not self.current_user["privileges"]["edit_inventario"]:
            return rx.toast("No tiene permisos para ajustar inventario.", duration=3000)
        description = self.inventory_adjustment_item["description"].strip()
        if not description:
            return rx.toast("Seleccione un producto para ajustar.", duration=3000)
        product_id = description.lower()
        if product_id not in self.inventory:
            return rx.toast("Producto no encontrado en el inventario.", duration=3000)
        quantity = self.inventory_adjustment_item["adjust_quantity"]
        if quantity <= 0:
            return rx.toast("Ingrese la cantidad a ajustar.", duration=3000)
        available = self.inventory[product_id]["stock"]
        if quantity > available:
            return rx.toast(
                "La cantidad supera el stock disponible.", duration=3000
            )
        item_copy = self.inventory_adjustment_item.copy()
        item_copy["temp_id"] = str(uuid.uuid4())
        item_copy["adjust_quantity"] = self._normalize_quantity_value(
            item_copy.get("adjust_quantity", 0), item_copy.get("unit", "")
        )
        self.inventory_adjustment_items.append(item_copy)
        self._reset_inventory_adjustment_form()

    @rx.event
    def remove_inventory_adjustment_item(self, temp_id: str):
        self.inventory_adjustment_items = [
            item for item in self.inventory_adjustment_items if item["temp_id"] != temp_id
        ]

    @rx.event
    def add_item_to_sale(self):
        if not self.current_user["privileges"]["create_ventas"]:
            return rx.toast("No tiene permisos para crear ventas.", duration=3000)
        self.sale_receipt_ready = False
        if (
            not self.new_sale_item["description"]
            or self.new_sale_item["quantity"] <= 0
            or self.new_sale_item["price"] <= 0
        ):
            return rx.toast(
                "Por favor, busque un producto y complete los campos.", duration=3000
            )
        product_id = self.new_sale_item["description"].lower().strip()
        if product_id not in self.inventory:
            return rx.toast("Producto no encontrado en el inventario.", duration=3000)
        if self.inventory[product_id]["stock"] < self.new_sale_item["quantity"]:
            return rx.toast("Stock insuficiente para realizar la venta.", duration=3000)
        item_copy = self.new_sale_item.copy()
        item_copy["temp_id"] = str(uuid.uuid4())
        self._apply_item_rounding(item_copy)
        self.new_sale_items.append(item_copy)
        self._reset_sale_form()
        self._refresh_payment_feedback()

    @rx.event
    def remove_item_from_sale(self, temp_id: str):
        self.new_sale_items = [
            item for item in self.new_sale_items if item["temp_id"] != temp_id
        ]
        self.sale_receipt_ready = False
        self._refresh_payment_feedback()
        self.sale_receipt_ready = False

    @rx.event
    def clear_sale_items(self):
        self.new_sale_items = []
        self._reset_sale_form()
        self.sale_receipt_ready = False
        self._refresh_payment_feedback()

    @rx.event
    def confirm_sale(self):
        if not self.current_user["privileges"]["create_ventas"]:
            return rx.toast("No tiene permisos para crear ventas.", duration=3000)
        denial = self._require_cashbox_open()
        if denial:
            return denial
        reservation = (
            self._find_reservation_by_id(self.reservation_payment_id)
            if self.reservation_payment_id
            else None
        )
        if reservation and reservation["status"] in ["cancelado", "eliminado"]:
            return rx.toast("No se puede cobrar una reserva cancelada o eliminada.", duration=3000)
        reservation_balance = (
            self._round_currency(
                max(reservation["total_amount"] - reservation["paid_amount"], 0)
            )
            if reservation
            else 0
        )
        if not self.new_sale_items and reservation_balance <= 0:
            if reservation:
                return rx.toast("La reserva ya esta pagada.", duration=3000)
            return rx.toast("No hay productos en la venta.", duration=3000)
        if not self.payment_method:
            return rx.toast("Seleccione un metodo de pago.", duration=3000)

        product_snapshot: list[TransactionItem] = []
        for item in self.new_sale_items:
            snapshot_item = item.copy()
            self._apply_item_rounding(snapshot_item)
            product_snapshot.append(snapshot_item)
        sale_snapshot: list[TransactionItem] = list(product_snapshot)
        if reservation_balance > 0:
            sale_snapshot.insert(
                0,
                {
                    "temp_id": reservation["id"],
                    "barcode": reservation["id"],
                    "description": (
                        f"Alquiler {reservation['field_name']} "
                        f"({reservation['start_datetime']} - {reservation['end_datetime']})"
                    ),
                    "category": "Servicios",
                    "quantity": 1,
                    "unit": "Servicio",
                    "price": reservation_balance,
                    "sale_price": reservation_balance,
                    "subtotal": reservation_balance,
                },
            )
        sale_total = self._round_currency(sum(item["subtotal"] for item in sale_snapshot))
        if sale_total <= 0:
            return rx.toast("No hay importe para cobrar.", duration=3000)

        self._refresh_payment_feedback(total_override=sale_total)
        if self.payment_method_kind == "cash":
            if self.payment_cash_status not in ["exact", "change"]:
                message = (
                    self.payment_cash_message or "Ingrese un monto valido en efectivo."
                )
                return rx.toast(message, duration=3000)
        if self.payment_method_kind == "mixed":
            if self.payment_mixed_status not in ["exact", "change"]:
                message = (
                    self.payment_mixed_message
                    or "Complete los montos del pago mixto."
                )
                return rx.toast(message, duration=3000)
        # Verifica stock antes de aplicar descuentos.
        for item in product_snapshot:
            product_id = item["description"].lower().strip()
            if product_id not in self.inventory:
                return rx.toast(f"Producto {item['description']} no encontrado en inventario.", duration=3000)
            if self.inventory[product_id]["stock"] < item["quantity"]:
                return rx.toast(f"Stock insuficiente para {item['description']}.", duration=3000)

        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        payment_summary = self._generate_payment_summary()
        payment_label, payment_breakdown = self._payment_label_and_breakdown(sale_total)
        sale_id = str(uuid.uuid4())
        history_entries: list[Movement] = []
        for item in product_snapshot:
            product_id = item["description"].lower().strip()
            new_stock = self.inventory[product_id]["stock"] - item["quantity"]
            if new_stock < 0:
                new_stock = 0
            self.inventory[product_id]["stock"] = self._normalize_quantity_value(
                new_stock, self.inventory[product_id]["unit"]
            )
            history_entry = {
                "id": str(uuid.uuid4()),
                "timestamp": timestamp,
                "type": "Venta",
                "product_description": item["description"],
                "quantity": item["quantity"],
                "unit": item["unit"],
                "total": item["subtotal"],
                "payment_method": self.payment_method,
                "payment_kind": self.payment_method_kind,
                "payment_label": payment_label,
                "payment_breakdown": [p.copy() for p in payment_breakdown],
                "payment_details": payment_summary,
                "user": self.current_user["username"],
                "sale_id": sale_id,
            }
            history_entries.append(history_entry)

        applied_amount = 0.0
        if reservation_balance > 0:
            applied_amount = reservation_balance
            paid_before = reservation["paid_amount"]
            reservation["paid_amount"] = self._round_currency(reservation["paid_amount"] + applied_amount)
            self._update_reservation_status(reservation)
            entry_type = "pago" if reservation["paid_amount"] >= reservation["total_amount"] else "adelanto"
            balance_after = max(reservation["total_amount"] - reservation["paid_amount"], 0)
            reservation_note = (
                f"{payment_summary} | Total: {self._format_currency(reservation['total_amount'])} "
                f"| Adelanto: {self._format_currency(paid_before)} "
                f"| Pago: {self._format_currency(applied_amount)} "
                f"| Saldo: {self._format_currency(balance_after)}"
            )
            payment_summary = reservation_note
            self._log_service_action(
                reservation,
                entry_type,
                applied_amount,
                notes=reservation_note,
                status=reservation["status"],
            )
            self.reservation_payment_amount = ""
            self._set_last_reservation_receipt(reservation)
            self.last_sale_reservation_context = {
                "total": reservation["total_amount"],
                "paid_before": paid_before,
                "paid_now": applied_amount,
                "paid_after": reservation["paid_amount"],
                "balance_after": balance_after,
                "header": f"Alquiler {reservation['field_name']} ({reservation['start_datetime']} - {reservation['end_datetime']})",
                "products_total": self._round_currency(sum(item["subtotal"] for item in product_snapshot)),
                "charged_total": sale_total,
            }
        else:
            self.last_sale_reservation_context = None
        for entry in history_entries:
            entry["payment_details"] = payment_summary
        self.history.extend(history_entries)
        self.cashbox_sales.append(
            {
                "sale_id": sale_id,
                "timestamp": timestamp,
                "user": self.current_user["username"],
                "payment_method": self.payment_method,
                "payment_kind": self.payment_method_kind,
                "payment_label": payment_label,
                "payment_breakdown": [p.copy() for p in payment_breakdown],
                "payment_details": payment_summary,
                "total": sale_total,
                "service_total": sale_total,
                "items": [item.copy() for item in sale_snapshot],
                "is_deleted": False,
                "delete_reason": "",
            }
        )
        self.last_sale_receipt = sale_snapshot
        self.last_sale_total = sale_total
        self.last_sale_timestamp = timestamp
        self.last_payment_summary = payment_summary
        self.sale_receipt_ready = True
        self.new_sale_items = []
        self._reset_sale_form()
        self._reset_payment_fields()
        self._refresh_payment_feedback()
        return rx.toast("Venta confirmada.", duration=3000)

    @rx.event
    def print_sale_receipt(self):
        if not self.sale_receipt_ready or not self.last_sale_receipt:
            return rx.toast(
                "No hay comprobante disponible. Confirme una venta primero.",
                duration=3000,
            )
        rows = "".join(
            f"<tr><td colspan='2' style='font-weight:bold;text-align:center;font-size:13px;'>{item['description']}</td></tr>"
            f"<tr><td>{item['quantity']} {item['unit']} x {self._format_currency(item['price'])}</td><td style='text-align:right;'>{self._format_currency(item['subtotal'])}</td></tr>"
            for item in self.last_sale_receipt
        )
        summary_rows = ""
        if self.last_sale_reservation_context:
            ctx = self.last_sale_reservation_context
            header = ctx.get("header", "")
            header_row = ""
            if header:
                header_row = (
                    f"<tr><td colspan='2' style='text-align:center;font-weight:bold;font-size:13px;'>"
                    f"{header}"
                    f"</td></tr>"
                )
            products_total = ctx.get("products_total", 0)
            summary_rows = (
                header_row
                + "<tr><td colspan='2' style='height:4px;'></td></tr>"
                + f"<tr><td>Total reserva</td><td style='text-align:right;'>{self._format_currency(ctx['total'])}</td></tr>"
                + "<tr><td colspan='2' style='height:4px;'></td></tr>"
                + f"<tr><td>Adelanto previo</td><td style='text-align:right;'>{self._format_currency(ctx['paid_before'])}</td></tr>"
                + f"<tr><td style='font-weight:bold;'>Pago actual</td><td style='text-align:right;font-weight:bold;'>{self._format_currency(ctx['paid_now'])}</td></tr>"
                + (
                    f"<tr><td>Productos adicionales</td><td style='text-align:right;'>{self._format_currency(products_total)}</td></tr>"
                    if products_total > 0
                    else ""
                )
                + f"<tr><td>Saldo pendiente</td><td style='text-align:right;'>{self._format_currency(ctx.get('balance_after', 0))}</td></tr>"
                + "<tr><td colspan='2' style='height:6px;'></td></tr>"
            )
        display_rows = summary_rows + rows
        display_total = (
            self.last_sale_reservation_context.get("charged_total", self.last_sale_total)
            if self.last_sale_reservation_context
            else self.last_sale_total
        )
        html_content = f"""
        <html>
            <head>
                <meta charset='utf-8' />
                <title>Comprobante de Pago</title>
                <style>
                    @page {{
                        size: 58mm auto;
                        margin: 2mm;
                    }}
                    body {{
                        font-family: 'Courier New', monospace;
                        width: 56mm;
                        margin: 0 auto;
                        font-size: 11px;
                    }}
                    h1 {{
                        text-align: center;
                        font-size: 14px;
                        margin: 0 0 6px 0;
                    }}
                    .section {{
                        margin-bottom: 6px;
                    }}
                    table {{
                        width: 100%;
                        border-collapse: collapse;
                    }}
                    td {{
                        padding: 2px 0;
                        text-align: left;
                    }}
                    td:last-child {{
                        text-align: right;
                    }}
                    hr {{
                        border: 0;
                        border-top: 1px dashed #000;
                        margin: 6px 0;
                    }}
                </style>
            </head>
            <body>
                <h1>Comprobante de Pago</h1>
                <div class="section"><strong>Fecha:</strong> {self.last_sale_timestamp}</div>
                <hr />
                <table>
                    {display_rows}
                </table>
                <hr />
                <div class="section"><strong>Total General:</strong> <strong>{self._format_currency(display_total)}</strong></div>
                <div class="section"><strong>Metodo de Pago:</strong> {self.last_payment_summary}</div>
            </body>
        </html>
        """
        script = f"""
        const receiptWindow = window.open('', '_blank');
        receiptWindow.document.write({json.dumps(html_content)});
        receiptWindow.document.close();
        receiptWindow.focus();
        receiptWindow.print();
        """
        # Para cobros de reserva, libera seleccion despues de imprimir
        if self.last_sale_reservation_context:
            self.reservation_payment_id = ""
            self.reservation_payment_amount = ""
            self.last_sale_reservation_context = None
        self._reset_payment_fields()
        self._refresh_payment_feedback()
        return rx.call_script(script)

    @rx.event
    def set_history_page(self, page_num: int):
        if 1 <= page_num <= self.total_pages:
            self.current_page_history = page_num

    @rx.event
    def apply_history_filters(self):
        self.history_filter_type = self.staged_history_filter_type
        self.history_filter_product = self.staged_history_filter_product
        self.history_filter_start_date = self.staged_history_filter_start_date
        self.history_filter_end_date = self.staged_history_filter_end_date
        self.current_page_history = 1

    @rx.event
    def reset_history_filters(self):
        self.staged_history_filter_type = "Todos"
        self.staged_history_filter_product = ""
        self.staged_history_filter_start_date = ""
        self.staged_history_filter_end_date = ""
        self.apply_history_filters()

    @rx.event
    def set_staged_history_filter_type(self, value: str):
        self.staged_history_filter_type = value or "Todos"

    @rx.event
    def set_staged_history_filter_product(self, value: str):
        self.staged_history_filter_product = value or ""

    @rx.event
    def set_staged_history_filter_start_date(self, value: str):
        self.staged_history_filter_start_date = value or ""

    @rx.event
    def set_staged_history_filter_end_date(self, value: str):
        self.staged_history_filter_end_date = value or ""

    def _cashbox_guard(self):
        if not self.current_user["privileges"]["view_cashbox"]:
            return rx.toast("No tiene permisos para Gestion de Caja.", duration=3000)
        if not self.cashbox_is_open:
            return rx.toast(
                "Debe aperturar la caja para operar la gestion de caja.",
                duration=3000,
            )
        return None

    @rx.event
    def set_cashbox_open_amount_input(self, value: float | str):
        """Store the cashbox opening input as string regardless of input type."""
        self.cashbox_open_amount_input = str(value or "").strip()

    @rx.event
    def open_cashbox_session(self):
        if not self.current_user["privileges"]["manage_cashbox"]:
            return rx.toast("No tiene permisos para gestionar la caja.", duration=3000)
        username = self.current_user["username"]
        if self.current_user["role"].lower() == "cajero" and not self.token:
            return rx.toast("Inicie sesión para abrir caja.", duration=3000)
        amount = self._safe_amount(self.cashbox_open_amount_input or "0")
        if amount <= 0:
            return rx.toast("Ingrese un monto válido para la caja inicial.", duration=3000)
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        session = self._get_or_create_cashbox_session(username)
        session["opening_amount"] = amount
        session["opening_time"] = timestamp
        session["closing_time"] = ""
        session["is_open"] = True
        session["opened_by"] = username
        self.cashbox_sessions[username.lower()] = session
        self.cashbox_open_amount_input = ""
        self.cashbox_logs.append(
            {
                "id": str(uuid.uuid4()),
                "action": "apertura",
                "timestamp": timestamp,
                "user": username,
                "opening_amount": amount,
                "closing_total": 0.0,
                "totals_by_method": [],
                "notes": "Apertura de caja",
            }
        )
        self.history.append(
            {
                "id": str(uuid.uuid4()),
                "timestamp": timestamp,
                "type": "Apertura de Caja",
                "product_description": "Caja inicial",
                "quantity": 0,
                "unit": "-",
                "total": amount,
                "payment_method": "Caja",
                "payment_details": f"Caja inicial de {username}",
                "user": username,
                "sale_id": "",
            }
        )
        return rx.toast("Caja abierta. Jornada iniciada.", duration=3000)

    def _close_cashbox_session(self):
        username = self.current_user["username"]
        session = self._get_or_create_cashbox_session(username)
        session["is_open"] = False
        session["closing_time"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.cashbox_sessions[username.lower()] = session

    @rx.event
    def set_cashbox_staged_start_date(self, value: str):
        denial = self._cashbox_guard()
        if denial:
            return denial
        self.cashbox_staged_start_date = value or ""

    @rx.event
    def set_cashbox_staged_end_date(self, value: str):
        denial = self._cashbox_guard()
        if denial:
            return denial
        self.cashbox_staged_end_date = value or ""

    @rx.event
    def apply_cashbox_filters(self):
        denial = self._cashbox_guard()
        if denial:
            return denial
        self.cashbox_filter_start_date = self.cashbox_staged_start_date
        self.cashbox_filter_end_date = self.cashbox_staged_end_date
        self.cashbox_current_page = 1

    @rx.event
    def reset_cashbox_filters(self):
        denial = self._cashbox_guard()
        if denial:
            return denial
        self.cashbox_filter_start_date = ""
        self.cashbox_filter_end_date = ""
        self.cashbox_staged_start_date = ""
        self.cashbox_staged_end_date = ""
        self.cashbox_current_page = 1

    @rx.event
    def set_cashbox_log_staged_start_date(self, value: str):
        if not self.current_user["privileges"]["view_cashbox"]:
            return rx.toast("No tiene permisos para Gestion de Caja.", duration=3000)
        self.cashbox_log_staged_start_date = value or ""

    @rx.event
    def set_cashbox_log_staged_end_date(self, value: str):
        if not self.current_user["privileges"]["view_cashbox"]:
            return rx.toast("No tiene permisos para Gestion de Caja.", duration=3000)
        self.cashbox_log_staged_end_date = value or ""

    @rx.event
    def apply_cashbox_log_filters(self):
        if not self.current_user["privileges"]["view_cashbox"]:
            return rx.toast("No tiene permisos para Gestion de Caja.", duration=3000)
        self.cashbox_log_filter_start_date = self.cashbox_log_staged_start_date
        self.cashbox_log_filter_end_date = self.cashbox_log_staged_end_date

    @rx.event
    def reset_cashbox_log_filters(self):
        if not self.current_user["privileges"]["view_cashbox"]:
            return rx.toast("No tiene permisos para Gestion de Caja.", duration=3000)
        self.cashbox_log_filter_start_date = ""
        self.cashbox_log_filter_end_date = ""
        self.cashbox_log_staged_start_date = ""
        self.cashbox_log_staged_end_date = ""

    @rx.event
    def set_cashbox_page(self, page: int):
        denial = self._cashbox_guard()
        if denial:
            return denial
        if 1 <= page <= self.cashbox_total_pages:
            self.cashbox_current_page = page

    @rx.event
    def set_show_cashbox_advances(self, value: bool | str):
        denial = self._cashbox_guard()
        if denial:
            return denial
        if isinstance(value, str):
            value = value.lower() in ["true", "1", "on", "yes"]
        self.show_cashbox_advances = bool(value)

    @rx.event
    def prev_cashbox_page(self):
        denial = self._cashbox_guard()
        if denial:
            return denial
        if self.cashbox_current_page > 1:
            self.cashbox_current_page -= 1

    @rx.event
    def next_cashbox_page(self):
        denial = self._cashbox_guard()
        if denial:
            return denial
        total_pages = (
            (len(self.filtered_cashbox_sales) + self.cashbox_items_per_page - 1)
            // self.cashbox_items_per_page
        )
        total_pages = total_pages or 1
        if self.cashbox_current_page < total_pages:
            self.cashbox_current_page += 1

    @rx.event
    def open_cashbox_close_modal(self):
        denial = self._cashbox_guard()
        if denial:
            return denial
        today = datetime.datetime.now().strftime("%Y-%m-%d")
        day_sales = self._get_day_sales(today)
        if not day_sales:
            return rx.toast("No hay ventas registradas hoy.", duration=3000)
        self.cashbox_close_summary_totals = self._build_cashbox_summary(day_sales)
        self.cashbox_close_summary_sales = day_sales
        self.cashbox_close_summary_date = today
        self.cashbox_close_modal_open = True

    @rx.event
    def close_cashbox_close_modal(self):
        denial = self._cashbox_guard()
        if denial:
            return denial
        self._reset_cashbox_close_summary()

    @rx.event
    def export_to_excel(self):
        if not self.current_user["privileges"]["export_data"]:
            return rx.toast("No tiene permisos para exportar datos.", duration=3000)
        import openpyxl
        from io import BytesIO

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Historial Movimientos"
        headers = [
            "Fecha y Hora",
            "Tipo",
            "Descripcion",
            "Cantidad",
            "Unidad",
            "Total",
            "Metodo de Pago",
            "Detalle Pago",
        ]
        sheet.append(headers)
        for movement in self.filtered_history:
            method_display = self._normalize_wallet_label(
                movement.get("payment_method", "")
            )
            sheet.append(
                [
                    movement["timestamp"],
                    movement["type"],
                    movement["product_description"],
                    movement["quantity"],
                    movement["unit"],
                    movement["total"],
                    method_display,
                    movement.get("payment_details", ""),
                ]
            )
        file_stream = BytesIO()
        workbook.save(file_stream)
        file_stream.seek(0)
        return rx.download(
            data=file_stream.read(), filename="historial_movimientos.xlsx"
        )

    @rx.event
    def export_inventory_to_excel(self):
        if not self.current_user["privileges"]["export_data"]:
            return rx.toast("No tiene permisos para exportar datos.", duration=3000)
        import openpyxl
        from io import BytesIO

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Inventario Actual"
        headers = [
            "Codigo de Barra",
            "Descripcion",
            "Categoria",
            "Unidad",
            "Stock Sistema",
            "Precio Compra",
            "Precio Venta",
            "Valor Total",
            "Conteo Fisico",
            "Diferencia",
            "Notas Adicionales",
        ]
        sheet.append(headers)
        products = sorted(
            self.inventory.values(), key=lambda p: p.get("description", "").lower()
        )
        for product in products:
            barcode = product.get("barcode", "")
            description = product.get("description", "")
            category = product.get(
                "category", self.categories[0] if self.categories else ""
            )
            unit = product.get("unit", "Unidad")
            stock = product.get("stock", 0)
            purchase_price = product.get("purchase_price", 0)
            sale_price = product.get("sale_price", 0)
            total_value = stock * purchase_price
            sheet.append(
                [
                    barcode,
                    description,
                    category,
                    unit,
                    stock,
                    purchase_price,
                    sale_price,
                    total_value,
                    "",
                    "",
                    "",
                ]
            )
            current_row = sheet.max_row
            sheet.cell(row=current_row, column=10).value = f"=I{current_row}-E{current_row}"
        file_stream = BytesIO()
        workbook.save(file_stream)
        file_stream.seek(0)
        return rx.download(data=file_stream.read(), filename="inventario_actual.xlsx")

    @rx.event
    def export_cashbox_report(self):
        denial = self._cashbox_guard()
        if denial:
            return denial
        if not self.current_user["privileges"]["export_data"]:
            return rx.toast("No tiene permisos para exportar datos.", duration=3000)
        if not self.cashbox_sales:
            return rx.toast("No hay ventas para exportar.", duration=3000)
        import openpyxl
        from io import BytesIO

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Gestion de Caja"
        headers = [
            "Fecha y Hora",
            "Usuario",
            "Metodo",
            "Metodo Detallado",
            "Detalle Pago",
            "Total",
            "Productos",
        ]
        sheet.append(headers)
        for sale in self.filtered_cashbox_sales:
            if sale.get("is_deleted"):
                continue
            method_raw = self._normalize_wallet_label(sale.get("payment_method", ""))
            method_label = self._normalize_wallet_label(
                sale.get("payment_label", sale.get("payment_method", ""))
            )
            details = ", ".join(
                f"{item['description']} (x{item['quantity']})" for item in sale["items"]
            )
            sheet.append(
                [
                    sale["timestamp"],
                    sale["user"],
                    method_raw,
                    method_label,
                    sale["payment_details"],
                    sale["total"],
                    details,
                ]
            )
        file_stream = BytesIO()
        workbook.save(file_stream)
        file_stream.seek(0)
        return rx.download(data=file_stream.read(), filename="gestion_caja.xlsx")

    @rx.event
    def export_cashbox_sessions(self):
        if not self.current_user["privileges"]["view_cashbox"]:
            return rx.toast("No tiene permisos para Gestion de Caja.", duration=3000)
        if not self.current_user["privileges"]["export_data"]:
            return rx.toast("No tiene permisos para exportar datos.", duration=3000)
        logs = self.filtered_cashbox_logs
        if not logs:
            return rx.toast("No hay aperturas o cierres para exportar.", duration=3000)
        import openpyxl
        from io import BytesIO

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Aperturas y Cierres"
        headers = [
            "Fecha y Hora",
            "Accion",
            "Usuario",
            "Monto Apertura",
            "Monto Cierre",
            "Totales por Metodo",
            "Notas",
        ]
        sheet.append(headers)
        for log in logs:
            totals_detail = ", ".join(
                f"{item.get('method', '')}: {self._round_currency(item.get('amount', 0))}"
                for item in log.get("totals_by_method", [])
                if item.get("amount", 0)
            )
            sheet.append(
                [
                    log.get("timestamp", ""),
                    (log.get("action") or "").capitalize(),
                    log.get("user", ""),
                    self._round_currency(log.get("opening_amount", 0)),
                    self._round_currency(log.get("closing_total", 0)),
                    totals_detail or "",
                    log.get("notes", ""),
                ]
            )
        file_stream = BytesIO()
        workbook.save(file_stream)
        file_stream.seek(0)
        return rx.download(
            data=file_stream.read(), filename="aperturas_cierres_caja.xlsx"
        )

    @rx.event
    def show_cashbox_log(self, log_id: str):
        if not self.current_user["privileges"]["view_cashbox"]:
            return rx.toast("No tiene permisos para Gestion de Caja.", duration=3000)
        entry = next((log for log in self.cashbox_logs if log["id"] == log_id), None)
        if not entry:
            return rx.toast("Registro de caja no encontrado.", duration=3000)
        self.cashbox_log_selected = entry
        self.cashbox_log_modal_open = True

    @rx.event
    def close_cashbox_log_modal(self):
        self.cashbox_log_modal_open = False
        self.cashbox_log_selected = None

    @rx.event
    def open_sale_delete_modal(self, sale_id: str):
        denial = self._cashbox_guard()
        if denial:
            return denial
        self.sale_to_delete = sale_id
        self.sale_delete_reason = ""
        self.sale_delete_modal_open = True

    @rx.event
    def close_sale_delete_modal(self):
        denial = self._cashbox_guard()
        if denial:
            return denial
        self.sale_delete_modal_open = False
        self.sale_to_delete = ""
        self.sale_delete_reason = ""

    @rx.event
    def set_sale_delete_reason(self, value: str):
        denial = self._cashbox_guard()
        if denial:
            return denial
        self.sale_delete_reason = value

    @rx.event
    def delete_sale(self):
        denial = self._cashbox_guard()
        if denial:
            return denial
        if not self.current_user["privileges"]["delete_sales"]:
            return rx.toast("No tiene permisos para eliminar ventas.", duration=3000)
        sale_id = self.sale_to_delete
        reason = self.sale_delete_reason.strip()
        if not sale_id:
            return rx.toast("Seleccione una venta a eliminar.", duration=3000)
        if not reason:
            return rx.toast(
                "Ingrese el motivo de la eliminación de la venta.", duration=3000
            )
        sale = next((s for s in self.cashbox_sales if s["sale_id"] == sale_id), None)
        if not sale:
            return rx.toast("Venta no encontrada.", duration=3000)
        if sale.get("is_deleted"):
            return rx.toast("Esta venta ya fue eliminada.", duration=3000)
        for item in sale["items"]:
            product_id = item["description"].lower().strip()
            if product_id in self.inventory:
                self.inventory[product_id]["stock"] += item["quantity"]
        self.history = [m for m in self.history if m.get("sale_id") != sale_id]
        self.history.append(
            {
                "id": str(uuid.uuid4()),
                "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "type": "Venta Eliminada",
                "product_description": f"Venta anulada ({sale_id})",
                "quantity": 0,
                "unit": "-",
                "total": -sale["total"],
                "payment_method": "Anulado",
                "payment_details": reason,
                "user": self.current_user["username"],
                "sale_id": sale_id,
            }
        )
        sale["is_deleted"] = True
        sale["delete_reason"] = reason
        self.close_sale_delete_modal()
        return rx.toast("Venta eliminada correctamente.", duration=3000)

    @rx.event
    def reprint_sale_receipt(self, sale_id: str):
        denial = self._cashbox_guard()
        if denial:
            return denial
        sale = next((s for s in self.cashbox_sales if s["sale_id"] == sale_id), None)
        if not sale:
            return rx.toast("Venta no encontrada.", duration=3000)
        items = sale.get("items", [])
        rows = "".join(
            f"<tr><td colspan='2'><strong>{item.get('description', '')}</strong></td></tr>"
            f"<tr><td>{item.get('quantity', 0)} {item.get('unit', '')} x {self._format_currency(item.get('price', 0))}</td><td style='text-align:right;'>{self._format_currency(item.get('subtotal', 0))}</td></tr>"
            for item in items
        )
        payment_summary = sale.get("payment_details") or sale.get(
            "payment_method", ""
        )
        html_content = f"""
        <html>
            <head>
                <meta charset='utf-8' />
                <title>Comprobante de Pago</title>
                <style>
                    @page {{
                        size: 58mm auto;
                        margin: 2mm;
                    }}
                    body {{
                        font-family: 'Courier New', monospace;
                        width: 56mm;
                        margin: 0 auto;
                        font-size: 11px;
                    }}
                    h1 {{
                        text-align: center;
                        font-size: 14px;
                        margin: 0 0 6px 0;
                    }}
                    .section {{
                        margin-bottom: 6px;
                    }}
                    table {{
                        width: 100%;
                        border-collapse: collapse;
                    }}
                    td {{
                        padding: 2px 0;
                        text-align: left;
                    }}
                    td:last-child {{
                        text-align: right;
                    }}
                    hr {{
                        border: 0;
                        border-top: 1px dashed #000;
                        margin: 6px 0;
                    }}
                </style>
            </head>
            <body>
                <h1>Comprobante de Pago</h1>
                <div class="section"><strong>Fecha:</strong> {sale.get('timestamp', '')}</div>
                <div class="section"><strong>Usuario:</strong> {sale.get('user', '')}</div>
                <hr />
                <table>
                    {rows}
                </table>
                <hr />
                <div class="section"><strong>Total General:</strong> {self._format_currency(sale.get('service_total', sale.get('total', 0)))} </div>
                <div class="section"><strong>Metodo de Pago:</strong> {payment_summary}</div>
            </body>
        </html>
        """
        script = f"""
        const receiptWindow = window.open('', '_blank');
        receiptWindow.document.write({json.dumps(html_content)});
        receiptWindow.document.close();
        receiptWindow.focus();
        receiptWindow.print();
        """
        return rx.call_script(script)

    @rx.event
    def close_cashbox_day(self):
        if not self.current_user["privileges"]["manage_cashbox"]:
            return rx.toast("No tiene permisos para gestionar la caja.", duration=3000)
        denial = self._cashbox_guard()
        if denial:
            return denial
        date = self.cashbox_close_summary_date or datetime.datetime.now().strftime(
            "%Y-%m-%d"
        )
        day_sales = self.cashbox_close_summary_sales or self._get_day_sales(date)
        if not day_sales:
            return rx.toast("No hay ventas registradas hoy.", duration=3000)
        summary = self.cashbox_close_summary_totals or self._build_cashbox_summary(
            day_sales
        )
        closing_timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        totals_list = [
            {"method": method, "amount": self._round_currency(amount)}
            for method, amount in summary.items()
            if amount > 0
        ]
        closing_total = self._round_currency(sum(summary.values()))
        self.cashbox_logs.append(
            {
                "id": str(uuid.uuid4()),
                "action": "cierre",
                "timestamp": closing_timestamp,
                "user": self.current_user["username"],
                "opening_amount": self.cashbox_opening_amount,
                "closing_total": closing_total,
                "totals_by_method": totals_list,
                "notes": f"Cierre de caja {date}",
            }
        )
        summary_rows = "".join(
            f"<tr><td>{method}</td><td>{self._format_currency(amount)}</td></tr>"
            for method, amount in summary.items()
            if amount > 0
        )
        grand_total_row = f"<tr><td><strong>Total cierre</strong></td><td><strong>{self._format_currency(closing_total)}</strong></td></tr>"
        detail_rows = "".join(
            (
                lambda method_label, breakdown_text: f"<tr><td>{sale['timestamp']}</td><td>{sale['user']}</td><td>{method_label}{('<br><small>' + breakdown_text + '</small>') if breakdown_text else ''}</td><td>{self._format_currency(sale['total'])}</td></tr>"
            )(
                sale.get("payment_label", sale.get("payment_method", "")),
                " / ".join(
                    f"{item.get('label', '')}: {self._format_currency(item.get('amount', 0))}"
                    for item in sale.get("payment_breakdown", [])
                    if item.get("amount", 0)
                ),
            )
            for sale in day_sales
        )
        html_content = f"""
        <html>
            <head>
                <meta charset='utf-8' />
                <title>Resumen de Caja</title>
                <style>
                    body {{ font-family: Arial, sans-serif; padding: 24px; }}
                    h1 {{ text-align: center; }}
                    table {{ width: 100%; border-collapse: collapse; margin-top: 16px; }}
                    th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                    th {{ background-color: #f3f4f6; }}
                </style>
            </head>
            <body>
                <h1>Resumen Diario de Caja</h1>
                <p><strong>Fecha:</strong> {date}</p>
                <p><strong>Responsable:</strong> {self.current_user['username']}</p>
                <h2>Totales por método</h2>
                <table>
                    <thead>
                        <tr>
                            <th>Método</th>
                            <th>Monto</th>
                        </tr>
                    </thead>
                    <tbody>
                        {summary_rows}
                        {grand_total_row}
                    </tbody>
                </table>
                <h2>Detalle de ventas</h2>
                <table>
                    <thead>
                        <tr>
                            <th>Fecha y Hora</th>
                            <th>Usuario</th>
                            <th>Método</th>
                            <th>Total</th>
                        </tr>
                    </thead>
                    <tbody>
                        {detail_rows}
                    </tbody>
                </table>
            </body>
        </html>
        """
        script = f"""
        const cashboxWindow = window.open('', '_blank');
        cashboxWindow.document.write({json.dumps(html_content)});
        cashboxWindow.document.close();
        cashboxWindow.focus();
        cashboxWindow.print();
        """
        self._close_cashbox_session()
        self._reset_cashbox_close_summary()
        return rx.call_script(script)

    @rx.event
    def open_inventory_check_modal(self):
        if not self.current_user["privileges"]["edit_inventario"]:
            return rx.toast(
                "No tiene permisos para registrar inventario.", duration=3000
            )
        self.inventory_check_modal_open = True
        self.inventory_check_status = "perfecto"
        self.inventory_adjustment_notes = ""
        self.inventory_adjustment_items = []
        self._reset_inventory_adjustment_form()

    @rx.event
    def open_edit_product(self, product: Product):
        if not self.current_user["privileges"]["edit_inventario"]:
            return rx.toast("No tiene permisos para editar el inventario.", duration=3000)
        self.editing_product = product.copy()
        self.is_editing_product = True

    @rx.event
    def cancel_edit_product(self):
        self.is_editing_product = False
        self.editing_product = {
            "id": "",
            "barcode": "",
            "description": "",
            "category": "",
            "stock": 0,
            "unit": "",
            "purchase_price": 0,
            "sale_price": 0,
        }

    @rx.event
    def handle_edit_product_change(self, field: str, value: str):
        if field in ["stock", "purchase_price", "sale_price"]:
            try:
                if value == "":
                    self.editing_product[field] = 0
                else:
                    self.editing_product[field] = float(value)
            except ValueError:
                pass 
        else:
            self.editing_product[field] = value

    @rx.event
    def save_edited_product(self):
        if not self.current_user["privileges"]["edit_inventario"]:
            return rx.toast("No tiene permisos para editar el inventario.", duration=3000)
        
        old_id = self.editing_product["id"]
        new_description = self.editing_product["description"].strip()
        new_id = new_description.lower()
        
        if not new_description:
             return rx.toast("La descripción no puede estar vacía.", duration=3000)

        if new_id != old_id:
            if new_id in self.inventory:
                 return rx.toast("Ya existe un producto con esa descripción.", duration=3000)
            del self.inventory[old_id]
            self.editing_product["id"] = new_id
        
        self.inventory[new_id] = self.editing_product
        self.is_editing_product = False
        return rx.toast("Producto actualizado correctamente.", duration=3000)

    @rx.event
    def delete_product(self, product_id: str):
        if not self.current_user["privileges"]["edit_inventario"]:
             return rx.toast("No tiene permisos para eliminar productos.", duration=3000)
        
        if product_id in self.inventory:
            del self.inventory[product_id]
            return rx.toast("Producto eliminado.", duration=3000)

    @rx.event
    def close_inventory_check_modal(self):
        self.inventory_check_modal_open = False
        self.inventory_check_status = "perfecto"
        self.inventory_adjustment_notes = ""
        self.inventory_adjustment_items = []
        self._reset_inventory_adjustment_form()

    @rx.event
    def set_inventory_check_status(self, status: str):
        if status not in ["perfecto", "ajuste"]:
            return
        self.inventory_check_status = status
        if status == "perfecto":
            self.inventory_adjustment_items = []
            self._reset_inventory_adjustment_form()

    @rx.event
    def set_inventory_adjustment_notes(self, notes: str):
        self.inventory_adjustment_notes = notes

    @rx.event
    def submit_inventory_check(self):
        if not self.current_user["privileges"]["edit_inventario"]:
            return rx.toast(
                "No tiene permisos para registrar inventario.", duration=3000
            )
        status = (
            self.inventory_check_status
            if self.inventory_check_status in ["perfecto", "ajuste"]
            else "perfecto"
        )
        notes = self.inventory_adjustment_notes.strip()
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if status == "perfecto":
            self.history.append(
                {
                    "id": str(uuid.uuid4()),
                    "timestamp": timestamp,
                    "type": "Inventario Perfecto",
                    "product_description": "Inventario fisico verificado sin diferencias.",
                    "quantity": 0,
                    "unit": "-",
                    "total": 0,
                    "payment_method": "Inventario Perfecto",
                    "payment_details": "Sin diferencias",
                    "user": self.current_user["username"],
                    "sale_id": "",
                }
            )
        else:
            if not self.inventory_adjustment_items:
                return rx.toast(
                    "Agregue los productos que requieren re ajuste.", duration=3000
                )
            recorded = False
            for item in self.inventory_adjustment_items:
                description = item["description"].strip()
                if not description:
                    continue
                product_id = description.lower()
                if product_id not in self.inventory:
                    continue
                quantity = item.get("adjust_quantity", 0) or 0
                if quantity <= 0:
                    continue
                available = self.inventory[product_id]["stock"]
                qty = min(quantity, available)
                qty = self._normalize_quantity_value(
                    qty, item.get("unit") or self.inventory[product_id]["unit"]
                )
                new_stock = max(available - qty, 0)
                self.inventory[product_id]["stock"] = self._normalize_quantity_value(
                    new_stock, self.inventory[product_id]["unit"]
                )
                detail_parts = []
                if item.get("reason"):
                    detail_parts.append(item["reason"])
                if notes:
                    detail_parts.append(notes)
                details = (
                    " | ".join(part for part in detail_parts if part)
                    if detail_parts
                    else "Ajuste inventario"
                )
                self.history.append(
                    {
                        "id": str(uuid.uuid4()),
                        "timestamp": timestamp,
                        "type": "Re Ajuste Inventario",
                        "product_description": description,
                        "quantity": -qty,
                        "unit": item.get("unit") or "-",
                        "total": 0,
                        "payment_method": "Re Ajuste Inventario",
                        "payment_details": details,
                        "user": self.current_user["username"],
                        "sale_id": "",
                    }
                )
                recorded = True
            if not recorded:
                return rx.toast(
                    "No se pudo registrar el re ajuste. Verifique los productos.",
                    duration=3000,
                )
        self.inventory_check_modal_open = False
        self.inventory_check_status = "perfecto"
        self.inventory_adjustment_notes = ""
        self.inventory_adjustment_items = []
        self._reset_inventory_adjustment_form()
        return rx.toast("Registro de inventario guardado.", duration=3000)

    def _reset_new_user_form(self):
        self.new_user_data = {
            "username": "",
            "password": "",
            "confirm_password": "",
            "role": "Usuario",
            "privileges": self._role_privileges("Usuario"),
        }
        self.editing_user = None

    @rx.event
    def show_create_user_form(self):
        self._reset_new_user_form()
        self.show_user_form = True

    def _open_user_editor(self, user: User):
        merged_privileges = self._normalize_privileges(user.get("privileges", {}))
        role_key = self._find_role_key(user["role"]) or user["role"]
        if role_key not in self.role_privileges:
            self.role_privileges[role_key] = merged_privileges.copy()
            if role_key not in self.roles:
                self.roles.append(role_key)
        if user["username"] in self.users:
            self.users[user["username"]]["privileges"] = merged_privileges
        self.new_user_data = {
            "username": user["username"],
            "password": "",
            "confirm_password": "",
            "role": role_key,
            "privileges": merged_privileges,
        }
        self.editing_user = self.users.get(user["username"], user)
        self.show_user_form = True

    @rx.event
    def show_edit_user_form(self, user: User):
        self._open_user_editor(user)

    @rx.event
    def show_edit_user_form_by_username(self, username: str):
        key = (username or "").strip().lower()
        user = self.users.get(key)
        if not user:
            return rx.toast("Usuario a editar no encontrado.", duration=3000)
        self._open_user_editor(user)

    @rx.event
    def set_user_form_open(self, is_open: bool):
        self.show_user_form = bool(is_open)
        if not is_open:
            self._reset_new_user_form()

    @rx.event
    def hide_user_form(self):
        self.show_user_form = False
        self._reset_new_user_form()

    @rx.event
    def handle_new_user_change(self, field: str, value: str):
        if field == "role":
            self.new_user_data["role"] = value
            self.new_user_data["privileges"] = self._role_privileges(value)
            return
        if field == "username":
            self.new_user_data["username"] = value.lower()
            return
        self.new_user_data[field] = value

    @rx.event
    def toggle_privilege(self, privilege: str):
        privileges = self._normalize_privileges(self.new_user_data["privileges"])
        privileges[privilege] = not privileges[privilege]
        self.new_user_data["privileges"] = privileges

    @rx.event
    def apply_role_privileges(self):
        role = self.new_user_data.get("role") or "Usuario"
        self.new_user_data["privileges"] = self._role_privileges(role)

    @rx.event
    def update_new_role_name(self, value: str):
        self.new_role_name = value.strip()

    @rx.event
    def create_role_from_current_privileges(self):
        name = (self.new_role_name or "").strip()
        if not name:
            return rx.toast("Ingrese un nombre para el rol nuevo.", duration=3000)
        if name.lower() == "superadmin":
            return rx.toast("Superadmin ya existe como rol principal.", duration=3000)
        existing = self._find_role_key(name)
        if existing:
            return rx.toast("Ese rol ya existe.", duration=3000)
        privileges = self._normalize_privileges(self.new_user_data["privileges"])
        self.role_privileges[name] = privileges.copy()
        if name not in self.roles:
            self.roles.append(name)
        self.new_role_name = ""
        self.new_user_data["role"] = name
        self.new_user_data["privileges"] = privileges.copy()
        return rx.toast(f"Rol {name} creado con los privilegios actuales.", duration=3000)

    @rx.event
    def save_role_template(self):
        role = (self.new_user_data.get("role") or "").strip()
        if not role:
            return rx.toast("Seleccione un rol para guardar sus privilegios.", duration=3000)
        if role.lower() == "superadmin":
            return rx.toast("No se puede modificar los privilegios de Superadmin.", duration=3000)
        privileges = self._normalize_privileges(self.new_user_data["privileges"])
        role_key = self._find_role_key(role) or role
        self.role_privileges[role_key] = privileges.copy()
        if role_key not in self.roles:
            self.roles.append(role_key)
        return rx.toast(f"Plantilla de rol {role_key} actualizada.", duration=3000)

    @rx.event
    def save_user(self):
        if not self.current_user["privileges"]["manage_users"]:
            return rx.toast("No tiene permisos para gestionar usuarios.", duration=3000)
        username = self.new_user_data["username"].lower().strip()
        if not username:
            return rx.toast("El nombre de usuario no puede estar vacío.", duration=3000)
        self.new_user_data["privileges"] = self._normalize_privileges(
            self.new_user_data["privileges"]
        )
        if self.editing_user:
            user_to_update = self.users.get(self.editing_user["username"])
            if not user_to_update:
                return rx.toast("Usuario a editar no encontrado.", duration=3000)
            if self.new_user_data["password"]:
                if (
                    self.new_user_data["password"]
                    != self.new_user_data["confirm_password"]
                ):
                    return rx.toast("Las contraseñas no coinciden.", duration=3000)
                password_hash = bcrypt.hashpw(
                    self.new_user_data["password"].encode(), bcrypt.gensalt()
                ).decode()
                user_to_update["password_hash"] = password_hash
            user_to_update["role"] = self.new_user_data["role"]
            user_to_update["privileges"] = self.new_user_data["privileges"].copy()
            self.hide_user_form()
            return rx.toast(f"Usuario {username} actualizado.", duration=3000)
        else:
            if username in self.users:
                return rx.toast("El nombre de usuario ya existe.", duration=3000)
            if not self.new_user_data["password"]:
                return rx.toast("La contraseña no puede estar vacía.", duration=3000)
            if self.new_user_data["password"] != self.new_user_data["confirm_password"]:
                return rx.toast("Las contraseñas no coinciden.", duration=3000)
            password_hash = bcrypt.hashpw(
                self.new_user_data["password"].encode(), bcrypt.gensalt()
            ).decode()
            self.users[username] = {
                "username": username,
                "password_hash": password_hash,
                "role": self.new_user_data["role"],
                "privileges": self.new_user_data["privileges"].copy(),
            }
            self.hide_user_form()
            return rx.toast(f"Usuario {username} creado.", duration=3000)

    @rx.event
    def delete_user(self, username: str):
        if not self.current_user["privileges"]["manage_users"]:
            return rx.toast("No tiene permisos para eliminar usuarios.", duration=3000)
        if username == "admin":
            return rx.toast("No se puede eliminar al superadmin.", duration=3000)
        if username == self.current_user["username"]:
            return rx.toast("No puedes eliminar tu propio usuario.", duration=3000)
        if self.users.pop(username, None):
            return rx.toast(f"Usuario {username} eliminado.", duration=3000)
        return rx.toast(f"Usuario {username} no encontrado.", duration=3000)

    def _payment_method_by_identifier(self, identifier: str) -> PaymentMethodConfig | None:
        target = (identifier or "").strip().lower()
        if not target:
            return None
        for method in self.payment_methods:
            if method["id"].lower() == target or method["name"].lower() == target:
                return method
        return None

    def _enabled_payment_methods_list(self) -> list[PaymentMethodConfig]:
        return [m for m in self.payment_methods if m.get("enabled", True)]

    def _default_payment_method(self) -> PaymentMethodConfig | None:
        enabled = self._enabled_payment_methods_list()
        if enabled:
            return enabled[0]
        return None

    def _ensure_payment_method_selected(self):
        available = self._enabled_payment_methods_list()
        if not available:
            self.payment_method = ""
            self.payment_method_description = ""
            self.payment_method_kind = "other"
            return
        if not any(m["name"] == self.payment_method for m in available):
            self._set_payment_method(available[0])

    @rx.event
    def set_new_payment_method_name(self, value: str):
        self.new_payment_method_name = value

    @rx.event
    def set_new_payment_method_description(self, value: str):
        self.new_payment_method_description = value

    @rx.event
    def set_new_payment_method_kind(self, value: str):
        self.new_payment_method_kind = (value or "").strip().lower() or "other"

    @rx.event
    def add_payment_method(self):
        if not self.current_user["privileges"]["manage_config"]:
            return rx.toast("No tiene permisos para configurar el sistema.", duration=3000)
        name = (self.new_payment_method_name or "").strip()
        description = (self.new_payment_method_description or "").strip()
        kind = (self.new_payment_method_kind or "other").strip().lower()
        if not name:
            return rx.toast("Asigne un nombre al metodo de pago.", duration=3000)
        if kind not in ["cash", "card", "wallet", "mixed", "other"]:
            kind = "other"
        if any(m["name"].lower() == name.lower() for m in self.payment_methods):
            return rx.toast("Ya existe un metodo con ese nombre.", duration=3000)
        method: PaymentMethodConfig = {
            "id": str(uuid.uuid4()),
            "name": name,
            "description": description or "Sin descripcion",
            "kind": kind,
            "enabled": True,
        }
        self.payment_methods.append(method)
        self.new_payment_method_name = ""
        self.new_payment_method_description = ""
        self.new_payment_method_kind = "other"
        self._set_payment_method(method)
        return rx.toast(f"Metodo {name} agregado.", duration=2500)

    @rx.event
    def toggle_payment_method_enabled(self, method_id: str, enabled: bool | str):
        if not self.current_user["privileges"]["manage_config"]:
            return rx.toast("No tiene permisos para configurar el sistema.", duration=3000)
        if isinstance(enabled, str):
            enabled = enabled.lower() in ["true", "1", "on", "yes"]
        active_methods = self._enabled_payment_methods_list()
        for method in self.payment_methods:
            if method["id"] == method_id:
                if not enabled and method.get("enabled", True) and len(active_methods) <= 1:
                    return rx.toast("Debe haber al menos un metodo activo.", duration=3000)
                method["enabled"] = enabled
                break
        self._ensure_payment_method_selected()

    @rx.event
    def remove_payment_method(self, method_id: str):
        method = self._payment_method_by_identifier(method_id)
        if not method:
            return
        remaining_enabled = [
            m for m in self._enabled_payment_methods_list() if m["id"] != method_id
        ]
        if not remaining_enabled:
            return rx.toast("No puedes eliminar el unico metodo activo.", duration=3000)
        self.payment_methods = [m for m in self.payment_methods if m["id"] != method_id]
        self._ensure_payment_method_selected()
        return rx.toast(f"Metodo {method['name']} eliminado.", duration=2500)

    def _generate_payment_summary(self) -> str:
        method = self.payment_method or "No especificado"
        kind = (self.payment_method_kind or "other").lower()
        if kind == "cash":
            detail = f"Monto {self._format_currency(self.payment_cash_amount)}"
            if self.payment_cash_message:
                detail += f" ({self.payment_cash_message})"
            return f"{method} - {detail}"
        if kind == "card":
            return f"{method} - {self.payment_card_type}"
        if kind == "wallet":
            provider = (
                self.payment_wallet_provider
                or self.payment_wallet_choice
                or "Proveedor no especificado"
            )
            return f"{method} - {provider}"
        if kind == "mixed":
            parts = []
            if self.payment_mixed_cash > 0:
                parts.append(f"Efectivo {self._format_currency(self.payment_mixed_cash)}")
            if self.payment_mixed_card > 0:
                parts.append(
                    f"Tarjeta ({self.payment_card_type}) {self._format_currency(self.payment_mixed_card)}"
                )
            if self.payment_mixed_wallet > 0:
                provider = (
                    self.payment_wallet_provider
                    or self.payment_wallet_choice
                    or "Billetera"
                )
                parts.append(f"{provider} {self._format_currency(self.payment_mixed_wallet)}")
            if self.payment_mixed_notes:
                parts.append(self.payment_mixed_notes)
            if not parts:
                parts.append("Sin detalle")
            if self.payment_mixed_message:
                parts.append(self.payment_mixed_message)
            return f"{method} - {' / '.join(parts)}"
        return f"{method} - {self.payment_method_description}"

    def _safe_amount(self, value: str) -> float:
        try:
            amount = float(value) if value else 0
        except ValueError:
            amount = 0
        return self._round_currency(amount)

    def _update_cash_feedback(self, total_override: float | None = None):
        effective_total = (
            total_override
            if total_override is not None
            else (self.sale_total if self.sale_total > 0 else self.selected_reservation_balance)
        )
        amount = self.payment_cash_amount
        diff = amount - effective_total
        if amount <= 0:
            self.payment_cash_message = "Ingrese un monto valido."
            self.payment_cash_status = "warning"
        elif diff > 0:
            self.payment_cash_message = f"Vuelto {self._format_currency(diff)}"
            self.payment_cash_status = "change"
        elif diff < 0:
            self.payment_cash_message = f"Faltan {self._format_currency(abs(diff))}"
            self.payment_cash_status = "due"
        else:
            self.payment_cash_message = "Monto exacto."
            self.payment_cash_status = "exact"

    def _update_mixed_message(self, total_override: float | None = None):
        total = (
            total_override
            if total_override is not None
            else (self.sale_total if self.sale_total > 0 else self.selected_reservation_balance)
        )
        total = self._round_currency(total)
        
        paid_cash = self._round_currency(self.payment_mixed_cash)
        paid_card = self._round_currency(self.payment_mixed_card)
        paid_wallet = self._round_currency(self.payment_mixed_wallet)
        
        total_paid = self._round_currency(paid_cash + paid_card + paid_wallet)
        
        if total_paid <= 0:
            self.payment_mixed_message = "Ingrese montos para los metodos seleccionados."
            self.payment_mixed_status = "warning"
            return
            
        # Calculate difference with rounding to avoid floating point issues
        diff = self._round_currency(total - total_paid)
        
        if diff > 0:
            self.payment_mixed_message = f"Restan {self._format_currency(diff)}"
            self.payment_mixed_status = "due"
            return
            
        # If diff <= 0, it means we paid enough or too much
        change = abs(diff)
        
        if change > 0:
            self.payment_mixed_message = f"Vuelto {self._format_currency(change)}"
            self.payment_mixed_status = "change"
        else:
            self.payment_mixed_message = "Montos completos."
            self.payment_mixed_status = "exact"

    def _refresh_payment_feedback(self, total_override: float | None = None):
        if self.payment_method_kind == "cash":
            self._update_cash_feedback(total_override=total_override)
        elif self.payment_method_kind == "mixed":
            self._update_mixed_message(total_override=total_override)
        else:
            self.payment_cash_message = ""
            self.payment_mixed_message = ""

    def _set_payment_method(self, method: PaymentMethodConfig | None):
        if method:
            self.payment_method = method.get("name", "")
            self.payment_method_description = method.get("description", "")
            self.payment_method_kind = (method.get("kind", "other") or "other").lower()
        else:
            self.payment_method = ""
            self.payment_method_description = ""
            self.payment_method_kind = "other"
        self.payment_cash_amount = 0
        self.payment_cash_message = ""
        self.payment_cash_status = "neutral"
        self.payment_card_type = "Credito"
        self.payment_wallet_choice = "Yape"
        self.payment_wallet_provider = "Yape"
        self.payment_mixed_cash = 0
        self.payment_mixed_card = 0
        self.payment_mixed_wallet = 0
        self.payment_mixed_message = ""
        self.payment_mixed_status = "neutral"
        self.payment_mixed_notes = ""
        self._refresh_payment_feedback()

    def _reset_payment_fields(self):
        default_method = self._default_payment_method()
        self._set_payment_method(default_method)

    @rx.event
    def select_payment_method(self, method: str, description: str = ""):
        match = self._payment_method_by_identifier(method)
        if not match:
            return rx.toast("Metodo de pago no disponible.", duration=3000)
        if not match.get("enabled", True):
            return rx.toast("Este metodo esta inactivo.", duration=3000)
        self._set_payment_method(match)

    @rx.event
    def set_cash_amount(self, value: str):
        try:
            amount = float(value) if value else 0
        except ValueError:
            amount = 0
        self.payment_cash_amount = self._round_currency(amount)
        self._update_cash_feedback()

    @rx.event
    def set_card_type(self, card_type: str):
        self.payment_card_type = card_type

    @rx.event
    def choose_wallet_provider(self, provider: str):
        self.payment_wallet_choice = provider
        if provider == "Otro":
            self.payment_wallet_provider = ""
        else:
            self.payment_wallet_provider = provider

    @rx.event
    def set_wallet_provider_custom(self, value: str):
        self.payment_wallet_provider = value
        self.payment_wallet_choice = "Otro"

    @rx.event
    def set_mixed_notes(self, notes: str):
        self.payment_mixed_notes = notes

    @rx.event
    def set_mixed_cash_amount(self, value: str):
        self.payment_mixed_cash = self._safe_amount(value)
        self._update_mixed_message()

    @rx.event
    def set_mixed_card_amount(self, value: str):
        self.payment_mixed_card = self._safe_amount(value)
        self._update_mixed_message()

    @rx.event
    def set_mixed_wallet_amount(self, value: str):
        self.payment_mixed_wallet = self._safe_amount(value)
        self._update_mixed_message()

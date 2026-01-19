"""
Utilidades de exportacion a Excel.

Funciones para simplificar la creacion y el formato de archivos Excel.
"""
import io
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from typing import Any


# Estilos por defecto para exportaciones Excel
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def create_excel_workbook(title: str) -> tuple[Workbook, Worksheet]:
    """
    Crea un workbook de Excel con una hoja nombrada.
    
    Parametros:
        title: Titulo de la hoja activa
        
    Retorna:
        Tupla (workbook, hoja_activa)
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = title[:31]  # Nombres de hoja en Excel maximo 31 caracteres
    return workbook, sheet


def style_header_row(ws: Worksheet, row: int, columns: list[str]) -> None:
    """
    Agrega encabezados con estilo a una fila de la hoja.
    
    Parametros:
        ws: Hoja a modificar
        row: Numero de fila (1-indexed)
        columns: Lista de textos de encabezado
    """
    for col_idx, header in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def add_data_rows(
    ws: Worksheet,
    data: list[list[Any]],
    start_row: int,
    apply_border: bool = True,
) -> int:
    """
    Agrega multiples filas de datos a una hoja.
    
    Parametros:
        ws: Hoja a modificar
        data: Lista de filas (cada fila es una lista de valores)
        start_row: Fila inicial (1-indexed)
        apply_border: Si se aplican bordes a las celdas
        
    Retorna:
        Numero de fila despues de la ultima fila agregada
    """
    current_row = start_row
    for row_data in data:
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            if apply_border:
                cell.border = THIN_BORDER
        current_row += 1
    return current_row


def add_simple_headers(ws: Worksheet, headers: list[str]) -> None:
    """
    Agrega encabezados simples a la primera fila (sin estilo).
    
    Parametros:
        ws: Hoja a modificar
        headers: Lista de encabezados
    """
    ws.append(headers)


def auto_adjust_column_widths(ws: Worksheet, min_width: int = 10, max_width: int = 50) -> None:
    """
    Ajusta automaticamente el ancho de columnas segun el contenido.
    
    Parametros:
        ws: Hoja a modificar
        min_width: Ancho minimo de columna
        max_width: Ancho maximo de columna
    """
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        if adjusted_width < min_width:
            adjusted_width = min_width
        if adjusted_width > max_width:
            adjusted_width = max_width
        ws.column_dimensions[column_letter].width = adjusted_width


def create_pdf_report(
    buffer: io.BytesIO,
    title: str,
    data: list[list[Any]],
    headers: list[str],
    info_dict: dict[str, Any] | None = None,
) -> None:
    """
    Crea un reporte PDF (A4) con titulo, metadatos opcionales y una tabla de datos.

    Parametros:
        buffer: Buffer BytesIO donde se escribe el PDF
        title: Titulo del reporte
        data: Lista de filas, cada fila es una lista de valores
        headers: Encabezados de columnas de la tabla
        info_dict: Metadatos opcionales (ej. fecha de generacion, usuario)
    """
    info_dict = info_dict or {}
    page_size = A4 if A4 else letter
    if info_dict.get("page_size") == "letter":
        page_size = letter

    logo_path = info_dict.get("logo_path") or info_dict.get("logo")

    doc = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        leftMargin=36,
        rightMargin=36,
        topMargin=36,
        bottomMargin=36,
    )
    styles = getSampleStyleSheet()
    title_style = styles["Heading1"]
    info_style = ParagraphStyle(
        name="InfoStyle",
        parent=styles["Normal"],
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#374151"),
    )

    elements: list[Any] = []

    if logo_path:
        try:
            logo = Image(logo_path)
            logo.drawHeight = 50
            logo.drawWidth = 50
            header_table = Table(
                [[logo, Paragraph(title, title_style)]],
                colWidths=[60, doc.width - 60],
            )
            header_table.setStyle(
                TableStyle(
                    [
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 0),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ]
                )
            )
            elements.append(header_table)
        except Exception:
            elements.append(Paragraph(title, title_style))
    else:
        elements.append(Paragraph(title, title_style))

    elements.append(Spacer(1, 12))

    info_rows: list[list[Any]] = []
    for key, value in info_dict.items():
        if key in {"logo_path", "logo", "page_size"}:
            continue
        label = str(key).replace("_", " ").title()
        info_rows.append(
            [
                Paragraph(f"<b>{label}</b>", info_style),
                Paragraph(str(value), info_style),
            ]
        )

    if info_rows:
        info_table = Table(
            info_rows,
            colWidths=[140, doc.width - 140],
            hAlign="LEFT",
        )
        info_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.whitesmoke),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        elements.append(info_table)
        elements.append(Spacer(1, 12))

    table_data: list[list[Any]] = []
    if headers:
        table_data.append(headers)

    column_count = len(headers) if headers else max((len(row) for row in data), default=0)
    for row in data:
        normalized_row = list(row)
        if column_count:
            if len(normalized_row) < column_count:
                normalized_row += [""] * (column_count - len(normalized_row))
            elif len(normalized_row) > column_count:
                normalized_row = normalized_row[:column_count]
        table_data.append(
            ["" if value is None else value for value in normalized_row]
        )

    if not table_data:
        table_data = [["Sin datos"]]

    main_table = Table(
        table_data,
        hAlign="LEFT",
        repeatRows=1 if headers else 0,
    )
    table_style = TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]
    )

    for row_idx in range(1, len(table_data)):
        if row_idx % 2 == 0:
            table_style.add(
                "BACKGROUND", (0, row_idx), (-1, row_idx), colors.HexColor("#F3F4F6")
            )

    main_table.setStyle(table_style)
    elements.append(main_table)

    doc.build(elements)
    buffer.seek(0)

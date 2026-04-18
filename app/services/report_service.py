"""
Servicio de Reportes Contables y Financieros.

Genera reportes profesionales para evaluaciones administrativas,
contables y financieras con el nivel de detalle requerido.
"""
import io
import re
from datetime import datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from typing import Any

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from sqlmodel import select, func
from sqlalchemy import and_
from sqlalchemy.orm import selectinload

from app.models import Sale, SaleItem, Product, Client, SaleInstallment, CashboxLog, SalePayment, User
from app.enums import SaleStatus, PaymentMethodType
from app.i18n import MSG
from app.utils.tenant import set_tenant_context, tenant_bypass
from app.utils.exports import _safe_decimal, _sanitize_excel_value
from app.utils.timezone import format_local_datetime, to_local_datetime, utc_now_naive


# =============================================================================
# ESTILOS PROFESIONALES PARA REPORTES (Unificados con exports.py)
# =============================================================================

# Colores consistentes con el sistema - Púrpura/Índigo corporativo
TITLE_FONT = Font(bold=True, size=16, color="4F46E5")
SUBTITLE_FONT = Font(bold=True, size=12, color="6366F1")
HEADER_FONT = Font(bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="C7D2FE", end_color="C7D2FE", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
POSITIVE_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
NEGATIVE_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
CURRENCY_FORMAT = '"S/"#,##0.00'
PERCENT_FORMAT = '0.00%'
DATE_FORMAT = 'DD/MM/YYYY'
DATETIME_FORMAT = 'DD/MM/YYYY HH:MM:SS'
NUMBER_FORMAT = '#,##0.####'

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

THICK_BORDER_BOTTOM = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="medium"),
)


def _round_currency(value: float | Decimal) -> Decimal:
    """Redondea a 2 decimales para moneda."""
    return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _format_currency(value: float | Decimal, currency_symbol: str | None = None) -> str:
    """Formatea valor como moneda."""
    symbol = (currency_symbol or "").strip()
    if symbol:
        symbol = f"{symbol} "
    else:
        symbol = "S/ "
    return f"{symbol}{_round_currency(value):,.2f}"


def _currency_label(currency_symbol: str | None) -> str:
    symbol = (currency_symbol or "").strip()
    return symbol or "S/"


def _currency_format(currency_symbol: str | None) -> str:
    symbol = _currency_label(currency_symbol).replace('"', "")
    return f'"{symbol}"#,##0.00'


def _report_now(
    generated_at: datetime | None,
    country_code: str | None,
    timezone: str | None,
) -> datetime:
    base = generated_at or utc_now_naive()
    if not country_code and not timezone:
        return base
    localized = to_local_datetime(
        base,
        country_code,
        timezone=timezone,
    )
    return localized or base


def _format_report_datetime(
    value: datetime | None,
    fmt: str,
    country_code: str | None,
    timezone: str | None,
) -> str:
    return format_local_datetime(
        value,
        fmt,
        country_code,
        timezone=timezone,
    )


def _add_company_header(
    ws: Worksheet,
    company_name: str,
    report_title: str,
    period: str,
    columns: int = 8,
    generated_at: datetime | None = None,
    country_code: str | None = None,
    timezone: str | None = None,
) -> int:
    """
    Agrega encabezado de empresa profesional al reporte.

    Args:
        ws: Hoja de trabajo
        company_name: Nombre de la empresa
        report_title: Título del reporte
        period: Período del reporte
        columns: Número de columnas para el merge (default 8)

    Retorna la fila siguiente disponible.
    """
    end_col = get_column_letter(columns)

    # Fila 1: Logo / Nombre empresa con fondo
    ws.merge_cells(f"A1:{end_col}1")
    ws["A1"] = company_name.upper()
    ws["A1"].font = Font(bold=True, size=18, color="FFFFFF")
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Fila 2: Título del reporte
    ws.merge_cells(f"A2:{end_col}2")
    ws["A2"] = report_title
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 25

    # Fila 3: Período
    ws.merge_cells(f"A3:{end_col}3")
    ws["A3"] = f"📅 Período: {period}"
    ws["A3"].font = Font(size=11, color="374151")
    ws["A3"].alignment = Alignment(horizontal="center")

    # Fila 4: Fecha de generación
    ws.merge_cells(f"A4:{end_col}4")
    generated_label = _report_now(generated_at, country_code, timezone).strftime(
        "%d/%m/%Y a las %H:%M:%S"
    )
    ws["A4"] = f"🕐 Generado: {generated_label}"
    ws["A4"].font = Font(italic=True, size=9, color="6B7280")
    ws["A4"].alignment = Alignment(horizontal="center")

    # Línea separadora visual (fila 5 vacía con borde inferior)
    ws.merge_cells(f"A5:{end_col}5")
    ws["A5"].border = Border(bottom=Side(style="medium", color="4F46E5"))

    return 7  # Siguiente fila disponible (dejamos una fila de espacio)


def _style_header_row(ws: Worksheet, row: int, columns: list[str]) -> None:
    """Aplica estilo a fila de encabezados."""
    for col_idx, header in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _add_totals_row_with_formulas(
    ws: Worksheet,
    row: int,
    start_data_row: int,
    columns_config: list[dict],
) -> None:
    """
    Agrega fila de totales usando fórmulas de Excel para verificabilidad.

    Args:
        ws: Hoja de trabajo
        row: Fila donde agregar los totales
        start_data_row: Primera fila de datos
        columns_config: Lista de diccionarios con configuración por columna:
            - type: 'label' | 'sum' | 'count' | 'average' | 'formula' | 'text'
            - value: Valor para 'label', 'text' o fórmula personalizada
            - col_letter: Letra de columna para fórmulas
    """
    for col_idx, config in enumerate(columns_config, start=1):
        cell = ws.cell(row=row, column=col_idx)
        col_type = config.get("type", "text")

        if col_type == "label":
            cell.value = config.get("value", "TOTAL")
        elif col_type == "sum":
            col_letter = config.get("col_letter", get_column_letter(col_idx))
            cell.value = f"=SUM({col_letter}{start_data_row}:{col_letter}{row-1})"
        elif col_type == "count":
            col_letter = config.get("col_letter", get_column_letter(col_idx))
            cell.value = f"=COUNT({col_letter}{start_data_row}:{col_letter}{row-1})"
        elif col_type == "average":
            col_letter = config.get("col_letter", get_column_letter(col_idx))
            cell.value = f"=AVERAGE({col_letter}{start_data_row}:{col_letter}{row-1})"
        elif col_type == "formula":
            cell.value = config.get("value", "")
        elif col_type == "text":
            cell.value = config.get("value", "")

        cell.font = Font(bold=True)
        cell.fill = TOTAL_FILL
        cell.border = THICK_BORDER_BOTTOM

        # Aplicar formato numérico si se especifica
        if config.get("number_format"):
            cell.number_format = config["number_format"]


def _add_notes_section(ws: Worksheet, row: int, notes: list[str], columns: int = 8) -> int:
    """
    Agrega una sección de notas explicativas al final de la hoja.

    Args:
        ws: Hoja de trabajo
        row: Fila inicial
        notes: Lista de notas a agregar
        columns: Número de columnas para el merge

    Returns:
        Siguiente fila disponible
    """
    row += 2  # Espacio
    end_col = get_column_letter(columns)

    ws.merge_cells(f"A{row}:{end_col}{row}")
    ws[f"A{row}"] = "📋 NOTAS Y DEFINICIONES:"
    ws[f"A{row}"].font = Font(bold=True, size=10, color="374151")
    row += 1

    for note in notes:
        ws.merge_cells(f"A{row}:{end_col}{row}")
        ws[f"A{row}"] = f"• {note}"
        ws[f"A{row}"].font = Font(size=9, color="6B7280")
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)
        row += 1

    return row


def _safe_int(value: Any) -> int:
    """Convierte un valor a int de forma segura."""
    if value is None:
        return 0
    try:
        return int(value)
    except (ValueError, TypeError):
        return 0


def _safe_string(value: Any, default: str = "") -> str:
    """Convierte un valor a string de forma segura."""
    if value is None:
        return default
    try:
        raw = str(value).strip()
        return str(_sanitize_excel_value(raw))
    except:
        return default


def _build_user_lookup(
    session,
    records: list,
    company_id: int | None,
) -> dict[int, str]:
    """Obtiene usernames por user_id para cualquier lista de registros con .user_id.

    Funciona con Sales, CashboxLog, o cualquier modelo que tenga user_id.
    Evita filtros por sucursal usando tenant_bypass para cross-branch reporting.
    """
    user_ids = {
        int(record.user_id)
        for record in records
        if record is not None and getattr(record, "user_id", None) is not None
    }
    if not user_ids:
        return {}

    query = select(User.id, User.username).where(User.id.in_(user_ids))
    if company_id:
        query = query.where(User.company_id == company_id)

    with tenant_bypass():
        rows = session.exec(query.execution_options(tenant_bypass=True)).all()

    lookup: dict[int, str] = {}
    for row in rows:
        try:
            user_id, username = row[0], row[1]
        except Exception:
            user_id = getattr(row, "id", None)
            username = getattr(row, "username", None)
        safe_id = _safe_int(user_id)
        safe_name = _safe_string(username)
        if safe_id and safe_name:
            lookup[safe_id] = safe_name
    return lookup


# Aliases retrocompatibles (evitan romper callers existentes)
_build_sale_user_lookup = _build_user_lookup
_build_cashbox_log_user_lookup = _build_user_lookup


def _resolve_sale_username(
    sale: Sale,
    user_lookup: dict[int, str],
    default: str = MSG.REPORT_UNKNOWN,
) -> str:
    """Resuelve el usuario de una venta usando relación y fallback por user_id."""
    if sale.user and getattr(sale.user, "username", None):
        return _safe_string(sale.user.username, default)
    user_id = _safe_int(getattr(sale, "user_id", None))
    if user_id and user_id in user_lookup:
        return _safe_string(user_lookup.get(user_id), default)
    return default


def _translate_cashbox_action(action: str | None) -> str:
    raw = _safe_string(action).lower()
    mapping = {
        "apertura": MSG.REPORT_MOVEMENT_TYPES["apertura"],
        "cierre": MSG.REPORT_MOVEMENT_TYPES["cierre"],
        "venta": MSG.REPORT_MOVEMENT_TYPES["venta"],
        "reserva": MSG.REPORT_MOVEMENT_TYPES["reserva"],
        "adelanto": MSG.REPORT_MOVEMENT_TYPES["adelanto"],
        "cobranza": MSG.REPORT_MOVEMENT_TYPES["cobranza"],
        "inicial credito": MSG.REPORT_MOVEMENT_TYPES["inicial_credito"],
        "gasto_caja_chica": MSG.REPORT_MOVEMENT_TYPES["gasto_caja_chica"],
        "gasto caja chica": MSG.REPORT_MOVEMENT_TYPES["gasto_caja_chica"],
    }
    if raw in mapping:
        return mapping[raw]
    return raw.replace("_", " ").strip().title() or "Movimiento"


def _cashbox_action_nature(action: str | None) -> str:
    raw = _safe_string(action).lower()
    if raw == "apertura":
        return "Apertura"
    if raw == "cierre":
        return "Cierre"
    if "gasto" in raw or "egreso" in raw:
        return "Egreso"
    return "Ingreso"


def _variant_label_from_item(item: SaleItem) -> str:
    variant = item.product_variant
    if variant:
        parts: list[str] = []
        if variant.size:
            parts.append(f"Talla {str(variant.size).strip()}")
        if variant.color:
            parts.append(str(variant.color).strip())
        label = " / ".join([p for p in parts if p])
        return label or "-"
    snapshot = (item.product_name_snapshot or "").strip()
    if snapshot:
        match = re.search(r"\(([^)]+)\)$", snapshot)
        if match:
            return match.group(1).strip() or "-"
    return "-"


def _product_base_name_from_item(item: SaleItem, variant_label: str) -> str:
    if item.product and item.product.description:
        return item.product.description
    name = (item.product_name_snapshot or "").strip() or "Producto"
    if variant_label != "-" and name.endswith(")") and "(" in name:
        base = name.rsplit("(", 1)[0].strip()
        return base or name
    return name


def _translate_payment_method(method: str) -> str:
    """
    Traduce códigos de método de pago a español legible.

    Args:
        method: Código o nombre del método de pago

    Returns:
        Nombre en español del método de pago
    """
    _pm = MSG.REPORT_PAYMENT_METHODS
    translations = {
        # Códigos del enum PaymentMethodType
        "cash": _pm["efectivo"],
        "efectivo": _pm["efectivo"],
        "debit": _pm["tarjeta_debito"],
        "debito": _pm["tarjeta_debito"],
        "tarjeta de débito": _pm["tarjeta_debito"],
        "tarjeta de debito": _pm["tarjeta_debito"],
        "credit": _pm["tarjeta_credito"],
        "credito": _pm["tarjeta_credito"],
        "tarjeta de crédito": _pm["tarjeta_credito"],
        "tarjeta de credito": _pm["tarjeta_credito"],
        "card": "Tarjeta",
        "tarjeta": "Tarjeta",
        "yape": _pm["yape"],
        "plin": _pm["plin"],
        "transfer": _pm["transferencia"],
        "transferencia": _pm["transferencia"],
        "transferencia bancaria": _pm["transferencia"],
        "bank_transfer": _pm["transferencia"],
        "wallet": _pm["billetera_digital"],
        "billetera": _pm["billetera_digital"],
        "mixed": _pm["mixto"],
        "mixto": _pm["mixto"],
        "pago mixto": _pm["mixto"],
        "other": _pm["otro"],
        "otro": _pm["otro"],
        # Ventas a crédito/fiado (condición de pago, no método)
        "fiado": _pm["credito"],
        "venta a credito": _pm["credito"],
        "venta a crédito": _pm["credito"],
        "credito_fiado": _pm["credito"],
        # Códigos legacy
        "credit_card": _pm["tarjeta_credito"],
        "debit_card": _pm["tarjeta_debito"],
        "check": _pm["cheque"],
        "cheque": _pm["cheque"],
        "no especificado": _pm["no_especificado"],
        "": _pm["no_especificado"],
    }

    method_lower = method.lower().strip()
    return translations.get(method_lower, method.capitalize())


def _auto_adjust_columns(ws: Worksheet, min_width: int = 12, max_width: int = 50) -> None:
    """Ajusta automáticamente el ancho de columnas."""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted_width


# =============================================================================
# REPORTE DE VENTAS CONSOLIDADO
# =============================================================================

def generate_sales_report(
    session,
    start_date: datetime,
    end_date: datetime,
    company_name: str = "TUWAYKIAPP",
    include_cancelled: bool = False,
    currency_symbol: str | None = None,
    company_id: int | None = None,
    branch_id: int | None = None,
    country_code: str | None = None,
    timezone: str | None = None,
    generated_at: datetime | None = None,
) -> io.BytesIO:
    """
    Genera reporte de ventas consolidado con detalles contables.

    Incluye:
    - Resumen ejecutivo
    - Detalle de ventas por día
    - Desglose por categoría
    - Desglose por método de pago
    - Análisis de utilidad bruta
    - Listado detallado de transacciones
    """
    set_tenant_context(company_id, branch_id)
    wb = Workbook()
    currency_label = _currency_label(currency_symbol)
    currency_format = _currency_format(currency_symbol)
    header_kwargs = {
        "generated_at": generated_at,
        "country_code": country_code,
        "timezone": timezone,
    }

    # Consultar ventas del período
    query = (
        select(Sale)
        .where(
            and_(
                Sale.timestamp >= start_date,
                Sale.timestamp <= end_date,
            )
        )
        .options(
            selectinload(Sale.items).selectinload(SaleItem.product),
            selectinload(Sale.items).selectinload(SaleItem.product_variant),
            selectinload(Sale.payments),
            selectinload(Sale.user),
            selectinload(Sale.client),
        )
        .order_by(Sale.timestamp.desc())
    )
    if company_id:
        query = query.where(Sale.company_id == company_id)
    if branch_id:
        query = query.where(Sale.branch_id == branch_id)

    if not include_cancelled:
        query = query.where(Sale.status != SaleStatus.cancelled)

    sales = session.exec(query).all()
    # P1-03: Sale.user ya está prefetched (selectinload arriba) → evitar query redundante
    sale_user_lookup: dict[int, str] = {
        int(s.user_id): _safe_string(s.user.username)
        for s in sales
        if s.user_id is not None and s.user is not None and getattr(s.user, "username", None)
    }

    period_str = (
        f"{_format_report_datetime(start_date, '%d/%m/%Y', country_code, timezone)} - "
        f"{_format_report_datetime(end_date, '%d/%m/%Y', country_code, timezone)}"
    )

    # =================
    # HOJA 1: RESUMEN EJECUTIVO
    # =================
    ws_summary = wb.active
    ws_summary.title = MSG.REPORT_SUMMARY_SHEET

    row = _add_company_header(
        ws_summary,
        company_name,
        MSG.REPORT_TITLE,
        period_str,
        generated_at=generated_at,
        country_code=country_code,
        timezone=timezone,
    )

    # Calcular métricas
    total_ventas = Decimal("0")
    total_costo = Decimal("0")
    total_descuentos = Decimal("0")
    ventas_count = 0
    ventas_credito = 0
    ventas_contado = 0
    monto_credito = Decimal("0")
    monto_contado = Decimal("0")

    by_category: dict[str, dict] = {}
    by_payment: dict[str, dict] = {}
    by_day: dict[str, dict] = {}
    by_user: dict[str, dict] = {}

    for sale in sales:
        if sale.status == SaleStatus.cancelled:
            continue

        ventas_count += 1
        sale_total = Decimal(str(sale.total_amount or 0))
        total_ventas += sale_total

        # Clasificar por tipo
        is_credit = (
            bool(getattr(sale, "is_credit", False)) or
            (sale.payment_condition or "").lower() in {"credito", "credit"}
        )

        if is_credit:
            ventas_credito += 1
            monto_credito += sale_total
        else:
            ventas_contado += 1
            monto_contado += sale_total

        # Por día
        day_key = (
            _format_report_datetime(
                sale.timestamp,
                "%Y-%m-%d",
                country_code,
                timezone,
            )
            if sale.timestamp
            else "Sin fecha"
        )
        if day_key not in by_day:
            by_day[day_key] = {"count": 0, "total": Decimal("0"), "cost": Decimal("0")}
        by_day[day_key]["count"] += 1
        by_day[day_key]["total"] += sale_total

        # Por usuario
        user_name = _resolve_sale_username(sale, sale_user_lookup, MSG.REPORT_UNKNOWN)
        if user_name not in by_user:
            by_user[user_name] = {"count": 0, "total": Decimal("0")}
        by_user[user_name]["count"] += 1
        by_user[user_name]["total"] += sale_total

        # Por categoría y calcular costo
        for item in (sale.items or []):
            item_total = Decimal(str(item.subtotal or 0))
            # Obtener costo del producto relacionado
            cost_price = Decimal(str(item.product.purchase_price or 0)) if item.product else Decimal("0")
            item_cost = cost_price * Decimal(str(item.quantity or 0))
            total_costo += item_cost

            category = item.product_category_snapshot or "Sin categoría"
            if category not in by_category:
                by_category[category] = {"count": 0, "total": Decimal("0"), "cost": Decimal("0"), "qty": 0}
            by_category[category]["count"] += 1
            by_category[category]["total"] += item_total
            by_category[category]["cost"] += item_cost
            by_category[category]["qty"] += int(item.quantity or 0)

        # Por método de pago
        for payment in (sale.payments or []):
            method = payment.method_type.value if payment.method_type else "No especificado"
            amount = Decimal(str(payment.amount or 0))
            if method not in by_payment:
                by_payment[method] = {"count": 0, "total": Decimal("0")}
            by_payment[method]["count"] += 1
            by_payment[method]["total"] += amount

    utilidad_bruta = total_ventas - total_costo
    margen_bruto = (utilidad_bruta / total_ventas * 100) if total_ventas > 0 else Decimal("0")
    ticket_promedio = (total_ventas / ventas_count) if ventas_count > 0 else Decimal("0")

    # Escribir resumen
    row += 1
    ws_summary.cell(row=row, column=1, value=MSG.REPORT_KPI_HEADER).font = SUBTITLE_FONT
    row += 1

    indicators = [
        (MSG.REPORT_KPI_GROSS_SALES, _format_currency(total_ventas, currency_symbol)),
        ("(-) Costo de Ventas:", _format_currency(total_costo, currency_symbol)),
        ("(=) Utilidad Bruta:", _format_currency(utilidad_bruta, currency_symbol)),
        (MSG.REPORT_KPI_MARGIN, f"{margen_bruto:.2f}%"),
        ("", ""),
        (MSG.REPORT_KPI_TRANSACTIONS, ventas_count),
        (MSG.REPORT_KPI_AVG_TICKET, _format_currency(ticket_promedio, currency_symbol)),
        ("", ""),
        (MSG.REPORT_KPI_CASH_SALES, f"{ventas_contado} ({_format_currency(monto_contado, currency_symbol)})"),
        (MSG.REPORT_KPI_CREDIT_SALES, f"{ventas_credito} ({_format_currency(monto_credito, currency_symbol)})"),
    ]

    for label, value in indicators:
        ws_summary.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws_summary.cell(row=row, column=2, value=value)
        row += 1

    _add_notes_section(ws_summary, row, [
        "Ventas Brutas: Importe total facturado en el período seleccionado.",
        "Costo de Ventas: Costo de adquisición de los ítems efectivamente vendidos.",
        "Utilidad Bruta = Ventas Brutas - Costo de Ventas.",
        "Ticket Promedio = Ventas Brutas ÷ Nº de transacciones.",
        "Ventas a Crédito: Operaciones con saldo pendiente parcial o total.",
    ], columns=8)

    # =================
    # HOJA 2: VENTAS POR DÍA (con fórmulas de Excel)
    # =================
    ws_daily = wb.create_sheet(MSG.REPORT_DAILY_SHEET)
    row = _add_company_header(
        ws_daily,
        company_name,
        MSG.REPORT_DAILY_TITLE,
        period_str,
        columns=6,
        **header_kwargs,
    )

    headers = [
        "Fecha",
        "Nº Transacciones",
        f"Venta Bruta ({currency_label})",
        f"Costo ({currency_label})",
        f"Utilidad ({currency_label})",
        "Margen (%)",
    ]
    _style_header_row(ws_daily, row, headers)
    data_start_row = row + 1
    row += 1

    for day_key in sorted(by_day.keys()):
        day_data = by_day[day_key]

        ws_daily.cell(row=row, column=1, value=_safe_string(day_key))
        ws_daily.cell(row=row, column=2, value=day_data["count"])
        ws_daily.cell(row=row, column=3, value=day_data["total"]).number_format = currency_format
        ws_daily.cell(row=row, column=4, value=day_data["cost"]).number_format = currency_format
        # Utilidad = Fórmula Excel: Venta - Costo
        ws_daily.cell(row=row, column=5, value=f"=C{row}-D{row}").number_format = currency_format
        # Margen % = Fórmula Excel: (Utilidad / Venta) * 100
        ws_daily.cell(row=row, column=6, value=f"=IF(C{row}>0,E{row}/C{row},0)").number_format = PERCENT_FORMAT

        for col in range(1, 7):
            ws_daily.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # Fila de totales con fórmulas
    totals_row = row
    _add_totals_row_with_formulas(ws_daily, totals_row, data_start_row, [
        {"type": "label", "value": "TOTALES"},
        {"type": "sum", "col_letter": "B"},
        {"type": "sum", "col_letter": "C", "number_format": currency_format},
        {"type": "sum", "col_letter": "D", "number_format": currency_format},
        {"type": "sum", "col_letter": "E", "number_format": currency_format},
        {"type": "formula", "value": f"=IF(C{totals_row}>0,E{totals_row}/C{totals_row},0)", "number_format": PERCENT_FORMAT},
    ])

    # Agregar notas explicativas
    _add_notes_section(ws_daily, totals_row, [
        "Venta Bruta: Total facturado sin incluir descuentos aplicados.",
        "Costo: Precio de compra/adquisición de los productos vendidos.",
        "Utilidad = Venta Bruta - Costo (fórmula verificable en Excel).",
        "Margen % = Utilidad ÷ Venta Bruta × 100.",
    ], columns=6)

    _auto_adjust_columns(ws_daily)

    # =================
    # HOJA 3: VENTAS POR CATEGORÍA (con fórmulas de Excel)
    # =================
    ws_category = wb.create_sheet(MSG.REPORT_CATEGORY_SHEET)
    row = _add_company_header(
        ws_category,
        company_name,
        MSG.REPORT_CATEGORY_TITLE,
        period_str,
        columns=7,
        **header_kwargs,
    )

    headers = [
        "Categoría",
        "Unidades Vendidas",
        f"Venta Bruta ({currency_label})",
        f"Costo ({currency_label})",
        f"Utilidad ({currency_label})",
        "Margen (%)",
        "Participación (%)",
    ]
    _style_header_row(ws_category, row, headers)
    cat_data_start = row + 1
    row += 1

    sorted_categories = sorted(by_category.items(), key=lambda x: x[1]["total"], reverse=True)

    for cat_name, cat_data in sorted_categories:
        ws_category.cell(row=row, column=1, value=_safe_string(cat_name))
        ws_category.cell(row=row, column=2, value=cat_data["qty"])
        ws_category.cell(row=row, column=3, value=cat_data["total"]).number_format = currency_format
        ws_category.cell(row=row, column=4, value=cat_data["cost"]).number_format = currency_format
        # Utilidad = Fórmula: Venta - Costo
        ws_category.cell(row=row, column=5, value=f"=C{row}-D{row}").number_format = currency_format
        # Margen % = Fórmula: Utilidad / Venta
        ws_category.cell(row=row, column=6, value=f"=IF(C{row}>0,E{row}/C{row},0)").number_format = PERCENT_FORMAT
        # % del Total = se calculará con referencia al total
        ws_category.cell(row=row, column=7, value=cat_data["total"]).number_format = currency_format  # Temporal

        for col in range(1, 8):
            ws_category.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    cat_totals_row = row
    _add_totals_row_with_formulas(ws_category, cat_totals_row, cat_data_start, [
        {"type": "label", "value": "TOTALES"},
        {"type": "sum", "col_letter": "B"},
        {"type": "sum", "col_letter": "C", "number_format": currency_format},
        {"type": "sum", "col_letter": "D", "number_format": currency_format},
        {"type": "sum", "col_letter": "E", "number_format": currency_format},
        {"type": "formula", "value": f"=IF(C{cat_totals_row}>0,E{cat_totals_row}/C{cat_totals_row},0)", "number_format": PERCENT_FORMAT},
        {"type": "text", "value": "100.00%"},
    ])

    # Actualizar columna G con fórmulas de participación
    for r in range(cat_data_start, cat_totals_row):
        ws_category.cell(row=r, column=7, value=f"=IF($C${cat_totals_row}>0,C{r}/$C${cat_totals_row},0)").number_format = PERCENT_FORMAT

    _add_notes_section(ws_category, cat_totals_row, [
        "Unidades Vendidas: Cantidad total de productos vendidos de esta categoría.",
        "Utilidad = Venta Bruta - Costo.",
        "Margen = Utilidad ÷ Venta Bruta (indica rentabilidad por categoría).",
        "Participación = Venta de categoría ÷ Venta Total (peso relativo).",
    ], columns=7)

    _auto_adjust_columns(ws_category)

    # =================
    # HOJA 4: POR MÉTODO DE PAGO (con fórmulas)
    # =================
    ws_payment = wb.create_sheet("Por Método de Pago")
    row = _add_company_header(
        ws_payment,
        company_name,
        "ANÁLISIS POR MÉTODO DE PAGO",
        period_str,
        columns=5,
        **header_kwargs,
    )

    # ---- SECCIÓN 1: PAGOS RECIBIDOS ----
    row += 1
    ws_payment.cell(row=row, column=1, value="PAGOS RECIBIDOS EN EL PERÍODO").font = SUBTITLE_FONT
    ws_payment.cell(row=row, column=2, value="(Dinero efectivamente cobrado)").font = Font(italic=True, size=9, color="6B7280")
    row += 1

    headers = [
        "Método de Pago",
        "Nº Operaciones",
        f"Monto Recaudado ({currency_label})",
        "Participación (%)",
        "Observación",
    ]
    _style_header_row(ws_payment, row, headers)
    pay_data_start = row + 1
    row += 1

    # Ordenar métodos de pago por monto (mayor a menor)
    sorted_payments = sorted(by_payment.items(), key=lambda x: x[1]["total"], reverse=True)

    for method, method_data in sorted_payments:
        # Traducir métodos de pago a español
        method_es = _translate_payment_method(method)
        ws_payment.cell(row=row, column=1, value=_safe_string(method_es))
        ws_payment.cell(row=row, column=2, value=method_data["count"])
        ws_payment.cell(row=row, column=3, value=method_data["total"]).number_format = currency_format
        ws_payment.cell(row=row, column=4, value=method_data["total"]).number_format = currency_format  # Temporal

        # Observación según tipo
        method_lower = method.lower()
        if method_lower == "cash":
            obs = "Verificar con arqueo de caja"
        elif method_lower in {"debit", "credit", "card"}:
            obs = "Verificar con POS/banco"
        elif method_lower in {"yape", "plin", "wallet"}:
            obs = "Verificar en app billetera"
        elif method_lower == "transfer":
            obs = "Verificar en extracto bancario"
        elif method_lower == "mixed":
            obs = "Combina múltiples métodos"
        else:
            obs = ""
        ws_payment.cell(row=row, column=5, value=_safe_string(obs))

        for col in range(1, 6):
            ws_payment.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    pay_totals_row = row
    _add_totals_row_with_formulas(ws_payment, pay_totals_row, pay_data_start, [
        {"type": "label", "value": "TOTAL RECAUDADO"},
        {"type": "sum", "col_letter": "B"},
        {"type": "sum", "col_letter": "C", "number_format": currency_format},
        {"type": "text", "value": "100.00%"},
        {"type": "text", "value": ""},
    ])

    # Actualizar columna D con fórmulas de participación
    for r in range(pay_data_start, pay_totals_row):
        ws_payment.cell(row=r, column=4, value=f"=IF($C${pay_totals_row}>0,C{r}/$C${pay_totals_row},0)").number_format = PERCENT_FORMAT

    # ---- SECCIÓN 2: VENTAS A CRÉDITO/FIADO ----
    row += 3
    ws_payment.cell(row=row, column=1, value="VENTAS A CRÉDITO/FIADO").font = SUBTITLE_FONT
    ws_payment.cell(row=row, column=2, value="(Pendiente de cobro)").font = Font(italic=True, size=9, color="6B7280")
    row += 1

    headers_credit = [
        "Concepto",
        "Cantidad",
        f"Monto ({currency_label})",
        "% del Total Ventas",
        "Estado",
    ]
    _style_header_row(ws_payment, row, headers_credit)
    credit_data_start = row + 1
    row += 1

    # Calcular monto pendiente de créditos (total crédito - pagos recibidos de créditos)
    # Nota: monto_credito es el total de ventas a crédito
    total_pagos_recibidos = sum(d["total"] for d in by_payment.values())

    ws_payment.cell(row=row, column=1, value="Ventas a Crédito/Fiado")
    ws_payment.cell(row=row, column=2, value=ventas_credito)
    ws_payment.cell(row=row, column=3, value=monto_credito).number_format = currency_format
    if total_ventas > 0:
        ws_payment.cell(row=row, column=4, value=monto_credito / total_ventas).number_format = PERCENT_FORMAT
    else:
        ws_payment.cell(row=row, column=4, value=0).number_format = PERCENT_FORMAT
    ws_payment.cell(row=row, column=5, value="Cuentas por cobrar")
    ws_payment.cell(row=row, column=5).font = Font(color="B45309")  # Amber
    ws_payment.cell(row=row, column=5).fill = WARNING_FILL

    for col in range(1, 6):
        ws_payment.cell(row=row, column=col).border = THIN_BORDER
    row += 1

    # ---- SECCIÓN 3: RESUMEN GENERAL ----
    row += 2
    ws_payment.cell(row=row, column=1, value="RESUMEN GENERAL DE VENTAS").font = SUBTITLE_FONT
    row += 2

    ws_payment.cell(row=row, column=1, value="Total Ventas del Período:").font = Font(bold=True)
    ws_payment.cell(row=row, column=2, value=total_ventas).number_format = currency_format
    row += 1

    ws_payment.cell(row=row, column=1, value="(-) Pagos Recibidos:").font = Font(bold=True)
    ws_payment.cell(row=row, column=2, value=f"=C{pay_totals_row}").number_format = currency_format
    row += 1

    ws_payment.cell(row=row, column=1, value="(=) Pendiente de Cobro:").font = Font(bold=True, color="B45309")
    ws_payment.cell(row=row, column=2, value=f"={total_ventas}-C{pay_totals_row}").number_format = currency_format
    ws_payment.cell(row=row, column=2).fill = WARNING_FILL

    _add_notes_section(ws_payment, row, [
        "PAGOS RECIBIDOS: Dinero efectivamente cobrado, clasificado por método de pago.",
        "  • Efectivo: Billetes y monedas recibidos.",
        "  • Tarjeta de Débito/Crédito: Pagos procesados por POS.",
        "  • Yape/Plin: Pagos recibidos por billetera digital.",
        "  • Transferencia: Depósitos bancarios confirmados.",
        "  • Pago Mixto: Combinación de varios métodos en una sola venta.",
        "VENTAS A CRÉDITO: Ventas fiadas, pendientes de cobro. Ver módulo Cuentas por Cobrar.",
        "Pendiente de Cobro = Total Ventas - Pagos Recibidos.",
    ], columns=5)

    _auto_adjust_columns(ws_payment)

    # =================
    # HOJA 5: POR VENDEDOR (con fórmulas)
    # =================
    ws_user = wb.create_sheet(MSG.REPORT_BY_SELLER_SHEET)
    row = _add_company_header(
        ws_user,
        company_name,
        "RENDIMIENTO POR VENDEDOR",
        period_str,
        columns=5,
        **header_kwargs,
    )

    headers = [
        "Vendedor",
        "Nº Transacciones",
        f"Venta Total ({currency_label})",
        f"Ticket Promedio ({currency_label})",
        "Participación (%)",
    ]
    _style_header_row(ws_user, row, headers)
    user_data_start = row + 1
    row += 1

    sorted_users = sorted(by_user.items(), key=lambda x: x[1]["total"], reverse=True)

    for user_name, user_data in sorted_users:
        ws_user.cell(row=row, column=1, value=_safe_string(user_name))
        ws_user.cell(row=row, column=2, value=user_data["count"])
        ws_user.cell(row=row, column=3, value=user_data["total"]).number_format = currency_format
        # Ticket Promedio = Fórmula: Venta Total / Nº Transacciones
        ws_user.cell(row=row, column=4, value=f"=IF(B{row}>0,C{row}/B{row},0)").number_format = currency_format
        # Participación - se calculará con referencia al total
        ws_user.cell(row=row, column=5, value=user_data["total"]).number_format = currency_format  # Temporal

        for col in range(1, 6):
            ws_user.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    user_totals_row = row
    _add_totals_row_with_formulas(ws_user, user_totals_row, user_data_start, [
        {"type": "label", "value": "TOTAL EQUIPO"},
        {"type": "sum", "col_letter": "B"},
        {"type": "sum", "col_letter": "C", "number_format": currency_format},
        {"type": "formula", "value": f"=IF(B{user_totals_row}>0,C{user_totals_row}/B{user_totals_row},0)", "number_format": currency_format},
        {"type": "text", "value": "100.00%"},
    ])

    # Actualizar columna E con fórmulas de participación
    for r in range(user_data_start, user_totals_row):
        ws_user.cell(row=r, column=5, value=f"=IF($C${user_totals_row}>0,C{r}/$C${user_totals_row},0)").number_format = PERCENT_FORMAT

    _add_notes_section(ws_user, user_totals_row, [
        "Nº Transacciones: Cantidad de ventas realizadas por el vendedor.",
        "Ticket Promedio = Venta Total ÷ Nº Transacciones.",
        "Participación: Porcentaje de ventas del vendedor respecto al total del equipo.",
    ], columns=5)

    _auto_adjust_columns(ws_user)

    # =================
    # HOJA 6: DETALLE DE TRANSACCIONES (por item)
    # =================
    ws_detail = wb.create_sheet(MSG.REPORT_TX_DETAIL_SHEET)
    row = _add_company_header(
        ws_detail,
        company_name,
        "LISTADO DETALLADO DE TRANSACCIONES",
        period_str,
        columns=11,
        **header_kwargs,
    )

    headers = [
        "Fecha y Hora",
        "Nº Venta",
        "Cliente",
        "Vendedor",
        "Método de Pago",
        "Producto",
        "Variante",
        "Categoría",
        "Cantidad",
        f"Precio Unitario ({currency_label})",
        f"Subtotal ({currency_label})",
    ]
    _style_header_row(ws_detail, row, headers)
    detail_data_start = row + 1
    row += 1

    for sale in sales:
        # Método de pago
        payment_method = "No especificado"
        for payment in (sale.payments or []):
            if payment.method_type:
                payment_method = _translate_payment_method(payment.method_type.value)
                break

        is_credit = (
            bool(getattr(sale, "is_credit", False)) or
            (sale.payment_condition or "").lower() in {"credito", "credit"}
        )
        if is_credit:
            payment_method = "Crédito/Fiado"

        sale_items = sale.items or []
        if not sale_items:
            sale_items = [None]

        for item in sale_items:
            if item is None:
                product_name = "Sin productos"
                variant_label = "-"
                category = "Sin categoría"
                quantity = Decimal("0")
                unit_price = Decimal("0")
                subtotal = Decimal("0")
            else:
                variant_label = _variant_label_from_item(item)
                product_name = _product_base_name_from_item(item, variant_label)
                category = (
                    item.product_category_snapshot
                    or (item.product.category if item.product else "")
                    or ("Servicios" if item.product_id is None else "General")
                )
                quantity = _safe_decimal(item.quantity)
                unit_price = _safe_decimal(item.unit_price)
                subtotal = _safe_decimal(item.subtotal)

            ws_detail.cell(
                row=row,
                column=1,
                value=(
                    _format_report_datetime(
                        sale.timestamp,
                        "%d/%m/%Y %H:%M",
                        country_code,
                        timezone,
                    )
                    if sale.timestamp
                    else "Sin fecha"
                ),
            )
            ws_detail.cell(row=row, column=2, value=sale.id)
            ws_detail.cell(
                row=row,
                column=3,
                value=_safe_string(
                    sale.client.name if sale.client else None,
                    MSG.REPORT_GENERAL_CLIENT,
                ),
            )
            ws_detail.cell(
                row=row,
                column=4,
                value=_safe_string(
                    _resolve_sale_username(sale, sale_user_lookup, "Sistema"),
                    "Sistema",
                ),
            )
            ws_detail.cell(row=row, column=5, value=_safe_string(payment_method))
            ws_detail.cell(row=row, column=6, value=_safe_string(product_name))
            ws_detail.cell(row=row, column=7, value=_safe_string(variant_label, "-"))
            ws_detail.cell(row=row, column=8, value=_safe_string(category, "General"))
            ws_detail.cell(row=row, column=9, value=quantity).number_format = NUMBER_FORMAT
            ws_detail.cell(row=row, column=10, value=unit_price).number_format = currency_format
            ws_detail.cell(row=row, column=11, value=subtotal).number_format = currency_format

            for col in range(1, 12):
                ws_detail.cell(row=row, column=col).border = THIN_BORDER
            row += 1

    # Fila de totales con fórmulas
    detail_totals_row = row
    _add_totals_row_with_formulas(ws_detail, detail_totals_row, detail_data_start, [
        {"type": "label", "value": "TOTALES"},
        {"type": "text", "value": ""},
        {"type": "text", "value": ""},
        {"type": "text", "value": ""},
        {"type": "text", "value": ""},
        {"type": "text", "value": ""},
        {"type": "text", "value": ""},
        {"type": "text", "value": ""},
        {"type": "sum", "col_letter": "I", "number_format": NUMBER_FORMAT},
        {"type": "text", "value": ""},
        {"type": "sum", "col_letter": "K", "number_format": currency_format},
    ])

    _add_notes_section(ws_detail, detail_totals_row, [
        "Cada fila representa un ítem vendido dentro de una transacción.",
        "Nº Venta: correlativo interno de la operación comercial.",
        "Subtotal = Cantidad × Precio Unitario para cada ítem.",
        "Método de Pago muestra la forma de cobro aplicada a la venta.",
    ], columns=11)

    _auto_adjust_columns(ws_detail)

    # =================
    # HOJA 7: PRODUCTOS MÁS VENDIDOS (TOP 20, con fórmulas)
    # =================
    ws_top_products = wb.create_sheet(MSG.REPORT_TOP_PRODUCTS_SHEET)
    row = _add_company_header(
        ws_top_products,
        company_name,
        "RANKING DE PRODUCTOS MÁS VENDIDOS",
        period_str,
        columns=8,
        **header_kwargs,
    )

    # Calcular productos más vendidos
    by_product: dict[str, dict] = {}
    for sale in sales:
        if sale.status == SaleStatus.cancelled:
            continue
        for item in (sale.items or []):
            product_name = item.product_name_snapshot or "Producto sin nombre"
            product_id = item.product_id or 0
            key = f"{product_id}|{product_name}"
            qty = _safe_int(item.quantity)
            subtotal = _safe_decimal(item.subtotal)
            cost_price = _safe_decimal(item.product.purchase_price) if item.product else Decimal("0")
            item_cost = cost_price * Decimal(str(qty))

            if key not in by_product:
                by_product[key] = {
                    "name": product_name,
                    "category": item.product_category_snapshot or "Sin categoría",
                    "qty": 0,
                    "total": Decimal("0"),
                    "cost": Decimal("0"),
                    "transactions": 0,
                }
            by_product[key]["qty"] += qty
            by_product[key]["total"] += subtotal
            by_product[key]["cost"] += item_cost
            by_product[key]["transactions"] += 1

    headers = [
        "Producto", "Categoría", "Unidades Vendidas", "Nº Ventas",
        f"Ingresos ({currency_label})", f"Costo ({currency_label})", f"Utilidad ({currency_label})", "Margen (%)"
    ]
    _style_header_row(ws_top_products, row, headers)
    top_data_start = row + 1
    row += 1

    # Ordenar por total de venta (mayor primero) y limitar a top 20
    sorted_products = sorted(by_product.values(), key=lambda x: x["total"], reverse=True)[:20]

    for prod in sorted_products:
        ws_top_products.cell(row=row, column=1, value=_safe_string(prod["name"][:50]))
        ws_top_products.cell(row=row, column=2, value=_safe_string(prod["category"]))
        ws_top_products.cell(row=row, column=3, value=prod["qty"])
        ws_top_products.cell(row=row, column=4, value=prod["transactions"])
        ws_top_products.cell(row=row, column=5, value=prod["total"]).number_format = currency_format
        ws_top_products.cell(row=row, column=6, value=prod["cost"]).number_format = currency_format
        # Utilidad = Fórmula: Ingresos - Costo
        ws_top_products.cell(row=row, column=7, value=f"=E{row}-F{row}").number_format = currency_format
        # Margen = Fórmula: Utilidad / Ingresos
        ws_top_products.cell(row=row, column=8, value=f"=IF(E{row}>0,G{row}/E{row},0)").number_format = PERCENT_FORMAT

        for col in range(1, 9):
            ws_top_products.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # Totales con fórmulas
    top_totals_row = row
    _add_totals_row_with_formulas(ws_top_products, top_totals_row, top_data_start, [
        {"type": "label", "value": "TOTALES TOP 20"},
        {"type": "text", "value": ""},
        {"type": "sum", "col_letter": "C"},
        {"type": "sum", "col_letter": "D"},
        {"type": "sum", "col_letter": "E", "number_format": currency_format},
        {"type": "sum", "col_letter": "F", "number_format": currency_format},
        {"type": "sum", "col_letter": "G", "number_format": currency_format},
        {"type": "formula", "value": f"=IF(E{top_totals_row}>0,G{top_totals_row}/E{top_totals_row},0)", "number_format": PERCENT_FORMAT},
    ])

    _add_notes_section(ws_top_products, top_totals_row, [
        "Este ranking muestra los 20 productos con mayores ingresos en el período.",
        "Unidades Vendidas: Cantidad total de unidades vendidas del producto.",
        "Nº Ventas: Cantidad de transacciones donde aparece el producto.",
        "Margen = Utilidad ÷ Ingresos (rentabilidad del producto).",
    ], columns=8)

    _auto_adjust_columns(ws_top_products)

    # =================
    # HOJA 8: ANÁLISIS HORARIO (con fórmulas)
    # =================
    ws_hourly = wb.create_sheet(MSG.REPORT_HOURLY_SHEET)
    row = _add_company_header(
        ws_hourly,
        company_name,
        "DISTRIBUCIÓN DE VENTAS POR HORA",
        period_str,
        columns=5,
        **header_kwargs,
    )

    by_hour: dict[int, dict] = {}
    for sale in sales:
        if sale.status == SaleStatus.cancelled:
            continue
        if sale.timestamp:
            local_timestamp = to_local_datetime(
                sale.timestamp,
                country_code,
                timezone=timezone,
            ) or sale.timestamp
            hour = local_timestamp.hour
            if hour not in by_hour:
                by_hour[hour] = {"count": 0, "total": Decimal("0")}
            by_hour[hour]["count"] += 1
            by_hour[hour]["total"] += _safe_decimal(sale.total_amount)

    headers = [
        "Franja Horaria",
        "Nº Transacciones",
        f"Venta Total ({currency_label})",
        "Participación (%)",
        f"Ticket Promedio ({currency_label})",
    ]
    _style_header_row(ws_hourly, row, headers)
    hourly_data_start = row + 1
    row += 1

    for hour in sorted(by_hour.keys()):
        hour_data = by_hour[hour]

        ws_hourly.cell(row=row, column=1, value=f"{hour:02d}:00 - {hour:02d}:59")
        ws_hourly.cell(row=row, column=2, value=hour_data["count"])
        ws_hourly.cell(row=row, column=3, value=hour_data["total"]).number_format = currency_format
        # Participación - se calculará con referencia al total
        ws_hourly.cell(row=row, column=4, value=hour_data["total"]).number_format = currency_format  # Temporal
        # Ticket Promedio = Fórmula: Venta / Transacciones
        ws_hourly.cell(row=row, column=5, value=f"=IF(B{row}>0,C{row}/B{row},0)").number_format = currency_format

        for col in range(1, 6):
            ws_hourly.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    hourly_totals_row = row
    _add_totals_row_with_formulas(ws_hourly, hourly_totals_row, hourly_data_start, [
        {"type": "label", "value": "TOTAL DÍA"},
        {"type": "sum", "col_letter": "B"},
        {"type": "sum", "col_letter": "C", "number_format": currency_format},
        {"type": "text", "value": "100.00%"},
        {"type": "formula", "value": f"=IF(B{hourly_totals_row}>0,C{hourly_totals_row}/B{hourly_totals_row},0)", "number_format": currency_format},
    ])

    # Actualizar columna D con fórmulas de participación
    for r in range(hourly_data_start, hourly_totals_row):
        ws_hourly.cell(row=r, column=4, value=f"=IF($C${hourly_totals_row}>0,C{r}/$C${hourly_totals_row},0)").number_format = PERCENT_FORMAT

    _add_notes_section(ws_hourly, hourly_totals_row, [
        "Este análisis muestra las horas de mayor actividad comercial.",
        "Utilice esta información para optimizar horarios de personal.",
        "Franja Horaria: Período de 1 hora del día.",
        "Ticket Promedio = Venta Total ÷ Nº Transacciones.",
    ], columns=5)

    _auto_adjust_columns(ws_hourly)

    # Guardar
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# =============================================================================
# REPORTE DE INVENTARIO VALORIZADO
# =============================================================================

def generate_inventory_report(
    session,
    company_name: str = "TUWAYKIAPP",
    include_zero_stock: bool = True,
    currency_symbol: str | None = None,
    company_id: int | None = None,
    branch_id: int | None = None,
    country_code: str | None = None,
    timezone: str | None = None,
    generated_at: datetime | None = None,
) -> io.BytesIO:
    """
    Genera reporte de inventario valorizado profesional.

    Incluye:
    - Resumen de valorización
    - Detalle por categoría
    - Análisis ABC
    - Productos con stock crítico
    - Rotación estimada
    """
    set_tenant_context(company_id, branch_id)
    wb = Workbook()
    currency_label = _currency_label(currency_symbol)
    currency_format = _currency_format(currency_symbol)
    header_kwargs = {
        "generated_at": generated_at,
        "country_code": country_code,
        "timezone": timezone,
    }

    # Consultar productos
    query = (
        select(Product)
        .order_by(Product.category, Product.description)
        .options(selectinload(Product.variants))
    )
    if company_id:
        query = query.where(Product.company_id == company_id)
    if branch_id:
        query = query.where(Product.branch_id == branch_id)
    if not include_zero_stock:
        query = query.where(Product.stock > 0)

    products = session.exec(query).all()

    def _variant_label(variant: Any) -> str:
        parts: list[str] = []
        if getattr(variant, "size", None):
            parts.append(str(variant.size).strip())
        if getattr(variant, "color", None):
            parts.append(str(variant.color).strip())
        return " ".join([p for p in parts if p]).strip()

    inventory_rows: list[dict[str, Any]] = []
    for product in products:
        variants = list(product.variants or [])
        if variants:
            for variant in variants:
                stock = _safe_decimal(getattr(variant, "stock", 0))
                if not include_zero_stock and stock <= 0:
                    continue
                label = _variant_label(variant)
                description = _safe_string(product.description, "Sin descripción")
                if label:
                    description = f"{description} ({label})"
                inventory_rows.append(
                    {
                        "sku": _safe_string(variant.sku, product.barcode or "S/C"),
                        "description": description,
                        "category": _safe_string(product.category, "Sin categoría"),
                        "stock": stock,
                        "unit": _safe_string(product.unit, "Unid."),
                        "purchase_price": _safe_decimal(product.purchase_price),
                        "sale_price": _safe_decimal(product.sale_price),
                    }
                )
        else:
            stock = _safe_decimal(product.stock)
            if not include_zero_stock and stock <= 0:
                continue
            inventory_rows.append(
                {
                    "sku": _safe_string(product.barcode, "S/C"),
                    "description": _safe_string(product.description, "Sin descripción"),
                    "category": _safe_string(product.category, "Sin categoría"),
                    "stock": stock,
                    "unit": _safe_string(product.unit, "Unid."),
                    "purchase_price": _safe_decimal(product.purchase_price),
                    "sale_price": _safe_decimal(product.sale_price),
                }
            )

    report_now = _report_now(generated_at, country_code, timezone)
    today = report_now.strftime("%d/%m/%Y")

    # =================
    # HOJA 1: RESUMEN
    # =================
    ws_summary = wb.active
    ws_summary.title = MSG.REPORT_VALUATION_SHEET

    row = _add_company_header(
        ws_summary,
        company_name,
        "INVENTARIO VALORIZADO",
        f"Al {today}",
        **header_kwargs,
    )

    # Calcular métricas
    total_items = len(inventory_rows)
    total_units = sum(row["stock"] for row in inventory_rows)
    total_cost_value = sum(
        row["stock"] * row["purchase_price"] for row in inventory_rows
    )
    total_sale_value = sum(
        row["stock"] * row["sale_price"] for row in inventory_rows
    )
    potential_profit = total_sale_value - total_cost_value

    stock_zero = sum(1 for row in inventory_rows if row["stock"] == 0)
    stock_low = sum(1 for row in inventory_rows if 0 < row["stock"] <= 5)
    stock_medium = sum(1 for row in inventory_rows if 5 < row["stock"] <= 10)
    stock_ok = sum(1 for row in inventory_rows if row["stock"] > 10)

    by_category: dict[str, dict] = {}
    for row_data in inventory_rows:
        cat = row_data["category"] or "Sin categoría"
        if cat not in by_category:
            by_category[cat] = {
                "items": 0,
                "units": Decimal("0"),
                "cost": Decimal("0"),
                "sale": Decimal("0"),
            }
        by_category[cat]["items"] += 1
        by_category[cat]["units"] += row_data["stock"]
        by_category[cat]["cost"] += row_data["stock"] * row_data["purchase_price"]
        by_category[cat]["sale"] += row_data["stock"] * row_data["sale_price"]

    row += 1
    ws_summary.cell(row=row, column=1, value="RESUMEN DE VALORIZACIÓN").font = SUBTITLE_FONT
    row += 1

    summary_data = [
        ("Total de Productos (SKU):", total_items),
        ("Total de Unidades en Stock:", total_units),
        ("", ""),
        ("Valor al Costo:", _format_currency(total_cost_value, currency_symbol)),
        ("Valor a Precio Venta:", _format_currency(total_sale_value, currency_symbol)),
        ("Utilidad Potencial:", _format_currency(potential_profit, currency_symbol)),
        ("", ""),
        ("ESTADO DEL STOCK:", ""),
        ("   Sin stock (0 unidades):", stock_zero),
        ("   Stock crítico (1-5 unidades):", stock_low),
        ("   Stock bajo (6-10 unidades):", stock_medium),
        ("   Stock normal (>10 unidades):", stock_ok),
    ]

    for label, value in summary_data:
        ws_summary.cell(row=row, column=1, value=label).font = Font(bold=True) if not label.startswith("   ") else Font()
        ws_summary.cell(row=row, column=2, value=value)
        row += 1

    _add_notes_section(ws_summary, row, [
        "Valor al Costo: inversión total del inventario (stock × costo unitario).",
        "Valor a Precio Venta: potencial de facturación si se vende todo el stock.",
        "Utilidad Potencial = Valor a Precio Venta - Valor al Costo.",
        "Stock crítico: productos con 1 a 5 unidades disponibles.",
        "Stock bajo: productos con 6 a 10 unidades disponibles.",
    ], columns=8)

    _auto_adjust_columns(ws_summary)

    # =================
    # HOJA 2: POR CATEGORÍA (con fórmulas)
    # =================
    ws_category = wb.create_sheet(MSG.REPORT_CATEGORY_SHEET)
    row = _add_company_header(
        ws_category,
        company_name,
        "VALORIZACIÓN POR CATEGORÍA DE PRODUCTO",
        f"Al {today}",
        columns=7,
        **header_kwargs,
    )

    headers = [
        "Categoría",
        "Nº Productos",
        "Unidades en Stock",
        f"Valor al Costo ({currency_label})",
        f"Valor a Venta ({currency_label})",
        f"Utilidad Potencial ({currency_label})",
        "Participación (%)",
    ]
    _style_header_row(ws_category, row, headers)
    inv_cat_data_start = row + 1
    row += 1

    sorted_cats = sorted(by_category.items(), key=lambda x: x[1]["cost"], reverse=True)

    for cat_name, cat_data in sorted_cats:
        ws_category.cell(row=row, column=1, value=_safe_string(cat_name))
        ws_category.cell(row=row, column=2, value=cat_data["items"])
        ws_category.cell(row=row, column=3, value=cat_data["units"])
        ws_category.cell(row=row, column=4, value=cat_data["cost"]).number_format = currency_format
        ws_category.cell(row=row, column=5, value=cat_data["sale"]).number_format = currency_format
        # Utilidad Potencial = Fórmula: Valor Venta - Valor Costo
        ws_category.cell(row=row, column=6, value=f"=E{row}-D{row}").number_format = currency_format
        # Participación - temporal, se actualizará con fórmula
        ws_category.cell(row=row, column=7, value=cat_data["cost"]).number_format = currency_format

        for col in range(1, 8):
            ws_category.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    inv_cat_totals_row = row
    _add_totals_row_with_formulas(ws_category, inv_cat_totals_row, inv_cat_data_start, [
        {"type": "label", "value": "TOTAL INVENTARIO"},
        {"type": "sum", "col_letter": "B"},
        {"type": "sum", "col_letter": "C"},
        {"type": "sum", "col_letter": "D", "number_format": currency_format},
        {"type": "sum", "col_letter": "E", "number_format": currency_format},
        {"type": "sum", "col_letter": "F", "number_format": currency_format},
        {"type": "text", "value": "100.00%"},
    ])

    # Actualizar columna G con fórmulas de participación
    for r in range(inv_cat_data_start, inv_cat_totals_row):
        ws_category.cell(row=r, column=7, value=f"=IF($D${inv_cat_totals_row}>0,D{r}/$D${inv_cat_totals_row},0)").number_format = PERCENT_FORMAT

    _add_notes_section(ws_category, inv_cat_totals_row, [
        "Valor al Costo: Stock × Precio de Compra.",
        "Valor a Venta: Stock × Precio de Venta al Público.",
        "Utilidad Potencial = Valor a Venta - Valor al Costo (ganancia si se vende todo).",
        "Participación: Peso de la categoría sobre el valor total del inventario.",
    ], columns=7)

    _auto_adjust_columns(ws_category)

    # =================
    # HOJA 3: DETALLE COMPLETO (con fórmulas)
    # =================
    ws_detail = wb.create_sheet(MSG.REPORT_INVENTORY_SHEET)
    row = _add_company_header(
        ws_detail,
        company_name,
        "LISTADO DETALLADO DE PRODUCTOS EN INVENTARIO",
        f"Al {today}",
        columns=12,
        **header_kwargs,
    )

    headers = [
        "Código/SKU", "Descripción del Producto", "Categoría", "Stock Actual", "Unidad de Medida",
        f"Costo Unitario ({currency_label})", f"Precio Venta ({currency_label})", f"Margen Unitario ({currency_label})", "Margen (%)",
        f"Valor en Costo ({currency_label})", f"Valor en Venta ({currency_label})", "Estado Stock"
    ]
    _style_header_row(ws_detail, row, headers)
    inv_detail_start = row + 1
    row += 1

    for row_data in inventory_rows:
        stock = _safe_decimal(row_data["stock"])
        cost = _safe_decimal(row_data["purchase_price"])
        price = _safe_decimal(row_data["sale_price"])

        # Estado del stock
        if stock == 0:
            status = "SIN STOCK"
        elif stock <= 5:
            status = "⚠️ CRÍTICO"
        elif stock <= 10:
            status = "⚡ BAJO"
        else:
            status = "✅ NORMAL"

        ws_detail.cell(row=row, column=1, value=_safe_string(row_data["sku"], "S/C"))
        ws_detail.cell(row=row, column=2, value=_safe_string(row_data["description"], "Sin descripción"))
        ws_detail.cell(row=row, column=3, value=_safe_string(row_data["category"], "Sin categoría"))
        ws_detail.cell(row=row, column=4, value=stock)
        ws_detail.cell(row=row, column=5, value=_safe_string(row_data["unit"], "Unid."))
        ws_detail.cell(row=row, column=6, value=cost).number_format = currency_format
        ws_detail.cell(row=row, column=7, value=price).number_format = currency_format
        # Margen Unitario = Fórmula: Precio - Costo
        ws_detail.cell(row=row, column=8, value=f"=G{row}-F{row}").number_format = currency_format
        # Margen % = Fórmula: Margen / Costo (si costo > 0)
        ws_detail.cell(row=row, column=9, value=f"=IF(F{row}>0,H{row}/F{row},0)").number_format = PERCENT_FORMAT
        # Valor en Costo = Fórmula: Stock × Costo
        ws_detail.cell(row=row, column=10, value=f"=D{row}*F{row}").number_format = currency_format
        # Valor en Venta = Fórmula: Stock × Precio
        ws_detail.cell(row=row, column=11, value=f"=D{row}*G{row}").number_format = currency_format
        ws_detail.cell(row=row, column=12, value=_safe_string(status))

        # Color según estado
        status_cell = ws_detail.cell(row=row, column=12)
        if "SIN STOCK" in status:
            status_cell.fill = NEGATIVE_FILL
        elif "CRÍTICO" in status:
            status_cell.fill = WARNING_FILL
        elif "BAJO" in status:
            status_cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        else:
            status_cell.fill = POSITIVE_FILL

        for col in range(1, 13):
            ws_detail.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # Fila de totales con fórmulas
    inv_detail_totals = row
    _add_totals_row_with_formulas(ws_detail, inv_detail_totals, inv_detail_start, [
        {"type": "label", "value": "TOTALES"},
        {"type": "text", "value": ""},
        {"type": "text", "value": ""},
        {"type": "sum", "col_letter": "D"},
        {"type": "text", "value": ""},
        {"type": "text", "value": ""},
        {"type": "text", "value": ""},
        {"type": "text", "value": ""},
        {"type": "text", "value": ""},
        {"type": "sum", "col_letter": "J", "number_format": currency_format},
        {"type": "sum", "col_letter": "K", "number_format": currency_format},
        {"type": "text", "value": ""},
    ])

    _add_notes_section(ws_detail, inv_detail_totals, [
        "Código/SKU: Identificador único del producto (código de barras o interno).",
        "Margen Unitario = Precio Venta - Costo Unitario.",
        "Margen % = Margen Unitario ÷ Costo Unitario (rentabilidad sobre costo).",
        "Valor en Costo = Stock × Costo Unitario (inversión en inventario).",
        "Valor en Venta = Stock × Precio Venta (potencial de venta).",
        "Estados: ❌ SIN STOCK (0 unid.), ⚠️ CRÍTICO (1-5), ⚡ BAJO (6-10), ✅ NORMAL (>10).",
    ], columns=12)

    _auto_adjust_columns(ws_detail)

    # =================
    # HOJA 4: STOCK CRÍTICO (productos a reponer)
    # =================
    ws_critical = wb.create_sheet("Productos a Reponer")
    row = _add_company_header(
        ws_critical,
        company_name,
        "PRODUCTOS CON STOCK CRÍTICO - REQUIEREN REPOSICIÓN",
        f"Al {today}",
        columns=6,
        **header_kwargs,
    )

    headers = [
        "Código/SKU",
        "Descripción",
        "Categoría",
        "Stock Actual",
        f"Precio Venta ({currency_label})",
        f"Valor Disponible ({currency_label})",
    ]
    _style_header_row(ws_critical, row, headers)
    critical_data_start = row + 1
    row += 1

    critical_products = [
        row_data for row_data in inventory_rows if _safe_decimal(row_data["stock"]) <= 10
    ]
    critical_products.sort(key=lambda r: _safe_decimal(r["stock"]))

    for row_data in critical_products:
        stock = _safe_decimal(row_data["stock"])

        ws_critical.cell(row=row, column=1, value=_safe_string(row_data["sku"], "S/C"))
        ws_critical.cell(row=row, column=2, value=_safe_string(row_data["description"], "Sin descripción"))
        ws_critical.cell(row=row, column=3, value=_safe_string(row_data["category"], "Sin categoría"))
        ws_critical.cell(row=row, column=4, value=stock)
        ws_critical.cell(row=row, column=5, value=_safe_decimal(row_data["sale_price"])).number_format = currency_format
        # Valor Disponible = Fórmula: Stock × Precio
        ws_critical.cell(row=row, column=6, value=f"=D{row}*E{row}").number_format = currency_format

        for col in range(1, 7):
            ws_critical.cell(row=row, column=col).border = THIN_BORDER

        # Color según nivel de urgencia
        if stock == 0:
            for col in range(1, 7):
                ws_critical.cell(row=row, column=col).fill = NEGATIVE_FILL
        elif stock <= 5:
            for col in range(1, 7):
                ws_critical.cell(row=row, column=col).fill = WARNING_FILL

        row += 1

    critical_totals_row = row
    _add_totals_row_with_formulas(ws_critical, critical_totals_row, critical_data_start, [
        {"type": "label", "value": "TOTALES"},
        {"type": "text", "value": f"{len(critical_products)} productos"},
        {"type": "text", "value": ""},
        {"type": "sum", "col_letter": "D"},
        {"type": "text", "value": ""},
        {"type": "sum", "col_letter": "F", "number_format": currency_format},
    ])

    _add_notes_section(ws_critical, critical_totals_row, [
        "Esta lista muestra productos que necesitan reposición urgente.",
        "🔴 Rojo: Sin stock (0 unidades) - Requiere pedido inmediato.",
        "🟡 Amarillo: Stock crítico (1-5 unidades) - Prioridad alta.",
        "⚪ Sin color: Stock bajo (6-10 unidades) - Planificar reposición.",
        "Valor Disponible: Dinero en inventario de estos productos.",
    ], columns=6)

    _auto_adjust_columns(ws_critical)

    # Guardar
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# =============================================================================
# REPORTE DE CUENTAS POR COBRAR (ANTIGÜEDAD DE DEUDA)
# =============================================================================

def generate_receivables_report(
    session,
    company_name: str = "TUWAYKIAPP",
    currency_symbol: str | None = None,
    company_id: int | None = None,
    branch_id: int | None = None,
    country_code: str | None = None,
    timezone: str | None = None,
    generated_at: datetime | None = None,
) -> io.BytesIO:
    """
    Genera reporte de cuentas por cobrar con análisis de antigüedad.

    Incluye:
    - Resumen de cartera
    - Antigüedad de deuda (0-30, 31-60, 61-90, >90 días)
    - Detalle por cliente
    - Provisión sugerida para cobranza dudosa
    """
    set_tenant_context(company_id, branch_id)
    wb = Workbook()
    currency_label = _currency_label(currency_symbol)
    currency_format = _currency_format(currency_symbol)
    header_kwargs = {
        "generated_at": generated_at,
        "country_code": country_code,
        "timezone": timezone,
    }

    # Consultar cuotas pendientes
    query = (
        select(SaleInstallment)
        .where(SaleInstallment.status != "paid")
        .order_by(SaleInstallment.due_date)
    )
    if company_id:
        query = query.where(SaleInstallment.company_id == company_id)
    if branch_id:
        query = query.where(SaleInstallment.branch_id == branch_id)

    installments = session.exec(query).all()

    # Cargar ventas y clientes relacionados
    sale_ids = list(set(inst.sale_id for inst in installments if inst.sale_id))
    sales_map = {}
    clients_map = {}

    if sale_ids:
        sales_query = (
            select(Sale)
            .where(Sale.id.in_(sale_ids))
            .options(selectinload(Sale.client))
        )
        if company_id:
            sales_query = sales_query.where(Sale.company_id == company_id)
        if branch_id:
            sales_query = sales_query.where(Sale.branch_id == branch_id)
        sales_list = session.exec(sales_query).all()
        for sale in sales_list:
            sales_map[sale.id] = sale
            if sale.client:
                clients_map[sale.id] = sale.client

    today = _report_now(generated_at, country_code, timezone)
    today_str = today.strftime("%d/%m/%Y")

    # Clasificar por antigüedad
    aging_buckets = {
        "current": {"label": MSG.REPORT_AGING_LABELS["current"], "days": "0", "amount": Decimal("0"), "count": 0},
        "0-30": {"label": MSG.REPORT_AGING_LABELS["1_30"], "days": "1-30", "amount": Decimal("0"), "count": 0},
        "31-60": {"label": MSG.REPORT_AGING_LABELS["31_60"], "days": "31-60", "amount": Decimal("0"), "count": 0},
        "61-90": {"label": MSG.REPORT_AGING_LABELS["61_90"], "days": "61-90", "amount": Decimal("0"), "count": 0},
        "90+": {"label": MSG.REPORT_AGING_LABELS["90_plus"], "days": ">90", "amount": Decimal("0"), "count": 0},
    }

    # Provisiones sugeridas
    provision_rates = {
        "current": Decimal("0"),
        "0-30": Decimal("0.05"),  # 5%
        "31-60": Decimal("0.10"),  # 10%
        "61-90": Decimal("0.25"),  # 25%
        "90+": Decimal("0.50"),   # 50%
    }

    by_client: dict[str, dict] = {}
    installments_data = []

    for installment in installments:
        amount = Decimal(str(installment.amount or 0))
        paid = Decimal(str(installment.paid_amount or 0))
        pending = amount - paid

        if pending <= 0:
            continue

        sale = sales_map.get(installment.sale_id)
        client = clients_map.get(installment.sale_id)
        client_name = client.name if client else "Sin cliente"
        due_date = installment.due_date

        # Calcular antigüedad - normalizar a date
        if due_date:
            localized_due_date = to_local_datetime(
                due_date,
                country_code,
                timezone=timezone,
            )
            if localized_due_date and hasattr(localized_due_date, "date"):
                due_date_normalized = localized_due_date.date()
            elif hasattr(due_date, "date"):
                due_date_normalized = due_date.date()
            else:
                due_date_normalized = due_date
            days_overdue = (today.date() - due_date_normalized).days
        else:
            days_overdue = 0

        if days_overdue <= 0:
            bucket = "current"
        elif days_overdue <= 30:
            bucket = "0-30"
        elif days_overdue <= 60:
            bucket = "31-60"
        elif days_overdue <= 90:
            bucket = "61-90"
        else:
            bucket = "90+"

        aging_buckets[bucket]["amount"] += pending
        aging_buckets[bucket]["count"] += 1

        # Por cliente
        if client_name not in by_client:
            by_client[client_name] = {
                "current": Decimal("0"),
                "0-30": Decimal("0"),
                "31-60": Decimal("0"),
                "61-90": Decimal("0"),
                "90+": Decimal("0"),
                "total": Decimal("0"),
            }
        by_client[client_name][bucket] += pending
        by_client[client_name]["total"] += pending

        installments_data.append({
            "client": client_name,
            "sale_id": sale.id if sale else 0,
            "installment_num": installment.number,
            "due_date": _format_report_datetime(
                due_date,
                "%d/%m/%Y",
                country_code,
                timezone,
            ) if due_date else "",
            "days_overdue": max(0, days_overdue),
            "amount": amount,
            "paid": paid,
            "pending": pending,
            "bucket": bucket,
        })

    total_receivables = sum(b["amount"] for b in aging_buckets.values())
    total_provision = sum(
        aging_buckets[k]["amount"] * provision_rates[k]
        for k in aging_buckets
    )

    # =================
    # HOJA 1: RESUMEN (mejorado)
    # =================
    ws_summary = wb.active
    ws_summary.title = MSG.REPORT_PORTFOLIO_SHEET

    row = _add_company_header(
        ws_summary,
        company_name,
        "ANÁLISIS DE CUENTAS POR COBRAR",
        f"Al {today_str}",
        columns=6,
        **header_kwargs,
    )

    row += 1
    ws_summary.cell(row=row, column=1, value="RESUMEN DE CARTERA DE CRÉDITOS").font = SUBTITLE_FONT
    row += 1

    summary = [
        ("Total Cuentas por Cobrar:", _format_currency(total_receivables, currency_symbol)),
        ("Número de Cuotas Pendientes:", sum(b["count"] for b in aging_buckets.values())),
        ("Clientes con Deuda Activa:", len(by_client)),
        ("", ""),
        ("Provisión Sugerida (Cobranza Dudosa):", _format_currency(total_provision, currency_symbol)),
        ("Cartera Neta Estimada:", _format_currency(total_receivables - total_provision, currency_symbol)),
    ]

    for label, value in summary:
        ws_summary.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws_summary.cell(row=row, column=2, value=value)
        row += 1

    row += 2
    ws_summary.cell(row=row, column=1, value="ANTIGÜEDAD DE CARTERA (Análisis de Vencimiento)").font = SUBTITLE_FONT
    row += 1

    headers = [
        "Período de Vencimiento",
        "Nº Cuotas",
        f"Monto Pendiente ({currency_label})",
        "Participación (%)",
        "Tasa Provisión",
        f"Provisión ({currency_label})",
    ]
    _style_header_row(ws_summary, row, headers)
    aging_data_start = row + 1
    row += 1

    for bucket_key in ["current", "0-30", "31-60", "61-90", "90+"]:
        bucket = aging_buckets[bucket_key]
        prov_rate = provision_rates[bucket_key]

        ws_summary.cell(row=row, column=1, value=bucket["label"])
        ws_summary.cell(row=row, column=2, value=bucket["count"])
        ws_summary.cell(row=row, column=3, value=bucket["amount"]).number_format = currency_format
        # Participación - temporal
        ws_summary.cell(row=row, column=4, value=bucket["amount"]).number_format = currency_format
        ws_summary.cell(row=row, column=5, value=prov_rate).number_format = PERCENT_FORMAT
        # Provisión = Fórmula: Monto × Tasa
        ws_summary.cell(row=row, column=6, value=f"=C{row}*E{row}").number_format = currency_format

        for col in range(1, 7):
            ws_summary.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    aging_totals_row = row
    _add_totals_row_with_formulas(ws_summary, aging_totals_row, aging_data_start, [
        {"type": "label", "value": "TOTAL CARTERA"},
        {"type": "sum", "col_letter": "B"},
        {"type": "sum", "col_letter": "C", "number_format": currency_format},
        {"type": "text", "value": "100.00%"},
        {"type": "text", "value": ""},
        {"type": "sum", "col_letter": "F", "number_format": currency_format},
    ])

    # Actualizar participación con fórmulas
    for r in range(aging_data_start, aging_totals_row):
        ws_summary.cell(row=r, column=4, value=f"=IF($C${aging_totals_row}>0,C{r}/$C${aging_totals_row},0)").number_format = PERCENT_FORMAT

    _add_notes_section(ws_summary, aging_totals_row, [
        "Vigente: Cuotas aún no vencidas (fecha de pago futura).",
        "Provisión: Reserva estimada para deudas de difícil cobro.",
        "Tasa de Provisión: 0% vigente, 5% (1-30d), 10% (31-60d), 25% (61-90d), 50% (>90d).",
        "Cartera Neta = Total por Cobrar - Provisión.",
    ], columns=6)

    _auto_adjust_columns(ws_summary)

    # =================
    # HOJA 2: POR CLIENTE (con fórmulas)
    # =================
    ws_client = wb.create_sheet(MSG.REPORT_BY_CLIENT_SHEET)
    row = _add_company_header(
        ws_client,
        company_name,
        "DEUDA DETALLADA POR CLIENTE",
        f"Al {today_str}",
        columns=7,
        **header_kwargs,
    )

    headers = [
        "Cliente",
        f"Vigente ({currency_label})",
        f"1-30 días ({currency_label})",
        f"31-60 días ({currency_label})",
        f"61-90 días ({currency_label})",
        f">90 días ({currency_label})",
        f"Total Deuda ({currency_label})",
    ]
    _style_header_row(ws_client, row, headers)
    client_data_start = row + 1
    row += 1

    sorted_clients = sorted(by_client.items(), key=lambda x: x[1]["total"], reverse=True)

    for client_name, client_data in sorted_clients:
        ws_client.cell(row=row, column=1, value=_safe_string(client_name))
        ws_client.cell(row=row, column=2, value=client_data["current"]).number_format = currency_format
        ws_client.cell(row=row, column=3, value=client_data["0-30"]).number_format = currency_format
        ws_client.cell(row=row, column=4, value=client_data["31-60"]).number_format = currency_format
        ws_client.cell(row=row, column=5, value=client_data["61-90"]).number_format = currency_format
        ws_client.cell(row=row, column=6, value=client_data["90+"]).number_format = currency_format
        # Total = Fórmula: Suma de columnas B a F
        ws_client.cell(row=row, column=7, value=f"=SUM(B{row}:F{row})").number_format = currency_format

        for col in range(1, 8):
            ws_client.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    client_totals_row = row
    _add_totals_row_with_formulas(ws_client, client_totals_row, client_data_start, [
        {"type": "label", "value": "TOTAL CLIENTES"},
        {"type": "sum", "col_letter": "B", "number_format": currency_format},
        {"type": "sum", "col_letter": "C", "number_format": currency_format},
        {"type": "sum", "col_letter": "D", "number_format": currency_format},
        {"type": "sum", "col_letter": "E", "number_format": currency_format},
        {"type": "sum", "col_letter": "F", "number_format": currency_format},
        {"type": "sum", "col_letter": "G", "number_format": currency_format},
    ])

    _add_notes_section(ws_client, client_totals_row, [
        "Los clientes están ordenados por monto total de deuda (mayor a menor).",
        "Total Deuda = Suma de todas las cuotas pendientes del cliente.",
        "Priorizar cobranza de clientes con deuda vencida (columnas 31-60, 61-90, >90 días).",
    ], columns=7)

    _auto_adjust_columns(ws_client)

    # =================
    # HOJA 3: DETALLE (mejorado)
    # =================
    ws_detail = wb.create_sheet(MSG.REPORT_INSTALLMENTS_SHEET)
    row = _add_company_header(
        ws_detail,
        company_name,
        "LISTADO DETALLADO DE CUOTAS PENDIENTES",
        f"Al {today_str}",
        columns=9,
        **header_kwargs,
    )

    headers = [
        "Cliente",
        "Nº Venta",
        "Nº Cuota",
        "Fecha Vencimiento",
        "Días Vencido",
        f"Monto Cuota ({currency_label})",
        f"Abonado ({currency_label})",
        f"Pendiente ({currency_label})",
        "Estado",
    ]
    _style_header_row(ws_detail, row, headers)
    cuota_data_start = row + 1
    row += 1

    # Ordenar por días vencido (mayor primero)
    installments_data.sort(key=lambda x: x["days_overdue"], reverse=True)

    for inst in installments_data:
        ws_detail.cell(row=row, column=1, value=_safe_string(inst["client"]))
        ws_detail.cell(row=row, column=2, value=inst["sale_id"])
        ws_detail.cell(row=row, column=3, value=inst["installment_num"])
        ws_detail.cell(row=row, column=4, value=inst["due_date"])
        ws_detail.cell(row=row, column=5, value=inst["days_overdue"])
        ws_detail.cell(row=row, column=6, value=inst["amount"]).number_format = currency_format
        ws_detail.cell(row=row, column=7, value=inst["paid"]).number_format = currency_format
        # Pendiente = Fórmula: Monto - Abonado
        ws_detail.cell(row=row, column=8, value=f"=F{row}-G{row}").number_format = currency_format
        ws_detail.cell(row=row, column=9, value=aging_buckets[inst["bucket"]]["label"])

        # Color según antigüedad
        if inst["bucket"] == "90+":
            fill = NEGATIVE_FILL
        elif inst["bucket"] == "61-90":
            fill = PatternFill(start_color="FDBA74", end_color="FDBA74", fill_type="solid")
        elif inst["bucket"] == "31-60":
            fill = WARNING_FILL
        elif inst["bucket"] == "0-30":
            fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
        else:
            fill = POSITIVE_FILL

        for col in range(1, 10):
            cell = ws_detail.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.fill = fill

        row += 1

    cuota_totals_row = row
    _add_totals_row_with_formulas(ws_detail, cuota_totals_row, cuota_data_start, [
        {"type": "label", "value": "TOTALES"},
        {"type": "text", "value": ""},
        {"type": "text", "value": ""},
        {"type": "text", "value": ""},
        {"type": "text", "value": ""},
        {"type": "sum", "col_letter": "F", "number_format": currency_format},
        {"type": "sum", "col_letter": "G", "number_format": currency_format},
        {"type": "sum", "col_letter": "H", "number_format": currency_format},
        {"type": "text", "value": ""},
    ])

    _add_notes_section(ws_detail, cuota_totals_row, [
        "Las cuotas están ordenadas por días de vencimiento (más antiguas primero).",
        "🔴 Rojo: Más de 90 días vencido - Riesgo alto de incobrabilidad.",
        "🟠 Naranja: 61-90 días - Requiere gestión de cobranza urgente.",
        "🟡 Amarillo: 31-60 días - Seguimiento prioritario.",
        "🟢 Verde: Vigente - Sin vencimiento.",
        "Pendiente = Monto de la Cuota - Monto Abonado.",
    ], columns=9)

    _auto_adjust_columns(ws_detail)

    # Guardar
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# =============================================================================
# REPORTE DE CAJA CONSOLIDADO
# =============================================================================

# Acciones de cobranzas que representan ingresos de cuotas/créditos y servicios
# IMPORTANTE: Mantener sincronizado con:
#   - credit_service.py: pay_installment() usa "Cobranza"
#   - sale_service.py: process_sale() usa "Inicial Credito"
#   - services_state.py: apply_reservation_payment/pay_reservation_with_payment_method
#                        usan "Adelanto" y "Reserva"
CASHBOX_COLLECTION_ACTIONS = {
    # Cobranza de cuotas de crédito - SOLO estos porque NO crean SalePayment
    # credit_service.py: pay_installment() crea CashboxLog pero NO SalePayment
    # Por eso se deben sumar aparte de las ventas
    "Cobranza",
    "Cobro de Cuota",
    "Pago Cuota",
    "Cobro Cuota",
    "Ingreso Cuota",
    "Amortizacion",
    "Pago Credito",
    # NOTA: NO incluir "Inicial Credito", "Adelanto" ni "Reserva" porque
    # esos flujos SÍ crean SalePayment y ya se cuentan en Sale.payments
}


def _normalize_payment_method(method_label: str) -> str:
    """Normaliza el método de pago para agrupar correctamente."""
    import unicodedata
    raw = (method_label or "").strip().lower()
    if not raw:
        return "other"
    normalized = unicodedata.normalize("NFKD", raw)
    normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))

    if "mixto" in normalized or "mixed" in normalized:
        return "mixed"
    if "yape" in normalized:
        return "yape"
    if "plin" in normalized:
        return "plin"
    if "transfer" in normalized or "banco" in normalized:
        return "transfer"
    if "debito" in normalized or "debit" in normalized:
        return "debit"
    if "credito" in normalized or "credit" in normalized or "tarjeta" in normalized:
        return "credit"
    if "efectivo" in normalized or normalized == "cash":
        return "cash"
    return "other"


def generate_cashbox_report(
    session,
    start_date: datetime,
    end_date: datetime,
    company_name: str = "TUWAYKIAPP",
    currency_symbol: str | None = None,
    company_id: int | None = None,
    branch_id: int | None = None,
    country_code: str | None = None,
    timezone: str | None = None,
    generated_at: datetime | None = None,
) -> io.BytesIO:
    """
    Genera reporte de caja consolidado con flujo de caja REAL.

    Incluye:
    - Resumen de movimientos
    - Detalle de aperturas y cierres
    - Ingresos por método de pago (Ventas + Cobranzas de cuotas)
    - Diferencias detectadas
    - Desglose de ingresos por origen (Ventas vs Cobranzas)
    """
    set_tenant_context(company_id, branch_id)
    wb = Workbook()
    currency_label = _currency_label(currency_symbol)
    currency_format = _currency_format(currency_symbol)
    header_kwargs = {
        "generated_at": generated_at,
        "country_code": country_code,
        "timezone": timezone,
    }

    period_str = (
        f"{_format_report_datetime(start_date, '%d/%m/%Y', country_code, timezone)} - "
        f"{_format_report_datetime(end_date, '%d/%m/%Y', country_code, timezone)}"
    )

    # Consultar logs de caja
    query = (
        select(CashboxLog)
        .where(
            and_(
                CashboxLog.timestamp >= start_date,
                CashboxLog.timestamp <= end_date,
            )
        )
        .order_by(CashboxLog.timestamp.desc())
    )
    if company_id:
        query = query.where(CashboxLog.company_id == company_id)
    if branch_id:
        query = query.where(CashboxLog.branch_id == branch_id)

    logs = session.exec(query).all()
    log_user_lookup = _build_cashbox_log_user_lookup(session, logs, company_id)

    # Consultar ventas del período
    sales_query = (
        select(Sale)
        .where(
            and_(
                Sale.timestamp >= start_date,
                Sale.timestamp <= end_date,
                Sale.status != SaleStatus.cancelled,
            )
        )
        .options(selectinload(Sale.payments))
    )
    if company_id:
        sales_query = sales_query.where(Sale.company_id == company_id)
    if branch_id:
        sales_query = sales_query.where(Sale.branch_id == branch_id)

    sales = session.exec(sales_query).all()

    # =========================================================================
    # NUEVO: Consultar cobros de cuotas/cobranzas para flujo de caja completo
    # =========================================================================
    collection_logs = [
        log for log in logs
        if log.action in CASHBOX_COLLECTION_ACTIONS and not log.is_voided
    ]

    # Calcular métricas usando funciones seguras
    total_openings = sum(1 for log in logs if log.action == "apertura")
    total_closings = sum(1 for log in logs if log.action == "cierre")
    total_opening_amount = sum(_safe_decimal(log.amount) for log in logs if log.action == "apertura")
    total_closing_amount = sum(_safe_decimal(log.amount) for log in logs if log.action == "cierre")

    total_sales = sum(_safe_decimal(s.total_amount) for s in sales)
    total_collections = sum(_safe_decimal(log.amount) for log in collection_logs)

    # Ingresos por método de pago - VENTAS
    by_payment_sales: dict[str, Decimal] = {}
    by_payment_sales_count: dict[str, int] = {}
    for sale in sales:
        for payment in (sale.payments or []):
            method = payment.method_type.value if payment.method_type else "No especificado"
            amount = _safe_decimal(payment.amount)
            by_payment_sales[method] = by_payment_sales.get(method, Decimal("0")) + amount
            by_payment_sales_count[method] = by_payment_sales_count.get(method, 0) + 1

    # Ingresos por método de pago - COBRANZAS (cuotas de crédito)
    by_payment_collections: dict[str, Decimal] = {}
    by_payment_collections_count: dict[str, int] = {}
    for log in collection_logs:
        method_raw = getattr(log, "payment_method", "") or "cash"
        method = _normalize_payment_method(method_raw)
        method_es = _translate_payment_method(method)
        amount = _safe_decimal(log.amount)
        by_payment_collections[method_es] = by_payment_collections.get(method_es, Decimal("0")) + amount
        by_payment_collections_count[method_es] = by_payment_collections_count.get(method_es, 0) + 1

    # Consolidar todo el flujo de caja (Ventas + Cobranzas)
    by_payment: dict[str, Decimal] = {}
    by_payment_count: dict[str, int] = {}

    # Agregar ventas
    for method, amount in by_payment_sales.items():
        method_es = _translate_payment_method(method)
        by_payment[method_es] = by_payment.get(method_es, Decimal("0")) + amount
        by_payment_count[method_es] = by_payment_count.get(method_es, 0) + by_payment_sales_count.get(method, 0)

    # Agregar cobranzas
    for method_es, amount in by_payment_collections.items():
        by_payment[method_es] = by_payment.get(method_es, Decimal("0")) + amount
        by_payment_count[method_es] = by_payment_count.get(method_es, 0) + by_payment_collections_count.get(method_es, 0)

    # =================
    # HOJA 1: RESUMEN (mejorado)
    # =================
    ws_summary = wb.active
    ws_summary.title = MSG.REPORT_CASH_SHEET

    row = _add_company_header(
        ws_summary,
        company_name,
        "REPORTE CONSOLIDADO DE CAJA",
        period_str,
        columns=4,
        **header_kwargs,
    )

    row += 1
    ws_summary.cell(row=row, column=1, value="MOVIMIENTOS DE CAJA EN EL PERÍODO").font = SUBTITLE_FONT
    row += 1

    total_cash_flow = total_sales + total_collections

    summary = [
        ("Número de Aperturas de Caja:", total_openings),
        ("Número de Cierres de Caja:", total_closings),
        ("", ""),
        ("Total Monto en Aperturas:", _format_currency(total_opening_amount, currency_symbol)),
        ("Total Monto en Cierres:", _format_currency(total_closing_amount, currency_symbol)),
        ("", ""),
        ("FLUJO DE CAJA - INGRESOS:", ""),
        ("  • Ventas del Período:", _format_currency(total_sales, currency_symbol)),
        ("  • Cobros de Cuotas/Créditos:", _format_currency(total_collections, currency_symbol)),
        ("  • TOTAL INGRESOS:", _format_currency(total_cash_flow, currency_symbol)),
        ("", ""),
        ("Número de Transacciones de Venta:", len(sales)),
        ("Número de Cobros de Cuotas:", len(collection_logs)),
    ]

    for label, value in summary:
        cell = ws_summary.cell(row=row, column=1, value=label)
        if "TOTAL INGRESOS" in label:
            cell.font = Font(bold=True, color="4F46E5")
        elif label.startswith("FLUJO"):
            cell.font = SUBTITLE_FONT
        else:
            cell.font = Font(bold=True)
        ws_summary.cell(row=row, column=2, value=value)
        row += 1

    row += 2
    ws_summary.cell(row=row, column=1, value="RECAUDACIÓN TOTAL POR MÉTODO DE PAGO").font = SUBTITLE_FONT
    ws_summary.cell(row=row, column=2, value="(Ventas + Cobros de Cuotas)").font = Font(italic=True, size=9, color="6B7280")
    row += 1

    headers = [
        "Método de Pago",
        "Nº Operaciones",
        f"Monto Recaudado ({currency_label})",
        "Participación (%)",
    ]
    _style_header_row(ws_summary, row, headers)
    caja_pay_start = row + 1
    row += 1

    sorted_payments = sorted(by_payment.items(), key=lambda x: x[1], reverse=True)
    total_payments = sum(by_payment.values())

    for method_es, amount in sorted_payments:
        count = by_payment_count.get(method_es, 0)

        ws_summary.cell(row=row, column=1, value=_safe_string(method_es))
        ws_summary.cell(row=row, column=2, value=count)
        ws_summary.cell(row=row, column=3, value=amount).number_format = currency_format
        # Participación - temporal, se reemplazará con fórmula
        ws_summary.cell(row=row, column=4, value=amount).number_format = currency_format

        for col in range(1, 5):
            ws_summary.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    caja_pay_totals = row
    _add_totals_row_with_formulas(ws_summary, caja_pay_totals, caja_pay_start, [
        {"type": "label", "value": "TOTAL RECAUDADO"},
        {"type": "sum", "col_letter": "B"},
        {"type": "sum", "col_letter": "C", "number_format": currency_format},
        {"type": "text", "value": "100.00%"},
    ])

    # Actualizar participación con fórmulas
    for r in range(caja_pay_start, caja_pay_totals):
        ws_summary.cell(row=r, column=4, value=f"=IF($C${caja_pay_totals}>0,C{r}/$C${caja_pay_totals},0)").number_format = PERCENT_FORMAT

    _add_notes_section(ws_summary, caja_pay_totals, [
        "Ingresos por Ventas: cobros de operaciones comerciales del período.",
        "Cobros de Cuotas: recuperaciones de ventas a crédito de períodos actuales o previos.",
        "Total Ingresos = Ventas + Cobros de Cuotas (dinero real ingresado en caja).",
        "Participación %: peso de cada método de pago sobre el total recaudado.",
    ], columns=4)

    _auto_adjust_columns(ws_summary)

    # =================
    # HOJA 2: MOVIMIENTOS DE CAJA
    # =================
    ws_logs = wb.create_sheet(MSG.REPORT_CASH_MOVES_SHEET)
    row = _add_company_header(
        ws_logs,
        company_name,
        "BITÁCORA DE MOVIMIENTOS DE CAJA",
        period_str,
        columns=7,
        **header_kwargs,
    )

    headers = [
        "Fecha/Hora",
        "Tipo Movimiento",
        "Responsable",
        "Naturaleza",
        f"Monto ({currency_label})",
        "Método de Pago",
        "Referencia / Notas",
    ]
    _style_header_row(ws_logs, row, headers)
    row += 1

    for log in logs:
        user_id = _safe_int(getattr(log, "user_id", None))
        user_name = _safe_string(log_user_lookup.get(user_id), "Sistema")
        action_label = _translate_cashbox_action(getattr(log, "action", ""))
        nature = _cashbox_action_nature(getattr(log, "action", ""))
        method_raw = _safe_string(getattr(log, "payment_method", ""))
        method_label = (
            _translate_payment_method(_normalize_payment_method(method_raw))
            if method_raw
            else "No especificado"
        )

        ws_logs.cell(
            row=row,
            column=1,
            value=(
                _format_report_datetime(
                    log.timestamp,
                    "%d/%m/%Y %H:%M:%S",
                    country_code,
                    timezone,
                )
                if log.timestamp
                else ""
            ),
        )
        ws_logs.cell(row=row, column=2, value=_safe_string(action_label))
        ws_logs.cell(row=row, column=3, value=user_name)
        ws_logs.cell(row=row, column=4, value=nature)
        ws_logs.cell(row=row, column=5, value=_safe_decimal(log.amount)).number_format = currency_format
        ws_logs.cell(row=row, column=6, value=_safe_string(method_label))
        notes_str = _safe_string(log.notes, "Sin observaciones")
        if "#" in notes_str and ", " in notes_str:
            notes_str = notes_str.replace(", ", "\n")
        notes_cell = ws_logs.cell(row=row, column=7, value=notes_str)
        notes_cell.alignment = Alignment(wrap_text=True, vertical="top")

        nature_cell = ws_logs.cell(row=row, column=4)
        if nature == "Ingreso":
            nature_cell.fill = POSITIVE_FILL
        elif nature == "Egreso":
            nature_cell.fill = NEGATIVE_FILL
        else:
            nature_cell.fill = WARNING_FILL

        for col in range(1, 8):
            ws_logs.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    _add_notes_section(ws_logs, row, [
        "Naturaleza Ingreso: dinero que entra a caja (ventas, reservas, cobranzas, adelantos).",
        "Naturaleza Egreso: dinero que sale de caja (gastos, devoluciones, ajustes).",
        "Apertura: fondo inicial al inicio de jornada; Cierre: arqueo final del día.",
        "Responsable: usuario que registró el movimiento en el sistema.",
    ], columns=7)

    ws_logs.column_dimensions["G"].width = 65
    _auto_adjust_columns(ws_logs)

    # =================
    # HOJA 3: DESGLOSE POR ORIGEN (Ventas vs Cobranzas)
    # =================
    ws_origin = wb.create_sheet("Ingresos por Origen")
    row = _add_company_header(
        ws_origin,
        company_name,
        "DESGLOSE DE INGRESOS POR ORIGEN",
        period_str,
        columns=5,
        **header_kwargs,
    )

    row += 1
    ws_origin.cell(row=row, column=1, value="INGRESOS POR VENTAS DIRECTAS").font = SUBTITLE_FONT
    row += 1

    headers = [
        "Método de Pago",
        "Nº Operaciones",
        f"Monto ({currency_label})",
        "% del Total Ventas",
        "Observación",
    ]
    _style_header_row(ws_origin, row, headers)
    ventas_start = row + 1
    row += 1

    sorted_sales = sorted(by_payment_sales.items(), key=lambda x: x[1], reverse=True)
    for method, amount in sorted_sales:
        method_es = _translate_payment_method(method)
        count = by_payment_sales_count.get(method, 0)

        ws_origin.cell(row=row, column=1, value=_safe_string(method_es))
        ws_origin.cell(row=row, column=2, value=count)
        ws_origin.cell(row=row, column=3, value=amount).number_format = currency_format
        ws_origin.cell(row=row, column=4, value=amount).number_format = currency_format  # Temporal
        ws_origin.cell(row=row, column=5, value="Pago de venta directa")

        for col in range(1, 6):
            ws_origin.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    ventas_totals = row
    _add_totals_row_with_formulas(ws_origin, ventas_totals, ventas_start, [
        {"type": "label", "value": "SUBTOTAL VENTAS"},
        {"type": "sum", "col_letter": "B"},
        {"type": "sum", "col_letter": "C", "number_format": currency_format},
        {"type": "text", "value": "100.00%"},
        {"type": "text", "value": ""},
    ])

    # Participación ventas
    for r in range(ventas_start, ventas_totals):
        ws_origin.cell(row=r, column=4, value=f"=IF($C${ventas_totals}>0,C{r}/$C${ventas_totals},0)").number_format = PERCENT_FORMAT

    row += 3
    ws_origin.cell(row=row, column=1, value="INGRESOS POR COBROS DE CUOTAS/CRÉDITOS").font = SUBTITLE_FONT
    row += 1

    headers = [
        "Método de Pago",
        "Nº Operaciones",
        f"Monto ({currency_label})",
        "% del Total Cobros",
        "Observación",
    ]
    _style_header_row(ws_origin, row, headers)
    cobros_start = row + 1
    row += 1

    if by_payment_collections:
        sorted_collections = sorted(by_payment_collections.items(), key=lambda x: x[1], reverse=True)
        for method_es, amount in sorted_collections:
            count = by_payment_collections_count.get(method_es, 0)

            ws_origin.cell(row=row, column=1, value=_safe_string(method_es))
            ws_origin.cell(row=row, column=2, value=count)
            ws_origin.cell(row=row, column=3, value=amount).number_format = currency_format
            ws_origin.cell(row=row, column=4, value=amount).number_format = currency_format  # Temporal
            ws_origin.cell(row=row, column=5, value="Pago de cuota de crédito")

            for col in range(1, 6):
                ws_origin.cell(row=row, column=col).border = THIN_BORDER
            row += 1
    else:
        ws_origin.cell(row=row, column=1, value="Sin cobros de cuotas en el período")
        ws_origin.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws_origin.cell(row=row, column=1).font = Font(italic=True, color="6B7280")
        row += 1

    cobros_totals = row
    _add_totals_row_with_formulas(ws_origin, cobros_totals, cobros_start, [
        {"type": "label", "value": "SUBTOTAL COBROS"},
        {"type": "sum", "col_letter": "B"},
        {"type": "sum", "col_letter": "C", "number_format": currency_format},
        {"type": "text", "value": "100.00%"},
        {"type": "text", "value": ""},
    ])

    # Participación cobros
    for r in range(cobros_start, cobros_totals):
        if ws_origin.cell(row=r, column=2).value is not None and isinstance(ws_origin.cell(row=r, column=2).value, int):
            ws_origin.cell(row=r, column=4, value=f"=IF($C${cobros_totals}>0,C{r}/$C${cobros_totals},0)").number_format = PERCENT_FORMAT

    row += 3
    ws_origin.cell(row=row, column=1, value="RESUMEN CONSOLIDADO").font = SUBTITLE_FONT
    row += 2

    ws_origin.cell(row=row, column=1, value="Total Ingresos por Ventas:").font = Font(bold=True)
    ws_origin.cell(row=row, column=2, value=f"=C{ventas_totals}").number_format = currency_format
    row += 1

    ws_origin.cell(row=row, column=1, value="Total Ingresos por Cobros:").font = Font(bold=True)
    ws_origin.cell(row=row, column=2, value=f"=C{cobros_totals}").number_format = currency_format
    row += 1

    ws_origin.cell(row=row, column=1, value="TOTAL FLUJO DE CAJA:").font = Font(bold=True, color="4F46E5", size=11)
    ws_origin.cell(row=row, column=2, value=f"=C{ventas_totals}+C{cobros_totals}").number_format = currency_format
    ws_origin.cell(row=row, column=2).font = Font(bold=True, size=11)
    ws_origin.cell(row=row, column=2).fill = POSITIVE_FILL

    _add_notes_section(ws_origin, row, [
        "Ventas Directas: Ingresos por ventas del período (contado + inicial de créditos).",
        "Cobros de Cuotas: Ingresos por pagos de cuotas de ventas a crédito anteriores.",
        "Flujo de Caja Total: Suma de ambos conceptos = dinero real ingresado.",
        "Este desglose es esencial para la conciliación contable y fiscal.",
    ], columns=5)

    _auto_adjust_columns(ws_origin)

    # =================
    # HOJA 4: VENTAS POR DÍA (con fórmulas)
    # =================
    ws_daily = wb.create_sheet(MSG.REPORT_DAILY_SALES_SHEET)
    row = _add_company_header(
        ws_daily,
        company_name,
        "RESUMEN DE VENTAS DIARIAS",
        period_str,
        columns=5,
        **header_kwargs,
    )

    # Agrupar ventas por día
    by_day: dict[str, dict] = {}
    for sale in sales:
        if sale.timestamp:
            day_key = _format_report_datetime(
                sale.timestamp,
                "%Y-%m-%d",
                country_code,
                timezone,
            )
            if day_key not in by_day:
                by_day[day_key] = {"count": 0, "total": Decimal("0"), "cash": Decimal("0"), "other": Decimal("0")}
            by_day[day_key]["count"] += 1
            by_day[day_key]["total"] += _safe_decimal(sale.total_amount)

            for payment in (sale.payments or []):
                amount = _safe_decimal(payment.amount)
                method = (payment.method_type.value if payment.method_type else "").lower()
                if method in {"efectivo", "cash"}:
                    by_day[day_key]["cash"] += amount
                else:
                    by_day[day_key]["other"] += amount

    headers = [
        "Fecha",
        "Nº Transacciones",
        f"Efectivo ({currency_label})",
        f"Otros Medios ({currency_label})",
        f"Total del Día ({currency_label})",
    ]
    _style_header_row(ws_daily, row, headers)
    caja_daily_start = row + 1
    row += 1

    for day_key in sorted(by_day.keys()):
        day_data = by_day[day_key]

        ws_daily.cell(row=row, column=1, value=_safe_string(day_key))
        ws_daily.cell(row=row, column=2, value=day_data["count"])
        ws_daily.cell(row=row, column=3, value=day_data["cash"]).number_format = currency_format
        ws_daily.cell(row=row, column=4, value=day_data["other"]).number_format = currency_format
        # Total = Fórmula: Efectivo + Otros
        ws_daily.cell(row=row, column=5, value=f"=C{row}+D{row}").number_format = currency_format

        for col in range(1, 6):
            ws_daily.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    caja_daily_totals = row
    _add_totals_row_with_formulas(ws_daily, caja_daily_totals, caja_daily_start, [
        {"type": "label", "value": "TOTALES"},
        {"type": "sum", "col_letter": "B"},
        {"type": "sum", "col_letter": "C", "number_format": currency_format},
        {"type": "sum", "col_letter": "D", "number_format": currency_format},
        {"type": "sum", "col_letter": "E", "number_format": currency_format},
    ])

    _add_notes_section(ws_daily, caja_daily_totals, [
        "Efectivo: Pagos recibidos en billetes y monedas.",
        "Otros Medios: Tarjetas, transferencias, Yape, Plin, etc.",
        "Total del Día = Efectivo + Otros Medios.",
    ], columns=5)

    _auto_adjust_columns(ws_daily)

    # =================
    # HOJA 5: CONCILIACIÓN DE CAJA (con fórmulas)
    # =================
    ws_conciliation = wb.create_sheet("Conciliación")
    row = _add_company_header(
        ws_conciliation,
        company_name,
        "CUADRE Y CONCILIACIÓN DE CAJA",
        period_str,
        columns=4,
        **header_kwargs,
    )

    row += 1
    ws_conciliation.cell(row=row, column=1, value="ANÁLISIS DE FLUJO DE EFECTIVO").font = SUBTITLE_FONT
    row += 2

    # Calcular efectivo esperado (ventas + cobros de cuotas en efectivo)
    cash_from_sales = sum(by_day[d]["cash"] for d in by_day)

    # Cobros en efectivo de cuotas
    cash_from_collections = Decimal("0")
    for log in collection_logs:
        method_raw = getattr(log, "payment_method", "") or "cash"
        method_key = _normalize_payment_method(method_raw)
        if method_key == "cash":
            cash_from_collections += _safe_decimal(log.amount)

    # Escribir datos con referencias para fórmulas
    data_start_row = row
    ws_conciliation.cell(row=row, column=1, value="(+) Monto Inicial (Aperturas):").font = Font(bold=True)
    ws_conciliation.cell(row=row, column=2, value=total_opening_amount).number_format = currency_format
    ws_conciliation.cell(row=row, column=3, value="Suma de todas las aperturas de caja")
    row += 1

    ws_conciliation.cell(row=row, column=1, value="(+) Efectivo por Ventas:").font = Font(bold=True)
    ws_conciliation.cell(row=row, column=2, value=cash_from_sales).number_format = currency_format
    ws_conciliation.cell(row=row, column=3, value="Ventas cobradas en efectivo")
    sales_cash_row = row
    row += 1

    ws_conciliation.cell(row=row, column=1, value="(+) Efectivo por Cobros de Cuotas:").font = Font(bold=True)
    ws_conciliation.cell(row=row, column=2, value=cash_from_collections).number_format = currency_format
    ws_conciliation.cell(row=row, column=3, value="Cuotas de créditos cobradas en efectivo")
    collections_cash_row = row
    row += 1

    ws_conciliation.cell(row=row, column=1, value="(=) Efectivo Esperado al Cierre:").font = Font(bold=True, color="4F46E5")
    ws_conciliation.cell(row=row, column=2, value=f"=B{data_start_row}+B{sales_cash_row}+B{collections_cash_row}").number_format = currency_format
    ws_conciliation.cell(row=row, column=3, value="Monto que debería haber en caja")
    expected_row = row
    row += 2

    ws_conciliation.cell(row=row, column=1, value="(-) Monto Real (Cierres):").font = Font(bold=True)
    ws_conciliation.cell(row=row, column=2, value=total_closing_amount).number_format = currency_format
    ws_conciliation.cell(row=row, column=3, value="Suma de todos los cierres de caja")
    actual_row = row
    row += 2

    # Diferencia con fórmula
    ws_conciliation.cell(row=row, column=1, value="DIFERENCIA (Real - Esperado):").font = Font(bold=True, size=11)
    diff_cell = ws_conciliation.cell(row=row, column=2, value=f"=B{actual_row}-B{expected_row}")
    diff_cell.number_format = currency_format
    diff_cell.font = Font(bold=True, size=11)

    # Resultado interpretativo
    row += 1
    ws_conciliation.cell(row=row, column=1, value="Interpretación:").font = Font(bold=True)
    expected_total = total_opening_amount + cash_from_sales + cash_from_collections
    actual_difference = Decimal(str(total_closing_amount)) - Decimal(str(expected_total))

    if actual_difference > 0:
        ws_conciliation.cell(row=row, column=2, value="SOBRANTE en caja").font = Font(bold=True, color="16A34A")
        ws_conciliation.cell(row=row, column=2).fill = POSITIVE_FILL
        ws_conciliation.cell(row=row, column=3, value="Hay más dinero del esperado. Verificar si hubo ingresos no registrados.")
    elif actual_difference < 0:
        ws_conciliation.cell(row=row, column=2, value="FALTANTE en caja").font = Font(bold=True, color="DC2626")
        ws_conciliation.cell(row=row, column=2).fill = NEGATIVE_FILL
        ws_conciliation.cell(row=row, column=3, value="Falta dinero. Revisar gastos no registrados o errores en cobro.")
    else:
        ws_conciliation.cell(row=row, column=2, value="CAJA CUADRADA").font = Font(bold=True, color="4F46E5")
        ws_conciliation.cell(row=row, column=2).fill = POSITIVE_FILL
        ws_conciliation.cell(row=row, column=3, value="El efectivo real coincide con el esperado. ¡Excelente!")

    _add_notes_section(ws_conciliation, row, [
        "Monto Inicial: Efectivo con el que se abrió caja cada día.",
        "Efectivo por Ventas: Ventas del período cobradas en efectivo.",
        "Efectivo por Cobros de Cuotas: Pagos de cuotas de créditos en efectivo.",
        "Efectivo Esperado = Monto Inicial + Efectivo Ventas + Efectivo Cobros.",
        "Diferencia = Monto Real de Cierre - Efectivo Esperado.",
        "Diferencia positiva (Sobrante): Puede indicar ingresos no registrados.",
        "Diferencia negativa (Faltante): Puede indicar gastos no registrados o errores.",
    ], columns=4)

    _auto_adjust_columns(ws_conciliation)

    # Guardar
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
